"""Import UPCs from 'UPC_CULTURE_KINGS.xlsx' into wms_upc_catalog.

Sheet layout (one row per UPC, already normalized — no pivot needed):
    STYLE | COLOR | SIZE | UPC

Mapping:
    upc                -> UPC column (digits-only, normalized uppercase)
    customer           -> "CULTURE KINGS"
    style              -> STYLE column (trimmed, uppercase)
    color              -> COLOR column (trimmed, uppercase)
    size               -> SIZE column (trimmed, uppercase)
    sku                -> "{style}-{color}-{size}" (matches Receiving auto-gen)
    brand, mfg, desc, country, fabric -> empty

Idempotent: a UPC already in wms_upc_catalog is skipped.

Run:
    python import_upc_culture_kings.py        # dry-run
    APPLY=1 python import_upc_culture_kings.py  # apply
"""
import asyncio
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from deps import db

XLSX = Path(__file__).parent / "UPC_CULTURE_KINGS.xlsx"
SHEET = "UPC CULTURE KINGS 2026"

CUSTOMER = "CULTURE KINGS"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm_upper(v) -> str:
    return norm(v).upper()


def norm_upc(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return re.sub(r"[^0-9A-Z]", "", s.upper())


def make_sku(style: str, color: str, size: str) -> str:
    parts = [
        re.sub(r"\s+", "-", style.strip().upper()),
        re.sub(r"\s+", "-", color.strip().upper())[:10] if color else "",
        size.strip().upper(),
    ]
    return "-".join(p for p in parts if p)


def gen_id() -> str:
    return f"upc_{secrets.token_hex(6)}"


def read_xlsx() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"[!] No se encontro {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"[!] Sheet '{SHEET}' no existe en {XLSX.name}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    col_idx = {h: i for i, h in enumerate(header) if h}
    style_col = col_idx.get("STYLE", 0)
    color_col = col_idx.get("COLOR", 1)
    size_col = col_idx.get("SIZE", 2)
    upc_col = col_idx.get("UPC", 3)

    out = []
    for row in rows[1:]:
        if row is None:
            continue
        upc = norm_upc(row[upc_col])
        if not upc:
            continue
        style = norm_upper(row[style_col])
        color = norm_upper(row[color_col])
        size = norm_upper(row[size_col])
        if not style:
            continue
        out.append({
            "upc": upc,
            "style": style,
            "color": color,
            "size": size,
        })
    return out


async def main():
    apply = os.environ.get("APPLY") == "1"
    print(f"Mode: {'APPLY' if apply else 'DRY-RUN'}")

    candidates = read_xlsx()
    print(f"[i] Candidatos leidos del XLSX: {len(candidates)}")

    existing_codes = set()
    if candidates:
        async for doc in db.wms_upc_catalog.find(
            {"upc": {"$in": [c["upc"] for c in candidates]}}, {"_id": 0, "upc": 1}
        ):
            existing_codes.add(doc["upc"])
    print(f"[i] Ya en wms_upc_catalog: {len(existing_codes)}")

    seen = set()
    duplicates_in_file = 0
    to_insert = []
    for c in candidates:
        if c["upc"] in seen:
            duplicates_in_file += 1
            continue
        seen.add(c["upc"])
        if c["upc"] in existing_codes:
            continue
        doc = {
            "catalog_id": gen_id(),
            "upc": c["upc"],
            "customer": CUSTOMER,
            "manufacturer": "",
            "brand": "",
            "style": c["style"],
            "color": c["color"],
            "size": c["size"],
            "description": "",
            "country_of_origin": "",
            "fabric_content": "",
            "sku": make_sku(c["style"], c["color"], c["size"]),
            "created_at": now_iso(),
            "created_by_name": "import_upc_culture_kings",
            "source_file": XLSX.name,
        }
        to_insert.append(doc)

    print(f"[i] Duplicados dentro del XLSX (ignorados): {duplicates_in_file}")
    print(f"[+] Por insertar: {len(to_insert)}")
    if to_insert:
        print("\nSample (primeros 5):")
        for d in to_insert[:5]:
            print(f"  upc={d['upc']}  style={d['style']!r}  color={d['color']!r}  size={d['size']}")
        print(f"\nSample (ultimos 3):")
        for d in to_insert[-3:]:
            print(f"  upc={d['upc']}  style={d['style']!r}  color={d['color']!r}  size={d['size']}")

    # Stats per style
    from collections import Counter
    style_counts = Counter(d["style"] for d in to_insert)
    print(f"\nUPCs per style:")
    for s, n in sorted(style_counts.items()):
        print(f"  {n:>4}  {s}")

    if not apply:
        print("\n[DRY-RUN] No se escribio nada. APPLY=1 para aplicar.")
        return

    if not to_insert:
        print("\n[OK] Nada por insertar.")
        return

    BATCH = 500
    inserted = 0
    for i in range(0, len(to_insert), BATCH):
        chunk = to_insert[i:i + BATCH]
        await db.wms_upc_catalog.insert_many(chunk, ordered=False)
        inserted += len(chunk)
        print(f"   ...{inserted}/{len(to_insert)}")
    print(f"\n[OK] Insertados {inserted} UPCs al catalogo.")


if __name__ == "__main__":
    asyncio.run(main())
