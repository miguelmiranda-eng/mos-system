"""Import UPCs from 'UPCs_Spektrum CLCA.xlsx' into wms_upc_catalog.

Sheet layout (one row per style+color combo):
    Row Labels | STYLE DESCRIPTION | Color | XS | S | M | L | XL | 2XL | 3XL

Each size cell holds a 12-digit UPC. We pivot to one document per UPC.

Mapping:
    upc                -> the cell value (digits-only, normalized uppercase)
    customer           -> "SPEKTRUM"
    brand              -> "CLCA"
    style              -> Row Labels cleaned (whitespace collapsed, trimmed)
    description        -> STYLE DESCRIPTION column ("SS", "HOODIE", "JOGGER", ...)
    color              -> Color column (uppercase, trimmed)
    size               -> XS / S / M / L / XL / 2XL / 3XL
    sku                -> "{style}-{color}-{size}" (matches Receiving auto-gen)
    country_of_origin  -> empty (not in source)
    fabric_content     -> empty (not in source)
    manufacturer       -> empty (not in source)

Idempotent: a UPC already in wms_upc_catalog is skipped. Re-running after a
partial apply just inserts the missing rows.

Run:
    python import_upc_spektrum_clca.py        # dry-run
    APPLY=1 python import_upc_spektrum_clca.py  # apply
"""
import asyncio
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from deps import db

XLSX = Path(__file__).parent / "UPCs_Spektrum CLCA.xlsx"
SHEET = "Per Style - Color (2)"

CUSTOMER = "SPEKTRUM"
BRAND = "CLCA"

SIZE_COLUMNS = ["XS", "S", "M", "L", "XL", "2XL", "3XL"]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def norm_upper(v) -> str:
    return norm(v).upper()


def norm_upc(v) -> str:
    """Convert a UPC cell value (could be int 845938030404 or str) to a clean
    digit-string. Drops anything non-alphanumeric — UPCs here are 12 digits."""
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"[^0-9A-Z]", "", s.upper())
    return s


def make_sku(style: str, color: str, size: str) -> str:
    parts = [
        re.sub(r"\s+", "-", style.strip().upper()),
        re.sub(r"\s+", "-", color.strip().upper())[:10] if color else "",
        size.strip().upper(),
    ]
    return "-".join(p for p in parts if p)


def gen_id() -> str:
    return f"upc_{secrets.token_hex(6)}"


def read_xlsx() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"[!] No se encontro {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"[!] Sheet '{SHEET}' no existe en {XLSX.name}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = rows[0]
    # Locate column indices defensively (in case col order changes)
    col_idx = {h: i for i, h in enumerate(header) if h}
    style_col = col_idx.get("Row Labels", 0)
    desc_col = col_idx.get("STYLE DESCRIPTION", 1)
    color_col = col_idx.get("Color", 2)
    size_cols = {sz: col_idx[sz] for sz in SIZE_COLUMNS if sz in col_idx}

    out = []
    for row in rows[1:]:
        if row is None:
            continue
        style_raw = row[style_col]
        if not style_raw:
            continue  # blank separator row
        style = norm_upper(style_raw)
        desc = norm_upper(row[desc_col]) if desc_col is not None and row[desc_col] else ""
        color = norm_upper(row[color_col]) if color_col is not None and row[color_col] else ""
        if not color:
            continue  # rows without color are summary rows (e.g. row 35 — just a style heading)
        for sz, ci in size_cols.items():
            cell = row[ci]
            upc = norm_upc(cell)
            if not upc:
                continue
            out.append({
                "upc": upc,
                "style": style,
                "description": desc,
                "color": color,
                "size": sz,
            })
    return out


async def main():
    apply = os.environ.get("APPLY") == "1"
    print(f"Mode: {'APPLY' if apply else 'DRY-RUN'}")

    candidates = read_xlsx()
    print(f"[i] Candidatos leidos del XLSX: {len(candidates)}")

    # Conflict detection
    existing_codes = set()
    if candidates:
        async for doc in db.wms_upc_catalog.find(
            {"upc": {"$in": [c["upc"] for c in candidates]}}, {"_id": 0, "upc": 1}
        ):
            existing_codes.add(doc["upc"])
    print(f"[i] Ya en wms_upc_catalog: {len(existing_codes)}")

    # Dedup within the file itself (xlsx had real duplicates as we saw in
    # inspection — same UPC appearing in multiple style/color rows)
    seen = set()
    duplicates_in_file = 0
    to_insert = []
    for c in candidates:
        if c["upc"] in seen:
            duplicates_in_file += 1
            continue
        seen.add(c["upc"])
        if c["upc"] in existing_codes:
            continue
        doc = {
            "catalog_id": gen_id(),
            "upc": c["upc"],
            "customer": CUSTOMER,
            "manufacturer": "",
            "brand": BRAND,
            "style": c["style"],
            "color": c["color"],
            "size": c["size"],
            "description": c["description"],
            "country_of_origin": "",
            "fabric_content": "",
            "sku": make_sku(c["style"], c["color"], c["size"]),
            "created_at": now_iso(),
            "created_by_name": "import_upc_spektrum_clca",
            "source_file": XLSX.name,
        }
        to_insert.append(doc)

    print(f"[i] Duplicados dentro del XLSX (ignorados): {duplicates_in_file}")
    print(f"[+] Por insertar: {len(to_insert)}")
    if to_insert:
        print("\nSample (primeros 5):")
        for d in to_insert[:5]:
            print(f"  upc={d['upc']}  style={d['style']!r}  color={d['color']!r}  size={d['size']}  desc={d['description']!r}")
        print(f"\nSample (ultimos 3):")
        for d in to_insert[-3:]:
            print(f"  upc={d['upc']}  style={d['style']!r}  color={d['color']!r}  size={d['size']}  desc={d['description']!r}")

    if not apply:
        print("\n[DRY-RUN] No se escribio nada. APPLY=1 para aplicar.")
        return

    if not to_insert:
        print("\n[OK] Nada por insertar.")
        return

    # Bulk insert in chunks
    BATCH = 500
    inserted = 0
    for i in range(0, len(to_insert), BATCH):
        chunk = to_insert[i:i + BATCH]
        await db.wms_upc_catalog.insert_many(chunk, ordered=False)
        inserted += len(chunk)
        print(f"   ...{inserted}/{len(to_insert)}")
    print(f"\n[OK] Insertados {inserted} UPCs al catalogo.")


if __name__ == "__main__":
    asyncio.run(main())
