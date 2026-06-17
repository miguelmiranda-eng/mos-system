"""Import the NARRO locations (NA01-A01 .. NA22-D48) from
Locaciones_NA01_NA22 (1).xlsx into wms_locations under tab="narro".

Positions 49..76 in the source Excel are NOT real warehouse slots — each
rack level (A/B/C/D) physically has 48 positions only. They are capped on
import (see MAX_POSITION below), so a re-run never recreates them.

Schema written here matches the rest of wms_locations:
  - location_id: "loc" + 12 hex chars
  - name: e.g. "NA01-A01"  (uppercase, unique)
  - zone: rack code ("NA01" .. "NA22")
  - type: "rack"
  - tab: "narro"   ← new field that the frontend Locations module uses to
                     split a third tab "NARRO" out of the existing
                     custom/system buckets
  - active: True
  - is_custom: False  (these are system-wide warehouse racks)
  - created_at: ISO timestamp

Idempotent: any name that already exists in wms_locations is skipped (with
a note in the dry-run summary). Re-running after a partial apply just
inserts the missing rows.

Run as:
    python import_narro_locations.py        # dry-run, no DB writes
    APPLY=1 python import_narro_locations.py  # actually insert
"""
import asyncio
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

# Each rack level (A/B/C/D) only has 48 physical positions (01..48). The source
# Excel still lists 01..76, so we cap on import — positions 49..76 don't exist in
# the warehouse and were removed from wms_locations. Names look like "NA01-A07".
MAX_POSITION = 48
_NAME_RE = re.compile(r"^(NA\d+)-([A-D])(\d+)$")

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

from deps import db  # noqa: E402

XLSX_PATH = ROOT / "Locaciones_NA01_NA22 (1).xlsx"
SHEET = "Locaciones NA"
TAB = "narro"
BATCH = 500


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def gen_id(prefix: str) -> str:
    return f"{prefix}{secrets.token_hex(6)}"


def read_xlsx() -> list[dict]:
    if not XLSX_PATH.exists():
        raise SystemExit(f"[!]No se encontró {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"[!]Sheet '{SHEET}' no existe en {XLSX_PATH.name}")
    ws = wb[SHEET]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        rack, _nivel, _pos, locacion, _ = row
        if not locacion:
            continue
        name = str(locacion).strip().upper()
        zone = str(rack).strip().upper() if rack else ""
        # Skip positions 49..76 — they aren't real warehouse slots.
        m = _NAME_RE.match(name)
        if m and int(m.group(3)) > MAX_POSITION:
            continue
        rows.append({"name": name, "zone": zone})
    return rows


async def main():
    apply = os.environ.get("APPLY") == "1"
    parsed = read_xlsx()
    print(f"[i]Filas leídas del XLSX: {len(parsed)}")
    racks = sorted({r['zone'] for r in parsed})
    print(f"[i]Racks: {racks[0]}..{racks[-1]} ({len(racks)} total)")

    # Skip names already present.
    existing_names = set()
    async for doc in db.wms_locations.find(
        {"name": {"$in": [r["name"] for r in parsed]}},
        {"_id": 0, "name": 1},
    ):
        existing_names.add(doc["name"])
    print(f"[i]Ya existían en Mongo: {len(existing_names)}")

    to_insert = []
    for r in parsed:
        if r["name"] in existing_names:
            continue
        to_insert.append({
            "location_id": gen_id("loc"),
            "name": r["name"],
            "zone": r["zone"],
            "type": "rack",
            "tab": TAB,
            "active": True,
            "is_custom": False,
            "created_at": now_iso(),
        })

    print(f"[+]Por insertar: {len(to_insert)}")
    if to_insert:
        print(f"   Primero: {to_insert[0]['name']} (zona {to_insert[0]['zone']})")
        print(f"   Último:  {to_insert[-1]['name']} (zona {to_insert[-1]['zone']})")

    if not apply:
        print("\n[DRY]DRY-RUN — no se escribió nada. Ejecuta con APPLY=1 para aplicar.")
        return

    if not to_insert:
        print("\n[OK]Nada por insertar.")
        return

    inserted = 0
    for i in range(0, len(to_insert), BATCH):
        chunk = to_insert[i:i + BATCH]
        await db.wms_locations.insert_many(chunk, ordered=False)
        inserted += len(chunk)
        print(f"   …{inserted}/{len(to_insert)}")
    print(f"\n[OK]Insertadas {inserted} locaciones bajo tab='{TAB}'.")


if __name__ == "__main__":
    asyncio.run(main())
