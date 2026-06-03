"""Import the 836 PALLET STORAGE locations (PS01-A01 .. PS22-A38) from
Pallet_Storage_Locations_PS01_PS22.xlsx into wms_locations under tab="pallet".

Same shape as import_narro_locations.py: one row per location code, rack
prefix becomes the zone, the row label becomes the name.

Run as:
    python import_pallet_locations.py        # dry-run, no DB writes
    APPLY=1 python import_pallet_locations.py  # actually insert
"""
import asyncio
import os
import re
import secrets
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

from deps import db  # noqa: E402

XLSX_PATH = ROOT / "Pallet_Storage_Locations_PS01_PS22.xlsx"
SHEET = "Pallet Storage Locations"
TAB = "pallet"
BATCH = 500


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def gen_id(prefix: str) -> str:
    return f"{prefix}{secrets.token_hex(6)}"


def read_xlsx() -> list[dict]:
    if not XLSX_PATH.exists():
        raise SystemExit(f"[!] No se encontro {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"[!] Sheet '{SHEET}' no existe en {XLSX_PATH.name}")
    ws = wb[SHEET]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header "Location"
        if not row or not row[0]:
            continue
        name = str(row[0]).strip().upper()
        # Rack prefix is everything before the first dash, e.g. PS01-A01 -> PS01.
        m = re.match(r"^([A-Z0-9]+)-", name)
        zone = m.group(1) if m else ""
        rows.append({"name": name, "zone": zone})
    return rows


async def main():
    apply = os.environ.get("APPLY") == "1"
    parsed = read_xlsx()
    print(f"[i]Filas leidas del XLSX: {len(parsed)}")
    racks = sorted({r["zone"] for r in parsed if r["zone"]})
    print(f"[i]Racks: {racks[0]}..{racks[-1]} ({len(racks)} total)")

    # Skip names already present.
    existing_names = set()
    async for doc in db.wms_locations.find(
        {"name": {"$in": [r["name"] for r in parsed]}},
        {"_id": 0, "name": 1},
    ):
        existing_names.add(doc["name"])
    print(f"[i]Ya existian en Mongo: {len(existing_names)}")

    to_insert = []
    for r in parsed:
        if r["name"] in existing_names:
            continue
        to_insert.append({
            "location_id": gen_id("loc"),
            "name": r["name"],
            "zone": r["zone"],
            "type": "rack",
            "tab": TAB,
            "active": True,
            "is_custom": False,
            "created_at": now_iso(),
        })

    print(f"[+]Por insertar: {len(to_insert)}")
    if to_insert:
        print(f"   Primero: {to_insert[0]['name']} (zona {to_insert[0]['zone']})")
        print(f"   Ultimo:  {to_insert[-1]['name']} (zona {to_insert[-1]['zone']})")

    if not apply:
        print("\n[DRY]DRY-RUN - no se escribio nada. Ejecuta con APPLY=1 para aplicar.")
        return

    if not to_insert:
        print("\n[OK]Nada por insertar.")
        return

    inserted = 0
    for i in range(0, len(to_insert), BATCH):
        chunk = to_insert[i:i + BATCH]
        await db.wms_locations.insert_many(chunk, ordered=False)
        inserted += len(chunk)
        print(f"   ...{inserted}/{len(to_insert)}")
    print(f"\n[OK]Insertadas {inserted} locaciones bajo tab='{TAB}'.")


if __name__ == "__main__":
    asyncio.run(main())
