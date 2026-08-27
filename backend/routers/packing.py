from fastapi import APIRouter, Request, HTTPException, Response
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from deps import require_auth
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
import io
import os
import json

router = APIRouter(prefix="/api/packing")
templates = Jinja2Templates(directory="templates/packing")


async def _parse_payload(request: Request):
    """Accept the packing payload either as a JSON body (fetch) or as a
    form-encoded 'data' field (the print/pallet forms submitted in a new tab).
    The urlencoded body is parsed manually so we don't depend on python-multipart."""
    content_type = request.headers.get('content-type', '')
    if 'application/json' in content_type:
        return await request.json()

    body = await request.body()
    text = body.decode('utf-8') if body else ''
    if 'application/x-www-form-urlencoded' in content_type or text.startswith('data='):
        from urllib.parse import parse_qs
        parsed = parse_qs(text, keep_blank_values=True)
        if 'data' in parsed and parsed['data']:
            return json.loads(parsed['data'][0])
    return json.loads(text or '{}')

def build_excel(data, base_dir="."):
    meta = data.get('meta', {})
    rows = data.get('rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # --- Column Widths (set once for whole sheet) ---
    col_widths = [
        14, 18, 10, 14, 14, 8, 8, 8, 12, 10, 8, 8, 10, 8, 10, 20, 12, 8, 20
    ]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # --- Styles ---
    header_fill = PatternFill(start_color='4A69A5', end_color='4A69A5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    
    label_fill = PatternFill(start_color='E0E8F7', end_color='E0E8F7', fill_type='solid')
    label_font = Font(bold=True, size=9)
    
    border_side = Side(style='thin', color='000000')
    all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    center_align = Alignment(horizontal='center', vertical='center', wrapText=True)
    left_align = Alignment(horizontal='left', vertical='center', indent=1, wrapText=True)
    right_align = Alignment(horizontal='right', vertical='center', wrapText=True)

    # --- Header Section ---
    ws.merge_cells('A1:S1')
    ws.row_dimensions[1].height = 65
    
    for col in range(1, 20):
        ws.cell(row=1, column=col).fill = header_fill

    header_img_path = os.path.join(base_dir, "static", "images", "header_hq.png")
    if os.path.exists(header_img_path):
        try:
            img = ExcelImage(header_img_path)
            img.height = 80 
            img.width = 80 * (img.width / img.height)
            ws.add_image(img, 'C1')
        except Exception as e:
            print(f"Header Image error: {e}")

    def set_border(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = all_border

    meta_rows = [
        ('VENDOR PO', meta.get('vendorPO'), 'Client', meta.get('client')),
        ('CLIENT PO', meta.get('clientPO'), 'Date Packed:', meta.get('datePacked')),
        ('STORE NAME', meta.get('storeName'), 'Packer Name:', meta.get('packerName'))
    ]

    for i, (l_lbl, l_val, r_lbl, r_val) in enumerate(meta_rows, 2):
        ws.row_dimensions[i].height = 20
        l_lbl_cell = ws.cell(row=i, column=1, value=l_lbl)
        l_lbl_cell.fill = label_fill
        l_lbl_cell.font = label_font
        l_lbl_cell.alignment = left_align
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=1)
        set_border(ws, f"A{i}:A{i}")

        l_val_cell = ws.cell(row=i, column=2, value=l_val)
        l_val_cell.alignment = left_align
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        set_border(ws, f"B{i}:D{i}")

        r_lbl_cell = ws.cell(row=i, column=14, value=r_lbl)
        r_lbl_cell.fill = label_fill
        r_lbl_cell.font = label_font
        r_lbl_cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws.merge_cells(start_row=i, start_column=14, end_row=i, end_column=16)
        set_border(ws, f"N{i}:P{i}")

        r_val_cell = ws.cell(row=i, column=17, value=r_val)
        r_val_cell.alignment = left_align
        ws.merge_cells(start_row=i, start_column=17, end_row=i, end_column=19)
        set_border(ws, f"Q{i}:S{i}")

    current_row = 6
    ws.row_dimensions[current_row].height = 30
    headers = [
        'STORE PO', 'Design Code', 'Garment Type', 'Garment Color',
        'XS or OSFA', 'SM', 'MD', 'LG', 'XL', '2XL', '3XL', '4XL', 'TOTAL',
        'Box Qty', 'Pcs per box', 'Box Dimension (LxWxH)', 'Box Weight (lbs)', 'Pallet #', 'Pallet Dimension (LxWxH)'
    ]
    for i, h in enumerate(headers):
        cell = ws.cell(row=current_row, column=i+1, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = all_border
    
    ws.freeze_panes = 'A7'

    current_row += 1
    for r_data in rows:
        for i, k in enumerate(['storePO', 'designCode', 'garmentType', 'garmentColor']):
            c = ws.cell(row=current_row, column=i+1, value=r_data.get(k, ''))
            c.border = all_border
            if i in [2, 3]: 
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                c.alignment = left_align
        
        sz_keys = ['xs', 'sm', 'md', 'lg', 'xl', 'xxl', 'xxxl', 'xxxxl']
        for i, k in enumerate(sz_keys):
            val = r_data.get(k)
            try: val = int(val) if val else ''
            except: pass
            c = ws.cell(row=current_row, column=5+i, value=val)
            c.border = all_border
            c.alignment = center_align
        
        r_total = r_data.get('total')
        try: r_total = int(r_total) if r_total else 0
        except: r_total = 0
        c_tot = ws.cell(row=current_row, column=13, value=r_total)
        c_tot.border = all_border
        c_tot.alignment = center_align
        c_tot.font = Font(bold=True)
        
        ship_keys = ['boxQty', 'pcsPerBox', 'boxDim', 'boxWeight', 'palletNo', 'palletDim']
        for i, k in enumerate(ship_keys):
            val = r_data.get(k, '')
            try:
                if val != "":
                    val = float(val) if '.' in str(val) else int(val)
            except:
                pass
            c = ws.cell(row=current_row, column=14+i, value=val)
            c.border = all_border
            c.alignment = center_align
        
        current_row += 1
    
    summaryGridData = data.get('summaryGridData', [])
    
    total_shipped = next((r for r in summaryGridData if r['id'] == 'totalShipped'), None)
    if total_shipped:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        lbl = ws.cell(row=current_row, column=1, value="TOTAL SHIPPED:")
        lbl.fill = label_fill
        lbl.font = Font(bold=True, size=10)
        lbl.alignment = right_align
        for c in range(1, 5): ws.cell(row=current_row, column=c).border = all_border
        
        sz_keys = ['xs', 'sm', 'md', 'lg', 'xl', 'xxl', 'xxxl', 'xxxxl', 'total']
        for i, k in enumerate(sz_keys):
            val = total_shipped.get(k)
            try: 
                if val is None or val == "" or val == 0: val = ""
                else: val = int(val)
            except: val = ""
            cell = ws.cell(row=current_row, column=5+i, value=val)
            cell.fill = label_fill
            cell.font = Font(bold=True, size=10)
            cell.alignment = center_align
            cell.border = all_border
            
        for c in range(14, 20):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = label_fill
            cell.font = Font(bold=True, size=10)
            cell.alignment = center_align
            cell.border = all_border
            if c == 14 and current_row > 7:
                cell.value = f"=SUM(N7:N{current_row-1})"
                
        current_row += 1
        
    current_row += 1
    
    sizes = ['XS or OSFA', 'SM', 'MD', 'LG', 'XL', '2XL', '3XL', '4XL', 'TOTAL']
    for i, sz in enumerate(sizes):
        c = ws.cell(row=current_row, column=5+i, value=sz)
        c.fill = header_fill
        c.font = Font(color='FFFFFF', bold=True, size=9)
        c.alignment = center_align
        c.border = all_border
    
    current_row += 1
    
    for s_row in summaryGridData:
        if s_row['id'] == 'totalShipped': continue
        
        is_blue_row = s_row['id'] in ['ejemplosLicense', 'orderedQty', 'difference', 'damagedMisprint', 'damagedGarment', 'extra']
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        lbl = ws.cell(row=current_row, column=1, value=s_row.get('label'))
        lbl.font = Font(bold=is_blue_row, size=9)
        lbl.alignment = right_align
        if is_blue_row: lbl.fill = label_fill
        
        for c in range(1, 5): 
            ws.cell(row=current_row, column=c).border = all_border
            if is_blue_row: ws.cell(row=current_row, column=c).fill = label_fill
        
        sz_keys = ['xs', 'sm', 'md', 'lg', 'xl', 'xxl', 'xxxl', 'xxxxl', 'total']
        for i, k in enumerate(sz_keys):
            val = s_row.get(k)
            try: 
                if val is None or val == "" or val == 0: val = ""
                else: val = int(val)
            except: val = ""
            
            c = ws.cell(row=current_row, column=5+i, value=val)
            c.border = all_border
            c.alignment = center_align
            if is_blue_row:
                c.font = Font(bold=True)
        
        if s_row['id'] == 'ejemplosLicense':
            ws.merge_cells(start_row=current_row, start_column=17, end_row=current_row, end_column=18)
            nb_lbl = ws.cell(row=current_row, column=17, value="NEW BOXES")
            nb_lbl.fill = label_fill
            nb_lbl.font = Font(bold=True, size=9)
            nb_lbl.alignment = right_align
            for c in range(17, 19): ws.cell(row=current_row, column=c).border = all_border
            nb_val = ws.cell(row=current_row, column=19, value=meta.get('newBoxes', ''))
            nb_val.alignment = center_align
            nb_val.border = all_border

        elif s_row['id'] == 'orderedQty':
            ws.merge_cells(start_row=current_row, start_column=17, end_row=current_row, end_column=18)
            rb_lbl = ws.cell(row=current_row, column=17, value="RECYCLED BOXES")
            rb_lbl.fill = label_fill
            rb_lbl.font = Font(bold=True, size=9)
            rb_lbl.alignment = right_align
            for c in range(17, 19): ws.cell(row=current_row, column=c).border = all_border
            rb_val = ws.cell(row=current_row, column=19, value=meta.get('recycledBoxes', ''))
            rb_val.alignment = center_align
            rb_val.border = all_border

        current_row += 1

    current_row += 2 

    note_data = next((r for r in summaryGridData if r['id'] == 'note'), None)
    if note_data:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value="NOTE:").fill = label_fill
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=13)
        note_val = ws.cell(row=current_row, column=5, value=meta.get('note', ''))
        note_val.fill = label_fill
        for c in range(1, 14): ws.cell(row=current_row, column=c).border = all_border
        current_row += 3

    sizeTablesData = data.get('sizeTablesData', [])
    if sizeTablesData:
        lbl_date = ws.cell(row=current_row, column=1, value="PICKING DATE:")
        lbl_date.font = Font(bold=True, size=10)
        val_date = ws.cell(row=current_row, column=3, value=meta.get('pickingDate', ''))
        
        lbl_name = ws.cell(row=current_row, column=7, value="PICKER NAME:")
        lbl_name.font = Font(bold=True, size=10)
        val_name = ws.cell(row=current_row, column=10, value=meta.get('pickerName', ''))
        
        current_row += 2
        start_row_breakdowns = current_row
        col_starts = [1, 7, 15]
        
        for idx, table_data in enumerate(sizeTablesData):
            row_group = idx // 3
            col_group = idx % 3
            c_start = col_starts[col_group]
            t_row = start_row_breakdowns + (row_group * 7)
            is_medium = (col_group == 1)
            is_wide   = (col_group == 2) 
            
            if is_wide:
                ws.merge_cells(start_row=t_row, start_column=15, end_row=t_row, end_column=19)
                title = ws.cell(row=t_row, column=15, value=f"SIZE {table_data.get('size', '')}")
                title.fill = header_fill
                title.font = Font(color='FFFFFF', bold=True, size=10)
                title.alignment = center_align
                for c in range(15, 20): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                wide_headers = [(15, 16, "Percentage"), (17, 18, "Country of Origin"), (19, 19, "Qty")]
                for cs, ce, h in wide_headers:
                    if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                    cell = ws.cell(row=t_row, column=cs, value=h)
                    cell.fill = label_fill
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = center_align
                    for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                for row_idx in range(3):
                    r_data = table_data.get('rows', [])[row_idx] if row_idx < len(table_data.get('rows', [])) else {}
                    vals = [r_data.get('perc', ''), r_data.get('orig', ''), r_data.get('qty', '')]
                    for (cs, ce, _), v in zip(wide_headers, vals):
                        if v == '-' or v is None: v = ''
                        if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                        cell = ws.cell(row=t_row, column=cs, value=v)
                        cell.alignment = center_align
                        cell.font = Font(size=9)
                        for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                    t_row += 1
                
                ws.merge_cells(start_row=t_row, start_column=15, end_row=t_row, end_column=18)
                total_lbl = ws.cell(row=t_row, column=15, value="Total:")
                total_lbl.fill = label_fill
                total_lbl.font = Font(bold=True, size=9)
                total_lbl.alignment = right_align
                for c in range(15, 19): ws.cell(row=t_row, column=c).fill = label_fill
                
                f_qty = table_data.get('total', 0)
                try: f_qty = int(f_qty) if f_qty else 0
                except: f_qty = 0
                total_val = ws.cell(row=t_row, column=19, value=f_qty)
                total_val.fill = label_fill
                total_val.font = Font(bold=True, size=10)
                total_val.alignment = center_align
                for c in range(15, 20): ws.cell(row=t_row, column=c).border = all_border

            elif is_medium:
                ws.merge_cells(start_row=t_row, start_column=7, end_row=t_row, end_column=13)
                title = ws.cell(row=t_row, column=7, value=f"SIZE {table_data.get('size', '')}")
                title.fill = header_fill
                title.font = Font(color='FFFFFF', bold=True, size=10)
                title.alignment = center_align
                for c in range(7, 14): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                medium_headers = [(7, 9, "Percentage"), (10, 12, "Country of Origin"), (13, 13, "Qty")]
                for cs, ce, h in medium_headers:
                    if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                    cell = ws.cell(row=t_row, column=cs, value=h)
                    cell.fill = label_fill
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = center_align
                    for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                for row_idx in range(3):
                    r_data = table_data.get('rows', [])[row_idx] if row_idx < len(table_data.get('rows', [])) else {}
                    vals = [r_data.get('perc', ''), r_data.get('orig', ''), r_data.get('qty', '')]
                    for (cs, ce, _), v in zip(medium_headers, vals):
                        if v == '-' or v is None: v = ''
                        if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                        cell = ws.cell(row=t_row, column=cs, value=v)
                        cell.alignment = center_align
                        cell.font = Font(size=9)
                        for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                    t_row += 1
                
                ws.merge_cells(start_row=t_row, start_column=7, end_row=t_row, end_column=12)
                total_lbl = ws.cell(row=t_row, column=7, value="Total:")
                total_lbl.fill = label_fill
                total_lbl.font = Font(bold=True, size=9)
                total_lbl.alignment = right_align
                for c in range(7, 13): ws.cell(row=t_row, column=c).fill = label_fill
                
                f_qty = table_data.get('total', 0)
                try: f_qty = int(f_qty) if f_qty else 0
                except: f_qty = 0
                total_val = ws.cell(row=t_row, column=13, value=f_qty)
                total_val.fill = label_fill
                total_val.font = Font(bold=True, size=10)
                total_val.alignment = center_align
                for c in range(7, 14): ws.cell(row=t_row, column=c).border = all_border

            else:
                ws.merge_cells(start_row=t_row, start_column=1, end_row=t_row, end_column=5)
                title = ws.cell(row=t_row, column=1, value=f"SIZE {table_data.get('size', '')}")
                title.fill = header_fill
                title.font = Font(color='FFFFFF', bold=True, size=10)
                title.alignment = center_align
                for c in range(1, 6): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                narrow_headers = [(1, 2, "Percentage"), (3, 4, "Country of Origin"), (5, 5, "Qty")]
                for cs, ce, h in narrow_headers:
                    if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                    cell = ws.cell(row=t_row, column=cs, value=h)
                    cell.fill = label_fill
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = center_align
                    for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                
                t_row += 1
                for row_idx in range(3):
                    r_data = table_data.get('rows', [])[row_idx] if row_idx < len(table_data.get('rows', [])) else {}
                    vals = [r_data.get('perc', ''), r_data.get('orig', ''), r_data.get('qty', '')]
                    for (cs, ce, _), v in zip(narrow_headers, vals):
                        if v == '-' or v is None: v = ''
                        if cs != ce: ws.merge_cells(start_row=t_row, start_column=cs, end_row=t_row, end_column=ce)
                        cell = ws.cell(row=t_row, column=cs, value=v)
                        cell.alignment = center_align
                        cell.font = Font(size=9)
                        for c in range(cs, ce+1): ws.cell(row=t_row, column=c).border = all_border
                    t_row += 1
                
                ws.merge_cells(start_row=t_row, start_column=1, end_row=t_row, end_column=4)
                total_lbl = ws.cell(row=t_row, column=1, value="Total:")
                total_lbl.fill = label_fill
                total_lbl.font = Font(bold=True, size=9)
                total_lbl.alignment = right_align
                for c in range(1, 5): ws.cell(row=t_row, column=c).fill = label_fill
                
                f_qty = table_data.get('total', 0)
                try: f_qty = int(f_qty) if f_qty else 0
                except: f_qty = 0
                total_val = ws.cell(row=t_row, column=5, value=f_qty)
                total_val.fill = label_fill
                total_val.font = Font(bold=True, size=10)
                total_val.alignment = center_align
                for c in range(1, 6): ws.cell(row=t_row, column=c).border = all_border
            
            if t_row + 2 > current_row:
                current_row = t_row + 2

    breakdowns = data.get('breakdowns', [])
    if breakdowns:
        current_row += 2
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
        imp_title = ws.cell(row=current_row, column=1, value="SUMMARY (IMPORT & EXPORT)")
        imp_title.fill = header_fill
        imp_title.font = Font(color='FFFFFF', bold=True, size=11)
        imp_title.alignment = center_align
        for c in range(1, 14): ws.cell(row=current_row, column=c).border = all_border
        
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        c1 = ws.cell(row=current_row, column=1, value="Percentage")
        c1.fill = label_fill; c1.font = Font(bold=True, size=9); c1.alignment = center_align
        for c in range(1, 3): ws.cell(row=current_row, column=c).border = all_border
        
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        c2 = ws.cell(row=current_row, column=3, value="Country of Origin")
        c2.fill = label_fill; c2.font = Font(bold=True, size=9); c2.alignment = center_align
        for c in range(3, 5): ws.cell(row=current_row, column=c).border = all_border
        
        default_sz_labels = ["XS or OSFA", "SM", "MD", "LG", "XL", "2XL", "3XL", "4XL"]
        sz_labels = data.get('sizeLabels') or default_sz_labels
        sz_headers = list(sz_labels) + ["TOTAL"]
        for i, h in enumerate(sz_headers):
            cell = ws.cell(row=current_row, column=5+i, value=h)
            cell.fill = label_fill; cell.font = Font(bold=True, size=9); cell.alignment = center_align
            cell.border = all_border
        
        current_row += 1
        for bd in breakdowns:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            c1 = ws.cell(row=current_row, column=1, value=bd.get('perc'))
            c1.alignment = center_align
            for c in range(1, 3): ws.cell(row=current_row, column=c).border = all_border
            
            ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
            c2 = ws.cell(row=current_row, column=3, value=bd.get('orig'))
            c2.alignment = center_align
            for c in range(3, 5): ws.cell(row=current_row, column=c).border = all_border
            
            sz_vals = [
                bd.get('xs'), bd.get('sm'), bd.get('md'), bd.get('lg'),
                bd.get('xl'), bd.get('xxl'), bd.get('xxxl'), bd.get('xxxxl'), bd.get('total')
            ]
            for i, v in enumerate(sz_vals):
                try:
                    if v and str(v).strip() != "":
                        v = float(v) if '.' in str(v) else int(v)
                except: pass
                cell = ws.cell(row=current_row, column=5+i, value=v)
                cell.border = all_border
                cell.alignment = center_align
            
            current_row += 1
            
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        lbl = ws.cell(row=current_row, column=1, value="Total Qty:")
        lbl.font = Font(bold=True)
        lbl.alignment = right_align
        for c in range(1, 13):
            ws.cell(row=current_row, column=c).border = all_border
        
        g_qty = data.get('grandTotalQty', 0)
        try: g_qty = int(g_qty) if g_qty else 0
        except: g_qty = 0
        
        val_cell = ws.cell(row=current_row, column=13, value=g_qty)
        val_cell.font = Font(bold=True)
        val_cell.alignment = center_align
        val_cell.border = all_border
        
        current_row += 1

    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 

    dynamic_title = meta.get('dynamicTitle')
    if dynamic_title:
        filename = f"{dynamic_title}.xlsx"
    else:
        filename = "Packing_List.xlsx"

    return wb, filename

@router.post("/export")
async def export_excel(request: Request):
    await require_auth(request)
    try:
        data = await _parse_payload(request)
        wb, filename = build_excel(data, base_dir=".")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        print(f"Export Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/preview", response_class=HTMLResponse)
async def print_preview(request: Request):
    await require_auth(request)
    try:
        data = await _parse_payload(request)
        return templates.TemplateResponse("print_preview.html", {"request": request, "data": data})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _pallet_printer_name(user):
    """Usuario que dispara la impresión, para estampar 'Imprimió' en la papeleta."""
    return (user or {}).get("name") or (user or {}).get("email") or (user or {}).get("user_id") or "—"


def _pallet_print_stamp():
    """dd/mm/yyyy HH:MM en hora del almacén (America/Tijuana)."""
    try:
        from datetime import datetime, timezone
        from zoneinfo import ZoneInfo
        return datetime.now(timezone.utc).astimezone(ZoneInfo("America/Tijuana")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        from datetime import datetime
        return datetime.now().strftime("%d/%m/%Y %H:%M")


@router.post("/pallet_label", response_class=HTMLResponse)
async def pallet_label(request: Request):
    user = await require_auth(request)
    try:
        data = await _parse_payload(request)
        pallets = {}
        rows = data.get('rows', [])
        meta = data.get('meta', {})
        size_tables = data.get('sizeTablesData', [])
        
        breakdown_map = {}
        for st in size_tables:
            size_name = st.get('size')
            breakdown_map[size_name] = []
            for row in st.get('rows', []):
                if int(row.get('qty', 0) or 0) > 0:
                    breakdown_map[size_name].append({
                        'perc': row.get('perc'),
                        'orig': row.get('orig')
                    })

        last_p_no = None
        for row in rows:
            p_no_raw = str(row.get('palletNo', '')).strip()
            if p_no_raw:
                last_p_no = p_no_raw
            
            p_no = last_p_no
            if not p_no: continue
            
            if p_no not in pallets:
                pallets[p_no] = {
                    'palletNo': p_no,
                    'totalPcs': 0,
                    'totalBoxes': 0,
                    'client': meta.get('client'),
                    'storeName': meta.get('storeName'),
                    'storePO': str(row.get('storePO', '')).strip(),
                    'clientPO': str(meta.get('clientPO', '')).strip(),
                    'style': (str(row.get('designCode', '')) or str(row.get('garmentType', ''))).strip(),
                    'color': str(row.get('garmentColor', '')).strip(),
                    'shippingDate': meta.get('datePacked'),
                    'compositions_set': set(),
                    'origins_set': set(),
                    'styles_set': set(),
                    'colors_set': set(),
                    'sizes': {
                        'xs': 0, 'sm': 0, 'md': 0, 'lg': 0, 'xl': 0, 
                        'xxl': 0, 'xxxl': 0, 'xxxxl': 0
                    }
                }
            
            p = pallets[p_no]
            p['totalPcs'] += int(row.get('total', 0) or 0)
            p['totalBoxes'] += int(row.get('boxQty', 0) or 0)
            
            row_s_po = str(row.get('storePO', '')).strip()
            if row_s_po: p['storePO'] = row_s_po
            
            row_style = (str(row.get('designCode', '')) or str(row.get('garmentType', ''))).strip()
            if row_style: 
                p['style'] = row_style
                p['styles_set'].add(row_style)
            
            row_color = str(row.get('garmentColor', '')).strip()
            if row_color: 
                p['color'] = row_color
                p['colors_set'].add(row_color)
            
            size_keys = ['xs', 'sm', 'md', 'lg', 'xl', 'xxl', 'xxxl', 'xxxxl']
            display_names = ['XS or OSFA', 'SM', 'MD', 'LG', 'XL', '2XL', '3XL', '4XL']
            
            for idx, sz in enumerate(size_keys):
                qty = int(row.get(sz, 0) or 0)
                if qty > 0:
                    p['sizes'][sz] += qty
                    d_name = display_names[idx]
                    if d_name in breakdown_map:
                        for b in breakdown_map[d_name]:
                            if b['perc'] and b['perc'] not in ['-', '', None]:
                                p['compositions_set'].add(b['perc'])
                            if b['orig'] and b['orig'] not in ['-', '', None]:
                                p['origins_set'].add(b['orig'])

        for p in pallets.values():
            p['compositions'] = " / ".join(sorted(p['compositions_set'])) if p['compositions_set'] else '-'
            p['origins'] = " / ".join(sorted(p['origins_set'])) if p['origins_set'] else '-'
            if p['styles_set']:
                p['style'] = " / ".join(sorted(p['styles_set']))
            if p['colors_set']:
                p['color'] = " / ".join(sorted(p['colors_set']))

        sorted_pallets = sorted(pallets.values(), key=lambda x: str(x['palletNo']))
        
        if sorted_pallets:
            master = sorted_pallets[0]
            master_client_po = master.get('clientPO') or meta.get('clientPO') or '-'
            master_store_po = master.get('storePO') or meta.get('vendorPO') or '-'
            master_style = master.get('style') or '-'
            master_color = master.get('color') or '-'

            for p in sorted_pallets:
                if p == master:
                    p['clientPO'] = master_client_po
                    p['storePO'] = master_store_po
                    p['style'] = master_style
                    p['color'] = master_color
                    continue

                raw_store_po = p.get('storePO', '').strip()
                if raw_store_po and raw_store_po != '-':
                    p['storePO'] = raw_store_po
                    p['style'] = p.get('style') if (p.get('style') and p['style'] != '-') else master_style
                    p['color'] = p.get('color') if (p.get('color') and p['color'] != '-') else master_color
                    p['clientPO'] = p.get('clientPO') if (p.get('clientPO') and p['clientPO'] != '-') else master_client_po
                else:
                    p['clientPO'] = master_client_po
                    p['storePO'] = master_store_po
                    p['style'] = master_style
                    p['color'] = master_color

        if not sorted_pallets:
            sorted_pallets = [{
                'palletNo': '',
                'totalPcs': 0,
                'totalBoxes': 0,
                'client': meta.get('client', ''),
                'storeName': meta.get('storeName', ''),
                'storePO': meta.get('vendorPO', '') or '-',
                'clientPO': meta.get('clientPO', '') or '-',
                'style': '-',
                'color': '-',
                'shippingDate': meta.get('datePacked', ''),
                'compositions': '',
                'origins': '',
                'sizes': {k: 0 for k in ['xs', 'sm', 'md', 'lg', 'xl', 'xxl', 'xxxl', 'xxxxl']},
                'current_pallet': 1,
                'total_pallets_count': 1
            }]
        else:
            total_pallets = len(sorted_pallets)
            for idx, p in enumerate(sorted_pallets):
                p['current_pallet'] = idx + 1
                p['total_pallets_count'] = total_pallets

        return templates.TemplateResponse("pallet_label.html", {
            "request": request, "pallets": sorted_pallets, "meta": meta,
            "printed_by": _pallet_printer_name(user), "printed_at": _pallet_print_stamp(),
        })
    except Exception as e:
        print(f"Pallet Label Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
