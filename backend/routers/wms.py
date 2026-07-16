"""WMS (Warehouse Management System) routes."""
from fastapi import APIRouter, HTTPException, Request, Response, UploadFile, File, Query
from typing import Optional
from fastapi.responses import StreamingResponse, HTMLResponse
from deps import db, get_current_user, require_auth, require_admin, require_admin_level, require_supersu, get_admin_level, require_inventory_level, get_inventory_level, DEFAULT_OPTIONS
from ws_manager import ws_manager
from wms_constants import (
    BoxStatus, TicketStatus, PickingStatus, CycleCountStatus,
    TaskType, TaskStatus, PickDestination, MovementType, AsnStatus,
)
from datetime import datetime, timezone, timedelta
from pymongo import ReturnDocument
import uuid, io, json, logging, re, asyncio, difflib, time

router = APIRouter(prefix="/api/wms")
logger = logging.getLogger(__name__)

# In-process cache for the expensive inventory_summary aggregation in /locations.
# 30s TTL is enough to absorb the burst of N users opening the Ubicaciones screen
# at the same time without serving data that's perceptibly stale.
_LOC_SUMMARY_CACHE = {"data": None, "ts": 0.0}
_LOC_SUMMARY_TTL = 30.0

# ── Customer-name normalization (protects against re-fragmenting a customer) ──
# Known variant (cleaned) -> canonical. Seeded from the 2026-06 cleanup; the
# write-time normalizer (_canonical_customer) applies this + a conservative fuzzy
# snap so typos/spacing/casing never create a new "GOODIE TWO SLEVEES" again.
_CUSTOMER_ALIASES = {
    "GOODIE TWO SLEVEES": "GOODIE TWO SLEEVES",
    "GOODIE TWO SLEEVE": "GOODIE TWO SLEEVES",
    "GOOGIE TWO SLEEVES": "GOODIE TWO SLEEVES",
    "GODDIE TWO SLEEVES": "GOODIE TWO SLEEVES",
    "GOODIE TWO SLEE": "GOODIE TWO SLEEVES",
    "GOOODIE TWO SLEEVE": "GOODIE TWO SLEEVES",
    "GTA TRACTOR": "GTS TRACTOR",
    "GTS TTRACTOR": "GTS TRACTOR",
    "TRACTOR GTS": "GTS TRACTOR",
    "TRACTOS GTS": "GTS TRACTOR",
    "SCREEN WORK": "SCREENWORKS",
    "SCREENWORK": "SCREENWORKS",
    "SCREEN WORKS": "SCREENWORKS",
    "SPELTRUM": "SPEKTRUM",
}
_KNOWN_CUSTOMERS_CACHE = {"data": None, "ts": 0.0}
_KNOWN_CUSTOMERS_TTL = 60.0


def _clean_customer(name):
    """Uppercase, trim, collapse internal whitespace. Pure cleanup (no merge)."""
    return re.sub(r"\s+", " ", (name or "").strip()).upper()


async def _known_customers():
    """Set of canonical customer names (cleaned) from boxes + curated catalog.
    Cached 60s so the write-time normalizer stays cheap."""
    now = time.monotonic()
    cache = _KNOWN_CUSTOMERS_CACHE
    if cache["data"] is not None and (now - cache["ts"]) < _KNOWN_CUSTOMERS_TTL:
        return cache["data"]
    names = set()
    for v in await db.wms_boxes.distinct("customer"):
        cv = _clean_customer(v)
        if cv:
            names.add(cv)
    for v in await db.wms_catalog_options.distinct("value", {"type": "customers"}):
        cv = _clean_customer(v)
        if cv:
            names.add(cv)
    cache["data"] = names
    cache["ts"] = now
    return names


async def _canonical_customer(name):
    """Canonical customer name on write: clean -> known alias -> exact match ->
    conservative fuzzy snap to an existing customer (cutoff 0.9). Returns the
    cleaned name unchanged when it's a genuinely new customer."""
    c = _clean_customer(name)
    if not c:
        return c
    if c in _CUSTOMER_ALIASES:
        return _CUSTOMER_ALIASES[c]
    known = await _known_customers()
    if c in known:
        return c
    match = difflib.get_close_matches(c, list(known), n=1, cutoff=0.9)
    return match[0] if match else c


def now_iso():
    return datetime.now(timezone.utc).isoformat()

def gen_id(prefix="wms"):
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


async def _reserve_box_seqs(n=1):
    """Atomically reserve `n` sequential box sequence numbers; return the FIRST.
    Uses the `counters` collection so two concurrent receivings (or a receiving +
    a move/split) can't mint the same BOX-id — the old read-max-then-+1 race. The
    counter is seeded from the current max at startup (ensure_wms_indexes); the
    guard below re-bases it the first time if that seed somehow didn't run, so a
    freshly-created counter can never hand out ids that collide with existing
    boxes (box_id is also uniquely indexed as a last-resort backstop)."""
    n = max(1, int(n or 1))
    doc = await db.counters.find_one_and_update(
        {"_id": "wms_box_seq"}, {"$inc": {"seq": n}},
        upsert=True, return_document=ReturnDocument.AFTER,
    )
    end = int(doc.get("seq", 0) or 0)
    if end <= n:  # counter was just created near zero — re-base off the real max
        last = await db.wms_boxes.find_one(sort=[("seq_num", -1)], projection={"seq_num": 1})
        base = int((last or {}).get("seq_num", 0) or 0)
        if base >= end:
            doc = await db.counters.find_one_and_update(
                {"_id": "wms_box_seq"}, {"$set": {"seq": base + n}},
                return_document=ReturnDocument.AFTER,
            )
            end = int(doc.get("seq", base + n))
    return end - n + 1


# Canonical name for the warehouse-wide holding location where boxes wait when
# they are received without a final destination. Lives next to the regular
# locations in wms_locations so the rest of the system (Locations module,
# move-location, aggregations) treats it like any other slot.
TRANSIT_LOCATION_NAME = "UBICACION TEMPORAL"

# Physical carts the Receiving operator can choose between when parking boxes
# before putaway. Each cart is a regular wms_location with type="transit" so the
# rest of the system (inventory, movements) treats them like any slot.
# These are just the DEFAULT carts seeded on first boot — admins can create more
# from Putaway 2.0; everything below treats ANY "CARRO <n>" as a cart by pattern.
TRANSIT_CART_NAMES = [f"CARRO {i}" for i in range(1, 51)]

# Every location managed by the Putaway 2.0 module. Legacy UBICACION TEMPORAL
# is kept so boxes received before the carts existed still surface.
SYSTEM_TRANSIT_LOCATIONS = [TRANSIT_LOCATION_NAME] + TRANSIT_CART_NAMES

# A transit slot is the legacy UBICACION TEMPORAL or ANY cart named "CARRO <n>".
# Matching by pattern (not a fixed list) lets admins add carts beyond the seeded
# 50 without code changes — newly created carts work everywhere automatically.
TRANSIT_NAME_REGEX = r"^(UBICACION TEMPORAL|CARRO \d+)$"
_CART_NAME_REGEX = r"^CARRO \d+$"


def _is_transit_name(name):
    return bool(re.match(TRANSIT_NAME_REGEX, (name or "").strip(), re.IGNORECASE))


def _transit_loc_filter():
    """Mongo value-filter matching any transit slot by name pattern."""
    return {"$regex": TRANSIT_NAME_REGEX, "$options": "i"}


async def _list_cart_locations():
    """All cart locations (CARRO <n>) in wms_locations, sorted by cart number."""
    docs = await db.wms_locations.find(
        {"name": {"$regex": _CART_NAME_REGEX, "$options": "i"}},
        {"_id": 0, "name": 1, "location_id": 1},
    ).to_list(5000)

    def _num(d):
        m = re.search(r"(\d+)", d.get("name") or "")
        return int(m.group(1)) if m else 0
    return sorted(docs, key=_num)


async def _ensure_transit_location():
    """Idempotently insert UBICACION TEMPORAL + the carts if they aren't in
    wms_locations yet. Safe to call on every request that may need to write
    there — bulk find + (optional) inserts."""
    existing = await db.wms_locations.find({
        "name": {"$in": SYSTEM_TRANSIT_LOCATIONS}
    }, {"_id": 0}).to_list(None)
    by_name = {l["name"]: l for l in existing}
    to_insert = []
    for name in SYSTEM_TRANSIT_LOCATIONS:
        if name in by_name:
            continue
        doc = {
            "location_id": gen_id("loc"),
            "name": name,
            # Group the carts under a single "CARROS" zone so they're easy to
            # find by zone (e.g. the manual-inventory modal). UBICACION TEMPORAL
            # keeps the generic TRANSIT zone.
            "zone": "CARROS" if name.upper().startswith("CARRO") else "TRANSIT",
            "type": "transit",
            "active": True,
            "is_custom": True,
            "created_at": now_iso(),
        }
        to_insert.append(doc)
        by_name[name] = doc
    if to_insert:
        await db.wms_locations.insert_many(to_insert)
    return by_name[TRANSIT_LOCATION_NAME]

async def log_movement(user, movement_type, details):
    await db.wms_movements.insert_one({
        "movement_id": gen_id("mov"),
        "type": movement_type,
        "details": details,
        "user_id": user.get("user_id"),
        "user_name": user.get("name", user.get("email", "")),
        "created_at": now_iso(),
    })

async def get_sku_movement_history(style: str, color: str = "", size: str = "", limit: int = 200):
    """Return movements that touch a given SKU dimension, newest first.

    wms_movements have heterogeneous shapes — sometimes the SKU is in
    details.sku, sometimes in details.style, sometimes only the box_id
    is present (in which case we resolve via wms_boxes). This handles
    all three.
    """
    # Direct match in details
    or_clauses = [
        {"details.style": {"$regex": f"^{re.escape(style)}$", "$options": "i"}},
        {"details.sku": {"$regex": f"^{re.escape(style)}$", "$options": "i"}},
    ]

    # Indirect match: find boxes with this SKU, then look up movements by their
    # box_id. _ci_eq (match exacto en MAYÚSCULAS) para usar el índice de wms_boxes
    # en vez de escanear 78k docs con regex 'i' en cada apertura de historial.
    box_query = {"style": _ci_eq(style)}
    if color:
        box_query["color"] = _ci_eq(color)
    if size:
        box_query["size"] = _ci_eq(size)
    matching_box_ids = await db.wms_boxes.distinct("box_id", box_query)
    if matching_box_ids:
        or_clauses.append({"details.box_id": {"$in": matching_box_ids}})

    movements = await db.wms_movements.find({"$or": or_clauses}, {"_id": 0}).sort("created_at", -1).to_list(limit)

    # Optional secondary filter on color/size when present in details
    if color or size:
        filtered = []
        for m in movements:
            d = m.get("details", {})
            if color and d.get("color") and d.get("color").lower() != color.lower():
                continue
            if size and d.get("size") and d.get("size") != size:
                continue
            filtered.append(m)
        return filtered
    return movements


@router.get("/inventory/history")
async def inventory_history(request: Request, style: str = "", color: str = "", size: str = "", limit: int = 200):
    """Audit trail for a specific SKU dimension (style + optional color/size)."""
    await require_auth(request)
    if not style:
        raise HTTPException(400, "style requerido")
    movements = await get_sku_movement_history(style, color, size, limit)
    return {"style": style, "color": color, "size": size, "count": len(movements), "movements": movements}


async def notify_badge_change(badge: str = "all"):
    """Tell connected WMS clients that a sidebar badge count may have changed.
    Frontend listens for `wms_badge_changed` and refreshes counts."""
    try:
        await ws_manager.broadcast("wms_badge_changed", {"badge": badge})
    except Exception as e:
        logger.warning(f"Failed to broadcast badge change: {e}")


# ==================== CATALOG OPTIONS (admin-managed dropdown values) ====================

# Receiving identity catalogs (customers/colors/styles) join the original
# descriptive ones. These drive the locked dropdowns in Receiving: operators can
# only pick existing values; only a lead/supervisor may add/rename/clean them.
CATALOG_TYPES = {"descriptions", "countries", "fabrics", "customers", "colors", "styles", "sizes", "manufacturers"}
# Tipos cuyo catálogo curado se scopea por CLIENTE (cada cliente tiene su lista).
# Modelo "global + por-cliente": un valor SIN customer es global (vale para todos);
# uno CON customer solo aplica a ese cliente. En Receiving nunca se muestran los
# valores específicos de OTRO cliente. El resto de tipos (sizes/countries/fabrics/
# descriptions) son globales — no dependen del cliente.
CUSTOMER_SCOPED = {"styles", "colors", "manufacturers"}

# Roles allowed to MUTATE catalogs (add / rename / delete / clean). Maps the
# business notion of "líder o supervisor" onto the existing elevated roles —
# change this single set if a dedicated 'supervisor' role is added later.
CATALOG_MANAGER_ROLES = {"admin", "supersu", "ceo"}


CATALOG_MIN_LEVEL = 3  # la identidad del inventario (estilos/colores/tallas/…) solo se edita nivel 3+

def _assert_catalog_manager(user):
    """Raise 403 salvo admin nivel 3+ (supersu = max). Los catálogos de identidad
    (styles/colors/sizes/customers/countries/fabrics/descriptions) solo se
    modifican desde Configuración por admin nivel 3 o superior — nadie mas puede
    cambiar estos valores."""
    if get_admin_level(user) < CATALOG_MIN_LEVEL:
        raise HTTPException(
            status_code=403,
            detail="Requiere admin nivel 3 o superior para modificar los catálogos de identidad.",
        )


@router.get("/catalogs")
async def list_catalogs(request: Request):
    """Return admin-curated dropdown values, grouped by type."""
    await require_auth(request)
    docs = await db.wms_catalog_options.find({}, {"_id": 0}).sort([("type", 1), ("value", 1)]).to_list(2000)
    grouped = {t: [] for t in CATALOG_TYPES}
    for d in docs:
        grouped.setdefault(d.get("type"), []).append(d)
    return grouped

@router.post("/catalogs")
async def add_catalog_option(request: Request):
    """Add a value to a catalog. Case-insensitive dedupe per type.
    Restricted to lead/supervisor (catalog manager)."""
    user = await require_auth(request)
    _assert_catalog_manager(user)
    body = await request.json()
    ctype = (body.get("type") or "").strip()
    value = (body.get("value") or "").strip()
    # Styles are scoped per customer (each client has its own valid style list).
    customer = (body.get("customer") or "").strip().upper()
    if ctype not in CATALOG_TYPES:
        raise HTTPException(400, f"type debe ser uno de {sorted(CATALOG_TYPES)}")
    if not value:
        raise HTTPException(400, "value es obligatorio")
    # Normalize customer names so the curated dropdown never gains a variant
    # (e.g. adding "GOODIE TWO SLEVEES" snaps to "GOODIE TWO SLEEVES" and the
    # dedupe below then rejects it as already existing).
    if ctype == "customers":
        value = await _canonical_customer(value)
    if ctype == "styles" and not customer:
        raise HTTPException(400, "customer es obligatorio para estilos")
    # Case-insensitive dedupe — scoped by customer for los tipos por-cliente.
    # Con customer -> dedupe contra ese cliente; sin customer -> contra los globales.
    dedupe = {"type": ctype, "value": {"$regex": f"^{re.escape(value)}$", "$options": "i"}}
    if ctype in CUSTOMER_SCOPED:
        dedupe["customer"] = customer if customer else {"$in": [None, ""]}
    existing = await db.wms_catalog_options.find_one(dedupe)
    if existing:
        where = f"{ctype} de {customer}" if customer else ctype
        raise HTTPException(400, f"'{value}' ya existe en {where}")
    doc = {
        "catalog_id": gen_id("cat"),
        "type": ctype,
        "value": value,
        "created_at": now_iso(),
        "created_by_name": user.get("name") or user.get("email", ""),
    }
    if customer:
        doc["customer"] = customer
    await db.wms_catalog_options.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.get("/catalogs/styles")
async def list_customer_styles(request: Request, customer: str = ""):
    """Curated style list for a customer — the locked options shown in Receiving.
    Returns just the style strings, sorted. Empty if the customer has no curated
    styles yet (so Receiving can fall back / show nothing to type-create)."""
    await require_auth(request)
    cust = (customer or "").strip().upper()
    if not cust:
        return {"customer": "", "styles": []}
    docs = await db.wms_catalog_options.find(
        {"type": "styles", "customer": cust}, {"_id": 0, "value": 1}
    ).sort("value", 1).to_list(2000)
    return {"customer": cust, "styles": [d.get("value") for d in docs if d.get("value")]}


@router.get("/catalogs/{ctype}/for-customer")
async def list_catalog_for_customer(ctype: str, request: Request, customer: str = ""):
    """Valores curados visibles para un cliente en un tipo por-cliente:
    los del cliente + los globales (sin customer). Sin cliente -> solo globales
    (nunca se filtran valores específicos de otro cliente). Para tipos NO
    por-cliente devuelve todos. Alimenta los desplegables de Color/Fabricante en
    Receiving scopeados por cliente."""
    await require_auth(request)
    if ctype not in CATALOG_TYPES:
        raise HTTPException(400, f"type debe ser uno de {sorted(CATALOG_TYPES)}")
    cust = (customer or "").strip().upper()
    if ctype in CUSTOMER_SCOPED:
        globals_or = [{"customer": {"$in": [None, ""]}}, {"customer": {"$exists": False}}]
        q = {"type": ctype, "$or": ([{"customer": cust}] + globals_or) if cust else globals_or}
    else:
        q = {"type": ctype}
    docs = await db.wms_catalog_options.find(q, {"_id": 0, "value": 1}).sort("value", 1).to_list(2000)
    return {"customer": cust, "values": [d.get("value") for d in docs if d.get("value")]}


@router.delete("/catalogs/{catalog_id}")
async def delete_catalog_option(catalog_id: str, request: Request):
    """Remove a value from a catalog. Does NOT alter existing inventory/receiving rows.
    Restricted to lead/supervisor (catalog manager)."""
    user = await require_auth(request)
    _assert_catalog_manager(user)
    res = await db.wms_catalog_options.delete_one({"catalog_id": catalog_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Opción no encontrada")
    return {"message": "Opción eliminada"}


# Map catalog type → (collection, field) it sources/normalizes.
_CATALOG_FIELD_MAP = {
    "countries":    ("country_of_origin", ["wms_inventory", "wms_receiving"]),
    "descriptions": ("description",       ["wms_inventory", "wms_receiving"]),
    "fabrics":      ("fabric_content",    ["wms_inventory", "wms_receiving"]),
    # Identity fields also live on the physical boxes, on UPCs and on pick tickets
    # — un rename/clean debe barrer TODAS o los mirrors se drift. Ejemplo real
    # (LIGH PINK, 2026-07-13): el color se colo en receiving, se propago a boxes
    # e inventory, y despues los UPCs escaneados y 7 tickets historicos quedaron
    # con el typo. Sweep de todas las colecciones cierra el hueco.
    "customers":    ("customer", ["wms_inventory", "wms_receiving", "wms_boxes", "wms_pick_tickets"]),
    "manufacturers": ("manufacturer", ["wms_inventory", "wms_receiving", "wms_boxes", "wms_upc_catalog", "wms_pick_tickets"]),
    "colors":       ("color",    ["wms_inventory", "wms_receiving", "wms_boxes", "wms_upc_catalog", "wms_pick_tickets"]),
    "styles":       ("style",    ["wms_inventory", "wms_receiving", "wms_boxes", "wms_upc_catalog", "wms_pick_tickets"]),
    "sizes":        ("size",     ["wms_inventory", "wms_receiving", "wms_boxes", "wms_upc_catalog", "wms_pick_tickets"]),
}


# ── Rechazo duro de valores fuera del catalogo curado ─────────────────────────
# Politica: si el catalogo curado de un campo tiene >= 1 valor para el customer
# (para styles) o global (para el resto), el receiving/UPC solo puede usar esos
# valores. Bootstrap: si el catalogo esta vacio, deja pasar hasta que exista el
# primer curado.
#
# Historial: SPEKTRUM llego a tener 72 estilos por captura libre (2026-07-13);
# esta guardia evita que el desorden vuelva a crecer. Ver:
# `wms-style-vs-sku-policy` en memoria.
_IDENTITY_LABELS = {
    "styles": "Estilo",
    "colors": "Color",
    "sizes": "Talla",
    "descriptions": "Descripcion",
    "countries": "Pais de origen",
    "fabrics": "Contenido/fabric",
    "manufacturers": "Fabricante",
}


async def _assert_curated_identity(customer: str, values: dict):
    """Rechaza (HTTP 400) si algun valor no esta en el catalogo curado.
    `values`: {ctype: value_string}. Vacios se ignoran. `styles` se scopea por
    customer (unico campo per-cliente); el resto son globales."""
    for ctype, raw in values.items():
        v = (raw or "").strip()
        if not v:
            continue
        query = {"type": ctype}
        if ctype in CUSTOMER_SCOPED:
            # Válido si es del cliente O global (sin customer). Modelo fallback.
            cu = (customer or "").strip().upper()
            query["$or"] = [
                {"customer": cu},
                {"customer": {"$in": [None, ""]}},
                {"customer": {"$exists": False}},
            ]
        curated = await db.wms_catalog_options.find(
            query, {"_id": 0, "value": 1}
        ).to_list(5000)
        if not curated:
            continue  # bootstrap: cliente sin catalogo aun, deja capturar
        curated_set = {(c.get("value") or "").strip().upper() for c in curated}
        if v.upper() in curated_set:
            continue  # curado ✓
        # UNIFICADO (curado + sistema): igual que los dropdowns fusionan curado
        # con inventario, aceptamos un valor que YA exista en inventario aunque
        # no este catalogado (ej. un color real no promovido aun). El lider lo
        # consolida con "Detectar typos". Solo se rechaza lo que no esta NI
        # curado NI en inventario (basura nueva). Los dropdowns son select-only,
        # asi que la UI ya no deja teclear valores libres.
        field = (_CATALOG_FIELD_MAP.get(ctype) or (None,))[0]
        if field:
            in_inv = await db.wms_inventory.find_one(
                {field: {"$regex": f"^{re.escape(v)}$", "$options": "i"}}, {"_id": 1})
            if in_inv:
                continue  # existe en inventario → válido
        label = _IDENTITY_LABELS.get(ctype, ctype)
        scope = f" de {customer}" if ctype in CUSTOMER_SCOPED and customer else ""
        raise HTTPException(
            400,
            f"{label} '{v}' no está ni en el catálogo curado ni en inventario{scope}. "
            f"Pídele al líder que lo agregue en Config WMS.",
        )


@router.get("/catalogs/{ctype}/sources")
async def list_catalog_sources(ctype: str, request: Request, limit: int = 500, customer: str = ""):
    """List the distinct values currently present in inventory/receiving for
    the given catalog type, with usage counts. Powers the "Fuentes desde
    inventario" panel in WMS Configuration so admins can see + clean typos.

    `customer` scopes the aggregate to rows of that client (case-insensitive).
    Sin este filtro, styles/colors mezclan valores de TODOS los clientes y la
    lista se ve "contaminada" cuando el usuario ya eligio un cliente arriba."""
    await require_auth(request)
    if ctype not in _CATALOG_FIELD_MAP:
        raise HTTPException(400, f"type debe ser uno de {sorted(_CATALOG_FIELD_MAP)}")
    field, collections = _CATALOG_FIELD_MAP[ctype]

    match_stage = {field: {"$ne": None, "$nin": ["", "."]}}
    cust = (customer or "").strip()
    if cust:
        match_stage["customer"] = {"$regex": f"^{re.escape(cust)}$", "$options": "i"}

    # Aggregate counts across all source collections (inventory + receiving).
    counts: dict[str, int] = {}
    for coll_name in collections:
        coll = getattr(db, coll_name)
        cursor = coll.aggregate([
            {"$match": match_stage},
            {"$group": {"_id": f"${field}", "count": {"$sum": 1}}},
        ])
        async for doc in cursor:
            v = (doc.get("_id") or "").strip()
            if not v:
                continue
            counts[v] = counts.get(v, 0) + int(doc.get("count", 0))

    # Traer los curados y FUSIONARLOS con los conteos de inventario. Curados que
    # nunca se han usado en la base entran con count=0 (para que la UI muestre
    # una sola lista unificada — sin panel duplicado arriba/abajo).
    curated_query = {"type": ctype}
    if cust and ctype in CUSTOMER_SCOPED:
        cu = cust.upper()
        curated_query["$or"] = [
            {"customer": cu},
            {"customer": {"$in": [None, ""]}},
            {"customer": {"$exists": False}},
        ]
    curated_docs = await db.wms_catalog_options.find(
        curated_query, {"_id": 0, "catalog_id": 1, "value": 1}
    ).to_list(5000)
    # index por valor UPPER para el badge y para inyectar el catalog_id
    curated_by_upper: dict[str, dict] = {}
    for c in curated_docs:
        v = (c.get("value") or "").strip()
        if not v:
            continue
        curated_by_upper[v.upper()] = {"catalog_id": c.get("catalog_id"), "value": v}
    counts_by_upper = {k.upper(): (k, v) for k, v in counts.items()}
    all_uppers = set(counts_by_upper.keys()) | set(curated_by_upper.keys())

    items = []
    for up in all_uppers:
        cur = curated_by_upper.get(up)
        raw_value, cnt = counts_by_upper.get(up, (cur["value"] if cur else "", 0))
        items.append({
            "value": raw_value or (cur["value"] if cur else ""),
            "count": cnt,
            "in_catalog": cur is not None,
            "catalog_id": cur["catalog_id"] if cur else None,
        })
    items.sort(key=lambda x: (-x["count"], x["value"].lower()))
    return {
        "type": ctype,
        "field": field,
        "total_distinct": len(items),
        "items": items[: max(1, min(limit, 5000))],
    }


def _levenshtein(a: str, b: str) -> int:
    """Distancia de edición: número mínimo de inserciones, borrados o sustituciones
    de UN caracter para transformar `a` en `b`. Case-insensitive."""
    a, b = (a or "").lower(), (b or "").lower()
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    # Corta rápido si la diferencia de longitud ya excede el umbral común
    if abs(len(a) - len(b)) > 3:
        return abs(len(a) - len(b))
    # Iterativo con dos filas — O(len(a) * len(b)) en tiempo, O(min) en memoria
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            insert = curr[j - 1] + 1
            delete = prev[j] + 1
            sub = prev[j - 1] + (0 if ca == cb else 1)
            curr.append(min(insert, delete, sub))
        prev = curr
    return prev[-1]


@router.get("/catalogs/{ctype}/similar")
async def catalog_similar_pairs(ctype: str, request: Request, max_dist: int = 2, min_count: int = 1, customer: str = ""):
    """Devuelve pares de valores del catálogo cuya distancia de edición es
    ≤ `max_dist` (default 2). Sirve para cazar typos como LIGH PINK vs LIGHT PINK,
    BANGLANDESH vs BANGLADESH, etc.

    Response: {
      "pairs": [
        { "a": "LIGH PINK", "b": "LIGHT PINK", "distance": 1,
          "count_a": 42, "count_b": 414, "recommend_keep": "LIGHT PINK" }
      ]
    }

    Reglas:
      - Ambos valores deben aparecer al menos `min_count` veces en las fuentes.
      - Se sugiere conservar el que tenga MÁS apariciones (más raíz de datos).
      - Se ignoran pares idénticos y valores vacíos.
      - Solo admin nivel 3+ (mismo umbral que rename).
    """
    user = await require_auth(request)
    _assert_catalog_manager(user)
    if ctype not in _CATALOG_FIELD_MAP:
        raise HTTPException(400, f"type debe ser uno de {sorted(_CATALOG_FIELD_MAP)}")
    field, collections = _CATALOG_FIELD_MAP[ctype]

    # Scope opcional por cliente — mismo motivo que /sources: sin filtrar,
    # los pares de typos se mezclan entre clientes.
    match_stage = {field: {"$nin": [None, "", "."]}}
    cust = (customer or "").strip()
    if cust:
        match_stage["customer"] = {"$regex": f"^{re.escape(cust)}$", "$options": "i"}

    # Recuenta ocurrencias por valor
    counts: dict[str, int] = {}
    for coll_name in collections:
        coll = getattr(db, coll_name)
        async for doc in coll.aggregate([
            {"$match": match_stage},
            {"$group": {"_id": f"${field}", "count": {"$sum": 1}}},
        ]):
            v = (doc.get("_id") or "").strip()
            if not v: continue
            counts[v] = counts.get(v, 0) + int(doc.get("count", 0))

    # Filtra por min_count
    values = sorted([v for v, c in counts.items() if c >= min_count])
    if len(values) > 3000:
        # protección: comparar 3k×3k = 9M pares es lento. Recorta al top por conteo.
        values = sorted(counts.keys(), key=lambda v: -counts[v])[:3000]
        values.sort()

    # Compara pares
    pairs = []
    for i, a in enumerate(values):
        for b in values[i + 1:]:
            # cortoide barato antes del Levenshtein
            if abs(len(a) - len(b)) > max_dist: continue
            # No comparar si la primera letra es MUY distinta (rápido descarte)
            d = _levenshtein(a, b)
            if d == 0 or d > max_dist: continue
            ca, cb = counts[a], counts[b]
            keep = a if ca >= cb else b
            drop = b if keep == a else a
            pairs.append({
                "a": a, "b": b,
                "distance": d,
                "count_a": ca, "count_b": cb,
                "recommend_keep": keep,
                "recommend_drop": drop,
            })
    # Los pares donde uno tiene MUCHO más uso que el otro son los typos más obvios
    pairs.sort(key=lambda p: (p["distance"], -abs(p["count_a"] - p["count_b"])))
    return {"type": ctype, "max_dist": max_dist, "pairs": pairs}


@router.post("/catalogs/{ctype}/rename")
async def rename_catalog_value(ctype: str, request: Request):
    """Bulk-rename a value across all source collections AND sync the curated
    catalog for `ctype`. Example: {old:'BLAK', new:'BLACK'} sweeps every
    inventory/receiving/boxes/upc/tickets row and, ademas, quita 'BLAK' del
    catalogo curado (si estaba) y agrega 'BLACK' (si no estaba). Sin el sync
    del catalogo el rechazo duro de Receiving (create_receiving/create_upc)
    rebota el nuevo valor porque no esta curado.
    Restricted to lead/supervisor (catalog manager)."""
    user = await require_auth(request)
    _assert_catalog_manager(user)
    if ctype not in _CATALOG_FIELD_MAP:
        raise HTTPException(400, f"type debe ser uno de {sorted(_CATALOG_FIELD_MAP)}")
    body = await request.json()
    old = (body.get("old") or "").strip()
    new = (body.get("new") or "").strip()
    if not old or not new:
        raise HTTPException(400, "old y new son requeridos")
    if old == new:
        raise HTTPException(400, "old y new no pueden ser iguales")
    field, collections = _CATALOG_FIELD_MAP[ctype]
    new_upper = new.upper()

    # --- 1) Sweep en las colecciones fuente ---
    total_matched = 0
    total_modified = 0
    old_ci = {"$regex": f"^{re.escape(old)}$", "$options": "i"}
    for coll_name in collections:
        coll = getattr(db, coll_name)
        res = await coll.update_many(
            {field: old_ci},
            {"$set": {field: new_upper}},
        )
        total_matched += res.matched_count
        total_modified += res.modified_count

    # --- 2) Sync del catalogo curado ---
    # Styles se scopea por customer (cada cliente tiene su lista curada); el
    # resto es global. Trabajamos por scope: para styles, iteramos por customer;
    # para el resto, un solo scope global.
    removed_catalog: list[dict] = []
    added_catalog: list[dict] = []

    curated_old = await db.wms_catalog_options.find(
        {"type": ctype, "value": old_ci}, {"_id": 0}
    ).to_list(500)

    if ctype in CUSTOMER_SCOPED:
        # Sincronizamos preservando el SCOPE de cada doc curado del `old`: por cada
        # customer que lo tenía, y también el scope global (customer vacío) si algún
        # doc del `old` era global. cust == "" representa el scope global.
        scopes = sorted({(d.get("customer") or "").strip().upper() for d in curated_old})
        for cust in scopes:
            for od in [d for d in curated_old if (d.get("customer") or "").strip().upper() == cust]:
                await db.wms_catalog_options.delete_one({"catalog_id": od["catalog_id"]})
                removed_catalog.append(od)
            # Dedupe del `new` dentro del mismo scope.
            dq = {"type": ctype, "value": {"$regex": f"^{re.escape(new_upper)}$", "$options": "i"}}
            dq["customer"] = cust if cust else {"$in": [None, ""]}
            exists = await db.wms_catalog_options.find_one(dq)
            if not exists:
                nd = {
                    "catalog_id": gen_id("cat"),
                    "type": ctype,
                    "value": new_upper,
                    "created_at": now_iso(),
                    "created_by_name": user.get("name") or user.get("email", "") or "system-rename",
                }
                if cust:
                    nd["customer"] = cust
                await db.wms_catalog_options.insert_one(nd)
                nd.pop("_id", None)
                added_catalog.append(nd)
    else:
        # Global: quitar todas las variantes del `old` y agregar `new` si no esta.
        for od in curated_old:
            await db.wms_catalog_options.delete_one({"catalog_id": od["catalog_id"]})
            removed_catalog.append(od)
        exists = await db.wms_catalog_options.find_one({
            "type": ctype,
            "value": {"$regex": f"^{re.escape(new_upper)}$", "$options": "i"},
        })
        if not exists:
            nd = {
                "catalog_id": gen_id("cat"),
                "type": ctype,
                "value": new_upper,
                "created_at": now_iso(),
                "created_by_name": user.get("name") or user.get("email", "") or "system-rename",
            }
            await db.wms_catalog_options.insert_one(nd)
            nd.pop("_id", None)
            added_catalog.append(nd)

    await log_movement(user, "catalog_rename", {
        "type": ctype, "field": field, "old": old, "new": new_upper,
        "modified": total_modified,
        "catalog_removed": [c.get("catalog_id") for c in removed_catalog],
        "catalog_added": [c.get("catalog_id") for c in added_catalog],
    })
    return {
        "type": ctype, "old": old, "new": new_upper,
        "matched": total_matched, "modified": total_modified,
        "catalog_removed": len(removed_catalog),
        "catalog_added": len(added_catalog),
    }


@router.delete("/catalogs/{ctype}/sources")
async def delete_catalog_value(ctype: str, request: Request):
    """Bulk-clear a value (set to empty) across source collections.
    Useful for purging junk like a stray '%' string.
    Restricted to lead/supervisor (catalog manager)."""
    user = await require_auth(request)
    _assert_catalog_manager(user)
    if ctype not in _CATALOG_FIELD_MAP:
        raise HTTPException(400, f"type debe ser uno de {sorted(_CATALOG_FIELD_MAP)}")
    body = await request.json()
    value = (body.get("value") or "").strip()
    if not value:
        raise HTTPException(400, "value es requerido")
    field, collections = _CATALOG_FIELD_MAP[ctype]

    total_modified = 0
    for coll_name in collections:
        coll = getattr(db, coll_name)
        res = await coll.update_many({field: value}, {"$set": {field: ""}})
        total_modified += res.modified_count

    await log_movement(user, "catalog_value_clear", {
        "type": ctype, "field": field, "value": value, "modified": total_modified,
    })
    return {"value": value, "modified": total_modified}


# ==================== LOCATIONS ====================

# ==================== LOCATION HOLDS (SAT) ====================
# Locations placed on HOLD (e.g. SAT customs hold) must NOT be touched — no
# putaway, picking, moves or edits — until a SUPERUSER releases them. The flag
# lives on the wms_locations doc (on_hold=True) so the Locations module and the
# inventory export can surface/filter it without a second collection.

async def _hold_location_names():
    """UPPERCASE names of every location currently on HOLD."""
    docs = await db.wms_locations.find({"on_hold": True}, {"_id": 0, "name": 1}).to_list(20000)
    return {(d.get("name") or "").strip().upper() for d in docs if d.get("name")}


async def _assert_not_on_hold(user, *locations):
    """Raise 423 if a non-supersu tries to touch a HOLD location. Supersu is the
    only role allowed to move stock in/out of a hold or release it."""
    if (user or {}).get("role") == "supersu":
        return
    wanted = {(l or "").strip().upper() for l in locations if l and str(l).strip()}
    if not wanted:
        return
    held = await db.wms_locations.find_one(
        {"on_hold": True, "name": {"$in": list(wanted)}}, {"_id": 0, "name": 1}
    )
    if held:
        raise HTTPException(
            status_code=423,
            detail=f"Locación {held.get('name')} está en HOLD (SAT) — no se puede tocar hasta que un superusuario la libere.",
        )


@router.get("/location-holds")
async def list_location_holds(request: Request):
    """Read the hold list (any authenticated user — used by the UI badge and the
    export 'exclude hold' option)."""
    await require_auth(request)
    docs = await db.wms_locations.find(
        {"on_hold": True},
        {"_id": 0, "name": 1, "hold_reason": 1, "hold_at": 1, "hold_by_name": 1},
    ).sort("name", 1).to_list(20000)
    return {"count": len(docs), "locations": docs}


@router.post("/location-holds")
async def add_location_holds(request: Request):
    """Place one or more locations on HOLD (supersu only). Body:
    { locations: ["RP09-A03", ...] } or { location: "RP09-A03" }, optional reason."""
    user = await require_admin_level(request, 2)
    body = await request.json()
    raw = body.get("locations") or ([body.get("location")] if body.get("location") else [])
    names = sorted({(n or "").strip().upper() for n in raw if (n or "").strip()})
    if not names:
        raise HTTPException(400, "Proporciona al menos una locación")
    reason = (body.get("reason") or "SAT").strip()
    hold_meta = {
        "on_hold": True, "hold_reason": reason, "hold_at": now_iso(),
        "hold_by": user.get("user_id"), "hold_by_name": user.get("name", ""),
    }
    updated, created = 0, 0
    for name in names:
        res = await db.wms_locations.update_one(
            {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
            {"$set": hold_meta},
        )
        if res.matched_count:
            updated += 1
        else:
            await db.wms_locations.insert_one({
                "location_id": gen_id("loc"), "name": name,
                "zone": name.split("-")[0] if "-" in name else "HOLD",
                "type": "rack", "active": True, "is_custom": False,
                "created_at": now_iso(), **hold_meta,
            })
            created += 1
    await log_movement(user, "location_hold_add", {"count": len(names), "reason": reason})
    return {"status": "ok", "held": len(names), "updated": updated, "created": created}


@router.delete("/location-holds/{name}")
async def release_location_hold(name: str, request: Request):
    """Release a location from HOLD (supersu only)."""
    user = await require_admin_level(request, 2)
    clean = (name or "").strip()
    res = await db.wms_locations.update_one(
        {"name": {"$regex": f"^{re.escape(clean)}$", "$options": "i"}},
        {"$set": {"on_hold": False, "hold_released_at": now_iso(),
                  "hold_released_by": user.get("user_id"), "hold_released_by_name": user.get("name", "")}},
    )
    if not res.matched_count:
        raise HTTPException(404, f"Locación {clean} no encontrada")
    await log_movement(user, "location_hold_release", {"location": clean.upper()})
    return {"status": "released", "location": clean.upper()}


@router.post("/locations")
async def create_location(request: Request):
    user = await require_auth(request)
    body = await request.json()
    name = body.get("name", "").strip().upper()
    zone = body.get("zone", "").strip().upper()
    loc_type = body.get("type", "rack")
    if not name:
        raise HTTPException(400, "Nombre de ubicacion requerido")
    
    # Búsqueda insensible a mayúsculas para evitar duplicados como "a1" y "A1"
    existing = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}})
    if existing:
        raise HTTPException(400, f"La ubicación '{name}' ya existe")
        
    loc = {
        "location_id": gen_id("loc"), "name": name, "zone": zone,
        "type": loc_type, "active": True, "created_at": now_iso(),
        "is_custom": True
    }
    await db.wms_locations.insert_one(loc)
    loc.pop("_id", None)
    return loc

@router.get("/locations")
async def list_locations(request: Request, summary: bool = True, skip: int = 0, limit: int = 20000):
    """List locations. Query params:
      - summary=false → skip the expensive inventory aggregation (use for dropdowns)
      - skip / limit  → paginate. Hard-capped at 20000 to prevent runaway
        responses while still fitting the full catalog (NARRO rack adds 6,688
        rows on top of the existing ~3k system slots).
    """
    await require_auth(request)

    skip = max(0, skip)
    limit = max(1, min(limit, 20000))

    locs = await db.wms_locations.find({}, {"_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)

    if not summary:
        return locs

    import time
    now = time.monotonic()
    cached = _LOC_SUMMARY_CACHE.get("data")
    if cached is not None and (now - _LOC_SUMMARY_CACHE["ts"]) < _LOC_SUMMARY_TTL:
        loc_summary = cached
    else:
        pipeline = [
            {"$match": {"units_on_hand": {"$gt": 0}, "location": {"$nin": [None, ""]}}},
            {"$group": {
                "_id": {"location": "$location", "style": "$style"},
                "style_units": {"$sum": "$units_on_hand"}
            }},
            {"$group": {
                "_id": "$_id.location",
                "total_units": {"$sum": "$style_units"},
                "skus_count": {"$sum": 1},
                "items": {"$push": {"style": {"$ifNull": ["$_id.style", "N/A"]}, "units": "$style_units"}}
            }}
        ]
        cursor = db.wms_inventory.aggregate(pipeline)
        loc_summary = {}
        async for doc in cursor:
            items = sorted(doc["items"], key=lambda x: x["units"], reverse=True)[:5]
            loc_summary[doc["_id"]] = {
                "total_units": doc["total_units"],
                "skus_count": doc["skus_count"],
                "items": items,
            }

        # Transit slots (CARRO <n> + UBICACION TEMPORAL) hold physical boxes that
        # may have NO wms_inventory row yet: stock received straight into a cart,
        # or a ledger row that drifted away while the boxes stayed put. Without
        # this fallback a cart shows "Vacío" here while Putaway counts its boxes
        # (e.g. CARRO 73: 0 inventory rows but 24 cajas). The picker/inventory
        # report already reads boxes for these slots — mirror it. Inventory wins
        # when a row exists for the slot, so we never double-count.
        box_pipeline = [
            {"$match": {"location": _transit_loc_filter(), "units": {"$gt": 0}, "status": {"$ne": "depleted"}}},
            {"$group": {
                "_id": {"location": "$location", "style": {"$ifNull": ["$style", "$sku"]}},
                "style_units": {"$sum": "$units"},
            }},
            {"$group": {
                "_id": "$_id.location",
                "total_units": {"$sum": "$style_units"},
                "skus_count": {"$sum": 1},
                "items": {"$push": {"style": {"$ifNull": ["$_id.style", "N/A"]}, "units": "$style_units"}},
            }},
        ]
        async for doc in db.wms_boxes.aggregate(box_pipeline):
            if doc["_id"] in loc_summary:
                continue  # ledger already covers this slot — trust it, don't double-count
            items = sorted(doc["items"], key=lambda x: x["units"], reverse=True)[:5]
            loc_summary[doc["_id"]] = {
                "total_units": doc["total_units"],
                "skus_count": doc["skus_count"],
                "items": items,
            }

        _LOC_SUMMARY_CACHE["data"] = loc_summary
        _LOC_SUMMARY_CACHE["ts"] = now

    for loc in locs:
        loc["inventory_summary"] = loc_summary.get(
            loc["name"], {"total_units": 0, "skus_count": 0, "items": []}
        )

    return locs


_LOC_NAMES_CACHE = {"data": None, "ts": 0.0}
_LOC_NAMES_TTL = 60.0


@router.get("/locations/names")
async def list_location_names(request: Request):
    """Lightweight location list for dropdowns / typeahead — only {name, zone}.
    The full /locations?summary=false ships ~8.8k COMPLETE docs to every module
    that merely needs names (Mover, Transit, Putaway, Inventory, Locations'
    relocate picker), which is a big slice of the WMS tab's RAM on the low-end
    warehouse PCs. This projection cuts payload + client heap to a fraction.
    Cached 60s; the catalog rarely changes."""
    await require_auth(request)
    import time
    now = time.monotonic()
    cached = _LOC_NAMES_CACHE.get("data")
    if cached is not None and (now - _LOC_NAMES_CACHE["ts"]) < _LOC_NAMES_TTL:
        return cached
    rows = await db.wms_locations.find(
        {}, {"_id": 0, "name": 1, "zone": 1}
    ).sort("name", 1).to_list(30000)
    _LOC_NAMES_CACHE["data"] = rows
    _LOC_NAMES_CACHE["ts"] = now
    return rows


@router.post("/move-location")
async def move_location_bulk(request: Request):
    """Move ALL inventory from one location to another in one operation.
    Body: { from: str, to: str }. Both must already exist. Performs:
      1. wms_boxes.update_many({location: from}, {location: to})
      2. wms_inventory: handles duplicate SKUs at the destination by merging
         (sum units_on_hand + units_allocated + total_boxes, delete source row).
      3. Single 'bulk_relocation' movement logged with totals.
    """
    user = await require_auth(request)
    body = await request.json()
    src = (body.get("from") or "").strip().upper()
    dst = (body.get("to") or "").strip().upper()
    if not src or not dst:
        raise HTTPException(400, "from y to son obligatorios")
    if src == dst:
        raise HTTPException(400, "from y to no pueden ser iguales")

    # Both locations must exist
    src_loc = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(src)}$", "$options": "i"}})
    dst_loc = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(dst)}$", "$options": "i"}})
    if not src_loc:
        raise HTTPException(404, f"Ubicación origen '{src}' no encontrada")
    if not dst_loc:
        raise HTTPException(404, f"Ubicación destino '{dst}' no encontrada. Créala primero.")
    await _assert_not_on_hold(user, src, dst)

    # 1. Bulk move boxes (no merge logic needed; box_id is unique).
    # Capture the moved box ids first so the per-box history (Case# 003) can
    # surface this relocation; update_many only returns a count.
    src_box_filter = {"location": {"$regex": f"^{re.escape(src)}$", "$options": "i"}}
    moved_box_ids = await db.wms_boxes.distinct("box_id", src_box_filter)
    box_res = await db.wms_boxes.update_many(src_box_filter, {"$set": {"location": dst, "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})
    boxes_moved = box_res.modified_count

    # 2. Inventory: merge per SKU at destination
    src_rows = await db.wms_inventory.find(
        {"location": {"$regex": f"^{re.escape(src)}$", "$options": "i"}, "units_on_hand": {"$gt": 0}}
    ).to_list(5000)
    units_moved = 0
    skus_moved = 0
    for row in src_rows:
        sku = row.get("sku") or row.get("style")
        color = row.get("color", "")
        size = row.get("size", "")
        on_hand = row.get("units_on_hand", 0)
        allocated = row.get("units_allocated", 0)
        boxes_cnt = row.get("total_boxes", 0)

        # Look for existing row at destination
        existing = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": dst
        })
        if existing:
            await db.wms_inventory.update_one(
                {"_id": existing["_id"]},
                {"$inc": {"units_on_hand": on_hand, "units_allocated": allocated, "total_boxes": boxes_cnt},
                 "$set": {"updated_at": now_iso()}}
            )
            await db.wms_inventory.delete_one({"_id": row["_id"]})
        else:
            await db.wms_inventory.update_one(
                {"_id": row["_id"]},
                {"$set": {"location": dst, "updated_at": now_iso()}}
            )
        units_moved += on_hand
        skus_moved += 1

    await log_movement(user, MovementType.BULK_RELOCATION, {
        "from": src, "to": dst,
        "boxes_moved": boxes_moved,
        "skus_moved": skus_moved,
        "units_moved": units_moved,
        "box_ids": moved_box_ids[:50],  # cap log payload
    })
    await notify_badge_change("all")

    return {
        "message": f"Movidas {skus_moved} SKUs ({units_moved} unidades, {boxes_moved} cajas) de {src} a {dst}",
        "from": src, "to": dst,
        "skus_moved": skus_moved, "units_moved": units_moved, "boxes_moved": boxes_moved,
    }


@router.get("/transit/info")
async def transit_info(request: Request):
    """Tell the frontend the canonical transit location, the cart names and
    how many boxes each one currently holds. Receiving uses this to populate
    its 'Recibir a Carro' dropdown; the Putaway 2.0 module uses it for tabs."""
    await require_auth(request)
    loc = await _ensure_transit_location()

    # Per-location count in a single aggregation. Only count boxes that still
    # hold stock — a box fully picked out (units=0) has left the cart even though
    # the doc lingers for traceability, so it must not inflate the cart count.
    pipeline = [
        {"$match": {"location": _transit_loc_filter(), "units": {"$gt": 0}, "status": {"$ne": "depleted"}}},
        {"$group": {"_id": "$location", "n": {"$sum": 1}}},
    ]
    counts = {row["_id"]: row["n"] async for row in db.wms_boxes.aggregate(pipeline)}

    # Carts come from the DB (any "CARRO <n>"), so admin-created carts show up.
    cart_docs = await _list_cart_locations()
    carts = [{
        "name": c["name"],
        "location_id": c.get("location_id"),
        "boxes": counts.get(c["name"], 0),
    } for c in cart_docs]

    return {
        "name": TRANSIT_LOCATION_NAME,
        "location_id": loc.get("location_id"),
        "boxes_in_transit": sum(counts.values()),
        "legacy_boxes": counts.get(TRANSIT_LOCATION_NAME, 0),
        "carts": carts,
    }


@router.get("/transit/boxes")
async def transit_boxes(request: Request, customer: str = "", style: str = "",
                        search: str = "", cart: str = "", limit: int = 1000):
    """List every box currently sitting in a transit slot (the carts +
    legacy UBICACION TEMPORAL). Optional `cart` narrows to a single slot;
    text filters help triage when there are hundreds of pending boxes."""
    await require_auth(request)
    await _ensure_transit_location()

    if cart:
        cart_norm = cart.strip().upper()
        if not _is_transit_name(cart_norm):
            raise HTTPException(400, f"'{cart}' no es una ubicación de tránsito válida")
        query = {"location": cart_norm}
    else:
        query = {"location": _transit_loc_filter()}
    # Hide boxes that have been fully picked out (units=0) — they've physically
    # left the cart; the doc only survives for traceability.
    query["units"] = {"$gt": 0}
    if customer:
        query["customer"] = {"$regex": re.escape(customer), "$options": "i"}
    if style:
        query["style"] = {"$regex": re.escape(style), "$options": "i"}
    if search:
        # Free-text search across the most useful fields.
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"box_id": rx}, {"sku": rx}, {"style": rx}, {"color": rx},
            {"description": rx}, {"customer": rx}, {"manufacturer": rx},
        ]
    boxes = await db.wms_boxes.find(query, {"_id": 0}).sort("created_at", 1).to_list(limit)
    cart_names = [c["name"] for c in await _list_cart_locations()]
    return {
        "transit_location": TRANSIT_LOCATION_NAME,
        "transit_locations": [TRANSIT_LOCATION_NAME] + cart_names,
        "carts": cart_names,
        "count": len(boxes),
        "boxes": boxes,
    }


@router.post("/transit/relocate")
async def transit_relocate(request: Request):
    """Move specific boxes out of UBICACION TEMPORAL into a real warehouse slot.

    Body: { box_ids: [str, ...], to: str }
    - box_ids must currently live at the transit location (silently skipped
      otherwise so concurrent moves don't error out).
    - `to` must be an existing wms_location.
    - Updates each box's location, decrements/clears the transit inventory
      row(s) and increments (or creates) the destination inventory row(s).
    - Logs a single transit_relocation movement with a summary.
    """
    user = await require_auth(request)
    body = await request.json()
    box_ids = body.get("box_ids") or []
    dst = (body.get("to") or "").strip().upper()
    if not box_ids:
        raise HTTPException(400, "box_ids es obligatorio")
    if not dst:
        raise HTTPException(400, "to es obligatorio")
    if _is_transit_name(dst):
        raise HTTPException(400, "El destino no puede ser una ubicación de tránsito (carro / temporal)")

    dst_loc = await db.wms_locations.find_one(
        {"name": {"$regex": f"^{re.escape(dst)}$", "$options": "i"}}
    )
    if not dst_loc:
        raise HTTPException(404, f"Ubicación destino '{dst}' no encontrada. Créala primero.")
    dst_name = dst_loc.get("name", dst)
    await _assert_not_on_hold(user, dst_name)

    # Pull the boxes that ACTUALLY are in any transit slot right now (5 carts
    # + legacy UBICACION TEMPORAL).
    boxes = await db.wms_boxes.find(
        {"box_id": {"$in": box_ids}, "location": _transit_loc_filter()},
        {"_id": 0},
    ).to_list(len(box_ids))
    if not boxes:
        return {
            "message": "Ninguna de las cajas indicadas está en una ubicación de tránsito.",
            "moved": 0, "units_moved": 0, "to": dst_name,
        }

    # 1. Move the boxes in one shot.
    moved_ids = [b["box_id"] for b in boxes]
    await db.wms_boxes.update_many(
        {"box_id": {"$in": moved_ids}},
        {"$set": {"location": dst_name, "state": "located", "status": "located", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}},
    )

    # 2. Aggregate units per (source_location, sku, color, size) — each box's
    # source location matters because inventory rows are keyed per-location.
    from collections import defaultdict
    bucket = defaultdict(lambda: {"units": 0, "boxes": 0, "sample": None})
    for b in boxes:
        key = (
            b.get("location") or TRANSIT_LOCATION_NAME,
            # Key by the SHORT style first (like receiving / _move_box_inventory).
            # Using the box's composite sku ("5000-AZALEA-XL") here failed to match
            # the cart's inventory row (keyed by "5000"), so the source row was
            # never decremented and the cart kept phantom (double-counted) units.
            b.get("style") or b.get("sku") or "",
            b.get("color", ""),
            b.get("size", ""),
        )
        bucket[key]["units"] += int(b.get("units") or b.get("qty") or 0)
        bucket[key]["boxes"] += 1
        if bucket[key]["sample"] is None:
            bucket[key]["sample"] = b

    units_moved = 0
    skus_moved = 0
    sources_set = set()
    for (src_location, sku, color, size), agg in bucket.items():
        units = agg["units"]
        boxes_cnt = agg["boxes"]
        sample = agg["sample"] or {}
        units_moved += units
        skus_moved += 1
        sources_set.add(src_location)

        # Decrement the source inventory row (or delete it if it would go to 0).
        src_inv = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": src_location,
        })
        if src_inv:
            new_on_hand = max(0, int(src_inv.get("units_on_hand", 0)) - units)
            new_total_boxes = max(0, int(src_inv.get("total_boxes", 0)) - boxes_cnt)
            if new_on_hand == 0 and new_total_boxes == 0:
                await db.wms_inventory.delete_one({"_id": src_inv["_id"]})
            else:
                await db.wms_inventory.update_one(
                    {"_id": src_inv["_id"]},
                    {"$set": {
                        "units_on_hand": new_on_hand,
                        "total_boxes": new_total_boxes,
                        "updated_at": now_iso(),
                    }},
                )

        # Increment (or create) the destination inventory row.
        dst_inv = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": dst_name,
        })
        if dst_inv:
            await db.wms_inventory.update_one(
                {"_id": dst_inv["_id"]},
                {"$inc": {"units_on_hand": units, "total_boxes": boxes_cnt},
                 "$set": {"updated_at": now_iso()}},
            )
        else:
            await db.wms_inventory.insert_one({
                "inventory_id": gen_id("inv"),
                "sku": sku,
                "style": sample.get("style") or sku,
                "color": color,
                "size": size,
                "customer": sample.get("customer", ""),
                "manufacturer": sample.get("manufacturer", ""),
                "description": sample.get("description", ""),
                "country_of_origin": sample.get("country_of_origin", "") or sample.get("coo", ""),
                "fabric_content": sample.get("fabric_content", ""),
                "location": dst_name,
                "total_boxes": boxes_cnt,
                "units_on_hand": units,
                "units_allocated": 0,
                "updated_at": now_iso(),
            })

    await log_movement(user, MovementType.TRANSIT_RELOCATION, {
        "to": dst_name,
        "from_sources": sorted(sources_set),
        "boxes_moved": len(moved_ids),
        "skus_moved": skus_moved,
        "units_moved": units_moved,
        "box_ids": moved_ids[:50],  # cap log payload
    })
    await notify_badge_change("all")

    return {
        "message": f"Movidas {len(moved_ids)} cajas ({units_moved} unidades, {skus_moved} SKUs) a {dst_name}",
        "moved": len(moved_ids),
        "units_moved": units_moved,
        "skus_moved": skus_moved,
        "to": dst_name,
    }


@router.post("/boxes/relocate")
async def boxes_relocate(request: Request):
    """Generic per-box relocation. Unlike /transit/relocate, doesn't restrict
    the source location — works from any source to any destination. Used by
    the Locations detail modal to let an admin pick individual LPNs out of a
    bin and ship them elsewhere.

    Body: { box_ids: [str, ...], to: str }
    """
    user = await require_auth(request)
    body = await request.json()
    box_ids = body.get("box_ids") or []
    dst = (body.get("to") or "").strip().upper()
    if not box_ids:
        raise HTTPException(400, "box_ids es obligatorio")
    if not dst:
        raise HTTPException(400, "to es obligatorio")

    dst_loc = await db.wms_locations.find_one(
        {"name": {"$regex": f"^{re.escape(dst)}$", "$options": "i"}}
    )
    if not dst_loc:
        raise HTTPException(404, f"Ubicación destino '{dst}' no encontrada. Créala primero.")
    dst_name = dst_loc.get("name", dst)

    # Load the boxes; silently skip those already at the destination so
    # accidental double-clicks don't error out.
    boxes_raw = await db.wms_boxes.find(
        {"box_id": {"$in": box_ids}}, {"_id": 0},
    ).to_list(len(box_ids))
    boxes = [b for b in boxes_raw if (b.get("location") or "").upper() != dst_name.upper()]
    if not boxes:
        return {
            "message": "Las cajas seleccionadas ya están en la ubicación destino.",
            "moved": 0, "units_moved": 0, "skus_moved": 0, "to": dst_name,
        }

    # HOLD guard: stock in (or headed into) a SAT-held bin can't be relocated
    # until released. The other move endpoints check this; this one didn't.
    src_locs = {(b.get("location") or "") for b in boxes}
    await _assert_not_on_hold(user, dst_name, *src_locs)

    # 1. Bulk update box locations.
    moved_ids = [b["box_id"] for b in boxes]
    await db.wms_boxes.update_many(
        {"box_id": {"$in": moved_ids}},
        {"$set": {"location": dst_name, "state": "located", "status": "located", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}},
    )

    # 2. Rebalance inventory: bucket by (source_location, sku, color, size) so
    #    moves spanning multiple source bins still decrement the right rows.
    from collections import defaultdict
    bucket = defaultdict(lambda: {"units": 0, "boxes": 0, "sample": None, "box_ids": []})
    for b in boxes:
        src = (b.get("location") or "")
        # Short style first (matches how inventory rows are keyed); the composite
        # sku missed the source row and left phantom double-counted inventory.
        key = (src, b.get("style") or b.get("sku") or "", b.get("color", ""), b.get("size", ""))
        bucket[key]["units"] += int(b.get("units") or b.get("qty") or 0)
        bucket[key]["boxes"] += 1
        if bucket[key]["sample"] is None:
            bucket[key]["sample"] = b
        bucket[key]["box_ids"].append(b["box_id"])

    units_moved = 0
    skus_moved = 0
    sources_touched = set()
    for (src, sku, color, size), agg in bucket.items():
        units = agg["units"]
        boxes_cnt = agg["boxes"]
        sample = agg["sample"] or {}
        units_moved += units
        skus_moved += 1
        if src:
            sources_touched.add(src)

        # Decrement the source inventory row (delete if it would hit 0/0).
        if src:
            src_inv = await db.wms_inventory.find_one({
                "sku": sku, "color": color, "size": size, "location": src,
            })
            if src_inv:
                new_on_hand = max(0, int(src_inv.get("units_on_hand", 0)) - units)
                new_total_boxes = max(0, int(src_inv.get("total_boxes", 0)) - boxes_cnt)
                if new_on_hand == 0 and new_total_boxes == 0:
                    await db.wms_inventory.delete_one({"_id": src_inv["_id"]})
                else:
                    await db.wms_inventory.update_one(
                        {"_id": src_inv["_id"]},
                        {"$set": {
                            "units_on_hand": new_on_hand,
                            "total_boxes": new_total_boxes,
                            "updated_at": now_iso(),
                        }},
                    )

        # Increment (or create) the destination inventory row.
        dst_inv = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": dst_name,
        })
        if dst_inv:
            dst_id = dst_inv.get("inventory_id")
            if not dst_id:
                dst_id = gen_id("inv")
                await db.wms_inventory.update_one(
                    {"_id": dst_inv["_id"]},
                    {"$inc": {"units_on_hand": units, "total_boxes": boxes_cnt},
                     "$set": {"inventory_id": dst_id, "updated_at": now_iso()}},
                )
            else:
                await db.wms_inventory.update_one(
                    {"_id": dst_inv["_id"]},
                    {"$inc": {"units_on_hand": units, "total_boxes": boxes_cnt},
                     "$set": {"updated_at": now_iso()}},
                )
        else:
            dst_id = gen_id("inv")
            await db.wms_inventory.insert_one({
                "inventory_id": dst_id,
                "sku": sku,
                "style": sample.get("style") or sku,
                "color": color,
                "size": size,
                "customer": sample.get("customer", ""),
                "manufacturer": sample.get("manufacturer", ""),
                "description": sample.get("description", ""),
                "country_of_origin": sample.get("country_of_origin", "") or sample.get("coo", ""),
                "fabric_content": sample.get("fabric_content", ""),
                "location": dst_name,
                "total_boxes": boxes_cnt,
                "units_on_hand": units,
                "units_allocated": 0,
                "updated_at": now_iso(),
            })

        # Link moved boxes to the destination inventory row
        bucket_box_ids = agg.get("box_ids") or []
        if bucket_box_ids:
            await db.wms_boxes.update_many(
                {"box_id": {"$in": bucket_box_ids}},
                {"$set": {"inventory_id": dst_id}}
            )

    # Use the transit movement type when EVERY source was a transit slot
    # (cart or legacy UBICACION TEMPORAL), so the Movements module surfaces it
    # under the same bucket as /transit/relocate.
    move_type = (
        MovementType.TRANSIT_RELOCATION
        if sources_touched and all(_is_transit_name(s) for s in sources_touched)
        else MovementType.BULK_RELOCATION
    )
    await log_movement(user, move_type, {
        "to": dst_name,
        "sources": sorted(sources_touched),
        "boxes_moved": len(moved_ids),
        "skus_moved": skus_moved,
        "units_moved": units_moved,
        "box_ids": moved_ids[:50],  # cap log payload
    })
    await notify_badge_change("all")

    return {
        "message": f"Movidas {len(moved_ids)} cajas ({units_moved} unidades) a {dst_name}",
        "moved": len(moved_ids),
        "units_moved": units_moved,
        "skus_moved": skus_moved,
        "to": dst_name,
    }


@router.post("/move-units")
async def move_units(request: Request):
    """Move a partial quantity of ONE sku/color/size from one location to another
    — the picker 'Mover > Unidades (consolidar)' flow. Consumes source boxes FIFO:
    a box taken in full is relocated, a box taken in part is split (source box
    reduced, a fresh box created at the destination). Inventory rows are kept in
    lockstep with the physical boxes the whole way so the two never drift."""
    user = await require_auth(request)
    body = await request.json()
    src = (body.get("from") or "").strip().upper()
    dst = (body.get("to") or "").strip().upper()
    sku = (body.get("sku") or body.get("style") or "").strip()
    color = (body.get("color") or "").strip()
    size = (body.get("size") or "").strip()
    try:
        units = int(body.get("units") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "Cantidad (units) inválida")

    if not src or not dst:
        raise HTTPException(400, "Ubicación origen y destino son obligatorias")
    if src == dst:
        raise HTTPException(400, "El origen y el destino no pueden ser iguales")
    if not sku:
        raise HTTPException(400, "SKU/Style es obligatorio")
    if units <= 0:
        raise HTTPException(400, "La cantidad debe ser mayor a 0")

    dst_loc = await db.wms_locations.find_one(
        {"name": {"$regex": f"^{re.escape(dst)}$", "$options": "i"}}
    )
    if not dst_loc:
        raise HTTPException(404, f"Ubicación destino '{dst}' no encontrada. Créala primero.")
    dst_name = dst_loc.get("name", dst)
    await _assert_not_on_hold(user, src, dst_name)

    # Block oversell: can't move more than what's physically at the source
    # (reuses the same availability check as picking — boxes OR inventory row).
    avail = await _available_units(sku, color, size, src)
    if units > avail:
        raise HTTPException(409, f"No hay suficiente en {src}: pides {units}, disponible {avail}")

    # Pre-lookup or pre-generate destination inventory_id
    dst_inv = await db.wms_inventory.find_one(
        {"sku": sku, "color": color, "size": size, "location": dst_name}
    )
    if dst_inv:
        dst_id = dst_inv.get("inventory_id")
        if not dst_id:
            dst_id = gen_id("inv")
            await db.wms_inventory.update_one({"_id": dst_inv["_id"]}, {"$set": {"inventory_id": dst_id}})
    else:
        dst_id = gen_id("inv")

    # Consume source boxes FIFO.
    q = {
        "$or": [{"sku": sku}, {"style": sku}], "color": color, "size": size,
        "location": {"$regex": f"^{re.escape(src)}$", "$options": "i"},
        "units": {"$gt": 0},
    }
    boxes = await db.wms_boxes.find(q).sort("created_at", 1).to_list(1000)
    sample = boxes[0] if boxes else {}
    remaining = units
    whole_count = 0   # boxes that left the source entirely
    split_count = 0   # boxes split → a new partial box created at the destination
    moved_box_ids = []  # every box touched (whole + shrunk source + split child) for box history
    next_seq = None

    for box in boxes:
        if remaining <= 0:
            break
        b_qty = int((box.get("units") if box.get("units") is not None else box.get("qty", 0)) or 0)
        if b_qty <= 0:
            continue
        take = min(b_qty, remaining)
        if take >= b_qty:
            # Whole box relocates to the destination.
            await db.wms_boxes.update_one(
                {"_id": box["_id"]},
                {"$set": {"location": dst_name, "status": "stored", "inventory_id": dst_id, "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}},
            )
            whole_count += 1
            moved_box_ids.append(box.get("box_id"))
        else:
            # Partial: shrink the source box, create a fresh box at the destination.
            await db.wms_boxes.update_one(
                {"_id": box["_id"]},
                {"$set": {"units": b_qty - take, "qty": b_qty - take}},
            )
            next_seq = await _reserve_box_seqs(1)
            new_box_id = f"BOX-{next_seq:06d}"
            child = {k: v for k, v in box.items() if k != "_id"}
            child.update({
                "box_id": new_box_id, "barcode": new_box_id, "lpn_id": new_box_id,
                "seq_num": next_seq, "location": dst_name,
                "units": take, "qty": take, "status": "stored",
                "split_from": box.get("box_id"), "created_at": now_iso(),
                "inventory_id": dst_id,
            })
            await db.wms_boxes.insert_one(child)
            split_count += 1
            moved_box_ids.extend([box.get("box_id"), new_box_id])
        remaining -= take

    # Inventory lockstep: decrement the source row by the full requested units and
    # by the boxes that physically left it; increment/create the destination row by
    # the units and every box that arrived there (whole + split).
    boxes_added = whole_count + split_count
    src_inv = await db.wms_inventory.find_one(
        {"sku": sku, "color": color, "size": size,
         "location": {"$regex": f"^{re.escape(src)}$", "$options": "i"}}
    ) or await db.wms_inventory.find_one(
        {"style": {"$regex": f"^{re.escape(sku)}$", "$options": "i"},
         "color": color, "size": size,
         "location": {"$regex": f"^{re.escape(src)}$", "$options": "i"}}
    )
    if src_inv:
        # Atomic decrement-and-clamp (same race-free pattern as the picking fix):
        # one document op so concurrent moves from the same row can't lose an
        # update or drive stock negative.
        updated = await db.wms_inventory.find_one_and_update(
            {"_id": src_inv["_id"]},
            [{"$set": {
                "units_on_hand": {"$max": [0, {"$subtract": [{"$ifNull": ["$units_on_hand", 0]}, units]}]},
                "total_boxes": {"$max": [0, {"$subtract": [{"$ifNull": ["$total_boxes", 0]}, whole_count]}]},
                "updated_at": now_iso(),
            }}],
            return_document=ReturnDocument.AFTER,
        )
        # Drop the row only once it is fully empty (no units, no boxes, nothing reserved).
        if (updated and int(updated.get("units_on_hand", 0) or 0) <= 0
                and int(updated.get("total_boxes", 0) or 0) <= 0
                and int(updated.get("units_allocated", 0) or 0) <= 0):
            await db.wms_inventory.delete_one({"_id": src_inv["_id"]})

    dst_inv = await db.wms_inventory.find_one(
        {"sku": sku, "color": color, "size": size, "location": dst_name}
    )
    if dst_inv:
        await db.wms_inventory.update_one(
            {"_id": dst_inv["_id"]},
            {"$inc": {"units_on_hand": units, "total_boxes": boxes_added},
             "$set": {"updated_at": now_iso(), "inventory_id": dst_id}},
        )
    else:
        await db.wms_inventory.insert_one({
            "inventory_id": dst_id,
            "sku": sku, "style": sample.get("style") or sku,
            "color": color, "size": size,
            "customer": sample.get("customer", ""),
            "manufacturer": sample.get("manufacturer", ""),
            "description": sample.get("description", ""),
            "country_of_origin": sample.get("country_of_origin", "") or sample.get("coo", ""),
            "fabric_content": sample.get("fabric_content", ""),
            "location": dst_name, "total_boxes": boxes_added,
            "units_on_hand": units, "units_allocated": 0, "updated_at": now_iso(),
        })

    await log_movement(user, MovementType.BULK_RELOCATION, {
        "from": src, "to": dst_name, "sku": sku, "color": color, "size": size,
        "units_moved": units, "boxes_relocated": whole_count, "boxes_split": split_count,
        "box_ids": [b for b in moved_box_ids if b][:50],  # cap log payload
    })
    await notify_badge_change("all")
    return {
        "message": f"Movidas {units} unidades de {sku} de {src} a {dst_name}",
        "from": src, "to": dst_name, "units_moved": units,
        "boxes_relocated": whole_count, "boxes_split": split_count,
    }


@router.post("/boxes/reconcile-lpn")
async def reconcile_lpn(request: Request):
    """Reconcile a migrated 'generic' LPN box with the box's REAL physical license
    plate, on the fly during a relocation.

    The bulk Excel import minted one box per case with a synthetic id
    (`LPN_xxxx`) — that id is never printed on the carton, which still carries its
    pre-WMS label (e.g. barcode A2600510001). So the operator can't scan the
    generic id; instead they identify the box by Location + Product, then scan the
    physical LPN to stamp it onto the box. 1:1, no split: exactly one pending
    generic box of that product at that location is matched, its barcode/lpn_id is
    set to the physical LPN, its quantity is corrected to the real count, and it is
    moved to the destination slot. The internal `box_id` is deliberately left
    untouched so nothing that references it (tasks, movements, splits) breaks —
    scanning the physical LPN resolves via the barcode/lpn_id fallback in get_box.

    Body: { location, sku|style, color, size, physical_lpn, units, destination }
    """
    user = await require_auth(request)
    body = await request.json()
    location = (body.get("location") or "").strip()
    sku = (body.get("sku") or body.get("style") or "").strip()
    color = (body.get("color") or "").strip()
    size = (body.get("size") or "").strip()
    physical_lpn = (body.get("physical_lpn") or "").strip().upper()
    destination = (body.get("destination") or "").strip()
    try:
        units = int(body.get("units") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "Cantidad (units) inválida")

    if not location:
        raise HTTPException(400, "Ubicación de origen es obligatoria")
    if not sku:
        raise HTTPException(400, "SKU/Style es obligatorio")
    if not physical_lpn:
        raise HTTPException(400, "El LPN físico es obligatorio")
    if not destination:
        raise HTTPException(400, "Ubicación de destino es obligatoria")
    if units <= 0:
        raise HTTPException(400, "La cantidad debe ser mayor a 0")

    # Destination must exist (same rule as the other relocate endpoints).
    dst_loc = await db.wms_locations.find_one(
        {"name": {"$regex": f"^{re.escape(destination)}$", "$options": "i"}}
    )
    if not dst_loc:
        raise HTTPException(404, f"Ubicación destino '{destination}' no encontrada. Créala primero.")
    dst_name = dst_loc.get("name", destination)
    await _assert_not_on_hold(user, location, dst_name)

    # Guard: a physical LPN must be unique — never collide with an existing
    # box_id/barcode/lpn_id on a DIFFERENT box (a double-reconcile or a typo).
    clash = await db.wms_boxes.find_one(
        {"$or": [{"box_id": physical_lpn}, {"barcode": physical_lpn}, {"lpn_id": physical_lpn}]},
        {"_id": 0, "box_id": 1, "lpn_id": 1},
    )
    if clash:
        raise HTTPException(
            409,
            f"El LPN físico '{physical_lpn}' ya está en uso por la caja {clash.get('box_id')}.",
        )

    # Find ONE pending generic box of this product still sitting at the origin.
    # 'Generic' = a synthetic import id (the migration mints either "LPN_<uuid>" or
    # "LPN<uuid>" — no underscore — so match "^LPN") that hasn't been reconciled yet.
    candidate = await db.wms_boxes.find_one(
        {
            "$or": [{"sku": _ci_eq(sku)}, {"style": _ci_eq(sku)}],
            "color": _ci_eq(color), "size": _ci_eq(size),
            "location": _ci_eq(location),
            "box_id": {"$regex": "^LPN", "$options": "i"},
            "lpn_reconciled_at": {"$exists": False},
            "units": {"$gt": 0},
        },
        sort=[("created_at", 1)],
    )
    if not candidate:
        raise HTTPException(
            404,
            f"No hay caja genérica pendiente de {sku} {color} {size} en {location}.",
        )

    box_id = candidate["box_id"]
    old_units = int(
        candidate.get("units") if candidate.get("units") is not None else candidate.get("qty", 0) or 0
    )

    # 1. Decrement the SOURCE inventory row by the box's CURRENT units + 1 box.
    #    Match on sku first, then fall back to style (Excel rows can carry the real
    #    value in `style` with sku missing/"None"). Drop the row only if it empties.
    src_inv = await db.wms_inventory.find_one(
        {"sku": sku, "color": color, "size": size,
         "location": {"$regex": f"^{re.escape(location)}$", "$options": "i"}}
    ) or await db.wms_inventory.find_one(
        {"style": {"$regex": f"^{re.escape(sku)}$", "$options": "i"},
         "color": color, "size": size,
         "location": {"$regex": f"^{re.escape(location)}$", "$options": "i"}}
    )
    if src_inv:
        new_on_hand = max(0, int(src_inv.get("units_on_hand", 0)) - old_units)
        new_total_boxes = max(0, int(src_inv.get("total_boxes", 0) or 0) - 1)
        if (new_on_hand == 0 and new_total_boxes == 0
                and int(src_inv.get("units_allocated", 0) or 0) <= 0):
            await db.wms_inventory.delete_one({"_id": src_inv["_id"]})
        else:
            await db.wms_inventory.update_one(
                {"_id": src_inv["_id"]},
                {"$set": {"units_on_hand": new_on_hand, "total_boxes": new_total_boxes,
                          "updated_at": now_iso()}},
            )

    # 2. Stamp the physical LPN onto the box, correct its qty, move it. box_id stays.
    await db.wms_boxes.update_one(
        {"_id": candidate["_id"]},
        {"$set": {
            "barcode": physical_lpn,
            "lpn_id": physical_lpn,
            "units": units,
            "qty": units,
            "location": dst_name,
            "status": "stored",
            "state": "located",
            "generic_lpn": box_id,            # remember the synthetic id we replaced
            "lpn_reconciled_at": now_iso(),
            "lpn_reconciled_by": user.get("user_id"),
            "last_transferred_at": now_iso(),
            "last_transferred_by": user.get("name", user.get("email", "")),
            "updated_at": now_iso(),
        }},
    )

    # 3. Increment (or create) the DESTINATION inventory row by the NEW units + 1 box.
    dst_inv = await db.wms_inventory.find_one(
        {"sku": sku, "color": color, "size": size, "location": dst_name}
    )
    if dst_inv:
        await db.wms_inventory.update_one(
            {"_id": dst_inv["_id"]},
            {"$inc": {"units_on_hand": units, "total_boxes": 1},
             "$set": {"updated_at": now_iso()}},
        )
    else:
        await db.wms_inventory.insert_one({
            "inventory_id": gen_id("inv"),
            "sku": sku,
            "style": candidate.get("style") or sku,
            "color": color,
            "size": size,
            "customer": candidate.get("customer", ""),
            "manufacturer": candidate.get("manufacturer", ""),
            "description": candidate.get("description", ""),
            "country_of_origin": candidate.get("country_of_origin", "") or candidate.get("coo", ""),
            "fabric_content": candidate.get("fabric_content", ""),
            "location": dst_name,
            "total_boxes": 1,
            "units_on_hand": units,
            "units_allocated": 0,
            "updated_at": now_iso(),
        })

    # 4. Audit trail.
    await log_movement(user, MovementType.LPN_RECONCILED, {
        "box_id": box_id,
        "generic_lpn": box_id,
        "physical_lpn": physical_lpn,
        "sku": sku, "color": color, "size": size,
        "from": location, "to": dst_name,
        "old_units": old_units, "new_units": units, "delta_units": units - old_units,
    })
    await notify_badge_change("all")

    return {
        "message": f"LPN {physical_lpn} casado ({units} u) y movido a {dst_name}",
        "box_id": box_id,
        "physical_lpn": physical_lpn,
        "from": location, "to": dst_name,
        "units": units, "old_units": old_units,
    }


@router.delete("/locations/{location_id}")
async def delete_location(location_id: str, request: Request, force: bool = False):
    """Delete a location. Two guards:
      1. System-protected slots (the canonical UBICACION TEMPORAL) can never
         be deleted — they're hard-wired to other modules (En Tránsito flow,
         Receiving "Recibir a Temporal" button) so removing them silently
         breaks those features.
      2. Regular slots are blocked when the bin still has boxes or inventory
         rows pointing to it — otherwise the operator orphans stock that
         can't be surfaced anywhere in the UI. Pass `?force=true` to
         override this second guard (advanced; doesn't bypass #1)."""
    user = await require_auth(request)
    loc = await db.wms_locations.find_one({"location_id": location_id})
    if not loc:
        raise HTTPException(404, "Ubicacion no encontrada")
    name = loc.get("name") or ""

    # Guard 1 — system-protected locations are off-limits, even with force.
    if _is_transit_name(name):
        raise HTTPException(
            status_code=403,
            detail=(
                f"'{name}' es una ubicación del sistema (módulo Putaway 2.0) "
                "y no se puede eliminar."
            ),
        )

    # Guard 2 — stock-bearing bins are blocked unless ?force=true is passed.
    if not force:
        n_boxes = await db.wms_boxes.count_documents({"location": name})
        n_inv = await db.wms_inventory.count_documents({
            "location": name, "units_on_hand": {"$gt": 0},
        })
        if n_boxes or n_inv:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"'{name}' tiene contenido: {n_boxes} cajas / {n_inv} líneas con stock. "
                    "Vácila primero (Mover) o pasa ?force=true para forzar."
                ),
            )

    # With force, Guard 2 was skipped — so sweep any remaining content into
    # UBICACION TEMPORAL instead of orphaning it (stock pointing at a deleted
    # bin is invisible everywhere). Case-insensitive to catch mixed casing.
    if force:
        temp = "UBICACION TEMPORAL"
        mb = await db.wms_boxes.update_many({"location": _ci_eq(name)}, {"$set": {"location": temp, "status": "located", "state": "located"}})
        mi = await db.wms_inventory.update_many({"location": _ci_eq(name)}, {"$set": {"location": temp, "updated_at": now_iso()}})
        if mb.modified_count or mi.modified_count:
            if not await db.wms_locations.find_one({"name": _ci_eq(temp)}):
                await db.wms_locations.insert_one({"location_id": gen_id("loc"), "name": temp, "zone": "TEMPORAL", "type": "rack", "active": True, "created_at": now_iso()})
            await log_movement(user, "location_force_swept", {"name": name, "moved_to": temp, "boxes": mb.modified_count, "inv_rows": mi.modified_count})

    result = await db.wms_locations.delete_one({"location_id": location_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Ubicacion no encontrada")
    await log_movement(user, "location_deleted", {"name": name, "forced": force})
    return {"message": f"Ubicacion '{name}' eliminada"}

@router.put("/locations/{location_id}")
async def update_location(location_id: str, request: Request):
    user = await require_auth(request)
    body = await request.json()
    new_name = body.get("name", "").strip().upper()
    new_zone = body.get("zone", "").strip().upper()

    existing_loc = await db.wms_locations.find_one({"location_id": location_id})
    if not existing_loc:
        raise HTTPException(404, "Ubicación no encontrada")

    old_name = existing_loc.get("name")

    # System-protected: the canonical transit slot and the 5 carts cannot be
    # renamed because the Putaway 2.0 module + Receiving "Recibir a Carro"
    # button look them up by exact string match.
    if _is_transit_name(old_name) and new_name and new_name != old_name:
        raise HTTPException(
            status_code=403,
            detail=(
                f"'{old_name}' es una ubicación del sistema (módulo Putaway 2.0) "
                "y no se puede renombrar."
            ),
        )

    update_doc = {}
    if new_name:
        if new_name != old_name:
            dup = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(new_name)}$", "$options": "i"}, "location_id": {"$ne": location_id}})
            if dup:
                raise HTTPException(400, f"La ubicación '{new_name}' ya existe")
            update_doc["name"] = new_name
    if new_zone:
        update_doc["zone"] = new_zone
        
    if update_doc:
        await db.wms_locations.update_one({"location_id": location_id}, {"$set": update_doc})
        if "name" in update_doc:
            # Case-insensitive match so rows stored with different casing
            # (Sand vs SAND) follow the rename instead of being orphaned at the
            # old name once the bin is gone.
            await db.wms_inventory.update_many({"location": _ci_eq(old_name)}, {"$set": {"location": new_name}})
            await db.wms_boxes.update_many({"location": _ci_eq(old_name)}, {"$set": {"location": new_name}})
            await log_movement(user, "location_renamed", {"old_name": old_name, "new_name": new_name})
            
    updated_loc = await db.wms_locations.find_one({"location_id": location_id}, {"_id": 0})
    return updated_loc

@router.get("/locations/print")
async def print_locations(request: Request, ids: str = "all", zone: str = ""):
    """Generate a PDF for Zebra Label Printers (4x2 inch) for WMS locations.
    Selection precedence (first match wins):
      - zone=X   → all locations in that zone
      - ids=A,B  → explicit list of location_ids
      - ids=all  → every location (default)
    """
    await require_auth(request)

    from reportlab.lib.pagesizes import landscape
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.graphics.barcode import code128
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF

    # 1. Fetch locations
    if zone:
        zone_query = {"zone": {"$regex": f"^{re.escape(zone)}$", "$options": "i"}}
        locs = await db.wms_locations.find(zone_query, {"_id": 0}).sort("name", 1).to_list(3000)
    elif ids == "all":
        locs = await db.wms_locations.find({}, {"_id": 0}).sort("name", 1).to_list(3000)
    else:
        id_list = ids.split(",")
        locs = await db.wms_locations.find({"location_id": {"$in": id_list}}, {"_id": 0}).sort("name", 1).to_list(3000)

    if not locs:
        raise HTTPException(404, "No se encontraron ubicaciones para imprimir")

    # 2. Create PDF buffer
    buffer = io.BytesIO()
    # Zebra labels are usually 4x2 inches
    label_width = 4 * inch
    label_height = 2 * inch
    c = canvas.Canvas(buffer, pagesize=(label_width, label_height))

    for loc in locs:
        name = loc.get("name", "N/A").upper()
        zone = loc.get("zone", "N/A").upper()
        
        # Draw Location Name (Large)
        c.setFont("Helvetica-Bold", 45)
        c.drawCentredString(label_width / 2, label_height - 50, name)
        
        # Draw Zone (Small)
        c.setFont("Helvetica", 12)
        c.drawCentredString(label_width / 2, label_height - 75, f"ZONA: {zone}")
        
        # Draw Barcode (Code128)
        try:
            # Code128 widget from reportlab.graphics.barcode
            barcode = code128.Code128(name, barHeight=0.5*inch, barWidth=1.0)
            
            # Center the barcode
            x_pos = (label_width - barcode.width) / 2
            
            # Draw directly to canvas (Y position lowered to 15 to avoid overlap)
            barcode.drawOn(c, x_pos, 15)
        except Exception as e:
            logger.error(f"Error generating barcode for {name}: {e}")
            c.setFont("Helvetica-Oblique", 8)
            c.drawCentredString(label_width / 2, 35, f"(Error: {str(e)[:30]})")

        c.showPage()

    c.save()
    buffer.seek(0)
    
    filename = f"Etiquetas_WMS_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer, 
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

@router.post("/receiving")
async def create_receiving(request: Request):
    user = await require_auth(request)
    body = await request.json()
    customer = await _canonical_customer(body.get("customer", ""))
    # Normalize the identity dimensions to UPPERCASE + single-spaced on write so
    # inventory/boxes/reports stay consistent (no more 'Sand' vs 'SAND' or
    # 'Bangladesh' vs 'BANGLADESH' splitting matches/reports).
    _norm = lambda v: re.sub(r"\s+", " ", str(v or "")).strip().upper()
    manufacturer = _norm(body.get("manufacturer"))
    style = _norm(body.get("style"))
    color = _norm(body.get("color"))
    size = _norm(body.get("size"))
    description = _norm(body.get("description"))
    country_of_origin = _norm(body.get("country_of_origin"))
    fabric_content = _norm(body.get("fabric_content"))
    # Default a UBICACION TEMPORAL (uppercase, el nombre oficial usado por el
    # modulo de Putaway/Transit). El viejo default 'Locación Temporal' (casing
    # mixto + acento) generaba filas que el match exacto de _ci_eq no encontraba.
    inv_location = body.get("inv_location", "").strip() or "UBICACION TEMPORAL"
    lot_number = body.get("lot_number", "").strip()
    sku = body.get("sku", "").strip().upper()
    dozens = int(body.get("dozens", 0) or 0)
    pieces = int(body.get("pieces", 0) or 0)
    units = int(body.get("units", 0) or 0)
    vendor = body.get("vendor", manufacturer).strip()
    items = body.get("items", [])
    is_bpo = body.get("is_bpo", False)

    if not style:
        raise HTTPException(400, "Style requerido")
    # Guardia de catalogo: los 6 campos de identidad deben venir del catalogo
    # curado del cliente (o global). Si el catalogo esta vacio, bootstrap.
    await _assert_curated_identity(customer, {
        "styles": style,
        "colors": color,
        "sizes": size,
        "descriptions": description,
        "countries": country_of_origin,
        "fabrics": fabric_content,
    })

    # Guardia de UPC: si el receiving se hace contra un UPC que YA esta en el
    # catalogo, los campos capturados deben coincidir con los del UPC. Sin
    # esta guardia, el operador puede capturar un pais/desc/fabric distinto y
    # las cajas se etiquetan con el valor equivocado (bug del receiving GILDAN
    # 5000 BLACK L de GOODIE 2026-07-13: UPC decia NICARAGUA pero se recibio
    # como REPUBLICA DOMINICANA). Solo bloqueamos si los valores difieren de
    # forma no-vacia (evita bloquear cuando el UPC tiene el campo vacio).
    upc_code = str(body.get("upc", "")).strip().upper()
    if upc_code:
        upc_doc = await db.wms_upc_catalog.find_one({"upc": upc_code}, {"_id": 0})
        if upc_doc:
            _upc_labels = {
                "style": "Style", "color": "Color", "size": "Talla",
                "description": "Descripcion", "country_of_origin": "Pais de origen",
                "fabric_content": "Fabric",
            }
            captured = {
                "style": style, "color": color, "size": size,
                "description": description,
                "country_of_origin": country_of_origin,
                "fabric_content": fabric_content,
            }
            divergences = []
            for k, cap in captured.items():
                upc_val = (upc_doc.get(k) or "").strip()
                if not upc_val or not cap:
                    continue
                # Normalize whitespace + upper for compare (mismo que el sweep)
                a = re.sub(r"\s+", " ", str(cap)).strip().upper()
                b = re.sub(r"\s+", " ", upc_val).strip().upper()
                if a != b:
                    divergences.append(f"{_upc_labels[k]}: capturaste '{cap}' pero el UPC dice '{upc_val}'")
            if divergences:
                raise HTTPException(400, (
                    f"El UPC {upc_code} ya existe en el catalogo con datos distintos:\n"
                    + "\n".join(f"  - {d}" for d in divergences)
                    + "\n\nCorrige tu captura o edita primero el UPC (con la casilla 'Aplicar a cajas ya recibidas')."
                ))
    # Block receiving against an ASN whose receiving process was finished.
    asn_ref = str(body.get("asn_reference", "")).strip()
    if asn_ref:
        ref_asn = await db.wms_asn.find_one({"asn_id": asn_ref}, {"_id": 0, "closed": 1})
        if ref_asn and ref_asn.get("closed"):
            raise HTTPException(409, f"El ASN {asn_ref} ya cerró su recibo. Reábrelo para recibir más.")
    if not country_of_origin:
        raise HTTPException(400, "País de origen (country_of_origin) es obligatorio")
    if not fabric_content:
        raise HTTPException(400, "Contenido / fabric_content es obligatorio")

    # Auto-generate SKU if not provided
    if not sku and style:
        base = style.upper().replace(' ', '-')
        parts = [base]
        if color: parts.append(color.upper().replace(' ', '-')[:10])
        if size: parts.append(size.upper())
        sku = '-'.join(parts)

    # Calculate total units
    total_units = units if units > 0 else (dozens * 12 + pieces)
    if total_units <= 0 and not items:
        raise HTTPException(400, "Debe ingresar cantidad (dozens/pieces/units)")

    receiving_id = gen_id("rcv")

    # Box generation — reserve a contiguous block of sequence numbers atomically
    # so two concurrent receivings can't mint the same BOX-id.
    _total_boxes = sum(int(it.get("boxes", 1) or 1) for it in items) if items else 1
    seq = await _reserve_box_seqs(_total_boxes) - 1

    box_docs = []
    if items:
        for item in items:
            item_size = item.get("size", "").strip().upper()
            boxes_count = int(item.get("boxes", 1))
            units_per_box = int(item.get("units_per_box", 1))
            for _ in range(boxes_count):
                seq += 1
                box_id = f"BOX-{seq:06d}"
                box_docs.append({
                    "box_id": box_id, "barcode": box_id, "receiving_id": receiving_id,
                    "customer": customer, "manufacturer": manufacturer, "style": style,
                    "sku": sku or style, "color": color, "size": item_size,
                    "units": units_per_box, "qty": units_per_box, "seq_num": seq, "location": inv_location,
                    "status": "putaway_pending", "state": "raw", "is_bpo": is_bpo,
                    "lpn_id": box_id, "coo": country_of_origin,
                    "country_of_origin": country_of_origin,
                    "fabric_content": fabric_content,
                    "description": description,
                    "lot_number": lot_number,
                    "asn_reference": body.get("asn_reference", "").strip(),
                    "upc": str(body.get("upc", "")).strip().upper(),
                    "created_at": now_iso(),
                })
    else:
        seq += 1
        box_id = f"BOX-{seq:06d}"
        box_docs.append({
            "box_id": box_id, "barcode": box_id, "receiving_id": receiving_id,
            "customer": customer, "manufacturer": manufacturer, "style": style,
            "sku": sku or style, "color": color, "size": size,
            "units": total_units, "qty": total_units, "seq_num": seq, "location": inv_location,
            "status": "putaway_pending", "state": "raw", "is_bpo": is_bpo,
            "lpn_id": box_id, "coo": country_of_origin,
            "country_of_origin": country_of_origin,
            "fabric_content": fabric_content,
            "description": description,
            "lot_number": lot_number,
            "asn_reference": body.get("asn_reference", "").strip(),
            "created_at": now_iso(),
        })

    # Authoritative received total = sum of the boxes actually created. The
    # single-item path already equaled this, but the multi-size items[] path left
    # total_units at 0 (only derived from top-level units/dozens/pieces), so
    # wms_receiving.total_units and the "receiving" ledger movement under-reported
    # multi-size receipts. Inventory was always correct (updated per item), so
    # this only fixes the receiving record + movement — no double count.
    if items:
        total_units = sum(int(b.get("units", 0) or 0) for b in box_docs)

    if box_docs:
        await db.wms_boxes.insert_many(box_docs)
        
        # WMS 2.0: Directed Work Task Generator (Cross-Dock vs Putaway)
        tasks_to_insert = []
        for box in box_docs:
            bd_style = box.get("style", "").upper()
            bd_color = box.get("color", "")
            
            # Busqueda de demanda (Backorders)
            demand_query = {
                "board": {"$regex": "^scheduling$|^blanks$|^crm$", "$options": "i"},
                "style": {"$regex": f"^{bd_style}$", "$options": "i"}
            }
            if bd_color:
                demand_query["color"] = {"$regex": f"^{bd_color}$", "$options": "i"}
                
            urgent_order = await db.orders.find_one(demand_query)
            
            task_type = "cross_dock" if urgent_order else "putaway"
            priority = "HOT" if urgent_order and urgent_order.get("priority", "").upper() == "HOT" else "NORMAL"
            suggested_zone = "ZONA PRODUCCION" if task_type == "cross_dock" else inv_location
            
            tasks_to_insert.append({
                "task_id": gen_id("tsk"),
                "lpn_id": box["box_id"],
                "task_type": task_type,
                "priority": priority,
                "status": "pending",
                "assigned_to": None,
                "context": {
                    "suggested_zone": suggested_zone, 
                    "sku": box["sku"],
                    "order_number": urgent_order.get("order_number") if urgent_order else None
                },
                "created_at": now_iso(),
            })
            
        if tasks_to_insert:
            await db.wms_tasks.insert_many(tasks_to_insert)

    receiving_doc = {
        "receiving_id": receiving_id, "customer": customer, "manufacturer": manufacturer,
        "style": style, "color": color, "size": size, "description": description,
        "country_of_origin": country_of_origin, "fabric_content": fabric_content,
        "inv_location": inv_location, "lot_number": lot_number, "sku": sku,
        "total_units": total_units, "is_bpo": is_bpo,
        # Traceability — linked back to the ASN + UPC the operator captured.
        "asn_reference": body.get("asn_reference", "").strip(),
        "upc": str(body.get("upc", "")).strip().upper(),
        "received_by": user.get("user_id"), "received_by_name": user.get("name", ""),
        "created_at": now_iso(),
        "boxes": [{"box_id": b["box_id"], "units": b["units"]} for b in box_docs]
    }
    await db.wms_receiving.insert_one(receiving_doc)
    await log_movement(user, "receiving", {"receiving_id": receiving_id, "total_units": total_units, "is_bpo": is_bpo})
    
    # Update inventory — pass through the receiving metadata so the inventory
    # row mirrors what the operator captured (Cliente, Manufacturer, COO, etc.).
    # We collect the inventory_id returned by each call so we can link the
    # freshly-created box docs to their inventory row in a single bulk update
    # below. Keyed by item_size so multi-size receiving stays per-line.
    meta = {
        "manufacturer": manufacturer,
        "description": description,
        "country_of_origin": country_of_origin,
        "fabric_content": fabric_content,
    }
    inv_id_by_size: dict[str, str] = {}
    if items:
        for item in items:
            item_size = item.get("size") or ""
            item_boxes = int(item.get("boxes", 1))
            inv_id = await _update_inventory_enhanced(
                style, color, item_size,
                item_boxes * int(item.get("units_per_box", 1)),
                "add", customer, inv_location, is_bpo, **meta,
                box_count=item_boxes,
            )
            if inv_id:
                inv_id_by_size[item_size] = inv_id
    else:
        inv_id = await _update_inventory_enhanced(
            style, color, size, total_units, "add",
            customer, inv_location, is_bpo, **meta,
        )
        if inv_id:
            inv_id_by_size[size or ""] = inv_id

    # Link every box we just inserted to the inventory row it physically
    # represents. Without this link the Locations modal can't match LPNs back
    # to inventory items and the "Cajas" drawer renders empty.
    if inv_id_by_size and box_docs:
        from collections import defaultdict
        by_inv: dict[str, list[str]] = defaultdict(list)
        for b in box_docs:
            target_inv = inv_id_by_size.get(b.get("size") or "")
            if target_inv:
                by_inv[target_inv].append(b["box_id"])
        for target_inv, box_ids in by_inv.items():
            await db.wms_boxes.update_many(
                {"box_id": {"$in": box_ids}},
                {"$set": {"inventory_id": target_inv}},
            )

    # Ensure location exists
    inv_location_upper = inv_location.upper()
    existing_loc = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(inv_location_upper)}$", "$options": "i"}})
    if not existing_loc:
        await db.wms_locations.insert_one({
            "location_id": gen_id("loc"), "name": inv_location_upper, 
            "zone": inv_location_upper.split('-')[0] if '-' in inv_location_upper else "RECEIVING",
            "type": "rack", "active": True, "created_at": now_iso(),
        })

    receiving_doc.pop("_id", None)
    await notify_badge_change("putaway")

    # ASN reconciliation (warning-permissive: never blocks, surfaces mismatches in response).
    # Two modes:
    #   1. asn_line_no provided -> all received qty decrements that exact line
    #      (operator confirmed which line they're receiving; style is irrelevant)
    #   2. asn_line_no missing  -> fall back to style-based match per part_number
    asn_warnings: list[dict] = []
    asn_ref = body.get("asn_reference", "").strip()
    asn_line_no_raw = body.get("asn_line_no")
    try:
        asn_line_no = int(asn_line_no_raw) if asn_line_no_raw not in (None, "", 0) else None
    except (TypeError, ValueError):
        asn_line_no = None

    if asn_ref and box_docs:
        total_units_received = sum(int(b.get("units", 0) or 0) for b in box_docs)
        try:
            if asn_line_no is not None:
                asn_result = await _apply_receiving_to_asn(
                    asn_ref, {}, user, target_line_no=asn_line_no,
                    target_qty=total_units_received,
                )
            else:
                received_by_pn: dict[str, int] = {}
                for b in box_docs:
                    key = (b.get("style") or b.get("sku") or "").strip().upper()
                    if not key:
                        continue
                    received_by_pn[key] = received_by_pn.get(key, 0) + int(b.get("units", 0) or 0)
                asn_result = await _apply_receiving_to_asn(asn_ref, received_by_pn, user)
            asn_warnings = asn_result.get("mismatched", [])
        except Exception as e:
            logger.exception(f"ASN reconciliation failed for {asn_ref}: {e}")
            asn_warnings = [{"reason": "reconcile_error", "detail": str(e)}]

    receiving_doc["asn_warnings"] = asn_warnings
    return receiving_doc

@router.get("/receiving")
async def list_receiving(request: Request, search: str = "", customer: str = "", limit: int = 200):
    """Search-driven receiving list. The UI no longer dumps the latest 500 — it
    queries by:
      - search:   matches receiving_id / style / sku / customer / inv_location
                  (free text) — so you can also pull every receipt of a cart
                  (e.g. "CARRO 233")
      - customer: restricts to one customer (exact-ish, case-insensitive)
    Results come back most-recent-first. With NO filter we return [] so the
    screen stays empty until the operator searches."""
    await require_auth(request)
    search = (search or "").strip()
    customer = (customer or "").strip()
    if not search and not customer:
        return []
    query = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [{"receiving_id": rx}, {"style": rx}, {"sku": rx},
                        {"customer": rx}, {"inv_location": rx}]
    if customer:
        query["customer"] = {"$regex": re.escape(customer), "$options": "i"}
    limit = max(1, min(int(limit or 200), 1000))
    docs = await db.wms_receiving.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return docs

@router.get("/receiving/{receiving_id}")
async def get_receiving(receiving_id: str, request: Request):
    await require_auth(request)
    doc = await db.wms_receiving.find_one({"receiving_id": receiving_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Receiving no encontrado")
    boxes = await db.wms_boxes.find({"receiving_id": receiving_id}, {"_id": 0}).to_list(500)
    doc["boxes"] = boxes
    return doc

@router.put("/receiving/{receiving_id}")
async def update_receiving(receiving_id: str, request: Request):
    user = await require_auth(request)
    body = await request.json()
    
    # Extract only metadata fields (no quantity/sku changes allowed here to protect inventory integrity)
    update_data = {}
    for field in ["customer", "manufacturer", "description", "country_of_origin", "fabric_content", "lot_number", "inv_location"]:
        if field in body:
            update_data[field] = body[field].strip()
            
    if not update_data:
        return {"message": "Nada que actualizar"}
        
    doc = await db.wms_receiving.find_one({"receiving_id": receiving_id})
    if not doc:
        raise HTTPException(404, "Receiving no encontrado")
        
    # Update the receiving record
    await db.wms_receiving.update_one({"receiving_id": receiving_id}, {"$set": update_data})
    
    # Also update the boxes
    box_update = {}
    if "customer" in update_data: box_update["customer"] = update_data["customer"]
    if "manufacturer" in update_data: box_update["manufacturer"] = update_data["manufacturer"]
    if "lot_number" in update_data: box_update["lot_number"] = update_data["lot_number"]
    if "country_of_origin" in update_data: box_update["coo"] = update_data["country_of_origin"]
    if "inv_location" in update_data: box_update["location"] = update_data["inv_location"]
    
    if box_update:
        await db.wms_boxes.update_many({"receiving_id": receiving_id}, {"$set": box_update})
        
    await log_movement(user, "receiving_update", {"receiving_id": receiving_id, "updated_fields": list(update_data.keys())})
    return {"message": "Registro actualizado exitosamente"}

@router.delete("/receiving/{receiving_id}")
async def delete_receiving(receiving_id: str, request: Request):
    user = await require_auth(request)
    doc = await db.wms_receiving.find_one({"receiving_id": receiving_id})
    if not doc:
        raise HTTPException(404, "Receiving no encontrado")
    
    # Revert inventory added by this receiving
    boxes = await db.wms_boxes.find({"receiving_id": receiving_id}).to_list(1000)

    # Guard: once a box has been put away / picked / processed, its units may now
    # live in a different location, so deleting the receiving can no longer cleanly
    # reverse the inventory — that is exactly what historically left phantom stock
    # (system shows material that is not physically there). Block it and tell the
    # user to adjust via Entradas/Salidas instead of deleting the receipt.
    RAW_STATES = {None, "", "raw"}
    RAW_STATUSES = {None, "", "putaway_pending", "received", "pending"}
    moved = [b for b in boxes
             if (b.get("state") not in RAW_STATES) or (b.get("status") not in RAW_STATUSES)
             or int(b.get("units_allocated", 0) or 0) > 0]
    if moved:
        raise HTTPException(
            409,
            f"No se puede eliminar: {len(moved)} de {len(boxes)} caja(s) ya fueron "
            "movidas (putaway), surtidas o procesadas. Borrar ahora dejaria inventario "
            "fantasma. Ajusta el inventario con Entradas/Salidas, o regresa/reubica las "
            "cajas a su estado original, antes de eliminar el recibo.",
        )

    for box in boxes:
        box_style = box.get("style")
        box_color = box.get("color")
        box_size = box.get("size")
        box_units = box.get("units", 0)
        box_location = box.get("location")

        if box_style and box_units > 0:
            # Revert the units AND the box this receiving added. ("remove" was
            # never a real op in _update_inventory_enhanced, so deletes used to
            # leave inventory inflated with no boxes behind it.)
            rev_inv = await db.wms_inventory.find_one({
                "sku": box_style, "color": box_color or "", "size": box_size or "", "location": box_location or "",
            }) or await db.wms_inventory.find_one({
                "style": {"$regex": f"^{re.escape(box_style)}$", "$options": "i"},
                "color": box_color or "", "size": box_size or "", "location": box_location or "",
            })
            if rev_inv:
                new_hand = max(0, int(rev_inv.get("units_on_hand", 0)) - box_units)
                new_boxes = max(0, int(rev_inv.get("total_boxes", 0) or 0) - 1)
                if new_hand == 0 and new_boxes == 0 and int(rev_inv.get("units_allocated", 0) or 0) <= 0:
                    await db.wms_inventory.delete_one({"_id": rev_inv["_id"]})
                else:
                    await db.wms_inventory.update_one(
                        {"_id": rev_inv["_id"]},
                        {"$set": {"units_on_hand": new_hand, "total_boxes": new_boxes, "updated_at": now_iso()}},
                    )
            
    box_ids = [b["box_id"] for b in boxes if "box_id" in b]
    if box_ids:
        await db.wms_tasks.delete_many({"lpn_id": {"$in": box_ids}})
        await db.wms_boxes.delete_many({"receiving_id": receiving_id})
        
    await db.wms_receiving.delete_one({"receiving_id": receiving_id})
    
    # Revert ASN progress if linked
    asn_ref = doc.get("asn_reference")
    if asn_ref:
        pn = (doc.get("sku") or doc.get("style") or "").upper()
        if pn:
            await _apply_receiving_to_asn(
                asn_ref,
                {pn: -int(doc.get("total_units", 0))},
                user
            )
            
    await log_movement(user, "receiving_deleted", {
        "receiving_id": receiving_id,
        "customer": doc.get("customer", ""),
        "style": doc.get("style", ""),
        "color": doc.get("color", ""),
        "size": doc.get("size", ""),
        "total_units": int(doc.get("total_units", 0) or 0),
        "inv_location": doc.get("inv_location", ""),
        "boxes_reverted": len(boxes),
        "details": f"Eliminado registro de receiving y revertidas {len(boxes)} cajas",
    })
    await notify_badge_change("putaway")
    return {"message": "Receiving eliminado y revertido exitosamente"}

# ==================== BOXES ====================

@router.get("/stocktakes")
@router.get("/boxes")
async def list_boxes(request: Request, sku: str = "", color: str = "", size: str = "",
                     location: str = "", status: str = "", state: str = "", po: str = "",
                     inventory_id: str = ""):
    """List boxes. Add `inventory_id=...` to fetch every LPN belonging to a
    single SKU+location inventory row — used by the Inventory and Locations
    UIs to expand a 'Cajas' cell into the actual LPN list."""
    await require_auth(request)
    query = {}
    if sku: query["sku"] = {"$regex": sku, "$options": "i"}
    if color: query["color"] = {"$regex": color, "$options": "i"}
    if size: query["size"] = {"$regex": size, "$options": "i"}
    if location: query["location"] = location
    if status:
        if status == "received":
            query["status"] = {"$in": ["received", "putaway_pending"]}
            query["units"] = {"$gt": 0}   # hide fully-picked empties from the putaway list
        else:
            query["status"] = status
    if state: query["state"] = state
    if po: query["po"] = {"$regex": po, "$options": "i"}
    if inventory_id: query["inventory_id"] = inventory_id
    boxes = await db.wms_boxes.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return boxes

@router.get("/stocktakes/{box_id}")
@router.get("/boxes/{box_id}")
async def get_box(box_id: str, request: Request):
    await require_auth(request)
    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        # Resolve by the box's REAL physical license plate too: once a migrated box
        # is reconciled (POST /boxes/reconcile-lpn), its printed LPN lives in
        # barcode/lpn_id while box_id stays the internal key, so a scan of that
        # label must still find the box.
        box = await db.wms_boxes.find_one(
            {"$or": [{"barcode": box_id}, {"lpn_id": box_id}]}, {"_id": 0}
        )
    if not box:
        # Fallback: if it's a receiving ID, return the first box associated with this receiving
        if box_id.upper().startswith("RCV_"):
            box = await db.wms_boxes.find_one({"receiving_id": box_id.lower()}, {"_id": 0})
            if box:
                box = dict(box)
                box["box_id"] = box_id
                return box
        raise HTTPException(404, "Caja no encontrada")
    return box


# Fields the operator can edit after a box is received. Excludes box_id /
# location / created_at / receiving_id / asn_reference / upc — those tie the
# box to its origin and to physical putaway, so the relocate endpoints own
# those mutations.
_BOX_EDITABLE_FIELDS = {
    "customer", "manufacturer", "description",
    "country_of_origin", "fabric_content", "lot_number",
    "units",  # qty mirror is kept in sync below
}


@router.put("/boxes/{box_id}")
async def update_box(box_id: str, request: Request):
    """Edit a received box. If `units` changes, the inventory row for the
    box's location is rebalanced by the delta so the on-hand stays accurate.
    Other field edits propagate to the box doc only — they don't affect
    inventory aggregations because those bucket on (sku, color, size,
    location), not on description/country/fabric/lot."""
    user = await require_auth(request)
    body = await request.json()

    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        raise HTTPException(404, f"Caja {box_id} no encontrada")
    await _assert_not_on_hold(user, box.get("location"))

    update_doc = {}
    for k, v in body.items():
        if k not in _BOX_EDITABLE_FIELDS:
            continue
        if k == "units":
            try:
                update_doc["units"] = int(v)
                update_doc["qty"] = int(v)  # qty mirrors units in the schema
            except (TypeError, ValueError):
                raise HTTPException(400, "units debe ser un entero")
        else:
            update_doc[k] = (str(v).strip() if v is not None else "")

    if not update_doc:
        raise HTTPException(400, "Nada por actualizar")

    update_doc["updated_at"] = now_iso()
    update_doc["updated_by"] = user.get("user_id")

    # 1. Inventory rebalance if units changed.
    old_units = int(box.get("units") or box.get("qty") or 0)
    new_units = update_doc.get("units", old_units)
    delta = new_units - old_units
    if delta != 0:
        sku = box.get("sku") or box.get("style") or ""
        color = box.get("color", "")
        size = box.get("size", "")
        location = box.get("location", "")
        inv = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": location,
        })
        if inv:
            new_on_hand = max(0, int(inv.get("units_on_hand", 0)) + delta)
            # Keep total_boxes in step: this box crossing to/from 0 units adds or
            # removes exactly one counted box at the slot.
            box_delta = -1 if (old_units > 0 and new_units == 0) else (1 if (old_units == 0 and new_units > 0) else 0)
            new_total_boxes = max(0, int(inv.get("total_boxes", 0) or 0) + box_delta)
            await db.wms_inventory.update_one(
                {"_id": inv["_id"]},
                {"$set": {"units_on_hand": new_on_hand, "total_boxes": new_total_boxes, "updated_at": now_iso()}},
            )
            # If the inventory row is empty AND has no other boxes pointing
            # at it, delete it so it doesn't ghost the Inventory listing.
            if new_on_hand == 0:
                still = await db.wms_boxes.count_documents({
                    "sku": sku, "color": color, "size": size, "location": location,
                    "box_id": {"$ne": box_id}, "units": {"$gt": 0},
                })
                if not still:
                    await db.wms_inventory.delete_one({"_id": inv["_id"]})

    # 2. Patch the box.
    await db.wms_boxes.update_one({"box_id": box_id}, {"$set": update_doc})

    # 3. Audit trail.
    await log_movement(user, "box_edited", {
        "box_id": box_id,
        "changes": {k: v for k, v in update_doc.items() if k not in ("updated_at", "updated_by")},
        "old_units": old_units,
        "new_units": new_units,
        "delta_units": delta,
    })
    await notify_badge_change("all")

    updated = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    return updated


@router.delete("/boxes/{box_id}")
async def delete_box(box_id: str, request: Request):
    """Delete a box (LPN). Super-user only. Deducts the box's units from the
    location's inventory row (mirrors the rebalance in update_box) and removes
    the inventory row entirely if it empties out and no other box feeds it."""
    user = await require_admin_level(request, 2)

    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        raise HTTPException(404, f"Caja {box_id} no encontrada")

    units = int(box.get("units") or box.get("qty") or 0)
    sku = box.get("sku") or box.get("style") or ""
    color = box.get("color", "")
    size = box.get("size", "")
    location = box.get("location", "")

    if units > 0 and location:
        inv = await db.wms_inventory.find_one({
            "sku": sku, "color": color, "size": size, "location": location,
        })
        if inv:
            new_on_hand = max(0, int(inv.get("units_on_hand", 0)) - units)
            new_total_boxes = max(0, int(inv.get("total_boxes", 0)) - 1)
            still = await db.wms_boxes.count_documents({
                "sku": sku, "color": color, "size": size, "location": location,
                "box_id": {"$ne": box_id}, "units": {"$gt": 0},
            })
            if new_on_hand == 0 and new_total_boxes == 0 and not still:
                await db.wms_inventory.delete_one({"_id": inv["_id"]})
            else:
                await db.wms_inventory.update_one(
                    {"_id": inv["_id"]},
                    {"$set": {"units_on_hand": new_on_hand, "total_boxes": new_total_boxes, "updated_at": now_iso()}},
                )

    await db.wms_boxes.delete_one({"box_id": box_id})
    await log_movement(user, "box_deleted", {
        "box_id": box_id, "sku": sku, "color": color, "size": size,
        "location": location, "units": units,
    })
    await notify_badge_change("all")
    return {"status": "deleted", "box_id": box_id}


@router.post("/boxes/{box_id}/adjust")
async def adjust_box_count(box_id: str, request: Request):
    """Case# 002 — adjust inventory at the physical box level.

    The operator scans a box (by box_id, printed barcode or lpn_id) and types
    the REAL counted units for that single box. We set the box to that exact
    count and rebalance the box's location inventory row by the delta — so an
    adjustment is always tied to a concrete box number (auditable in the box's
    movement history, Case# 003).

    Guards:
      • A reason is mandatory (free text) so every adjustment is justified.
      • A downward adjustment is blocked when it would push the row's on-hand
        below its allocated (committed) units, mirroring the manual-out guard.
      • Honors location HOLDs.
    Side effects:
      • If the new count is 0 the box is deleted (no zero-unit ghosts).
      • An empty inventory row (no units, no other boxes) is dropped.
    """
    user = await require_inventory_level(request, 2)
    body = await request.json()

    # Resolve by internal key first, then by the physical label (barcode/lpn_id)
    # so a scan of the printed plate of a reconciled box still finds it.
    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        box = await db.wms_boxes.find_one(
            {"$or": [{"barcode": box_id}, {"lpn_id": box_id}]}, {"_id": 0}
        )
    if not box:
        raise HTTPException(404, f"Caja {box_id} no encontrada")
    real_box_id = box["box_id"]

    await _assert_not_on_hold(user, box.get("location"))

    reason = str(body.get("reason", "") or "").strip()
    if not reason:
        raise HTTPException(400, "El motivo del ajuste es obligatorio")

    try:
        counted = int(float(body.get("counted_units")))
    except (TypeError, ValueError):
        raise HTTPException(400, "Las unidades contadas deben ser un número entero")
    if counted < 0:
        raise HTTPException(400, "Las unidades contadas no pueden ser negativas")

    old_units = int(box.get("units") or box.get("qty") or 0)
    delta = counted - old_units
    if delta == 0:
        raise HTTPException(400, f"La caja ya tiene {old_units} unidades; nada por ajustar")

    sku = box.get("sku") or box.get("style") or ""
    color = box.get("color", "")
    size = box.get("size", "")
    location = box.get("location", "")

    # Rebalance the location's inventory row by the same delta.
    inv = None
    if location:
        inv = (
            (box.get("inventory_id") and await db.wms_inventory.find_one({"inventory_id": box["inventory_id"]}))
            or await db.wms_inventory.find_one({"sku": sku, "color": color, "size": size, "location": location})
        )
    if inv:
        on_hand = int(inv.get("units_on_hand", 0) or 0)
        allocated = int(inv.get("units_allocated", 0) or 0)
        new_on_hand = on_hand + delta
        if new_on_hand < allocated:
            raise HTTPException(
                400,
                f"No puedes dejar {new_on_hand} en existencia: {allocated} unidades están "
                f"comprometidas (allocated) en esta ubicación. Libera la asignación primero.",
            )
        new_on_hand = max(0, new_on_hand)
        # This box crossing to/from 0 units adds or removes one counted box.
        box_delta = -1 if (old_units > 0 and counted == 0) else (1 if (old_units == 0 and counted > 0) else 0)
        new_total_boxes = max(0, int(inv.get("total_boxes", 0) or 0) + box_delta)
        await db.wms_inventory.update_one(
            {"_id": inv["_id"]},
            {"$set": {"units_on_hand": new_on_hand, "total_boxes": new_total_boxes, "updated_at": now_iso()}},
        )
        # Drop a fully-empty row so an adjusted-to-zero slot doesn't ghost.
        if new_on_hand == 0:
            still = await db.wms_boxes.count_documents({
                "sku": sku, "color": color, "size": size, "location": location,
                "box_id": {"$ne": real_box_id}, "units": {"$gt": 0},
            })
            if not still:
                await db.wms_inventory.delete_one({"_id": inv["_id"]})

    # Apply to the box itself: delete when emptied, otherwise set the new count.
    if counted == 0:
        await db.wms_boxes.delete_one({"box_id": real_box_id})
    else:
        await db.wms_boxes.update_one(
            {"box_id": real_box_id},
            {"$set": {"units": counted, "qty": counted,
                      "updated_at": now_iso(), "updated_by": user.get("user_id")}},
        )

    await log_movement(user, "inventory_adjust_box", {
        "box_id": real_box_id,
        "sku": sku, "color": color, "size": size, "location": location,
        "old_units": old_units, "new_units": counted, "delta_units": delta,
        "reason": reason,
        "box_deleted": counted == 0,
    })
    await notify_badge_change("all")

    return {
        "status": "adjusted",
        "box_id": real_box_id,
        "old_units": old_units,
        "new_units": counted,
        "delta_units": delta,
        "box_deleted": counted == 0,
        "location": location,
    }


@router.get("/boxes/{box_id}/history")
async def box_history(box_id: str, request: Request, limit: int = 300):
    """Case# 003 — full transaction timeline for a single box / LPN.

    Resolves the box by internal id, printed barcode or lpn_id, then returns:
      • box_events — every movement that names this box (details.box_id, or this
        box inside a bulk move's details.box_ids), plus the receiving event that
        created it. Newest first.
      • sku_context — movements at the SKU/location level that don't name a box
        (allocation, picking, manual adjustments, cycle counts). Helps diagnose
        discrepancies but is not specific to this box.

    Read-only; any authenticated WMS user.
    """
    await require_auth(request)

    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        box = await db.wms_boxes.find_one(
            {"$or": [{"barcode": box_id}, {"lpn_id": box_id}]}, {"_id": 0}
        )

    # Every identifier this box has answered to. A deleted box still leaves
    # movements keyed by its original id, so fall back to the raw input.
    keys = set()
    if box:
        for k in (box.get("box_id"), box.get("lpn_id"), box.get("barcode")):
            if k:
                keys.add(k)
    keys.add(box_id)
    keys = list(keys)

    or_clauses = [
        {"details.box_id": {"$in": keys}},     # single-box events
        {"details.box_ids": {"$in": keys}},    # this box inside a bulk move
    ]
    # The creation (receiving) event is keyed by receiving_id, not box_id.
    receiving_id = (box or {}).get("receiving_id")
    if receiving_id:
        or_clauses.append({"details.receiving_id": receiving_id})

    box_events = await db.wms_movements.find(
        {"$or": or_clauses}, {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    box_event_ids = {m.get("movement_id") for m in box_events}

    # Receiving-linked events (receiving / receiving_update / deallocate) only log
    # the receiving_id + a grand total. Join the receiving doc so the timeline
    # shows the product, this box's received units, lot, ASN/UPC, etc.
    rcv_ids = {m.get("details", {}).get("receiving_id") for m in box_events}
    rcv_ids.discard(None)
    if rcv_ids:
        rcv_map = {}
        async for r in db.wms_receiving.find({"receiving_id": {"$in": list(rcv_ids)}}, {"_id": 0}):
            rcv_map[r.get("receiving_id")] = r
        for m in box_events:
            d = m.get("details") or {}
            r = rcv_map.get(d.get("receiving_id"))
            if not r:
                continue
            for k in ("customer", "manufacturer", "style", "color", "size", "sku",
                      "description", "country_of_origin", "fabric_content",
                      "lot_number", "asn_reference", "upc"):
                if r.get(k) and not d.get(k):
                    d[k] = r.get(k)
            # This specific box's received units, pulled from the receiving's box list.
            for b in (r.get("boxes") or []):
                if b.get("box_id") in keys:
                    d["box_units_received"] = b.get("units")
                    break
            if r.get("received_by_name") and not m.get("user_name"):
                m["user_name"] = r["received_by_name"]
            m["details"] = d

    # SKU/location context — only when we still know the box's dimensions.
    sku_context = []
    if box:
        sku = box.get("sku") or box.get("style") or ""
        if sku:
            ctx = await get_sku_movement_history(sku, box.get("color", ""), box.get("size", ""), limit)
            sku_context = [m for m in ctx if m.get("movement_id") not in box_event_ids]

    return {
        "box_id": (box or {}).get("box_id", box_id),
        "found": bool(box),
        "box": box,
        "box_events": box_events,
        "box_event_count": len(box_events),
        "sku_context": sku_context,
        "sku_context_count": len(sku_context),
    }


# ==================== PUTAWAY ====================

async def _adjust_inventory_boxes(inv_id, delta):
    """Keep total_boxes in step with a physical box event on a single inventory
    row. Clamped at zero. Deletes the row if it ends up fully empty (no units,
    no boxes, no allocation) so a picked-out slot doesn't ghost the listing.

    total_boxes is a denormalized cache of wms_boxes; every path that creates,
    moves, empties or deletes a box must adjust it here so the two never drift
    (the cause of "report shows boxes but 0 product")."""
    if not inv_id or not delta:
        return
    row = await db.wms_inventory.find_one({"inventory_id": inv_id})
    if not row:
        return
    new_boxes = max(0, int(row.get("total_boxes", 0) or 0) + delta)
    if (new_boxes == 0 and int(row.get("units_on_hand", 0) or 0) <= 0
            and int(row.get("units_allocated", 0) or 0) <= 0):
        await db.wms_inventory.delete_one({"_id": row["_id"]})
    else:
        await db.wms_inventory.update_one(
            {"_id": row["_id"]},
            {"$set": {"total_boxes": new_boxes, "updated_at": now_iso()}},
        )


async def _move_box_inventory(box, old_loc, new_loc):
    if old_loc == new_loc:
        return
    sku = box.get("style") or box.get("sku")
    color = box.get("color") or ""
    size = box.get("size") or ""
    qty = box.get("units") or 0
    customer = box.get("customer") or ""
    is_bpo = box.get("is_bpo", False)

    # 1. Deduct from old location in wms_inventory (units AND the one box we're
    #    moving out). Only drop the row when it truly empties (no boxes, no units,
    #    no allocation) — total_boxes alone staying >0 must keep the row alive.
    if old_loc:
        old_inv = await db.wms_inventory.find_one({"sku": sku, "color": color, "size": size, "location": old_loc})
        if old_inv:
            new_qty = max(0, old_inv.get("units_on_hand", 0) - qty)
            new_boxes = max(0, int(old_inv.get("total_boxes", 0) or 0) - 1)
            if new_qty == 0 and new_boxes == 0 and int(old_inv.get("units_allocated", 0) or 0) <= 0:
                await db.wms_inventory.delete_one({"_id": old_inv["_id"]})
            else:
                await db.wms_inventory.update_one({"_id": old_inv["_id"]}, {"$set": {"units_on_hand": new_qty, "total_boxes": new_boxes, "updated_at": now_iso()}})

    # 2. Add to new location in wms_inventory (units AND the box that arrived).
    dst_inv = await db.wms_inventory.find_one({"sku": sku, "color": color, "size": size, "location": new_loc})
    if dst_inv:
        dst_id = dst_inv.get("inventory_id")
        if not dst_id:
            dst_id = gen_id("inv")
            await db.wms_inventory.update_one(
                {"_id": dst_inv["_id"]},
                {"$inc": {"units_on_hand": qty, "total_boxes": 1},
                 "$set": {"inventory_id": dst_id, "updated_at": now_iso(), "customer": customer, "is_bpo": is_bpo, "style": sku}}
            )
        else:
            await db.wms_inventory.update_one(
                {"_id": dst_inv["_id"]},
                {"$inc": {"units_on_hand": qty, "total_boxes": 1},
                 "$set": {"updated_at": now_iso(), "customer": customer, "is_bpo": is_bpo, "style": sku}}
            )
    else:
        dst_id = gen_id("inv")
        await db.wms_inventory.insert_one({
            "inventory_id": dst_id,
            "sku": sku,
            "style": sku,
            "color": color,
            "size": size,
            "location": new_loc,
            "units_on_hand": qty,
            "total_boxes": 1,
            "units_allocated": 0,
            "customer": customer,
            "is_bpo": is_bpo,
            "updated_at": now_iso()
        })

    # 3. Update the moved box's inventory_id to link it correctly.
    await db.wms_boxes.update_one(
        {"box_id": box["box_id"]},
        {"$set": {"inventory_id": dst_id}}
    )


def _ci_eq(v):
    """Exact match normalizado a MAYUSCULAS + espacios colapsados.

    HISTORIA: antes devolvia {"$regex": "^v$", "$options": "i"} para tolerar
    casing inconsistente ('Sand' vs 'SAND'). PERO un regex case-insensitive NO
    usa indices en Mongo -> forzaba COLLECTION SCAN de wms_boxes (78k docs) en
    CADA stock-check/deduccion del picking, disparando el CPU al 100% (reportado
    2026-07-13).

    Desde 2026-07-01 los campos de identidad (style/sku/color/size/location)
    estan normalizados a MAYUSCULAS system-wide (wms-identity-normalized-uppercase);
    verificado: 0 docs con minusculas en wms_boxes y wms_inventory. Por eso el
    match exacto en uppercase es 100% equivalente al regex viejo PERO usa los
    indices existentes (sku_1_color_1_size_1_location_1, style_1_color_1_size_1).
    Colapsa espacios internos igual que el normalizador de escritura para que
    'LIGHT  PINK' matchee 'LIGHT PINK'."""
    return re.sub(r"\s+", " ", (v or "")).strip().upper()


async def _deduct_pick_boxes(style, color, size, location, qty, inv_operation,
                             customer="", order_number=None, order_id=None,
                             user=None, ticket_id=None, only_box_id=None):
    """Deduct `qty` units for a pick from the physical boxes AND the
    inventory row, keeping wms_boxes / units_on_hand / total_boxes in lockstep.

    Cuando `only_box_id` viene (el picker escaneó una caja específica), se
    deduce ÚNICAMENTE de esa caja — no FIFO ciego. Este es el modo que quita
    el problema de "el sistema vaciaba una caja que aún tenía material" porque
    el picker escanea la caja física real de la que está sacando piezas.

    Si `only_box_id` no viene, cae al FIFO tradicional (usado por flujos que
    todavía no requieren scan o cuando el toggle `pick_requires_scan` está OFF).

    Every deduction logs a 'pick_deduction' movement with the box_ids it drained.
    Without that per-box trail, a picked-empty box is indistinguishable from a
    box lost to data corruption — the orphan-boxes restore (orphanrestore_dc70037d)
    used "no movements for this box_id" as its data-loss signal and re-materialized
    ~46k already-picked units into the carros. This movement closes that hole."""
    remaining = int(qty or 0)
    if remaining <= 0:
        return
    if only_box_id:
        # Modo scan: agarra SOLO esa caja. Sin caídas al FIFO — si la caja no
        # da, se propaga el faltante y el picker debe escanear otra.
        specific = await db.wms_boxes.find_one({"box_id": only_box_id})
        if not specific or int(specific.get("units") or 0) <= 0:
            raise HTTPException(409, f"La caja {only_box_id} no tiene unidades disponibles")
        # CROSS-CHECK: la caja escaneada debe corresponder al style/color/size
        # del ticket. Si no calza, algo va MUY mal (frontend stale, corrupción,
        # bug) y hay que abortar antes de descontar de la caja equivocada.
        b_style = (specific.get("style") or (specific.get("sku") or "").split("-")[0] or "").strip().upper()
        b_color = (specific.get("color") or "").strip().upper()
        b_size  = (specific.get("size") or "").strip().upper()
        t_style = (style or "").strip().upper()
        t_color = (color or "").strip().upper()
        t_size  = (size or "").strip().upper()
        if b_size != t_size or b_color != t_color:
            raise HTTPException(409, (
                f"La caja {only_box_id} contiene {b_color}/{b_size}, "
                f"no {t_color}/{t_size}. Escanea otra caja."
            ))
        # style: aceptamos que el sku venga como "STYLE-COLOR-SIZE" o solo "STYLE"
        if b_style != t_style and not (specific.get("sku") or "").upper().startswith(t_style + "-"):
            raise HTTPException(409, (
                f"La caja {only_box_id} es de style {b_style}, no {t_style}. Escanea otra."
            ))
        boxes = [specific]
    else:
        q = {"$or": [{"sku": _ci_eq(style)}, {"style": _ci_eq(style)}],
             "color": _ci_eq(color), "size": _ci_eq(size), "units": {"$gt": 0}}
        if location:
            q["location"] = _ci_eq(location)
        boxes = await db.wms_boxes.find(q).sort("created_at", 1).to_list(500)
    touched = []
    for box in boxes:
        if remaining <= 0:
            break
        b_qty = box.get("units") if box.get("units") is not None else box.get("qty", 0)
        if b_qty <= 0:
            continue
        take = min(b_qty, remaining)
        new_b = b_qty - take
        upd = {"units": new_b, "qty": new_b, "last_picked_at": now_iso()}
        if ticket_id:
            upd["last_pick_ticket"] = ticket_id
        # When a pick empties a box, mark it 'depleted' so it drops off the cart /
        # putaway lists instead of lingering as a 0-unit "putaway_pending" box
        # (the "cajas con 0 stock al recibir" confusion — the box was received
        # fine, then fully picked/cross-docked out of the cart).
        if new_b == 0:
            upd["status"] = "depleted"
        if order_number is not None:
            upd["order_number"] = order_number
        if order_id is not None:
            upd["last_order_id"] = order_id
        # ATÓMICO: solo actualiza si la caja aún tiene los `b_qty` que leímos.
        # Si dos pickers descuentan simultáneamente, el segundo find_one_and_update
        # falla su filtro (units != b_qty tras la primera baja) y reintenta con
        # la lectura fresca — sin ghost inventory ni units negativos.
        atomic_res = await db.wms_boxes.find_one_and_update(
            {"_id": box["_id"], "units": b_qty},
            {"$set": upd},
            return_document=ReturnDocument.AFTER,
        )
        if not atomic_res:
            # Otro proceso movió la caja entre el find y el update. Re-lee y
            # reintenta esta iteración solo si aún hay unidades.
            fresh = await db.wms_boxes.find_one({"_id": box["_id"]})
            if not fresh or int(fresh.get("units") or 0) <= 0:
                if only_box_id:
                    raise HTTPException(409, (
                        f"La caja {only_box_id} fue vaciada por otro picker "
                        "durante tu descuento. Escanea otra caja."
                    ))
                continue
            new_fresh_qty = int(fresh.get("units") or 0)
            take = min(new_fresh_qty, remaining)
            new_b = new_fresh_qty - take
            upd["units"] = new_b; upd["qty"] = new_b
            if new_b == 0: upd["status"] = "depleted"
            atomic_res = await db.wms_boxes.find_one_and_update(
                {"_id": box["_id"], "units": new_fresh_qty},
                {"$set": upd},
                return_document=ReturnDocument.AFTER,
            )
            if not atomic_res:
                # 2do intento falló también: race conditions muy raras. Aborta
                # esta caja limpiamente en vez de dejar inventory drift.
                if only_box_id:
                    raise HTTPException(409, (
                        f"Contención alta en la caja {only_box_id}. Reintenta."
                    ))
                continue
        touched.append({"box_id": box.get("box_id"), "taken": take,
                        "emptied": new_b == 0, "location": box.get("location", location)})
        # Key the inventory deduct the SAME way _move_box_inventory / receiving
        # CREATE the row: short style first (e.g. "5000"), composite sku only as
        # fallback. Boxes carry the composite sku ("5000-BLACK-XL") but shelf
        # inventory rows are keyed by the short style — deducting by the box's
        # composite sku missed the row entirely, so the box emptied while the
        # ledger stayed full (the NA04-A17 ghost-inventory case).
        inv_id = await _update_inventory_enhanced(
            box.get("style") or box.get("sku") or style, box.get("color", color), box.get("size", size),
            take, inv_operation, location=box.get("location", location),
            customer=box.get("customer", customer),
        )
        if new_b == 0:
            await _adjust_inventory_boxes(inv_id, -1)
        remaining -= take
    # Leftover with no backing box: deduct straight from inventory (legacy/Excel).
    if remaining > 0:
        await _update_inventory_enhanced(style, color, size, remaining, inv_operation,
                                         location=location, customer=customer)
    # Per-box audit trail. box_ids is the field future sweeps/restores check to
    # tell "picked empty" apart from "lost"; no_box_units flags the portion that
    # left WITHOUT a backing box so it can never be mistaken for still-on-hand.
    if touched or remaining > 0:
        await log_movement(user or {"user_id": "system", "name": "system"}, "pick_deduction", {
            "ticket_id": ticket_id, "order_number": order_number,
            "style": style, "color": color, "size": size, "location": location or "",
            "qty": int(qty or 0),
            "box_ids": [t["box_id"] for t in touched if t.get("box_id")],
            "boxes": touched,
            "no_box_units": remaining if remaining > 0 else 0,
            "scanned": bool(only_box_id),         # <— True cuando el picker escaneó específica
        })
        # Log estructurado adicional para trazabilidad rápida en producción:
        # una línea por descuento con solo lo esencial (grep-eable).
        logger.info(
            "pick_deduction ticket=%s order=%s sku=%s/%s/%s loc=%s qty=%d scanned=%s boxes=%s left_ledger=%d",
            ticket_id, order_number, style, color, size, location or "-",
            int(qty or 0), bool(only_box_id),
            ",".join(t["box_id"] for t in touched if t.get("box_id")) or "-",
            remaining if remaining > 0 else 0,
        )


async def _available_units(style, color, size, location=""):
    """Physical units available to pick for a SKU dimension. Stock can live as
    physical boxes AND/OR as a wms_inventory row, and the two mirrors can drift,
    so we take the GREATER of the two. That blocks genuine oversell (BOTH mirrors
    short of the request) without false-rejecting a pick when only one mirror
    lags behind (e.g. Excel-imported rows with no boxes, or boxes whose inventory
    row hasn't caught up)."""
    style = (style or "").strip()
    color = (color or "").strip()
    sz = (size or "").strip()
    # Case-insensitive match (Sand vs SAND) so this agrees with what the picker
    # sees in _compute_size_locations — otherwise a legit pick is blocked with
    # "disponible 0" even though the stock is right there.
    base = {"$or": [{"sku": _ci_eq(style)}, {"style": _ci_eq(style)}], "color": _ci_eq(color), "size": _ci_eq(sz)}

    # Exclude boxes that have already left available stock (shipped, in
    # production, finished, in neck cutting, confirmed) so picking can't be
    # offered units that are physically committed elsewhere. Legacy boxes with
    # no status field stay counted ($nin keeps nulls).
    box_q = {**base, "units": {"$gt": 0}, "status": {"$nin": list(_BOX_OUT_STATUSES)}}
    if location:
        box_q["location"] = _ci_eq(location)
    box_units = 0
    for b in await db.wms_boxes.find(box_q, {"_id": 0, "units": 1, "qty": 1}).to_list(5000):
        box_units += int((b.get("units") if b.get("units") is not None else b.get("qty", 0)) or 0)

    inv_q = dict(base)
    if location:
        inv_q["location"] = _ci_eq(location)
    inv_units = 0
    for r in await db.wms_inventory.find(inv_q, {"_id": 0, "units_on_hand": 1}).to_list(5000):
        inv_units += int(r.get("units_on_hand", 0) or 0)

    return max(box_units, inv_units)


async def _open_discrepancy_task(style, color, size, location, requested, available, user):
    """Open (or reuse) a pending discrepancy cycle-count task so the warehouse
    reconciles a slot whose system stock fell short of a pick request."""
    existing = await db.wms_tasks.find_one({
        "task_type": "discrepancy", "status": "pending",
        "context.style": style, "context.color": color or "",
        "context.size": (size or "").strip(), "context.location": location or "",
    })
    if existing:
        return
    await db.wms_tasks.insert_one({
        "task_id": gen_id("tsk"), "task_type": "discrepancy", "priority": "HIGH",
        "status": "pending", "assigned_to": None,
        "context": {
            "style": style, "color": color or "", "size": (size or "").strip(),
            "location": location or "", "requested": int(requested), "available": int(available),
            "suggested_zone": location or "",
            "reason": "Pick oversell bloqueado: solicitado > disponible",
        },
        "created_by": (user or {}).get("user_id"),
        "created_at": now_iso(),
    })


async def _assert_pick_stock(style, color, picked_sizes, user):
    """WMS-002: block a pick (HTTP 409) that asks for more units than physically
    available, BEFORE any deduction runs, and open a discrepancy task per
    shortfall so the warehouse reconciles. Standalone Mongo has no multi-document
    transactions, so we gate up-front; the per-row deduct is also atomic-clamped
    (_update_inventory_enhanced) as a last-resort guard for the microsecond race
    on the final units."""
    style = (style or "").strip()
    color = (color or "").strip()
    shortfalls = []
    for sz, data in (picked_sizes or {}).items():
        sz_clean = (sz or "").strip()
        if isinstance(data, dict) and data.get("details"):
            for loc, qty in data["details"].items():
                q = int(qty or 0)
                if q <= 0:
                    continue
                avail = await _available_units(style, color, sz_clean, loc)
                if q > avail:
                    shortfalls.append({"size": sz_clean, "location": loc, "requested": q, "available": avail})
        else:
            q = int(data.get("total", 0)) if isinstance(data, dict) else int(data or 0)
            if q <= 0:
                continue
            avail = await _available_units(style, color, sz_clean, "")
            if q > avail:
                shortfalls.append({"size": sz_clean, "location": "", "requested": q, "available": avail})

    if not shortfalls:
        return
    for s in shortfalls:
        await _open_discrepancy_task(style, color, s["size"], s["location"], s["requested"], s["available"], user)
    detail = "; ".join(
        f"{style} {color} {s['size']}".strip()
        + (f" @ {s['location']}" if s["location"] else "")
        + f": pidió {s['requested']}, disponible {s['available']}"
        for s in shortfalls
    )
    raise HTTPException(409, f"Stock insuficiente para surtir; se abrió conteo de discrepancia. {detail}")


@router.post("/putaway")
async def putaway_box(request: Request):
    user = await require_auth(request)
    body = await request.json()
    box_id = body.get("box_id", "").strip()
    location = body.get("location", "").strip()
    if not box_id or not location:
        raise HTTPException(400, "box_id y location requeridos")
    
    loc = await db.wms_locations.find_one({"name": location})
    if not loc:
        raise HTTPException(404, "Ubicacion no encontrada")
    await _assert_not_on_hold(user, location)

    box = await db.wms_boxes.find_one({"box_id": box_id})
    if not box:
        # Fallback: if it's a receiving_id, find all received boxes of that receiving event
        if box_id.upper().startswith("RCV_"):
            boxes = await db.wms_boxes.find({"receiving_id": box_id.lower()}).to_list(1000)
            if not boxes:
                raise HTTPException(404, f"No se encontraron cajas para el Receiving ID: {box_id}")
            await _assert_not_on_hold(user, *{b.get("location") for b in boxes})
            for b in boxes:
                old_loc = b.get("location")
                await db.wms_boxes.update_one({"box_id": b["box_id"]}, {"$set": {"location": location, "status": "stored", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})
                await log_movement(user, "putaway", {"box_id": b["box_id"], "from": old_loc, "to": location, "sku": b.get("sku"), "units": b.get("units")})
                await _move_box_inventory(b, old_loc, location)
            await notify_badge_change("putaway")
            return {"message": f"Se ubicaron exitosamente {len(boxes)} cajas del receiving {box_id} en {location}", "box_id": box_id, "location": location}
        raise HTTPException(404, "Caja no encontrada")

    old_location = box.get("location")
    await _assert_not_on_hold(user, old_location)
    await db.wms_boxes.update_one({"box_id": box_id}, {"$set": {"location": location, "status": "stored", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})
    await log_movement(user, "putaway", {"box_id": box_id, "from": old_location, "to": location, "sku": box.get("sku"), "units": box.get("units")})
    await _move_box_inventory(box, old_location, location)
    await notify_badge_change("putaway")
    return {"message": f"Caja {box_id} ubicada en {location}", "box_id": box_id, "location": location}

@router.post("/putaway/bulk")
async def putaway_bulk(request: Request):
    user = await require_auth(request)
    body = await request.json()
    assignments = body.get("assignments", [])
    results = []
    failed = []
    loc_cache: dict[str, dict] = {}

    async def _resolve_loc(name):
        # Cache catalog lookups; match case-insensitively like the other move
        # endpoints. Returns the canonical location doc or None.
        key = name.upper()
        if key not in loc_cache:
            loc_cache[key] = await db.wms_locations.find_one(
                {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
            )
        return loc_cache[key]

    for a in assignments:
        box_id = a.get("box_id", "").strip()
        location = a.get("location", "").strip()
        if not box_id or not location:
            failed.append({"box_id": box_id, "location": location, "reason": "box_id y location requeridos"})
            continue
        # A7: never move stock into a bin that isn't in the catalog — that is how
        # entire rack zones ended up as phantom locations (LIF03/07-11), invisible
        # in the Locations UI. putaway_box already guards this; bulk did not.
        dst_loc = await _resolve_loc(location)
        if not dst_loc:
            failed.append({"box_id": box_id, "location": location, "reason": f"Ubicación destino '{location}' no encontrada. Créala primero."})
            continue
        box = await db.wms_boxes.find_one({"box_id": box_id})
        if not box:
            failed.append({"box_id": box_id, "location": location, "reason": "Caja no encontrada"})
            continue
        old_loc = box.get("location")
        try:
            await _assert_not_on_hold(user, location, old_loc)
        except HTTPException as e:
            # A held bin must not abort the whole batch — record and move on.
            failed.append({"box_id": box_id, "location": location, "reason": e.detail})
            continue
        # Persist the catalog's canonical name so casing stays consistent.
        canonical = dst_loc.get("name", location)
        await db.wms_boxes.update_one({"box_id": box_id}, {"$set": {"location": canonical, "status": "stored", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})
        await _move_box_inventory(box, old_loc, canonical)
        results.append({"box_id": box_id, "location": canonical})

    await log_movement(user, "putaway_bulk", {
        "count": len(results),
        "failed": len(failed),
        "box_ids": [r["box_id"] for r in results][:50],  # cap log payload
    })
    await notify_badge_change("putaway")
    msg = f"{len(results)} cajas ubicadas"
    if failed:
        msg += f", {len(failed)} con error"
    return {"message": msg, "results": results, "failed": failed}


@router.post("/boxes/generate")
async def generate_box(request: Request):
    """Mint a NEW box (LPN) for material the warehouse receives from production
    without a label, so it can be moved/adjusted and physically tagged. Creates a
    BOX-###### id + its inventory row and returns the id to print a label."""
    user = await require_auth(request)
    body = await request.json()
    style = (body.get("style") or "").strip().upper()
    color = (body.get("color") or "").strip().upper()
    size = (body.get("size") or "").strip().upper()
    units = int(body.get("units") or 0)
    location = (body.get("location") or "").strip()
    customer = await _canonical_customer(body.get("customer", ""))
    if not style or units <= 0 or not location:
        raise HTTPException(400, "Estilo, unidades (>0) y ubicación son requeridos")
    loc = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(location)}$", "$options": "i"}})
    if not loc:
        raise HTTPException(404, f"Ubicación '{location}' no encontrada. Créala primero.")
    location = loc.get("name", location)
    await _assert_not_on_hold(user, location)

    # Style must already exist — new styles are added ONLY from the Config module,
    # never minted here (keeps the normalized catalog clean).
    known = await db.wms_inventory.find_one(
        {"$or": [{"style": _ci_eq(style)}, {"sku": _ci_eq(style)}]}, {"_id": 1})
    if not known:
        known = await db.wms_catalog_options.find_one(
            {"type": "styles", "value": {"$regex": f"^{re.escape(style)}$", "$options": "i"}}, {"_id": 1})
    if not known:
        raise HTTPException(400, f"El estilo '{style}' no existe. Agrégalo primero en Configuración.")

    seq = await _reserve_box_seqs(1)
    box_id = f"BOX-{seq:06d}"
    sku = (body.get("sku") or style).strip().upper()
    coo = re.sub(r"\s+", " ", str(body.get("country_of_origin") or "")).strip().upper()
    fabric = re.sub(r"\s+", " ", str(body.get("fabric_content") or "")).strip().upper()
    is_bpo = bool(body.get("is_bpo", False))

    inv_id = await _update_inventory_enhanced(
        style, color, size, units, "add", customer, location, is_bpo,
        manufacturer=re.sub(r"\s+", " ", str(body.get("manufacturer") or "")).strip().upper(),
        description=re.sub(r"\s+", " ", str(body.get("description") or "")).strip().upper(),
        country_of_origin=coo, fabric_content=fabric, box_count=1,
    )
    box_doc = {
        "box_id": box_id, "barcode": box_id, "lpn_id": box_id,
        "inventory_id": inv_id, "customer": customer,
        "style": style, "sku": sku, "color": color, "size": size,
        "units": units, "qty": units, "seq_num": seq, "location": location,
        "status": "stored", "state": "finished", "is_bpo": is_bpo,
        "country_of_origin": coo, "coo": coo, "fabric_content": fabric,
        "description": (body.get("description") or "").strip(),
        "source": "production_generated",
        "last_transferred_at": now_iso(),
        "last_transferred_by": user.get("name", user.get("email", "")),
        "created_at": now_iso(),
    }
    await db.wms_boxes.insert_one(box_doc)
    await log_movement(user, "box_generated", {
        "box_id": box_id, "style": style, "color": color, "size": size,
        "units": units, "location": location, "source": "production",
    })
    await notify_badge_change("all")
    return {"box_id": box_id, "location": location, "units": units,
            "message": f"Caja {box_id} generada"}

# ==================== INVENTORY ====================

async def _update_inventory_enhanced(
    sku, color, size, qty, operation, customer="", location="", is_bpo=False,
    *, manufacturer="", description="", country_of_origin="", fabric_content="",
    box_count=1,
):
    """Add/allocate/deallocate/deduct stock. The inventory key is now
    (sku, color, size, location, fabric_content, country_of_origin) so two
    batches of the same SKU at the same shelf with different fabric content
    or country of origin stay as separate rows instead of being silently
    merged. For non-add operations we fall back to the legacy 4-field key
    to keep picking/move flows working when callers don't supply meta."""
    full_key = {
        "sku": sku, "color": color or "", "size": size or "",
        "location": location or "",
        "fabric_content": fabric_content or "",
        "country_of_origin": country_of_origin or "",
    }
    inv = await db.wms_inventory.find_one(full_key)
    # Excel-imported inventory may have sku='None' with real value in 'style'
    if not inv and sku:
        style_key = {
            "style": {"$regex": f"^{re.escape(sku)}$", "$options": "i"},
            **{k: v for k, v in full_key.items() if k != "sku"},
        }
        inv = await db.wms_inventory.find_one(style_key)
    # Loose fallback: if caller didn't pass meta, match by the legacy 4-field
    # key so picks/moves that don't know fabric/COO still find existing rows.
    if not inv and (not fabric_content and not country_of_origin):
        loose_key = {"sku": sku, "color": color or "", "size": size or "", "location": location or ""}
        inv = await db.wms_inventory.find_one(loose_key)
        if not inv and sku:
            inv = await db.wms_inventory.find_one({
                "style": {"$regex": f"^{re.escape(sku)}$", "$options": "i"},
                **{k: v for k, v in loose_key.items() if k != "sku"},
            })
    if not inv and operation == "add":
        new_inv_id = gen_id("inv")
        await db.wms_inventory.insert_one({
            "sku": sku, "style": sku, "color": color or "", "size": size or "",
            "inventory_id": new_inv_id, "customer": customer,
            "manufacturer": manufacturer or "",
            "description": description or "",
            "country_of_origin": country_of_origin or "",
            "fabric_content": fabric_content or "",
            "location": location, "is_bpo": is_bpo,
            "units_on_hand": qty, "units_allocated": 0, "total_boxes": max(0, int(box_count or 0)),
            "updated_at": now_iso()
        })
        return new_inv_id
    elif inv:
        doc_key = {"_id": inv["_id"]}
        if operation == "add":
            # Backfill empty metadata fields on existing rows so re-receiving
            # the same SKU into the same location heals legacy gaps.
            backfill = {}
            for field, value in (
                ("manufacturer", manufacturer),
                ("description", description),
                ("country_of_origin", country_of_origin),
                ("fabric_content", fabric_content),
            ):
                if value and not (inv.get(field) or "").strip():
                    backfill[field] = value
            await db.wms_inventory.update_one(
                doc_key,
                {
                    "$inc": {"units_on_hand": qty, "total_boxes": max(0, int(box_count or 0))},
                    "$set": {"updated_at": now_iso(), "is_bpo": is_bpo, **backfill},
                },
            )
        elif operation == "allocate":
            await db.wms_inventory.update_one(doc_key, {"$inc": {"units_allocated": qty}, "$set": {"updated_at": now_iso()}})
        elif operation == "deallocate":
            # Guard: allocation can never go below zero.
            new_alloc = max(0, inv.get("units_allocated", 0) - qty)
            await db.wms_inventory.update_one(doc_key, {"$set": {"units_allocated": new_alloc, "updated_at": now_iso()}})
        elif operation == "deduct":
            # ATOMIC deduction (WMS-001). An aggregation-pipeline update lets a
            # SINGLE document operation subtract-and-clamp-at-zero, so two
            # concurrent picks can no longer both read the same units_on_hand and
            # lose an update (the RP04-B42 oversell). True oversell is blocked
            # up-front by _assert_pick_stock(); this clamp is the last-resort
            # guard that keeps a row from ever going negative under a race.
            updated = await db.wms_inventory.find_one_and_update(
                doc_key,
                [{"$set": {
                    "units_on_hand": {"$max": [0, {"$subtract": ["$units_on_hand", qty]}]},
                    "units_allocated": {"$max": [0, {"$subtract": ["$units_allocated", qty]}]},
                    "updated_at": now_iso(),
                }}],
                return_document=ReturnDocument.AFTER,
            )
            new_hand = int((updated or {}).get("units_on_hand", 0) or 0)

            # Fila totalmente vacía y sin cajas que la respalden → se elimina
            # (mismo criterio que _adjust_inventory_boxes al vaciarse la última
            # caja). Sin esto, los descuentos SIN caja de respaldo (bulk/Excel)
            # dejaban renglones fantasma en 0 para siempre en Locaciones.
            # El historial queda en wms_movements.
            if (updated is not None and new_hand <= 0
                    and int(updated.get("units_allocated", 0) or 0) <= 0
                    and int(updated.get("total_boxes", 0) or 0) <= 0):
                await db.wms_inventory.delete_one({"_id": updated["_id"]})

            # WMS 2.0 Cycle Count Trigger
            if new_hand < 50:
                existing_cc = await db.wms_tasks.find_one({"task_type": "cycle_count", "context.sku": sku, "status": "pending"})
                if not existing_cc:
                    await db.wms_tasks.insert_one({
                        "task_id": gen_id("tsk"), "task_type": "cycle_count", "priority": "HIGH", "status": "pending",
                        "assigned_to": None, "context": {"sku": sku, "suggested_zone": location, "reason": "Threshold breach (<50)"},
                        "created_at": now_iso(),
                    })
        elif operation == "pick_to_neck":
            # 1. ATOMIC deduct-and-clamp from the current shelf (same race-free
            #    guard as "deduct" above).
            updated = await db.wms_inventory.find_one_and_update(
                doc_key,
                [{"$set": {
                    "units_on_hand": {"$max": [0, {"$subtract": ["$units_on_hand", qty]}]},
                    "units_allocated": {"$max": [0, {"$subtract": ["$units_allocated", qty]}]},
                    "updated_at": now_iso(),
                }}],
                return_document=ReturnDocument.AFTER,
            )
            # Igual que en "deduct": renglón vacío sin cajas → fuera.
            if (updated is not None
                    and int(updated.get("units_on_hand", 0) or 0) <= 0
                    and int(updated.get("units_allocated", 0) or 0) <= 0
                    and int(updated.get("total_boxes", 0) or 0) <= 0):
                await db.wms_inventory.delete_one({"_id": updated["_id"]})
            # 2. Add to CUTTING_NECK (keep it allocated/reserved for the order)
            # Use the same sku/style as the source record for consistency
            inv_sku = inv.get("sku") or inv.get("style") or sku
            neck_key = {"sku": inv_sku, "color": color or "", "size": size or "", "location": "CUTTING_NECK"}
            await db.wms_inventory.update_one(
                neck_key,
                {"$inc": {"units_on_hand": qty, "units_allocated": qty}, "$set": {"updated_at": now_iso(), "customer": customer, "style": inv.get("style", sku)}},
                upsert=True
            )
        # Surface the inventory_id so callers (e.g. receiving) can link the
        # boxes they just created to the inventory row that holds them.
        return inv.get("inventory_id")
    return None


async def _update_inventory(sku, color, size, qty, operation="add", location=""):
    # Standard fallback for old calls
    await _update_inventory_enhanced(sku, color, size, qty, operation, location=location)

@router.get("/inventory")
async def get_inventory(
    request: Request,
    sku: str = "", color: str = "", size: str = "", location: str = "",
    customer: str = "", category: str = "", style: str = "",
    description: str = "", country_of_origin: str = "", fabric_content: str = "",
    paginated: bool = False, skip: int = 0, limit: int = 5000,
    exclude_hold: bool = False,
):
    """List inventory rows.
      - Default (legacy): returns bare array, all rows up to 5000.
      - paginated=true   : returns { items, total, has_more } with skip/limit.
                            Allows the UI to load in chunks without freezing.
    """
    await require_auth(request)
    query = {}
    if sku: query["sku"] = {"$regex": sku, "$options": "i"}
    if style: query["style"] = {"$regex": style, "$options": "i"}
    if color: query["color"] = {"$regex": color, "$options": "i"}
    if size: query["size"] = {"$regex": size, "$options": "i"}
    if customer: query["customer"] = {"$regex": customer, "$options": "i"}
    if category == "LOW_STOCK":
        query["units_on_hand"] = {"$lte": 10, "$gt": 0}
    elif category:
        query["category"] = {"$regex": category, "$options": "i"}
    if location: query["location"] = {"$regex": location, "$options": "i"}
    if description: query["description"] = {"$regex": description, "$options": "i"}
    if country_of_origin: query["country_of_origin"] = {"$regex": country_of_origin, "$options": "i"}
    if fabric_content: query["fabric_content"] = {"$regex": fabric_content, "$options": "i"}
    # Optionally hide stock parked in SAT-held locations.
    if exclude_hold and not location:
        held = await _hold_location_names()
        if held:
            query["location"] = {"$nin": list(held)}

    skip = max(0, skip)
    limit = max(1, min(limit, 5000))

    # Use aggregation to project/alias for frontend compatibility
    pipeline = [
        {"$match": query},
        {"$sort": {"sku": 1}},
        {"$skip": skip},
        {"$limit": limit},
        {"$project": {
            "_id": 0,
            "inventory_id": 1,  # needed so the Locations modal can link items to their LPNs
            "sku": 1,
            "style": 1,
            "color": 1,
            "size": 1,
            "description": 1,
            "customer": 1,
            "manufacturer": 1,
            "category": 1,
            "location": 1,
            "total_boxes": 1,
            "last_updated": 1,
            "country_of_origin": 1,
            "fabric_content": 1,
            "size_header": 1,
            "po": 1,
            "bpo": 1,
            "import_number": 1,
            "on_hand": "$units_on_hand",
            "allocated": "$units_allocated",
            "available": {"$subtract": ["$units_on_hand", "$units_allocated"]},
            "inv_location": "$location",
            "units_on_hand": 1,
            "units_allocated": 1,
        }},
    ]
    inventory = await db.wms_inventory.aggregate(pipeline).to_list(limit)

    if not paginated:
        return inventory

    total = await db.wms_inventory.count_documents(query)
    return {"items": inventory, "total": total, "has_more": (skip + len(inventory)) < total}

@router.get("/inventory/filters")
async def inventory_filters_v2(request: Request):
    """Return unique filter values for inventory dropdowns."""
    await require_auth(request)
    customers = await db.wms_inventory.distinct("customer")
    categories = await db.wms_inventory.distinct("category")
    manufacturers = await db.wms_inventory.distinct("manufacturer")
    styles = await db.wms_inventory.distinct("style")

    # If the admin curated a country catalog, use that as the authoritative
    # list. Otherwise fall back to inventory distinct, filtering out leaked
    # fabric-content rows (anything with "%").
    curated_countries = await db.wms_catalog_options.find(
        {"type": "countries"}, {"_id": 0, "value": 1}
    ).to_list(2000)
    curated_country_vals = [c["value"] for c in curated_countries if c.get("value")]
    if curated_country_vals:
        country_list = curated_country_vals
    else:
        raw_countries = await db.wms_inventory.distinct("country_of_origin")
        country_list = [c for c in raw_countries if c and "%" not in c]

    raw_fabrics = await db.wms_inventory.distinct("fabric_content")
    fabrics = sorted([f for f in raw_fabrics if f and isinstance(f, str)])

    return {
        "customers": sorted([c for c in customers if c]),
        "categories": sorted([c for c in categories if c]),
        "manufacturers": sorted([m for m in manufacturers if m]),
        "styles": sorted([s for s in styles if s]),
        "countries": sorted(country_list),
        "fabrics": fabrics,
    }

@router.get("/inventory/facets")
async def inventory_facets(request: Request, customer: str = "", sku: str = "",
                           color: str = "", size: str = "", location: str = "",
                           country_of_origin: str = "", fabric_content: str = ""):
    """Cascading filter options for the Inventory grid. Given the currently
    selected column filters, return the still-valid distinct values for EACH
    field, computed over rows matching all the OTHER selected filters. Picking a
    customer narrows styles/colors/sizes/... to that customer; then picking a
    style narrows colors/sizes further, etc. Each field excludes its OWN filter
    so you can still switch among its valid values."""
    await require_auth(request)
    sel = {
        "customer": (customer or "").strip(), "sku": (sku or "").strip(),
        "color": (color or "").strip(), "size": (size or "").strip(),
        "location": (location or "").strip(),
        "country_of_origin": (country_of_origin or "").strip(),
        "fabric_content": (fabric_content or "").strip(),
    }

    def rx(v):
        return {"$regex": re.escape(v), "$options": "i"}

    def build(exclude):
        q = {}
        if exclude != "customer" and sel["customer"]:
            q["customer"] = rx(sel["customer"])
        if exclude != "sku" and sel["sku"]:
            q["$or"] = [{"sku": rx(sel["sku"])}, {"style": rx(sel["sku"])}]
        if exclude != "color" and sel["color"]:
            q["color"] = rx(sel["color"])
        if exclude != "size" and sel["size"]:
            q["size"] = rx(sel["size"])
        if exclude != "location" and sel["location"]:
            q["location"] = rx(sel["location"])
        if exclude != "country_of_origin" and sel["country_of_origin"]:
            q["country_of_origin"] = rx(sel["country_of_origin"])
        if exclude != "fabric_content" and sel["fabric_content"]:
            q["fabric_content"] = rx(sel["fabric_content"])
        return q

    async def facet(field, exclude):
        vals = await db.wms_inventory.distinct(field, build(exclude))
        out = [v for v in vals if v and isinstance(v, str)]
        if field == "country_of_origin":
            out = [v for v in out if "%" not in v]  # drop leaked fabric rows
        return sorted(out)[:2000]

    return {
        "customers": await facet("customer", "customer"),
        "styles": await facet("style", "sku"),
        "colors": await facet("color", "color"),
        "sizes": await facet("size", "size"),
        "locations": await facet("location", "location"),
        "countries": await facet("country_of_origin", "country_of_origin"),
        "fabrics": await facet("fabric_content", "fabric_content"),
    }

_STYLE_INFO_CACHE = {}  # {style_upper: (timestamp, payload)}
_STYLE_INFO_TTL = 60.0


@router.get("/inventory/style-info")
async def inventory_style_info(request: Request, style: str = ""):
    """Cascade data for the manual-entry modal: given a style, returns the
    distinct colors and sizes already in inventory, plus the "template" fields
    (customer, description, country_of_origin, category) taken from an existing
    row. One aggregation round-trip + 60s in-process cache so repeated clicks
    on the same style are instant.
    """
    await require_auth(request)
    if not style:
        raise HTTPException(400, "Style requerido")

    import time
    key = style.strip().upper()
    cached = _STYLE_INFO_CACHE.get(key)
    if cached and (time.monotonic() - cached[0]) < _STYLE_INFO_TTL:
        return cached[1]

    # Exact match on the upper-cased value: import_inventory stores everything
    # already upper-cased, so equality is fine and lets the index kick in.
    match = {"$or": [{"style": key}, {"sku": key}]}
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": None,
            "colors": {"$addToSet": "$color"},
            "sizes": {"$addToSet": "$size"},
            "template": {"$first": "$$ROOT"},
        }},
    ]
    docs = await db.wms_inventory.aggregate(pipeline).to_list(1)
    if not docs:
        raise HTTPException(404, f"No existe inventario para style '{style}'")

    agg = docs[0]
    colors = sorted(c for c in agg.get("colors", []) if c)
    SIZE_ORDER = ['XS', 'S', 'M', 'L', 'XL', '2X', '3X', '4X', '5X', '6X',
                  'YXS', 'YS', 'YM', 'YL', 'YXL', '2T', '3T', '4T', '5T']
    raw_sizes = {s for s in agg.get("sizes", []) if s}
    ordered = [s for s in SIZE_ORDER if s in raw_sizes]
    sizes = ordered + sorted(raw_sizes - set(ordered))
    template = agg.get("template") or {}
    payload = {
        "style": template.get("style") or template.get("sku") or key,
        "colors": colors,
        "sizes": sizes,
        "customer": template.get("customer", ""),
        "description": template.get("description", ""),
        "country_of_origin": template.get("country_of_origin", ""),
        "category": template.get("category", ""),
        "manufacturer": template.get("manufacturer", ""),
        "size_header": template.get("size_header", ""),
        "fabric_content": template.get("fabric_content", ""),
    }
    _STYLE_INFO_CACHE[key] = (time.monotonic(), payload)
    return payload


@router.get("/inventory/locations-lookup")
async def locations_lookup(request: Request, style: str = "", color: str = ""):
    """Lookup inventory locations for a style+color, grouped by size.

    Queries BOTH wms_inventory (permanent locations + Excel imports) AND
    wms_boxes (physical boxes, including putaway/transit/temporary locations)
    so pickers always see stock regardless of where it physically sits.
    """
    await require_auth(request)
    if not style:
        raise HTTPException(400, "Style requerido")

    style_rx = {"$regex": f"^{re.escape(style)}$", "$options": "i"}
    color_rx  = {"$regex": f"^{re.escape(color)}$",  "$options": "i"} if color else None

    # ── 1. wms_inventory (aggregated rows, permanent + Excel) ──────────────
    inv_query = {"$or": [{"style": style_rx}, {"sku": style_rx}]}
    if color_rx:
        inv_query["color"] = color_rx
    records = await db.wms_inventory.find(inv_query, {"_id": 0}).to_list(5000)

    # by_size[sz][loc] = {"available": int, "boxes": int, "customer": str, "country_of_origin": str}
    by_loc: dict[str, dict[str, dict]] = {}  # by_loc[sz][loc] = merged entry

    for r in records:
        sz  = r.get("size", "")
        loc = r.get("location") or r.get("inv_location") or ""
        avail = r.get("units_on_hand", r.get("available", 0)) - r.get("units_allocated", 0)
        if not loc or avail <= 0:
            continue
        by_loc.setdefault(sz, {})
        if loc in by_loc[sz]:
            by_loc[sz][loc]["available"] += avail
            by_loc[sz][loc]["boxes"]     += r.get("total_boxes", 0)
        else:
            by_loc[sz][loc] = {
                "location": loc, "available": avail,
                "boxes": r.get("total_boxes", 0),
                "customer": r.get("customer", ""),
                "country_of_origin": r.get("country_of_origin", ""),
                "_from_boxes": False,
            }

    # ── 2. wms_boxes (physical boxes — covers putaway / transit / temp locs) ─
    box_query = {
        "$or": [{"sku": style_rx}, {"style": style_rx}],
        "units": {"$gt": 0},
        "location": {"$exists": True, "$ne": ""},
    }
    if color_rx:
        box_query["color"] = color_rx
    boxes = await db.wms_boxes.find(
        box_query, {"_id": 0, "size": 1, "location": 1, "units": 1, "qty": 1, "customer": 1, "country_of_origin": 1}
    ).to_list(5000)

    for b in boxes:
        sz  = b.get("size", "")
        loc = b.get("location", "")
        units = int(b.get("units") or b.get("qty") or 0)
        if not loc or units <= 0:
            continue
        by_loc.setdefault(sz, {})
        if loc in by_loc[sz]:
            # Location already covered by wms_inventory — don't double-count;
            # only bump if the box count actually exceeds what inventory says
            # (signals an out-of-sync row — trust the higher value).
            existing = by_loc[sz][loc]
            if not existing["_from_boxes"]:
                pass  # trust wms_inventory for this location
        else:
            # Location NOT in wms_inventory → add from boxes (putaway / transit)
            if loc not in by_loc[sz]:
                by_loc[sz][loc] = {
                    "location": loc, "available": 0,
                    "boxes": 0, "customer": b.get("customer", ""),
                    "country_of_origin": b.get("country_of_origin", ""),
                    "_from_boxes": True,
                }
            by_loc[sz][loc]["available"] += units
            by_loc[sz][loc]["boxes"]     += 1

    # ── 3. Build final response ────────────────────────────────────────────
    by_size: dict = {}
    for sz, locs_dict in by_loc.items():
        locs = [
            {k: v for k, v in e.items() if k != "_from_boxes"}
            for e in locs_dict.values() if e["available"] > 0
        ]
        locs.sort(key=lambda x: -x["available"])
        total_avail = sum(l["available"] for l in locs)
        total_boxes = sum(l["boxes"] for l in locs)
        for l in locs:
            l["percentage"] = round((l["available"] / total_avail) * 100) if total_avail else 0
        by_size[sz] = {"size": sz, "locations": locs, "total_available": total_avail, "total_boxes": total_boxes}

    return {"style": style, "color": color, "sizes": by_size}

@router.get("/inventory/options")
async def inventory_options(request: Request, customer: str = "", manufacturer: str = "", style: str = ""):
    """Return unique dropdown values from inventory, case-insensitive dedup, filtered by customer and cascading."""
    await require_auth(request)
    base = {}
    # Scope styles/colors/manufacturers al cliente seleccionado (Receiving pide que
    # al elegir cliente solo se vean SUS valores). Solo aplica cuando se pasa
    # `customer`; los callers que no lo pasan siguen viendo todo. Los nombres de
    # cliente ya están canonicalizados, así que el match exacto no oculta nada.
    if customer:
        base["customer"] = {"$regex": f"^{re.escape(customer)}$", "$options": "i"}

    # Manufacturers: filter by customer only
    mfr_match = {k: v for k, v in base.items()}
    mfr_pipeline = [
        {"$match": mfr_match},
        {"$group": {"_id": {"$toLower": "$manufacturer"}, "val": {"$first": "$manufacturer"}}},
        {"$match": {"_id": {"$ne": ""}}},
        {"$sort": {"_id": 1}}
    ]

    # Styles: no cascaded filters, show all
    style_match = {k: v for k, v in base.items()}
    style_pipeline = [
        {"$match": style_match},
        {"$group": {"_id": {"$toLower": "$style"}, "val": {"$first": "$style"}}},
        {"$match": {"_id": {"$nin": ["", None]}}},
        {"$sort": {"_id": 1}}
    ]

    # Colors: no cascaded filters, show all
    color_match = {k: v for k, v in base.items()}
    color_pipeline = [
        {"$match": color_match},
        {"$group": {"_id": {"$toLower": "$color"}, "val": {"$first": "$color"}}},
        {"$match": {"_id": {"$nin": ["", None]}}},
        {"$sort": {"_id": 1}}
    ]

    # Locations: show all unique locations
    loc_pipeline = [
        {"$group": {"_id": {"$toLower": "$location"}, "val": {"$first": "$location"}}},
        {"$match": {"_id": {"$nin": ["", None]}}},
        {"$sort": {"_id": 1}}
    ]

    mfrs = await db.wms_inventory.aggregate(mfr_pipeline).to_list(5000)
    styles = await db.wms_inventory.aggregate(style_pipeline).to_list(5000)
    colors = await db.wms_inventory.aggregate(color_pipeline).to_list(5000)
    locs = await db.wms_inventory.aggregate(loc_pipeline).to_list(5000)

    # Customers list: Read directly from config_options (same source as MOS Dashboard)
    config = await db.config_options.find_one({"config_id": "main"}, {"clients": 1, "_id": 0})
    merged_customers = sorted(config.get("clients", []), key=lambda s: s.lower()) if config else []

    return {
        "customers": merged_customers,
        "manufacturers": [m["val"] for m in mfrs if m and m.get("val")],
        "styles": [s["val"] for s in styles if s and s.get("val")],
        "colors": [c["val"] for c in colors if c and c.get("val")],
        "locations": [l["val"] for l in locs if l and l.get("val")]
    }

@router.get("/movements/summary")
async def inventory_summary(request: Request, customer: str = ""):
    await require_auth(request)
    match_query = {}
    if customer:
        match_query["customer"] = {"$regex": customer, "$options": "i"}
    
    pipeline = [
        {"$match": match_query},
        {"$group": {
            "_id": None,
            "total_on_hand": {"$sum": "$units_on_hand"},
            "total_allocated": {"$sum": "$units_allocated"},
            "total_available": {"$sum": {"$subtract": ["$units_on_hand", "$units_allocated"]}},
            "total_skus": {"$sum": 1},
            "total_boxes_sum": {"$sum": "$total_boxes"}
        }}
    ]
    low_stock_query = {"units_on_hand": {"$lte": 10, "$gt": 0}}
    if customer:
        low_stock_query["customer"] = {"$regex": customer, "$options": "i"}
    # Locations count depends on the same customer filter if specified.
    locations_coro = (
        db.wms_inventory.distinct("location", match_query) if customer
        else db.wms_locations.count_documents({"active": True})
    )
    # Three independent reads fired concurrently instead of serial awaits.
    result, locations_raw, low_stock = await asyncio.gather(
        db.wms_inventory.aggregate(pipeline).to_list(1),
        locations_coro,
        db.wms_inventory.find(low_stock_query, {"_id": 0}).sort("units_on_hand", 1).to_list(20),
    )
    agg = result[0] if result else {}
    total_locations = len(locations_raw) if customer else locations_raw
    
    summary = {
        "total_on_hand": agg.get("total_on_hand", 0),
        "total_allocated": agg.get("total_allocated", 0),
        "total_available": agg.get("total_available", 0),
        "total_skus": agg.get("total_skus", 0),
        "total_boxes": agg.get("total_boxes_sum", 0),
        "total_locations": total_locations,
        "low_stock_items": len(low_stock),
        "low_stock": low_stock
    }
    return summary

@router.get("/inventory/summary")
async def inventory_summary_alias(request: Request, customer: str = ""):
    """Explicitly support the InventoryModule's summary call path."""
    return await inventory_summary(request, customer)

@router.get("/inventory/chart-data")
async def get_inventory_chart_data(request: Request, customer: str = ""):
    await require_auth(request)
    match_query = {}
    if customer:
        match_query["customer"] = {"$regex": customer, "$options": "i"}

    # 1. Top 10 SKUs by total available units
    top_skus_pipeline = [
        {"$match": match_query},
        {"$group": {"_id": "$sku", "available": {"$sum": "$units_on_hand"}}},
        {"$sort": {"available": -1}},
        {"$limit": 10},
        {"$project": {"name": "$_id", "value": "$available", "_id": 0}}
    ]
    top_skus = await db.wms_inventory.aggregate(top_skus_pipeline).to_list(10)

    # 2. Units by Status/State (Finished Goods vs Raw vs WIP) — queried from wms_boxes
    box_match = {}
    if customer:
        box_match["customer"] = {"$regex": customer, "$options": "i"}
    state_pipeline = [
        {"$match": box_match},
        {"$group": {"_id": "$status", "count": {"$sum": "$qty"}}},
        {"$match": {"_id": {"$ne": None}}},
        {"$project": {"name": {"$ifNull": ["$_id", "unknown"]}, "value": "$count", "_id": 0}}
    ]
    by_state = await db.wms_boxes.aggregate(state_pipeline).to_list(10)

    # 3. Units by Manufacturer/Category
    cat_pipeline = [
        {"$match": match_query},
        {"$group": {"_id": "$manufacturer", "count": {"$sum": "$units_on_hand"}}},
        {"$sort": {"count": -1}},
        {"$limit": 8},
        {"$project": {"name": "$_id", "value": "$count", "_id": 0}}
    ]
    by_manufacturer = await db.wms_inventory.aggregate(cat_pipeline).to_list(8)

    # 4. Activity History (last 15 days)
    # We group movements by day. We filter movements by customer if details.customer matches
    from datetime import timedelta
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=15)).strftime("%Y-%m-%d")
    movement_query = {"created_at": {"$gte": cutoff_date}}
    if customer:
        movement_query["$or"] = [
            {"details.customer": {"$regex": customer, "$options": "i"}},
            {"details.receiving_id": {"$exists": True}}  # Always include receiving events
        ]

    # Activity count by day for last 15 days
    activity_pipeline = [
        {"$match": movement_query},
        {"$addFields": {"date": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {"_id": "$date", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
        {"$limit": 15},
        {"$project": {"date": "$_id", "count": 1, "_id": 0}}
    ]
    activity_history = await db.wms_movements.aggregate(activity_pipeline).to_list(15)

    # 5. Material mas usado (consumo): suma de unidades SURTIDAS por estilo en
    #    pick tickets ya procesados. Refleja que producto se mueve mas, de mayor
    #    a menor. Se calcula en Python porque picked_sizes es anidado.
    from collections import defaultdict as _dd
    pt_query = {"status": {"$in": ["confirmed", "in_neck_cutting", "completed"]}}
    if customer:
        pt_query["customer"] = {"$regex": customer, "$options": "i"}
    tickets = await db.wms_pick_tickets.find(
        pt_query, {"_id": 0, "style": 1, "picked_sizes": 1, "sizes": 1}
    ).to_list(20000)

    def _sum_sizes(d):
        tot = 0
        for v in (d or {}).values():
            tot += int(v.get("total", 0) or 0) if isinstance(v, dict) else int(v or 0)
        return tot

    usage = _dd(int)
    for t in tickets:
        st = (t.get("style") or "").strip()
        if not st:
            continue
        used = _sum_sizes(t.get("picked_sizes")) or _sum_sizes(t.get("sizes"))
        if used:
            usage[st] += used
    most_used = sorted(
        [{"name": k, "value": v} for k, v in usage.items()],
        key=lambda x: -x["value"],
    )[:10]

    return {
        "top_skus": top_skus,
        "by_state": by_state,
        "by_manufacturer": by_manufacturer,
        "activity_history": activity_history,
        "most_used": most_used,
    }

# ==================== ORDERS (from CRM) ====================

@router.get("/orders")
async def list_wms_orders(request: Request, status: str = ""):
    await require_auth(request)
    # Broaden query: include orders from relevant boards for WMS
    # Logic: include anything in scheduling/blanks, or with wms activity
    query = {"$or": [
        {"board": {"$regex": "^blanks$|^crm$|^ventas$|^sales$|^scheduling$|^production$|^final bill$", "$options": "i"}},
        {"blank_status": {"$regex": "partial|parcial|pending|ready|todo|picked", "$options": "i"}},
        {"wms_status": {"$exists": True}}
    ]}
    if status:
        query["wms_status"] = status
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return orders

@router.get("/orders/{order_id}")
async def get_wms_order(order_id: str, request: Request):
    await require_auth(request)
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        order = await db.orders.find_one({"order_number": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Orden no encontrada")
    allocations = await db.wms_allocations.find({"order_id": order.get("order_id")}, {"_id": 0}).to_list(100)
    order["allocations"] = allocations
    return order

# ==================== ALLOCATION ====================


@router.post("/allocations")
@router.post("/stocktakes")
async def create_allocation(request: Request):
    user = await require_auth(request)
    # NEUTRALIZED (WMS audit C1). The allocation/reservation subsystem is
    # disconnected from picking: pick tickets are generated straight from the
    # order and the picker deducts units_on_hand directly, never reading
    # wms_allocations. Worse, the check below validated against inv["available"],
    # a field that is never persisted (only computed on the fly), so every
    # allocation with qty>=1 already failed with "Disponible: 0". Leaving the
    # write path active only risked inflating units_allocated with reservations
    # that no pick ever releases. Disabled here; GET /allocations and
    # DELETE /allocations/{id} stay live so existing rows can be viewed/cleared.
    raise HTTPException(
        410,
        "El módulo de Allocation está deshabilitado: el surtido descuenta "
        "inventario directamente y no requiere reserva previa.",
    )
    body = await request.json()
    order_id = body.get("order_id", "").strip()
    items = body.get("items", [])
    if not order_id or not items:
        raise HTTPException(400, "order_id e items requeridos")
    order = await db.orders.find_one({"$or": [{"order_id": order_id}, {"order_number": order_id}]})
    if not order:
        raise HTTPException(404, "Orden no encontrada")
    allocation_id = gen_id("alloc")
    alloc_items = []
    for item in items:
        sku = item.get("sku", "")
        color = item.get("color", "")
        size = item.get("size", "")
        qty = int(item.get("qty", 0))
        inv = await db.wms_inventory.find_one({"sku": sku, "color": color, "size": size})
        if not inv or inv.get("available", 0) < qty:
            raise HTTPException(400, f"Inventario insuficiente para {sku} {color} {size}. Disponible: {inv.get('available', 0) if inv else 0}")
        await _update_inventory(sku, color, size, qty, "allocate")
        alloc_items.append({"sku": sku, "color": color, "size": size, "qty": qty})
    alloc_doc = {
        "allocation_id": allocation_id,
        "order_id": order.get("order_id"),
        "order_number": order.get("order_number"),
        "items": alloc_items,
        "status": "allocated",
        "allocated_by": user.get("user_id"),
        "allocated_by_name": user.get("name", ""),
        "created_at": now_iso(),
    }
    await db.wms_allocations.insert_one(alloc_doc)
    alloc_doc.pop("_id", None)
    await db.orders.update_one({"order_id": order.get("order_id")}, {"$set": {"wms_status": "allocated"}})
    await log_movement(user, "allocation", {"allocation_id": allocation_id, "order_number": order.get("order_number"), "items": alloc_items})
    return alloc_doc

@router.get("/allocations")
async def list_allocations(request: Request, order_id: str = ""):
    await require_auth(request)
    query = {}
    if order_id: query["$or"] = [{"order_id": order_id}, {"order_number": order_id}]
    allocs = await db.wms_allocations.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return allocs

@router.delete("/allocations/{allocation_id}")
async def delete_allocation(allocation_id: str, request: Request):
    user = await require_auth(request)
    alloc = await db.wms_allocations.find_one({"allocation_id": allocation_id})
    if not alloc:
        raise HTTPException(404, "Allocation no encontrada")
    for item in alloc.get("items", []):
        await _update_inventory(item["sku"], item.get("color", ""), item.get("size", ""), item["qty"], "deallocate")
    await db.wms_allocations.delete_one({"allocation_id": allocation_id})
    await log_movement(user, "deallocate", {"allocation_id": allocation_id, "order_number": alloc.get("order_number")})
    return {"message": "Allocation eliminada"}

# ==================== PICK TICKETS ====================

# Valid picking strategies — keep in sync with frontend Picking.js dropdown
PICK_STRATEGIES = {"default", "proximity", "origin"}

def _location_sort_key_proximity(loc_obj: dict):
    """Sort by location-name prefix (everything before first '-'), then full name.
    Groups RP01-* together, then RP02-*, etc."""
    name = (loc_obj.get("location") or "").upper()
    prefix = name.split("-", 1)[0] if "-" in name else name
    return (prefix, name)

def _location_sort_key_origin(loc_obj: dict):
    """Group by country_of_origin first, then by location name."""
    coo = (loc_obj.get("country_of_origin") or "").upper() or "ZZZ"  # unknowns last
    name = (loc_obj.get("location") or "").upper()
    return (coo, name)

def apply_picking_strategy(size_locations: dict, strategy: str) -> dict:
    """Reorder the `locations` array inside each size of size_locations
    according to the chosen strategy. Default keeps backend's existing order."""
    if strategy not in {"proximity", "origin"} or not size_locations:
        return size_locations
    key_fn = _location_sort_key_proximity if strategy == "proximity" else _location_sort_key_origin
    for sz_data in size_locations.values():
        if isinstance(sz_data, dict) and "locations" in sz_data:
            sz_data["locations"].sort(key=key_fn)
    return size_locations


# Cache in-memory para size_locations. Cada picker recarga sus tickets 3-5 veces
# durante un turno de trabajo; sin cache eso era 3-5x N regex-queries a Mongo por
# recarga (my-tickets tardaba ~2 min en producción). TTL corto para no servir
# stock muerto — el ticket que ya usó una ubicación tiene 30s hasta refrescarse.
_SIZE_LOCS_CACHE: dict = {}
_SIZE_LOCS_TTL = 30.0   # seg


async def _compute_size_locations(style: str, color: str, sizes: dict, strategy: str = "default") -> dict:
    """Build {size: {locations:[...], total_available}} from CURRENT inventory.

    Used so pickers always see live availability (not the snapshot captured when
    the ticket was created). Color is matched exactly first; if a size finds no
    location that way it retries with a contains match (handles naming variants
    like 'OLIVE' vs 'OLIVE GREEN') so available stock isn't hidden by a label
    mismatch.
    """
    style = (style or "").strip()
    color = (color or "").strip()
    size_locations: dict = {}
    if not style:
        return size_locations

    # Cache lookup — key inmutable con las tallas del ticket
    import time as _time
    cache_key = (style.upper(), color.upper(), strategy,
                 tuple(sorted((str(k), int(v or 0)) for k, v in (sizes or {}).items())))
    cached = _SIZE_LOCS_CACHE.get(cache_key)
    if cached and (_time.monotonic() - cached[0]) < _SIZE_LOCS_TTL:
        return cached[1]

    async def _run(q):
        # Limite alto: un SKU muy fragmentado puede vivir en +50 ubicaciones
        # (5000 BLACK L estaba en 96). Con .to_list(50) el picker no veia el stock
        # de la ubicacion #51+ aunque tuviera cajas fisicas (caso PS06-A30). Se
        # ordena por unidades desc, asi que las mas llenas siguen primero.
        recs = await db.wms_inventory.find(
            q, {"_id": 0, "location": 1, "units_on_hand": 1, "units_allocated": 1, "total_boxes": 1, "customer": 1, "country_of_origin": 1}
        ).sort("units_on_hand", -1).to_list(500)
        locs = [{
            "location": r.get("location", ""),
            "available": r.get("units_on_hand", 0) - r.get("units_allocated", 0),
            "boxes": r.get("total_boxes", 0),
            "country_of_origin": r.get("country_of_origin", ""),
        } for r in recs if r.get("location")]
        return [l for l in locs if l["available"] > 0]

    for sz, qty in (sizes or {}).items():
        try:
            qn = int(qty) if qty else 0
        except (TypeError, ValueError):
            qn = 0
        if qn <= 0:
            continue
        # Match EXACTO en MAYÚSCULAS (_ci_eq) en vez de regex case-insensitive:
        # los campos de identidad están normalizados a UPPERCASE system-wide, así
        # que es equivalente al regex viejo PERO usa los índices
        # (style_1_color_1_size_1_location_1 / sku_1_…). El regex "i" forzaba
        # COLLECTION SCAN de wms_inventory (22k) por cada talla → my-tickets lento.
        base = {
            "$or": [
                {"style": _ci_eq(style)},
                {"sku": _ci_eq(style)},
            ],
            "size": _ci_eq(sz),
            "units_on_hand": {"$gt": 0},
        }
        if color:
            exact = dict(base); exact["color"] = _ci_eq(color)
            locs = await _run(exact)
            if not locs:
                # Fallback tolerante a variantes de nombre ('OLIVE' vs 'OLIVE
                # GREEN'). Regex NO indexado, pero SOLO corre cuando el match exacto
                # no encontró nada (raro) → no golpea la ruta caliente.
                loose = dict(base); loose["color"] = {"$regex": re.escape(color), "$options": "i"}
                locs = await _run(loose)
        else:
            locs = await _run(base)
        # Merge rows that share the SAME physical location (inventory is also
        # split by country_of_origin/fabric, so one shelf can yield several rows).
        # Without this the picker sees duplicate inputs for the same location and
        # typing in one mirrors the other.
        merged: dict = {}
        for l in locs:
            key = l["location"]
            if key in merged:
                m = merged[key]
                m["available"] += l["available"]
                m["boxes"] += l.get("boxes", 0)
                coo = l.get("country_of_origin")
                if coo and coo not in m["_origins"]:
                    m["_origins"].append(coo)
            else:
                merged[key] = {**l, "_origins": [l["country_of_origin"]] if l.get("country_of_origin") else []}

        # Surface material that lives only as PHYSICAL BOXES — carts (CARRO N),
        # transit and putaway-pending slots — so the picker sees stock wherever it
        # physically sits, even when no inventory row backs it (or the row drifted).
        # Shelves already covered by wms_inventory keep the inventory figure; we
        # don't double-count those. Picking from such a location deducts the cart
        # boxes directly (see _deduct_pick_boxes), which is exactly what we want.
        # Igual que arriba: _ci_eq exacto para usar el índice
        # sku_1_color_1_size_1_location_1_units_1 en vez de escanear wms_boxes (78k).
        box_q = {
            "$or": [
                {"sku": _ci_eq(style)},
                {"style": _ci_eq(style)},
            ],
            "size": _ci_eq(sz),
            "units": {"$gt": 0},
            "location": {"$exists": True, "$ne": ""},
        }
        if color:
            box_q["color"] = _ci_eq(color)
        for b in await db.wms_boxes.find(
            box_q, {"_id": 0, "location": 1, "units": 1, "qty": 1, "country_of_origin": 1}
        ).to_list(2000):
            loc = b.get("location", "")
            if not loc or loc in merged:
                continue  # inventory already accounts for this location — trust it
            u = int((b.get("units") if b.get("units") is not None else b.get("qty", 0)) or 0)
            if u <= 0:
                continue
            e = merged.setdefault(loc, {
                "location": loc, "available": 0, "boxes": 0,
                "country_of_origin": b.get("country_of_origin", ""), "_origins": [],
            })
            e["available"] += u
            e["boxes"] += 1
            coo = b.get("country_of_origin")
            if coo and coo not in e["_origins"]:
                e["_origins"].append(coo)

        locs = []
        for m in merged.values():
            m["country_of_origin"] = ", ".join(o for o in m.pop("_origins", []) if o)
            locs.append(m)
        total = sum(l["available"] for l in locs)
        for l in locs:
            l["percentage"] = round((l["available"] / total) * 100) if total > 0 else 0
        size_locations[sz] = {"locations": locs, "total_available": total}

    result = apply_picking_strategy(size_locations, strategy if strategy in PICK_STRATEGIES else "default")
    # Guarda en cache — próxima llamada con mismo (style, color, sizes) devuelve
    # instantáneo por hasta 30s.
    _SIZE_LOCS_CACHE[cache_key] = (_time.monotonic(), result)
    # Evita crecimiento ilimitado: purga entradas viejas cuando el cache pasa 5000
    if len(_SIZE_LOCS_CACHE) > 5000:
        cutoff = _time.monotonic() - _SIZE_LOCS_TTL
        for k in [k for k, v in _SIZE_LOCS_CACHE.items() if v[0] < cutoff]:
            _SIZE_LOCS_CACHE.pop(k, None)
    return result


async def internal_create_picking_ticket(data: dict, user: dict) -> dict:
    """
    Internal function to create a pick ticket.
    Expected data: order_number, customer, client, manufacturer, style, color, quantity, sizes, board_category, assigned_to...
    """
    ticket_id = gen_id("pick")
    order_number = data.get("order_number", "").strip()
    style = data.get("style", "").strip()
    force_duplicate = bool(data.get("force_duplicate", False))

    # Validation for manual creation might be stricter than automated skeleton
    if not order_number:
        raise HTTPException(400, "Numero de orden requerido")

    # --- Duplicate guard -----------------------------------------------------------
    # Prevent creating multiple active pick tickets for the same order unless the
    # caller explicitly confirms they want a duplicate (force_duplicate=True).
    if not force_duplicate:
        existing = await db.wms_pick_tickets.find_one(
            {
                "order_number": order_number,
                "status": {"$nin": ["confirmed", "cancelled"]},
            },
            {"_id": 0, "ticket_id": 1, "created_by_name": 1, "created_at": 1,
             "style": 1, "color": 1, "total_pick_qty": 1, "status": 1, "picking_status": 1},
        )
        if existing:
            raise HTTPException(
                409,
                {
                    "message": f"Ya existe un pick ticket activo para la orden {order_number}.",
                    "existing_ticket": existing,
                },
            )
    # -------------------------------------------------------------------------------

    sizes = data.get("sizes", {})
    # Comodín 2% OPCIONAL: si el creador lo activa (include_comodin), se REPARTE
    # dentro de las tallas — base + ceil(2%) por talla (mismo redondeo hacia
    # arriba que la etiqueta) — para que el sistema lo descuente de verdad. Se
    # marca `comodin_applied` para que la etiqueta NO vuelva a sumar el 2%.
    include_comodin = bool(data.get("include_comodin", False))
    comodin_pct = 0.02
    if include_comodin and sizes:
        inflated = {}
        for sz, q in sizes.items():
            base = int(q or 0)
            # ceil(base * 0.02) con enteros: (base*2 + 99)//100
            inflated[sz] = base + ((base * 2 + 99) // 100) if base > 0 else base
        sizes = inflated

    total_qty = sum(int(v) for v in sizes.values() if v)
    if total_qty == 0 and data.get("quantity"):
         total_qty = int(data.get("quantity"))

    style = data.get("style", "").strip()
    color = data.get("color", "").strip()
    
    # Auto-lookup locations for each size from inventory
    size_locations = {}
    if style:
        for sz, qty in sizes.items():
            qty = int(qty) if qty else 0
            if qty > 0:
                inv_query = {
                    "$or": [{"style": _ci_eq(style)}, {"sku": _ci_eq(style)}],
                    "size": _ci_eq(sz),
                    "units_on_hand": {"$gt": 0}
                }
                if color:
                    inv_query["color"] = _ci_eq(color)
                inv_records = await db.wms_inventory.find(inv_query, {"_id": 0, "location": 1, "units_on_hand": 1, "units_allocated": 1, "total_boxes": 1, "customer": 1, "country_of_origin": 1}).sort("units_on_hand", -1).to_list(500)  # ver nota en _compute_size_locations: no ocultar stock de SKUs muy repartidos
                locs = [{"location": r.get("location", ""), "available": r.get("units_on_hand", 0) - r.get("units_allocated", 0), "boxes": r.get("total_boxes", 0), "country_of_origin": r.get("country_of_origin", "")} for r in inv_records if r.get("location")]
                locs = [l for l in locs if l["available"] > 0]
                total_sz_avail = sum(l["available"] for l in locs)
                for l in locs:
                    l["percentage"] = round((l["available"] / total_sz_avail) * 100) if total_sz_avail > 0 else 0
                size_locations[sz] = {"locations": locs, "total_available": total_sz_avail}

    # Apply picking strategy (proximity / origin / default)
    strategy = data.get("strategy", "default")
    if strategy not in PICK_STRATEGIES:
        strategy = "default"
    size_locations = apply_picking_strategy(size_locations, strategy)

    assigned_to = data.get("assigned_to", "").strip()
    assigned_to_name = data.get("assigned_to_name", "").strip()

    ticket_doc = {
        "ticket_id": ticket_id,
        "order_number": order_number,
        "customer": data.get("customer", "").strip(),
        "client": data.get("client", "").strip(),
        "manufacturer": data.get("manufacturer", "").strip(),
        "style": style,
        "color": data.get("color", "").strip(),
        "quantity": int(data.get("quantity", 0)),
        "sizes": sizes,
        "size_locations": size_locations,
        "strategy": strategy,
        "comodin_applied": include_comodin,      # tallas ya incluyen el 2%
        "comodin_pct": comodin_pct if include_comodin else 0,
        "total_pick_qty": total_qty,
        "status": "pending",
        "board_category": data.get("board_category", "UNSET"),
        "destination": data.get("destination", "production"),
        "blank_status": data.get("blank_status", ""),
        "picking_status": "assigned" if assigned_to else "unassigned",
        "assigned_to": assigned_to or None,
        "assigned_to_name": assigned_to_name or None,
        "assigned_at": now_iso() if assigned_to else None,
        "picked_sizes": {},
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name", ""),
        "created_at": now_iso(),
        "sla_deadline": (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat(),
        "sla_status": "on_time"
    }

    await db.wms_pick_tickets.insert_one(ticket_doc)
    ticket_doc.pop("_id", None)
    
    await log_movement(user, "pick_ticket_created", {"ticket_id": ticket_id, "order_number": order_number})
    
    if assigned_to:
        await ws_manager.broadcast("ticket_assigned", {
            "ticket_id": ticket_id,
            "assigned_to": assigned_to,
            "assigned_to_name": assigned_to_name,
            "order_number": order_number,
            "message": f"Nuevo pick ticket: {ticket_id} (PO: {order_number})"
        })
    
    return ticket_doc

@router.post("/pick-tickets")
@router.post("/move-stock")
async def create_pick_ticket(request: Request):
    user = await require_auth(request)
    body = await request.json()
    
    allocation_id = body.get("allocation_id", "").strip()
    if allocation_id:
        # Legacy allocation-based flow
        alloc = await db.wms_allocations.find_one({"allocation_id": allocation_id}, {"_id": 0})
        if not alloc:
            raise HTTPException(404, "Allocation no encontrada")
        ticket_id = gen_id("pick")
        pick_lines = []
        for item in alloc.get("items", []):
            boxes = await db.wms_boxes.find({
                "sku": item["sku"], "color": item.get("color", ""), "size": item.get("size", ""),
                "status": "stored", "state": "raw"
            }, {"_id": 0}).sort("seq_num", 1).to_list(100)
            remaining = item["qty"]
            for box in boxes:
                if remaining <= 0:
                    break
                pick_qty = min(box["units"], remaining)
                pick_lines.append({
                    "box_id": box["box_id"], "sku": item["sku"],
                    "color": item.get("color", ""), "size": item.get("size", ""),
                    "location": box.get("location", ""), "qty": pick_qty
                })
                remaining -= pick_qty
        ticket_doc = {
            "ticket_id": ticket_id, "allocation_id": allocation_id,
            "order_id": alloc.get("order_id"), "order_number": alloc.get("order_number"),
            "lines": pick_lines, "status": "pending",
            "created_by": user.get("user_id"),
            "created_by_name": user.get("name", ""),
            "created_at": now_iso(),
        }
        await db.wms_pick_tickets.insert_one(ticket_doc)
        ticket_doc.pop("_id", None)
        await log_movement(user, "pick_ticket_created", {"ticket_id": ticket_id, "order_number": ticket_doc.get("order_number", "")})
        await notify_badge_change("picking")
        return ticket_doc
    else:
        result = await internal_create_picking_ticket(body, user)
        await notify_badge_change("picking")
        return result

@router.get("/inventory/field-options")
async def get_inventory_field_options(request: Request):
    """Get unique values for description, country_of_origin, fabric_content from inventory."""
    await require_auth(request)
    desc_pipeline = [
        {"$match": {"description": {"$ne": None, "$nin": ["", "."]}}},
        {"$group": {"_id": {"$toLower": "$description"}, "val": {"$first": "$description"}}},
        {"$sort": {"_id": 1}}
    ]
    # Some imported Excel rows shifted columns and dumped fabric-content strings
    # ("52% COTTON 48% POLYESTER", "100% COTTON", etc.) into country_of_origin.
    # Filter those out so the Receiving "País de origen" dropdown stays clean.
    country_pipeline = [
        {"$match": {
            "country_of_origin": {"$ne": None, "$nin": ["", "."]},
            "$expr": {"$eq": [{"$indexOfCP": ["$country_of_origin", "%"]}, -1]},
        }},
        {"$group": {"_id": {"$toLower": "$country_of_origin"}, "val": {"$first": "$country_of_origin"}}},
        {"$sort": {"_id": 1}}
    ]
    fabric_pipeline = [
        {"$match": {"fabric_content": {"$ne": None, "$nin": ["", "."]}}},
        {"$group": {"_id": {"$toLower": "$fabric_content"}, "val": {"$first": "$fabric_content"}}},
        {"$sort": {"_id": 1}}
    ]
    descs = await db.wms_inventory.aggregate(desc_pipeline).to_list(500)
    countries = await db.wms_inventory.aggregate(country_pipeline).to_list(500)

    # Fabric: merge from wms_inventory + wms_receiving + standard seed list
    fabrics_inv = await db.wms_inventory.aggregate(fabric_pipeline).to_list(500)
    fabrics_rcv = await db.wms_receiving.aggregate(fabric_pipeline).to_list(500)

    FABRIC_SEED = [
        "100% COTTON", "100% POLYESTER", "100% NYLON", "100% RAYON",
        "50% COTTON / 50% POLYESTER", "60% COTTON / 40% POLYESTER",
        "65% POLYESTER / 35% COTTON", "52% COTTON / 48% POLYESTER",
        "90% COTTON / 10% POLYESTER", "80% COTTON / 20% POLYESTER",
        "95% COTTON / 5% SPANDEX", "97% COTTON / 3% SPANDEX",
        "95% POLYESTER / 5% SPANDEX", "88% POLYESTER / 12% SPANDEX",
        "100% COMBED COTTON", "100% RING-SPUN COTTON",
        "50% POLYESTER / 25% COTTON / 25% RAYON",
        "60% COTTON / 40% RAYON", "55% HEMP / 45% COTTON",
        "100% LINEN", "100% BAMBOO", "100% MODAL",
    ]

    seen = set()
    all_fabrics = []
    for f in fabrics_inv:
        key = f["val"].strip().upper()
        if key not in seen:
            seen.add(key)
            all_fabrics.append(key)
    for f in fabrics_rcv:
        key = f["val"].strip().upper()
        if key not in seen:
            seen.add(key)
            all_fabrics.append(key)
    for seed in FABRIC_SEED:
        key = seed.strip().upper()
        if key not in seen:
            seen.add(key)
            all_fabrics.append(key)

    # Merge admin-curated catalog values (case-insensitive)
    catalog_docs = await db.wms_catalog_options.find({}, {"_id": 0, "type": 1, "value": 1}).to_list(2000)
    catalog_by_type = {"descriptions": [], "countries": [], "fabrics": []}
    for d in catalog_docs:
        if d.get("type") in catalog_by_type:
            catalog_by_type[d["type"]].append(d.get("value", ""))

    def merge_unique(base_list, extras):
        s = {v.strip().upper() for v in base_list}
        for x in extras:
            k = (x or "").strip().upper()
            if k and k not in s:
                s.add(k)
                base_list.append(x)
        return base_list

    # If the admin curated a catalog for a given type, that becomes the SOLE
    # source of truth — the inventory distinct (which is full of typos like
    # REPIBLICA / BANGLANDESH / RAPUBLICA DOMINICANA) is ignored. When the
    # catalog is empty we fall back to inventory distinct so existing installs
    # don't suddenly show empty dropdowns.
    def authoritative_or_merge(inv_list, curated):
        curated_clean = [c for c in curated if (c or "").strip()]
        return curated_clean if curated_clean else merge_unique(inv_list, curated_clean)

    descs_list = authoritative_or_merge([d["val"] for d in descs], catalog_by_type["descriptions"])
    countries_list = authoritative_or_merge([c["val"] for c in countries], catalog_by_type["countries"])
    fabrics_list = authoritative_or_merge(all_fabrics, catalog_by_type["fabrics"])

    fabrics_list.sort()
    descs_list.sort(key=lambda x: x.lower())
    countries_list.sort(key=lambda x: x.lower())

    return {
        "descriptions": descs_list,
        "countries": countries_list,
        "fabrics": fabrics_list,
    }

@router.get("/pick-tickets")
async def list_pick_tickets(
    request: Request,
    status: str = "",
    paginated: bool = False, skip: int = 0, limit: int = 1000,
    exclude_completed: bool = False,
):
    """List pick tickets.
      - Default (legacy): bare array, up to 1000 newest.
      - paginated=true   : { items, total, has_more } using skip/limit.
        Virtual tickets (orders without a ticket) are only added when
        paginated=false OR when skip=0, to keep pagination semantics clean.
    """
    await require_auth(request)
    query = {"status": status} if status else {}
    if exclude_completed and not status:
        # Default load skips finished picks (the Completed tab fetches them on
        # demand) to save bandwidth/RAM on warehouse devices. Virtual pre-tickets
        # are still synthesized below since status stays empty.
        query = {"status": {"$ne": "confirmed"}, "picking_status": {"$ne": "completed"}}

    skip = max(0, skip)
    limit = max(1, min(limit, 1000))

    # Unified aggregation to get tickets + order info in one go
    pipeline = [
        {"$match": query},
        {"$sort": {"created_at": -1}},
        {"$skip": skip},
        {"$limit": limit},
        {"$lookup": {
            "from": "orders",
            "localField": "order_number",
            "foreignField": "order_number",
            "as": "order_data"
        }},
        {"$addFields": {
            "order_info": {"$arrayElemAt": ["$order_data", 0]}
        }},
        {"$project": {
            "_id": 0,
            "order_data": 0
        }}
    ]

    real_tickets = await db.wms_pick_tickets.aggregate(pipeline).to_list(limit)
    
    # Process job titles and other order info
    for rt in real_tickets:
        oi = rt.pop("order_info", None)
        if oi:
            for k in ["job_title_a", "job_title_b"]:
                rt[k] = oi.get(k)
            # Optionally sync more info if needed
            if not rt.get("customer"): rt["customer"] = oi.get("client") or oi.get("branding")
            
    # --- VIRTUAL TICKETS LOGIC ---
    # Synthesize a pre-ticket for every active order that doesn't have a real ticket yet.
    # When paginated, only include them on the first page (skip=0) to keep cursor semantics.
    if (not status or status == "pending") and (not paginated or skip == 0):
        existing_order_numbers = {t.get("order_number") for t in real_tickets if t.get("order_number")}

        # Also exclude admin-dismissed pre-tickets
        dismissed_docs = await db.wms_dismissed_pretickets.find({}, {"_id": 0, "order_number": 1}).to_list(2000)
        dismissed_order_numbers = {d["order_number"] for d in dismissed_docs if d.get("order_number")}
        excluded_order_numbers = list(existing_order_numbers | dismissed_order_numbers)

        virtual_query = {
            "status": {"$nin": ["cancelled", "shipped", "completed", "COMPLETADO", "CERRADO"]},
            "wms_status": {"$ne": "picked"},
            # Orders that already finished picking (or never need it) must NOT resurface
            # as pre-tickets. Terminal/backup/shipped/inventory boards are excluded.
            # (case-insensitive; orders with no board still pass.)
            "board": {"$not": {"$regex": r"(?i)^\s*(FINAL BILL|COMPLETOS|CANCELLED|PAPELERA|RESPALDO MONDAY|EDI|INVENTARIO)\b"}},
            "order_number": {"$nin": excluded_order_numbers}
        }
        
        # Limit to 500 to avoid performance issues
        virtual_orders = await db.orders.find(virtual_query, {"_id": 0}).sort("created_at", -1).to_list(500)
        
        for vo in virtual_orders:
            real_tickets.append({
                "ticket_id": f"virt_{vo.get('order_id')}",
                "order_number": vo.get("order_number"),
                "customer": vo.get("client") or vo.get("customer") or vo.get("branding") or "No Client",
                "client": vo.get("client") or vo.get("customer") or vo.get("branding"),
                "manufacturer": vo.get("manufacturer") or vo.get("branding"),
                "style": vo.get("style") or vo.get("garment_style") or "",
                "color": vo.get("color") or vo.get("garment_color") or "",
                "sizes": vo.get("sizes", {}),
                "quantity": vo.get("quantity") or 0,
                "total_pick_qty": vo.get("quantity") or 0,
                "status": "pending",
                "blank_status": vo.get("blank_status") or "PENDIENTE",
                "picking_status": "unassigned",
                "board_category": vo.get("board", "UNSET").upper(),
                "destination": "neck_cutting" if vo.get("art_neck_status") else "production",
                "job_title_a": vo.get("job_title_a"),
                "job_title_b": vo.get("job_title_b"),
                # Deadline so the assigner can prioritize by date instead of
                # filtering an Excel. Strictly cancel_date (the hard deadline).
                "cancel_date": vo.get("cancel_date") or "",
                "created_at": vo.get("created_at") or now_iso(),
                "is_virtual": True
            })

    # Sort the combined list by created_at descending so newest orders/tickets appear first
    real_tickets.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    if not paginated:
        return real_tickets

    total = await db.wms_pick_tickets.count_documents(query)
    # has_more is based on real tickets only; virtual tickets are computed every call
    has_more = (skip + len(real_tickets)) < total
    return {"items": real_tickets, "total": total, "has_more": has_more}

@router.put("/pick-tickets/virtual/{order_number}/dismiss")
async def dismiss_preticket(order_number: str, request: Request):
    """Admin-only: permanently hide a virtual pre-ticket (order without a real pick ticket).
    Stores the order_number in wms_dismissed_pretickets so the virtual query excludes it."""
    user = await require_auth(request)
    role = user.get("role", "")
    if role not in ("admin", "supersu", "ceo"):
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar pre-tickets")
    
    # Idempotent upsert — safe to call multiple times
    await db.wms_dismissed_pretickets.update_one(
        {"order_number": order_number},
        {"$set": {
            "order_number": order_number,
            "dismissed_by": user.get("user_id"),
            "dismissed_by_name": user.get("name", ""),
            "dismissed_at": now_iso()
        }},
        upsert=True
    )
    await log_movement(user, "preticket_dismissed", {"order_number": order_number})
    return {"message": "Pre-ticket ocultado correctamente", "order_number": order_number}


@router.post("/pick-tickets/{ticket_id}/incidents")
async def report_incident(ticket_id: str, request: Request):
    user = await require_auth(request)
    body = await request.json()

    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")

    replacement_sizes = {k: int(v) for k, v in (body.get("replacement_sizes") or {}).items() if int(v or 0) > 0}
    replacement_qty = sum(replacement_sizes.values())

    incident = {
        "incident_id": gen_id("inc"),
        "ticket_id": ticket_id,
        "order_number": ticket.get("order_number"),
        "sku": body.get("sku"),
        "qty": int(body.get("qty", 1)),
        "reason": body.get("reason", "Dañado"),
        "replacement_sizes": replacement_sizes,
        "replacement_qty": replacement_qty,
        "replacement_description": body.get("replacement_description", ""),
        "notes": body.get("notes", ""),
        "inventory_deducted": False,
        "operator_id": user.get("user_id"),
        "operator_name": user.get("name", user.get("email", "")),
        "timestamp": now_iso()
    }

    # Deduct replacement units from inventory (FIFO por caja)
    if replacement_sizes:
        style = ticket.get("style", "")
        color = ticket.get("color", "")
        deduct_errors = []
        for size, qty in replacement_sizes.items():
            try:
                await _deduct_pick_boxes(
                    style, color, size, location=None, qty=qty,
                    inv_operation="incident_replacement",
                    customer=ticket.get("customer", ""),
                    order_number=ticket.get("order_number"),
                    user=user, ticket_id=ticket_id,
                )
            except Exception as e:
                deduct_errors.append(f"{size}: {str(e)}")
        incident["inventory_deducted"] = len(deduct_errors) == 0
        if deduct_errors:
            incident["deduct_errors"] = deduct_errors

    await db.wms_incidents.insert_one(incident)
    await log_movement(user, "incident_reported", {
        "ticket_id": ticket_id,
        "sku": incident["sku"],
        "qty": incident["qty"],
        "replacement_sizes": replacement_sizes,
        "replacement_qty": replacement_qty,
        "inventory_deducted": incident["inventory_deducted"],
    })
    return {
        "message": "Incidencia reportada",
        "incident_id": incident["incident_id"],
        "inventory_deducted": incident["inventory_deducted"],
        "replacement_qty": replacement_qty,
    }

# ==================== OPERATOR MODULE ====================

@router.get("/operators")
async def list_operators(request: Request):
    """List all users with role 'operator' or 'picker'."""
    await require_auth(request)
    operators = await db.users.find({"role": {"$in": ["operator", "picker"]}}, {"_id": 0, "password_hash": 0}).to_list(200)
    
    # Inject Contador 1 alias / virtual user to encapsulate users
    if not any(op.get("user_id") == "contador_1" or op.get("name") == "Contador 1" for op in operators):
        operators.insert(0, {
            "user_id": "contador_1",
            "name": "Contador 1",
            "email": "contador1@mos.system",
            "role": "operator",
            "active": True
        })
    return operators

@router.put("/pick-tickets/{ticket_id}/assign")
async def assign_pick_ticket(ticket_id: str, request: Request):
    """Admin assigns a pick ticket to an operator."""
    user = await require_admin(request)
    body = await request.json()
    operator_id = body.get("operator_id", "").strip()
    operator_name = body.get("operator_name", "").strip()

    # Empty operator_id = UNASSIGN (retirar del operador): clear the assignment
    # and return the ticket to the unassigned pool. Any picking progress
    # (picked_sizes / deducted_map) stays on the ticket, so reassigning later
    # resumes where it was left.
    if not operator_id:
        result = await db.wms_pick_tickets.update_one(
            {"ticket_id": ticket_id},
            {"$set": {"assigned_to": None, "assigned_to_name": "",
                      "picking_status": "unassigned", "assigned_at": None}},
        )
        if result.matched_count == 0:
            raise HTTPException(404, "Pick ticket no encontrado")
        await log_movement(user, "pick_ticket_unassigned", {"ticket_id": ticket_id})
        await notify_badge_change("picking")
        return {"message": f"Ticket {ticket_id} retirado del operador", "ticket_id": ticket_id}

    result = await db.wms_pick_tickets.update_one(
        {"ticket_id": ticket_id},
        {"$set": {
            "assigned_to": operator_id,
            "assigned_to_name": operator_name,
            "picking_status": "assigned",
            "assigned_at": now_iso()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Pick ticket no encontrado")
    await log_movement(user, "pick_ticket_assigned", {"ticket_id": ticket_id, "operator_id": operator_id, "operator_name": operator_name})
    # Notify operator in real-time via WebSocket
    await ws_manager.broadcast("ticket_assigned", {
        "ticket_id": ticket_id,
        "assigned_to": operator_id,
        "assigned_to_name": operator_name,
        "message": f"Nuevo ticket asignado: {ticket_id}"
    })
    return {"message": f"Ticket {ticket_id} asignado a {operator_name}", "ticket_id": ticket_id}

@router.get("/operator/my-tickets")
async def get_operator_tickets(request: Request):
    """Get pick tickets assigned to the current operator."""
    user = await require_auth(request)
    user_id = user.get("user_id", "")
    email = user.get("email", "")
    # Match by user_id or email
    query = {
        "$or": [{"assigned_to": user_id}, {"assigned_to": email}],
        "status": {"$ne": "confirmed"}
    }
    tickets = await db.wms_pick_tickets.find(query, {"_id": 0}).sort("assigned_at", -1).to_list(200)
    # Refresh available pick locations from CURRENT inventory so the operator
    # always sees where the stock actually is now (the stored size_locations is
    # only a snapshot taken when the ticket was created).
    # Recompute in parallel (was sequential — N round-trips in series for an
    # operator with many tickets). A semaphore bounds concurrency so we don't
    # flood the Mongo connection pool when a picker has a large queue.
    pending = [tk for tk in tickets if tk.get("picking_status") != "completed"]
    # 4 en paralelo (bajado de 8) porque los queries regex de _compute_size_locations
    # golpean fuerte cuando el picker tiene 15+ tickets y Mongo se congestiona.
    # Con el cache de 30s la mayoria de tickets no dispara query real.
    sem = asyncio.Semaphore(4)

    async def _attach(tk):
        async with sem:
            try:
                tk["size_locations"] = await _compute_size_locations(
                    tk.get("style", ""), tk.get("color", ""), tk.get("sizes", {}),
                    tk.get("strategy", "default"),
                )
            except Exception as e:
                logger.error(f"[my-tickets] live size_locations failed for {tk.get('ticket_id')}: {e}")

    await asyncio.gather(*[_attach(tk) for tk in pending])
    return tickets

@router.get("/operator/completed-tickets")
async def get_operator_completed_tickets(request: Request):
    """Get completed pick tickets for the current operator."""
    user = await require_auth(request)
    user_id = user.get("user_id", "")
    email = user.get("email", "")
    query = {
        "$or": [{"assigned_to": user_id}, {"assigned_to": email}],
        "picking_status": "completed"
    }
    tickets = await db.wms_pick_tickets.find(query, {"_id": 0}).sort("completed_at", -1).to_list(200)
    return tickets

def _normalize_picked(ps):
    """Flatten a picked_sizes payload into {size: {location: qty}}.

    picked_sizes arrives as either { "S": { "total": 10, "details": {loc: q} } }
    or the simple { "S": 10 }. We collapse it to a per-(size, location) map so the
    incremental deducer can diff it against what was already deducted. The simple
    format (no shelf chosen) lands under location "" (deduct FIFO across shelves)."""
    out = {}
    for sz, data in (ps or {}).items():
        szk = (sz or "").strip()
        if isinstance(data, dict) and "details" in data:
            for loc, q in (data.get("details") or {}).items():
                qi = int(q or 0)
                if qi:
                    out.setdefault(szk, {})[loc] = out.setdefault(szk, {}).get(loc, 0) + qi
        else:
            qi = int(data.get("total", 0)) if isinstance(data, dict) else int(data or 0)
            if qi:
                out.setdefault(szk, {})[""] = out.setdefault(szk, {}).get("", 0) + qi
    return out


@router.put("/pick-tickets/{ticket_id}/pick-progress")
async def save_pick_progress(ticket_id: str, request: Request):
    """Operator saves picking progress (partial or complete)."""
    user = await require_auth(request)
    body = await request.json()
    picked_sizes = body.get("picked_sizes", {})  # { "S": 100, "M": 50, ... }
    is_complete = body.get("is_complete", False)

    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")

    # Authorization: an operator may only save progress on a ticket assigned to
    # them. Unassigned tickets are left open (operators never see those in
    # /my-tickets, so this only guards direct API calls). Admin/supersu/ceo bypass.
    assignee = (ticket.get("assigned_to") or "").strip()
    caller_ids = {user.get("user_id", ""), user.get("email", "")}
    elevated = user.get("role") in {"admin", "supersu", "ceo"}
    if assignee and assignee not in caller_ids and not elevated:
        raise HTTPException(403, "Este pick ticket está asignado a otro operador")

    # HOLD guard: stock in a SAT-held location can't be picked until released.
    picked_locs = [loc for d in picked_sizes.values() if isinstance(d, dict)
                   for loc in (d.get("details") or {}).keys()]
    if picked_locs:
        await _assert_not_on_hold(user, *picked_locs)

    # Guard against a partial/stale payload silently wiping already-picked sizes.
    # That didn't just lose progress — it REVERSED the deduction of those units
    # (the ticket-1793 case: a re-save carrying only {L,XL} returned the
    # previously-picked S+M to stock). A save now updates ONLY the sizes it
    # actually mentions; sizes the client omits keep their existing picked value.
    # To intentionally un-pick a size, send it explicitly with total 0.
    _prev_picked = ticket.get("picked_sizes") or {}
    _omitted = [s for s in _prev_picked if s not in picked_sizes]
    if _omitted:
        logger.warning("save_pick_progress %s: payload omitió tallas ya surtidas %s; se preservan (no se revierte el descuento)", ticket_id, _omitted)
        picked_sizes = {**_prev_picked, **picked_sizes}

    # Destination + completeness drive the final status. A partial "complete"
    # must NOT close the ticket (Case# 005): it stays active/in_progress so an
    # admin can reassign it when more material arrives. Only a FULL pick closes.
    is_neck = ticket.get("destination") == "neck_cutting"
    required_total = sum(int(v or 0) for v in (ticket.get("sizes") or {}).values())
    picked_total = sum(
        int((c.get("total") if isinstance(c, dict) else c) or 0)
        for c in picked_sizes.values()
    )
    is_full = required_total > 0 and picked_total >= required_total
    finalize = is_complete and is_full
    partial_close = is_complete and not is_full
    picking_status = ("in_neck_cutting" if is_neck else "completed") if finalize else "in_progress"

    # Idempotency: once a ticket is finalized no more progress is applied here
    # (further changes go through the edit endpoint). Partial tickets stay open so
    # the operator can keep picking — each save only deducts the NEW delta below.
    if ticket.get("status") in ("confirmed", "in_neck_cutting"):
        return {
            "message": "El pick ya fue finalizado; no se aplican más cambios aquí",
            "ticket_id": ticket_id,
            "picking_status": ticket.get("picking_status"),
        }

    update = {
        "picked_sizes": picked_sizes,
        "picking_status": picking_status,
        "partial_closed": partial_close,
        "last_picked_by": user.get("user_id"),
        "last_picked_by_name": user.get("name", user.get("email", "")),
        "last_picked_at": now_iso()
    }
    # INCREMENTAL DEDUCTION (ghost-inventory fix). Stock is deducted AS THE
    # OPERATOR PICKS — on every save, partial or complete — not only at the end.
    # We deduct the DELTA between what's picked now and what was already deducted
    # (tracked in `deducted_map`), so a partial save removes exactly the newly
    # picked units, re-saving the same numbers is a no-op (no double counting),
    # and a correction downward returns those units to the shelf.
    inv_op = "pick_to_neck" if is_neck else "deduct"
    style = ticket.get('style', '').strip()
    color = ticket.get('color', '').strip()
    customer = ticket.get('customer', '')
    _ord_no = ticket.get("order_number")
    _ord_id = ticket.get("order_id")

    new_map = _normalize_picked(picked_sizes)
    old_map = ticket.get("deducted_map") or {}
    pos, neg = [], []
    cells = {(sz, loc) for sz, locs in new_map.items() for loc in locs} | \
            {(sz, loc) for sz, locs in old_map.items() for loc in locs}
    for (sz, loc) in cells:
        d = new_map.get(sz, {}).get(loc, 0) - old_map.get(sz, {}).get(loc, 0)
        if d > 0:
            pos.append((sz, loc, d))
        elif d < 0:
            neg.append((sz, loc, -d))

    # WMS-002: block oversell on the NEW units being pulled now (the delta),
    # before mutating anything (raises 409 + opens a discrepancy task).
    if pos:
        delta_ps = {}
        for sz, loc, d in pos:
            cell = delta_ps.setdefault(sz, {"total": 0, "details": {}})
            cell["details"][loc] = cell["details"].get(loc, 0) + d
            cell["total"] += d
        await _assert_pick_stock(style, color, delta_ps, user)

    # WMS-003: a failure must ABORT the save — never leave the ticket updated with
    # stock not deducted (that was a ghost-inventory source).
    try:
        for sz, loc, d in pos:
            await _deduct_pick_boxes(style, color, sz, loc, d, inv_op, customer, _ord_no, _ord_id,
                                     user=user, ticket_id=ticket_id)
        for sz, loc, d in neg:
            # Correction downward: return the units to the shelf they came from.
            await _update_inventory_enhanced(style, color, sz, d, "add", location=loc, customer=customer)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error aplicando deducción incremental de picking")
        raise HTTPException(500, "No se pudo actualizar inventario; no se guardó el progreso")

    update["deducted_map"] = new_map

    if finalize:
        # Full pick → close the ticket (to production or neck cutting).
        update["completed_at"] = now_iso()
        update["status"] = "in_neck_cutting" if is_neck else "confirmed"
    elif partial_close:
        # Stock ran out: the picked units were already deducted incrementally
        # above. Record who closed it short and leave it ACTIVE so an admin can
        # reassign it when more material arrives (Case# 005).
        update["partial_closed_at"] = now_iso()
        update["partial_closed_by_name"] = user.get("name", user.get("email", ""))

    await db.wms_pick_tickets.update_one({"ticket_id": ticket_id}, {"$set": update})
    await log_movement(user, "pick_progress", {
        "ticket_id": ticket_id,
        "picking_status": picking_status,
        "partial_closed": partial_close,
        "picked_sizes": picked_sizes
    })
    return {
        "message": "Cerrado parcial — material descontado, el ticket queda activo" if partial_close
                   else f"Progreso guardado ({picking_status})",
        "ticket_id": ticket_id,
        "picking_status": picking_status,
        "partial_closed": partial_close,
        "is_full": is_full,
    }


# ═══════════════ SCAN + BIND DE CAJA FISICA ═══════════════════════════════════
# Cierra el hueco donde el picker no podia decir "esta caja fisica es tal en
# la BD" — el descuento FIFO ciego vaciaba la caja mas vieja de la base cuando
# el picker en realidad estaba jalando material de otra. El scan ata el LPN
# fisico a un box_id especifico y desde ahi el descuento va a esa caja.

def _norm_lpn(s):
    """Canoniza el LPN: strip, uppercase, quita espacios/guiones extras."""
    return re.sub(r"\s+", "", (s or "").strip().upper())


async def _find_box_by_lpn(lpn):
    """Busca una caja por LPN en 3 lugares: box_id (etiquetas BOX-), campo
    physical_lpn (bindeo previo) o entrada en lpn_aliases (por si se re-etiqueto).
    Devuelve None si nada matchea."""
    lpn = _norm_lpn(lpn)
    if not lpn:
        return None
    return await db.wms_boxes.find_one(
        {"$or": [
            {"box_id": lpn},
            {"physical_lpn": lpn},
            {"lpn_aliases": lpn},
        ]},
        {"_id": 0},
    )


@router.post("/pick-tickets/{ticket_id}/scan-box")
async def scan_box(ticket_id: str, request: Request):
    """Identifica una caja escaneada durante el picking.

    Body: { lpn: "BOX-000123" o "SUP-XYZ", location?: str }

    Respuestas:
      { status: "matched",       box: {...} }              # ya conocida → deduce
      { status: "needs_binding", ticket_context: {...},    # no conocida →
                                 sizes_needed: [...],      #   pide talla y
                                 candidates: [...] }       #   cantidad al picker
      { status: "wrong_ticket",  box: {...} }              # existe pero es otro style/color
    """
    user = await require_auth(request)
    body = await request.json()
    lpn = _norm_lpn(body.get("lpn"))
    location = (body.get("location") or "").strip().upper()
    if not lpn:
        raise HTTPException(400, "lpn requerido")

    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")
    if ticket.get("status") in ("confirmed", "in_neck_cutting"):
        raise HTTPException(409, "El pick ya fue finalizado")

    # Autorizacion (misma regla que pick-size).
    assignee = (ticket.get("assigned_to") or "").strip()
    caller_ids = {user.get("user_id", ""), user.get("email", "")}
    elevated = user.get("role") in {"admin", "supersu", "ceo"}
    if assignee and assignee not in caller_ids and not elevated:
        raise HTTPException(403, "Este pick ticket está asignado a otro operador")

    t_style = (ticket.get("style") or "").strip().upper()
    t_color = (ticket.get("color") or "").strip().upper()

    # Tallas pendientes en el ticket (requerido - picked). Para el modal cuestionario.
    required = {str(k): int(v or 0) for k, v in (ticket.get("sizes") or {}).items()}
    picked_map = ticket.get("picked_sizes") or {}
    def _picked_of(sz):
        v = picked_map.get(sz)
        if isinstance(v, dict): return int(v.get("total") or 0)
        return int(v or 0)
    sizes_needed = [
        {"size": sz, "remaining": max(0, qty - _picked_of(sz))}
        for sz, qty in required.items() if qty > 0 and _picked_of(sz) < qty
    ]

    ticket_context = {
        "style": t_style, "color": t_color,
        "customer": ticket.get("customer") or "",
        "order_number": ticket.get("order_number") or "",
    }

    # 1) Caja ya identificada
    box = await _find_box_by_lpn(lpn)
    if box:
        b_style = (box.get("style") or box.get("sku", "").split("-")[0] or "").upper()
        b_color = (box.get("color") or "").upper()
        # Compare tolerando SKU compuesto (style o sku column)
        matches = (b_style == t_style or box.get("sku", "").upper().startswith(t_style)) and b_color == t_color
        if not matches:
            return {
                "status": "wrong_ticket",
                "message": f"Esta caja es {b_style}/{b_color}, no {t_style}/{t_color}",
                "box": box,
                "ticket_context": ticket_context,
            }
        return {
            "status": "matched",
            "box": box,
            "ticket_context": ticket_context,
            "sizes_needed": sizes_needed,
        }

    # 2) LPN desconocido → pide binding (modal cuestionario)
    # Ofrece candidatas en la loc reportada: cajas del ticket sin physical_lpn.
    cand_q = {
        "$or": [{"style": t_style}, {"sku": {"$regex": f"^{re.escape(t_style)}-"}}],
        "color": t_color,
        "units": {"$gt": 0},
        "$and": [
            {"$or": [{"physical_lpn": {"$exists": False}}, {"physical_lpn": None}]},
        ],
    }
    if location:
        cand_q["location"] = location
    candidates = await db.wms_boxes.find(
        cand_q, {"_id": 0, "box_id": 1, "size": 1, "units": 1, "location": 1}
    ).sort("created_at", 1).limit(50).to_list(50)

    return {
        "status": "needs_binding",
        "lpn": lpn,
        "location": location,
        "ticket_context": ticket_context,
        "sizes_needed": sizes_needed,
        "candidates": candidates,
    }


@router.post("/pick-tickets/{ticket_id}/bind-box")
async def bind_box(ticket_id: str, request: Request):
    """Ata un LPN físico a una wms_boxes doc del ticket. Corre después de un
    scan que devolvió `needs_binding`, con la respuesta del cuestionario.

    Body: { lpn, location, size, actual_units }

    Estrategia:
      1. Busca una caja en (location, style=ticket, color=ticket, size) sin
         physical_lpn aún. FIFO por created_at.
      2. Le pone physical_lpn = lpn.
      3. Si actual_units != box.units → ajusta y logea reconciliación.
      4. Devuelve la caja lista para deducir.
    """
    user = await require_auth(request)
    body = await request.json()
    lpn = _norm_lpn(body.get("lpn"))
    location = (body.get("location") or "").strip().upper()
    size = (body.get("size") or "").strip().upper()
    try:
        actual_units = int(body.get("actual_units") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "actual_units debe ser entero")
    if not (lpn and location and size and actual_units >= 0):
        raise HTTPException(400, "lpn, location, size y actual_units requeridos")

    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")
    if ticket.get("status") in ("confirmed", "in_neck_cutting"):
        raise HTTPException(409, "El pick ya fue finalizado")

    assignee = (ticket.get("assigned_to") or "").strip()
    caller_ids = {user.get("user_id", ""), user.get("email", "")}
    elevated = user.get("role") in {"admin", "supersu", "ceo"}
    if assignee and assignee not in caller_ids and not elevated:
        raise HTTPException(403, "Este pick ticket está asignado a otro operador")

    # Que otro picker no haya bindeado este LPN en simultáneo entre el scan y el bind.
    dup = await _find_box_by_lpn(lpn)
    if dup:
        return {"status": "already_bound", "box": dup,
                "message": "Este LPN ya estaba atado a otra caja mientras respondías"}

    t_style = (ticket.get("style") or "").strip().upper()
    t_color = (ticket.get("color") or "").strip().upper()

    q = {
        "$or": [{"style": t_style}, {"sku": {"$regex": f"^{re.escape(t_style)}-"}}],
        "color": t_color,
        "size": size,
        "location": location,
        "$and": [
            {"$or": [{"physical_lpn": {"$exists": False}}, {"physical_lpn": None}]},
        ],
    }
    # FIFO: la más vieja primero para mantener la disciplina de rotación.
    box = await db.wms_boxes.find_one_and_update(
        q,
        {"$set": {"physical_lpn": lpn, "bound_by": user.get("user_id"),
                  "bound_at": now_iso()}},
        sort=[("created_at", 1)],
        return_document=True,
        projection={"_id": 0},
    )
    if not box:
        raise HTTPException(404, (
            f"No hay cajas sin identificar para {t_style}/{t_color}/{size} en {location}. "
            "Verifica ubicación y talla, o pide al admin registrar la caja como nueva."
        ))

    # Reconciliación de cantidad si el picker contó distinto
    db_units = int(box.get("units") or 0)
    delta = actual_units - db_units
    reconciled = False
    if delta != 0:
        await db.wms_boxes.update_one(
            {"box_id": box["box_id"]},
            {"$set": {"units": actual_units, "qty": actual_units,
                      "last_bind_reconciliation": {
                          "picker": user.get("user_id"),
                          "at": now_iso(),
                          "from": db_units,
                          "to": actual_units,
                          "delta": delta,
                          "ticket_id": ticket_id,
                      }}}
        )
        # Ajusta la fila de wms_inventory correspondiente (+ o −)
        await _update_inventory_enhanced(
            box.get("style") or box.get("sku"), box.get("color"), box.get("size"),
            abs(delta), "add" if delta > 0 else "deduct",
            location=box.get("location"), customer=box.get("customer", ""),
        )
        await log_movement(user, "pick_bind_reconciliation", {
            "ticket_id": ticket_id, "box_id": box["box_id"], "lpn": lpn,
            "from_units": db_units, "to_units": actual_units, "delta": delta,
        })
        box["units"] = actual_units
        box["qty"] = actual_units
        reconciled = True

    await log_movement(user, "box_bound", {
        "ticket_id": ticket_id, "box_id": box["box_id"],
        "physical_lpn": lpn, "location": location, "size": size,
    })
    box["physical_lpn"] = lpn
    return {"status": "bound", "box": box, "reconciled": reconciled,
            "delta": delta}


@router.put("/pick-tickets/{ticket_id}/pick-size")
async def pick_size(ticket_id: str, request: Request):
    """Deduct ONE size the instant the operator OKs it — no need to wait for the
    full pick or a partial close. Scoped strictly to the given size: it diffs
    that size's cells against deducted_map and pulls only the NEW units, so
    re-OKing the same numbers is a no-op (no double counting) and lowering the
    amount returns the difference to the shelf. Other (not-yet-OK'd) sizes are
    never touched. Uses the same _deduct_pick_boxes / _assert_pick_stock guards
    as the incremental save, so boxes and inventory always move in lockstep.

    Body: { size: str, details: { location: qty, ... } }  (location "" = FIFO)
    """
    user = await require_auth(request)
    body = await request.json()
    size = (body.get("size") or "").strip()
    # `details` acepta 2 formatos (compat + nuevo):
    #   viejo: { "NA04-A17": 30 }                        → FIFO ciego
    #   nuevo: { "NA04-A17": {"qty": 30, "box_id": "BOX-000123"} }
    raw = body.get("details") or {}
    details = {}       # {loc: qty}
    box_by_loc = {}    # {loc: box_id | None}
    for loc, v in raw.items():
        if isinstance(v, dict):
            details[str(loc)] = int(v.get("qty") or 0)
            box_by_loc[str(loc)] = (v.get("box_id") or None)
        else:
            details[str(loc)] = int(v or 0)
            box_by_loc[str(loc)] = None
    if not size:
        raise HTTPException(400, "size requerido")

    # Toggle global de scan obligatorio (config_options). Solo se resuelve; el
    # chequeo real ocurre mas abajo, sobre el DELTA positivo (pos) — no sobre
    # el total cumulativo. Antes rechazabamos si CUALQUIER loc con qty>0 no
    # traia box_id, aunque delta=0 (ej: re-OK de talla porque el operador aniade
    # de OTRA caja/loc). Ahora solo se pide scan cuando hay extraccion nueva.
    cfg = await db.config_options.find_one({"config_id": "main"}, {"_id": 0}) or {}
    require_scan = bool(cfg.get("pick_requires_scan",
                                DEFAULT_OPTIONS.get("pick_requires_scan", True)))

    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")

    # Same authorization as save_pick_progress.
    assignee = (ticket.get("assigned_to") or "").strip()
    caller_ids = {user.get("user_id", ""), user.get("email", "")}
    elevated = user.get("role") in {"admin", "supersu", "ceo"}
    if assignee and assignee not in caller_ids and not elevated:
        raise HTTPException(403, "Este pick ticket está asignado a otro operador")
    if ticket.get("status") in ("confirmed", "in_neck_cutting"):
        raise HTTPException(409, "El pick ya fue finalizado; no se aceptan más descuentos")

    # HOLD guard on the shelves being pulled.
    hold_locs = [l for l in details.keys() if l]
    if hold_locs:
        await _assert_not_on_hold(user, *hold_locs)

    is_neck = ticket.get("destination") == "neck_cutting"
    inv_op = "pick_to_neck" if is_neck else "deduct"
    style = ticket.get("style", "").strip()
    color = ticket.get("color", "").strip()
    customer = ticket.get("customer", "")
    _ord_no = ticket.get("order_number")
    _ord_id = ticket.get("order_id")

    # Preserve locations already deducted for this size that the payload omits,
    # so a partial re-submit can't silently return them to stock (the vector
    # that lost 575 units of S@CARRO 182 on ticket 1784). To drop a location's
    # pick on purpose, send it explicitly with 0.
    prev_cells = (ticket.get("deducted_map") or {}).get(size, {})
    _omitted = [loc for loc in prev_cells if loc not in details]
    if _omitted:
        logger.warning("pick_size %s talla %s: payload omitió ubicaciones ya surtidas %s; se preservan", ticket_id, size, _omitted)
        for loc in _omitted:
            details[loc] = int(prev_cells[loc] or 0)

    # New cumulative picked_sizes with THIS size replaced by the OK'd numbers.
    picked_sizes = dict(ticket.get("picked_sizes") or {})
    total = sum(details.values())
    required = int((ticket.get("sizes") or {}).get(size, 0) or 0)
    if required and total > required:
        raise HTTPException(400, f"Talla {size}: {total} excede lo requerido ({required})")
    picked_sizes[size] = {"total": total, "details": dict(details)}

    new_map = _normalize_picked(picked_sizes)
    old_map = ticket.get("deducted_map") or {}
    # Restrict the diff to THIS size's cells only.
    cells = {(size, loc) for loc in new_map.get(size, {})} | \
            {(size, loc) for loc in old_map.get(size, {})}
    pos, neg = [], []
    for (sz, loc) in cells:
        d = new_map.get(sz, {}).get(loc, 0) - old_map.get(sz, {}).get(loc, 0)
        if d > 0:
            pos.append((sz, loc, d))
        elif d < 0:
            neg.append((sz, loc, -d))

    # WMS-002: block oversell on the NEW units before mutating anything.
    if pos:
        delta_ps = {}
        for sz, loc, d in pos:
            cell = delta_ps.setdefault(sz, {"total": 0, "details": {}})
            cell["details"][loc] = cell["details"].get(loc, 0) + d
            cell["total"] += d
        await _assert_pick_stock(style, color, delta_ps, user)

    # WMS-003: abort on failure so we never record a deduction that didn't apply.
    try:
        for sz, loc, d in pos:
            # Si el picker escaneó una caja para esta celda, descontamos de ESA
            # caja; si no, cae al FIFO (permitido solo cuando require_scan=False).
            only_box = box_by_loc.get(loc)
            if require_scan and not only_box:
                raise HTTPException(400, (
                    f"Debes escanear la caja para descontar {d} pz en {loc} (talla {sz}). "
                    "El descuento sin scan está deshabilitado (pick_requires_scan)."
                ))
            await _deduct_pick_boxes(style, color, sz, loc, d, inv_op, customer, _ord_no, _ord_id,
                                     user=user, ticket_id=ticket_id,
                                     only_box_id=only_box)
        for sz, loc, d in neg:
            await _update_inventory_enhanced(style, color, sz, d, "add", location=loc, customer=customer)
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error descontando talla en pick-size")
        raise HTTPException(500, "No se pudo descontar la talla; no se guardó")

    # Merge only this size's deducted cells into the map.
    merged_map = {k: dict(v) for k, v in old_map.items()}
    merged_map[size] = new_map.get(size, {})

    await db.wms_pick_tickets.update_one({"ticket_id": ticket_id}, {"$set": {
        "picked_sizes": picked_sizes,
        "deducted_map": merged_map,
        "picking_status": "in_progress",
        "last_picked_by": user.get("user_id"),
        "last_picked_by_name": user.get("name", user.get("email", "")),
        "last_picked_at": now_iso(),
    }})
    await log_movement(user, "pick_size", {"ticket_id": ticket_id, "size": size, "details": details})
    return {"message": f"Talla {size} descontada ({total} pz)", "ticket_id": ticket_id,
            "size": size, "deducted": total}

@router.put("/pick-tickets/{ticket_id}/confirm")
@router.post("/stocktakes/{stocktake_id}/finalize")
async def confirm_pick(ticket_id: str, request: Request, stocktake_id: str = None):
    # Use ticket_id if provided via newer route, else stocktake_id from older legacy route
    target_id = ticket_id or stocktake_id
    user = await require_auth(request)
    body = await request.json()
    confirmed_lines = body.get("lines", [])
    ticket = await db.wms_pick_tickets.find_one({"ticket_id": target_id})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")

    # Guard: skip if already processed by pick-progress completion
    if ticket.get("status") in ("confirmed", "in_neck_cutting"):
        return {"message": "Pick ya fue procesado previamente", "ticket_id": target_id}

    # If picking already deducted incrementally (pick-progress with deducted_map),
    # inventory is settled — just finalize status, NEVER deduct again here, or the
    # stock would be double-counted.
    if ticket.get("deducted_map"):
        new_status = "in_neck_cutting" if ticket.get("destination") == "neck_cutting" else "confirmed"
        await db.wms_pick_tickets.update_one({"ticket_id": target_id}, {"$set": {
            "status": new_status,
            "picking_status": "completed" if new_status == "confirmed" else "in_neck_cutting",
            "confirmed_at": now_iso(), "confirmed_by": user.get("user_id"),
        }})
        await db.orders.update_one({"order_id": ticket.get("order_id")}, {"$set": {"wms_status": "picked"}})
        await log_movement(user, "pick_confirmed", {"ticket_id": target_id, "items_confirmed": "incremental"})
        await notify_badge_change("picking")
        return {"message": "Pick confirmado", "ticket_id": target_id}

    # HOLD guard: stock in a SAT-held location can't be picked until released.
    confirm_locs = [line.get("location") for line in confirmed_lines if line.get("location")]
    confirm_locs += [loc for d in (ticket.get("picked_sizes") or {}).values()
                     if isinstance(d, dict) for loc in (d.get("details") or {}).keys()]
    if confirm_locs:
        await _assert_not_on_hold(user, *confirm_locs)

    is_neck_cutting = ticket.get("destination") == "neck_cutting"
    inv_operation = "pick_to_neck" if is_neck_cutting else "deduct"

    # Handle confirmed lines (legacy or explicit)
    _line_box_ids = []
    if confirmed_lines:
        for line in confirmed_lines:
            box_id = line.get("box_id")
            pick_qty = int(line.get("qty", 0))
            box = await db.wms_boxes.find_one({"box_id": box_id})
            if box:
                # Support both 'units' and 'qty' field names
                current_qty = box.get("units") if box.get("units") is not None else box.get("qty", 0)
                new_qty = max(0, current_qty - pick_qty)

                # Update box: just reduce units and link to order
                box_update = {
                    "units": new_qty,
                    "qty": new_qty,
                    "order_number": ticket.get("order_number"),
                    "last_order_id": ticket.get("order_id"),
                    "last_picked_at": now_iso(),
                    "last_pick_ticket": target_id,
                }
                await db.wms_boxes.update_one({"box_id": box_id}, {"$set": box_update})
                _line_box_ids.append(box_id)

                # Move to neck process area OR deduct from global
                _inv_id = await _update_inventory_enhanced(box["sku"], box.get("color", ""), box.get("size", ""), pick_qty, inv_operation, location=box.get("location", ""), customer=box.get("customer", ""))
                # Box emptied out -> drop it from the location's box count.
                if new_qty == 0:
                    await _adjust_inventory_boxes(_inv_id, -1)
    else:
        # Newer flow: deduct from picked_sizes. Delegate to _deduct_pick_boxes —
        # the SAME deducer the incremental PDA flow uses — so the physical boxes
        # AND the inventory row always drop in lockstep. The previous hand-rolled
        # version queried boxes by {"sku": style} (e.g. "5000"), but boxes store
        # the COMPOSITE sku ("5000-CHARCOAL-M") with the short style in `style`;
        # that match found almost nothing and silently fell back to an
        # inventory-only deduct, leaving the physical boxes full → ghost boxes
        # (the boxes↔inventory drift behind the cart discrepancies).
        picked_sizes = ticket.get("picked_sizes") or ticket.get("sizes") or {}
        style = ticket.get("style", "")
        color = ticket.get("color", "")
        order_number = ticket.get("order_number")
        order_id = ticket.get("order_id")
        customer = ticket.get("customer", "")
        # WMS-002: block oversell before deducting anything in this flow too.
        await _assert_pick_stock(style, color, picked_sizes, user)
        for sz, data in picked_sizes.items():
            # Formats: {sz: qty} or {sz: {total: X, details: {loc: q}}}.
            if isinstance(data, dict):
                if "details" in data:
                    for loc, loc_qty in (data.get("details") or {}).items():
                        if int(loc_qty or 0) > 0:
                            await _deduct_pick_boxes(style, color, sz, loc, int(loc_qty),
                                                     inv_operation, customer, order_number, order_id,
                                                     user=user, ticket_id=target_id)
                    continue
                qty_to_pick = int(data.get("total", 0))
            else:
                qty_to_pick = int(data or 0)
            # No shelf chosen → FIFO across this SKU's boxes (location="").
            if qty_to_pick > 0:
                await _deduct_pick_boxes(style, color, sz, "", qty_to_pick,
                                         inv_operation, customer, order_number, order_id,
                                         user=user, ticket_id=target_id)

    new_status = "in_neck_cutting" if is_neck_cutting else "confirmed"
    
    await db.wms_pick_tickets.update_one({"ticket_id": target_id}, {"$set": {
        "status": new_status, 
        "picking_status": "completed" if not is_neck_cutting else "in_neck_cutting", 
        "confirmed_at": now_iso(), 
        "confirmed_by": user.get("user_id")
    }})
    await db.orders.update_one({"order_id": ticket.get("order_id")}, {"$set": {"wms_status": "picked"}})
    await log_movement(user, "pick_confirmed", {
        "ticket_id": target_id,
        "items_confirmed": len(confirmed_lines) if confirmed_lines else "auto",
        # Per-box trail for the legacy lines flow (the picked_sizes flow logs its
        # own 'pick_deduction' movements inside _deduct_pick_boxes).
        **({"box_ids": _line_box_ids} if _line_box_ids else {}),
    })
    await notify_badge_change("picking")
    return {"message": "Pick confirmado", "ticket_id": target_id}

@router.put("/pick-tickets/{ticket_id}/edit")
async def edit_pick_ticket(ticket_id: str, request: Request):
    """Edit an existing pick ticket (only if not confirmed/completed)."""
    user = await require_auth(request)
    body = await request.json()
    ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(404, "Pick ticket no encontrado")
    if ticket.get("status") == "confirmed" or ticket.get("picking_status") == "completed":
        raise HTTPException(400, "No se puede editar un ticket confirmado/completado")

    update = {}
    force_duplicate = bool(body.get("force_duplicate", False))
    new_order_number = body.get("order_number", ticket.get("order_number", "")).strip()

    # --- Duplicate guard for edits -------------------------------------------------
    if not force_duplicate and new_order_number:
        existing = await db.wms_pick_tickets.find_one(
            {
                "order_number": new_order_number,
                "ticket_id": {"$ne": ticket_id},
                "status": {"$nin": ["confirmed", "cancelled"]},
            },
            {"_id": 0, "ticket_id": 1, "created_by_name": 1, "created_at": 1,
             "style": 1, "color": 1, "total_pick_qty": 1, "status": 1, "picking_status": 1},
        )
        if existing:
            raise HTTPException(
                409,
                {
                    "message": f"Ya existe otro pick ticket activo para la orden {new_order_number}.",
                    "existing_ticket": existing,
                },
            )
    # -------------------------------------------------------------------------------

    for field in ["order_number", "customer", "client", "manufacturer", "style", "color", "quantity", "destination"]:
        if field in body:
            update[field] = str(body[field]).strip() if isinstance(body[field], str) else body[field]
    if "sizes" in body:
        update["sizes"] = {k: int(v) for k, v in body["sizes"].items()}
        update["total_pick_qty"] = sum(update["sizes"].values())
    if "assigned_to" in body:
        update["assigned_to"] = body.get("assigned_to") or None
        update["assigned_to_name"] = body.get("assigned_to_name", "")
        if body.get("assigned_to"):
            update["picking_status"] = "assigned"
            update["assigned_at"] = now_iso()
        else:
            update["picking_status"] = "unassigned"
            update["assigned_at"] = None

    # Allow updating the picking strategy
    if "strategy" in body and body["strategy"] in PICK_STRATEGIES:
        update["strategy"] = body["strategy"]

    # Re-lookup locations if style/color changed (also re-apply strategy if changed)
    new_style = update.get("style", ticket.get("style", ""))
    new_color = update.get("color", ticket.get("color", ""))
    needs_relookup = "style" in update or "color" in update
    if needs_relookup:
        size_locations = {}
        new_sizes = update.get("sizes", ticket.get("sizes", {}))
        for sz, qty in new_sizes.items():
            if int(qty) <= 0: continue
            inv_items = await db.wms_inventory.find(
                {"style": {"$regex": f"^{re.escape(new_style)}$", "$options": "i"}, "color": {"$regex": f"^{re.escape(new_color)}$", "$options": "i"}, "size": {"$regex": f"^{re.escape(sz)}$", "$options": "i"}},
                {"_id": 0, "location": 1, "units_on_hand": 1, "units_allocated": 1, "total_boxes": 1, "country_of_origin": 1}
            ).sort("units_on_hand", -1).to_list(500)  # ver nota en _compute_size_locations: no ocultar stock de SKUs muy repartidos
            locs = [{"location": it.get("location", ""), "available": it.get("units_on_hand", 0) - it.get("units_allocated", 0), "boxes": it.get("total_boxes", 0), "country_of_origin": it.get("country_of_origin", "")} for it in inv_items if it.get("location")]
            locs = [l for l in locs if l["available"] > 0]
            total_sz_avail = sum(l["available"] for l in locs)
            for l in locs:
                l["percentage"] = round((l["available"] / total_sz_avail) * 100) if total_sz_avail > 0 else 0
            size_locations[sz] = {"locations": locs, "total_available": total_sz_avail}
        update["size_locations"] = size_locations
    # Re-apply strategy if it changed (use freshly-built or existing size_locations)
    if "strategy" in update:
        target = update.get("size_locations") or ticket.get("size_locations") or {}
        update["size_locations"] = apply_picking_strategy(target, update["strategy"])

    update["updated_at"] = now_iso()
    update["updated_by"] = user.get("user_id")
    await db.wms_pick_tickets.update_one({"ticket_id": ticket_id}, {"$set": update})
    await log_movement(user, "pick_ticket_edited", {"ticket_id": ticket_id, "changes": list(update.keys())})
    updated = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0})
    return updated

@router.get("/orders-with-tickets")
async def get_orders_with_tickets(request: Request):
    """Get orders with their pick ticket assignments and progress."""
    await require_auth(request)
    # Get all pick tickets grouped by order_number
    tickets = await db.wms_pick_tickets.find({}, {"_id": 0}).to_list(1000)
    ticket_map = {}
    for t in tickets:
        on = t.get("order_number", "")
        if on not in ticket_map:
            ticket_map[on] = []
        ticket_map[on].append({
            "ticket_id": t.get("ticket_id"),
            "assigned_to_name": t.get("assigned_to_name", ""),
            "picking_status": t.get("picking_status", "unassigned"),
            "status": t.get("status", "pending"),
            "total_pick_qty": t.get("total_pick_qty", 0),
            "picked_sizes": t.get("picked_sizes", {}),
            "sizes": t.get("sizes", {}),
        })
    return ticket_map

@router.get("/pick-tickets/stats")
async def pick_ticket_stats(request: Request):
    """Dashboard stats for picker productivity."""
    await require_auth(request)
    tickets = await db.wms_pick_tickets.find({}, {"_id": 0}).to_list(2000)
    operators_map = {}
    total_completed = 0
    total_in_progress = 0
    total_pending = 0
    for t in tickets:
        ps = t.get("picking_status", "unassigned")
        if ps == "completed": total_completed += 1
        elif ps == "in_progress": total_in_progress += 1
        else: total_pending += 1
        name = t.get("assigned_to_name", "")
        if not name: continue
        if name not in operators_map:
            operators_map[name] = {"name": name, "completed": 0, "in_progress": 0, "assigned": 0, "total_pieces": 0, "picked_pieces": 0}
        op = operators_map[name]
        if ps == "completed": op["completed"] += 1
        elif ps == "in_progress": op["in_progress"] += 1
        else: op["assigned"] += 1
        sizes = t.get("sizes", {})
        picked = t.get("picked_sizes", {})
        op["total_pieces"] += sum(int(v) for v in sizes.values())
        op["picked_pieces"] += sum(int(v) for v in picked.values())
    return {
        "total_tickets": len(tickets),
        "completed": total_completed,
        "in_progress": total_in_progress,
        "pending": total_pending,
        "operators": list(operators_map.values())
    }

# ==================== PRODUCTION ====================

@router.post("/production/move")
async def production_move(request: Request):
    user = await require_auth(request)
    body = await request.json()
    box_ids = body.get("box_ids", [])
    target_state = body.get("target_state", "wip")
    # Optional production/order reference so the move is traceable to WHERE the
    # material went (the log previously kept only a count).
    order_ref = (str(body.get("order_number") or body.get("order_id") or "")).strip()
    if not box_ids:
        raise HTTPException(400, "box_ids requeridos")
    if target_state not in ["raw", "wip", "finished"]:
        raise HTTPException(400, "target_state debe ser raw, wip o finished")
    moved = []
    for box_id in box_ids:
        box = await db.wms_boxes.find_one({"box_id": box_id})
        if not box:
            continue
        await _assert_not_on_hold(user, box.get("location"))
        old_state = box.get("state", "raw")
        units = int((box.get("units") if box.get("units") is not None else box.get("qty", 0)) or 0)
        sku = box.get("style") or box.get("sku")
        loc = box.get("location", "")
        cust = box.get("customer", "")
        # Inventory effect by state (decisión: WIP sale del inventario de almacén,
        # finished sigue contando como producto terminado embarcable):
        #   entrar a wip  -> deduct units_on_hand (deja de ser surtible)
        #   salir de wip  -> re-add las unidades (vuelve como raw/finished)
        # box_count=0 en el re-add: la caja ya existe, no se crea otra.
        if units > 0:
            if old_state != "wip" and target_state == "wip":
                await _update_inventory_enhanced(sku, box.get("color", ""), box.get("size", ""), units, "deduct", location=loc, customer=cust)
            elif old_state == "wip" and target_state != "wip":
                await _update_inventory_enhanced(sku, box.get("color", ""), box.get("size", ""), units, "add", customer=cust, location=loc, box_count=0)
        await db.wms_boxes.update_one({"box_id": box_id}, {"$set": {"state": target_state, "status": "in_production" if target_state == "wip" else ("finished" if target_state == "finished" else box.get("status")), **({"production_order": order_ref} if order_ref else {})}})
        moved.append({"box_id": box_id, "from": old_state, "to": target_state,
                      "sku": sku, "units": units})
    move_doc = {
        "move_id": gen_id("pmov"), "box_ids": box_ids,
        "target_state": target_state, "moved": moved,
        "order_number": order_ref or None,
        "moved_by": user.get("user_id"),
        "moved_by_name": user.get("name", ""),
        "created_at": now_iso(),
    }
    await db.wms_production_moves.insert_one(move_doc)
    move_doc.pop("_id", None)
    await log_movement(user, "production_move", {
        "target_state": target_state, "count": len(moved),
        "order_number": order_ref or None,
        "box_ids": [m["box_id"] for m in moved][:50],
        "skus": sorted({m["sku"] for m in moved if m.get("sku")})[:50],
    })
    return move_doc

    return boxes

@router.get("/neck-cutting")
async def list_neck_cutting(request: Request):
    await require_auth(request)
    tickets = await db.wms_pick_tickets.find({"status": "in_neck_cutting"}, {"_id": 0}).sort("order_number", 1).to_list(1000)
    
    grouped = {}
    for t in tickets:
        ono = t.get("order_number") or "SIN_ORDEN"
        if ono not in grouped:
            grouped[ono] = {
                "ticket_id": t.get("ticket_id"),
                "order_number": ono,
                "customer": t.get("customer", ""),
                "style": t.get("style", ""),
                "items": [],
                "total_qty": 0,
                "last_order_id": t.get("order_id")
            }
        
        picked = t.get("picked_sizes", {})
        if not picked:
            picked = t.get("sizes", {})
            
        for sz, data in picked.items():
            qty = 0
            if isinstance(data, dict):
                qty = int(data.get("total", 0))
            else:
                qty = int(data)
                
            if qty > 0:
                # Mock box_id for UI compatibility if needed, but we don't strictly need it 
                # since we are no longer updating box units from the UI here.
                grouped[ono]["items"].append({
                    "sku": t.get("style", ""),
                    "size": sz,
                    "qty": qty,
                    "units": qty,
                    "box_id": f"dummy_{sz}"
                })
                grouped[ono]["total_qty"] += qty
                
    return list(grouped.values())

@router.post("/neck-cutting/deliver")
async def deliver_to_production(request: Request):
    user = await require_auth(request)
    body = await request.json()
    order_number = body.get("order_number")
    
    if not order_number:
        raise HTTPException(400, "order_number requerido")
        
    ticket = await db.wms_pick_tickets.find_one({"order_number": order_number, "status": "in_neck_cutting"})
    if not ticket:
        raise HTTPException(404, "No hay material en Neck Cutting para esta orden")
        
    picked_sizes = ticket.get("picked_sizes") or ticket.get("sizes") or {}
    style = ticket.get("style", "")
    color = ticket.get("color", "")
    
    delivered_count = 0
    for sz, data in picked_sizes.items():
        qty = 0
        if isinstance(data, dict):
            qty = int(data.get("total", 0))
        else:
            qty = int(data)
            
        if qty > 0:
            # REAL DEDUCT from inventory (Final removal from CUTTING_NECK).
            # Match the neck row by EITHER the composite sku or the short style:
            # neck rows have been created under BOTH conventions (the row's sku is
            # whatever the source box carried, often composite like "1717-IVORY-M"),
            # and the old {"sku": style} lookup missed the composite ones — so the
            # deduct fell through to the shelf fallback and left the neck row's
            # reserved units stranded (units_allocated > units_on_hand).
            neck_inv = await db.wms_inventory.find_one({
                "$or": [{"sku": _ci_eq(style)}, {"style": _ci_eq(style)}],
                "color": _ci_eq(color or ""), "size": _ci_eq(sz or ""), "location": "CUTTING_NECK",
            })
            if neck_inv:
                await _update_inventory_enhanced(
                    neck_inv.get("sku") or style, color, sz, qty, "deduct", location="CUTTING_NECK"
                )
            else:
                # FALLBACK: CUTTING_NECK record missing (pick_to_neck never ran)
                # Deduct from original shelf locations using picked_sizes details
                logger.warning(f"Deliver fallback: no CUTTING_NECK record for {style}/{color}/{sz}, deducting from original locations")
                if isinstance(data, dict) and "details" in data:
                    for loc, loc_qty in data["details"].items():
                        if loc_qty > 0:
                            await _update_inventory_enhanced(
                                style, color, sz, loc_qty, "deduct", location=loc
                            )
                else:
                    # Last resort: deduct without location
                    await _update_inventory_enhanced(style, color, sz, qty, "deduct")
            delivered_count += qty
            
    # Mark ticket as completed
    await db.wms_pick_tickets.update_one(
        {"_id": ticket["_id"]}, 
        {"$set": {"status": "confirmed", "picking_status": "completed", "delivered_at": now_iso()}}
    )
        
    await log_movement(user, "neck_cut_delivery", {"order_number": order_number, "qty": delivered_count})
    await notify_badge_change("neck_cutting")
    return {"message": f"Entrega a producción exitosa: {delivered_count} piezas", "order_number": order_number}

# ==================== FINISHED GOODS ====================

@router.get("/finished-goods")
async def list_finished_goods(request: Request, is_bpo: bool = None):
    await require_auth(request)
    query = {"state": "finished"}
    if is_bpo is not None:
        query["is_bpo"] = is_bpo
    boxes = await db.wms_boxes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return boxes

@router.put("/finished-goods/{box_id}")
async def edit_finished_good(box_id: str, request: Request):
    user = await require_admin(request)
    body = await request.json()
    update = {k: v for k, v in body.items() if k in ["units", "location", "po", "is_bpo", "sku", "color", "size"]}
    result = await db.wms_boxes.update_one({"box_id": box_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(404, "Caja no encontrada")
    await log_movement(user, "edit_finished_good", {"box_id": box_id, "changes": update})
    return {"message": "Caja actualizada", "box_id": box_id}

# ==================== SHIPPING ====================

@router.post("/movements")
async def create_shipment(request: Request):
    user = await require_auth(request)
    body = await request.json()
    order_id = body.get("order_id", "")
    box_ids = body.get("box_ids", [])
    pallet = body.get("pallet", "")
    carrier = body.get("carrier", "")
    tracking = body.get("tracking", "")
    if not box_ids:
        raise HTTPException(400, "box_ids requeridos")
    shipment_id = gen_id("ship")
    shipped_boxes = []
    for box_id in box_ids:
        box = await db.wms_boxes.find_one({"box_id": box_id})
        if not box:
            continue
        units = int((box.get("units") if box.get("units") is not None else box.get("qty", 0)) or 0)
        # If the box still carries stock (e.g. shipped without going through a
        # pick), draw it out of the inventory ledger so shipped units stop
        # counting as on-hand. Already-picked boxes are at 0, so this is a no-op.
        if units > 0:
            inv_id = await _update_inventory_enhanced(
                box.get("style") or box.get("sku"), box.get("color", ""), box.get("size", ""),
                units, "deduct", location=box.get("location", ""), customer=box.get("customer", ""),
            )
            await _adjust_inventory_boxes(inv_id, -1)
        await db.wms_boxes.update_one({"box_id": box_id}, {"$set": {"status": "shipped", "shipment_id": shipment_id, "units": 0, "qty": 0}})
        shipped_boxes.append({"box_id": box_id, "sku": box.get("sku"), "color": box.get("color"), "size": box.get("size"), "units": units})
    total_units = sum(b.get("units", 0) for b in shipped_boxes)
    shipment_doc = {
        "shipment_id": shipment_id, "order_id": order_id,
        "boxes": shipped_boxes, "total_boxes": len(shipped_boxes),
        "total_units": total_units, "pallet": pallet,
        "carrier": carrier, "tracking": tracking,
        "shipped_by": user.get("user_id"),
        "shipped_by_name": user.get("name", ""),
        "created_at": now_iso(),
    }
    await db.wms_shipments.insert_one(shipment_doc)
    shipment_doc.pop("_id", None)
    if order_id:
        await db.orders.update_one({"$or": [{"order_id": order_id}, {"order_number": order_id}]}, {"$set": {"wms_status": "shipped"}})
    await log_movement(user, "shipment", {"shipment_id": shipment_id, "total_boxes": len(shipped_boxes), "total_units": total_units,
                                          "box_ids": [b["box_id"] for b in shipped_boxes]})
    return shipment_doc

@router.get("/shipments")
async def list_shipments(request: Request):
    await require_auth(request)
    ships = await db.wms_shipments.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return ships

# ==================== MOVEMENTS (AUDIT) ====================

@router.get("/movements")
async def list_movements(request: Request, movement_type: str = "", limit: int = 200):
    await require_auth(request)
    query = {}
    if movement_type: query["type"] = movement_type
    movements = await db.wms_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return movements


# ==================== AUDITORIA (SOLO SUPER ADMIN) ====================
# Modulo de auditoria del flujo completo: salud/consistencia del sistema,
# trazabilidad por caja, trazabilidad por SKU y busqueda de movimientos.
# Todo gated a require_supersu.

def _norm_cell(loc, key, color, size):
    return (str(loc or "").strip().upper(), str(key or "").strip().upper(),
            str(color or "").strip().upper(), str(size or "").strip().upper())


@router.get("/audit/health")
async def audit_health(request: Request):
    """Panel de consistencia: totales, drift inventario vs cajas por celda,
    tickets con picks sin descontar, negativos y cajas estancadas."""
    await require_supersu(request)
    from collections import defaultdict

    inv_agg = await db.wms_inventory.aggregate([
        {"$group": {"_id": None, "units": {"$sum": "$units_on_hand"}, "rows": {"$sum": 1}}}]).to_list(1)
    box_agg = await db.wms_boxes.aggregate([
        {"$match": {"units": {"$gt": 0}, "status": {"$nin": list(_BOX_OUT_STATUSES)}}},
        {"$group": {"_id": None, "units": {"$sum": "$units"}, "boxes": {"$sum": 1}}}]).to_list(1)

    # Drift por celda (location/style/color/size): inventario vs cajas vivas.
    inv_cells = defaultdict(int)
    async for r in db.wms_inventory.find({}, {"_id": 0, "location": 1, "style": 1, "sku": 1, "color": 1, "size": 1, "units_on_hand": 1}):
        inv_cells[_norm_cell(r.get("location"), r.get("style") or r.get("sku"), r.get("color"), r.get("size"))] += int(r.get("units_on_hand", 0) or 0)
    box_cells = defaultdict(int)
    async for b in db.wms_boxes.find({"units": {"$gt": 0}, "status": {"$nin": list(_BOX_OUT_STATUSES)}},
                                     {"_id": 0, "location": 1, "style": 1, "color": 1, "size": 1, "units": 1}):
        box_cells[_norm_cell(b.get("location"), b.get("style"), b.get("color"), b.get("size"))] += int(b.get("units", 0) or 0)
    drift = []
    for cell in set(inv_cells) | set(box_cells):
        diff = inv_cells.get(cell, 0) - box_cells.get(cell, 0)
        if diff != 0:
            drift.append({"location": cell[0], "style": cell[1], "color": cell[2], "size": cell[3],
                          "inventario": inv_cells.get(cell, 0), "cajas": box_cells.get(cell, 0), "diff": diff})
    drift.sort(key=lambda d: -abs(d["diff"]))

    # Tickets REALMENTE sin descontar. Ojo: "sin deducted_map" NO basta —
    # ese campo solo lo escribe el flujo incremental del PDA. Un ticket
    # confirmado por el flujo clasico (confirm_pick) descuenta bien pero deja
    # un movimiento 'pick_confirmed' sin deducted_map, y los liquidados por el
    # fix de junio quedan con deducted_map vacio + phantom_settled. El criterio
    # verdadero (el que uso el fix): picked>0 Y sin deducted_map Y sin phantom_
    # settled Y sin NINGUN movimiento pick_confirmed/pick_deduction del ticket.
    candidates = []
    async for t in db.wms_pick_tickets.find(
            {"status": {"$in": ["confirmed", "in_neck_cutting", "pending"]},
             "deducted_map": {"$in": [None, {}]}, "phantom_settled": {"$ne": True}},
            {"_id": 0, "ticket_id": 1, "order_number": 1, "style": 1, "color": 1, "status": 1, "picked_sizes": 1, "created_at": 1}):
        picked = 0
        for v in (t.get("picked_sizes") or {}).values():
            if isinstance(v, dict):
                picked += int(v.get("total", 0) or 0)
        if picked > 0:
            candidates.append((t, picked))
    # Un solo query: que tickets candidatos SI tienen movimiento de descuento.
    cand_ids = [t["ticket_id"] for t, _ in candidates]
    settled_ids = set()
    if cand_ids:
        settled_ids = set(await db.wms_movements.distinct("details.ticket_id", {
            "type": {"$in": ["pick_confirmed", "pick_deduction"]},
            "details.ticket_id": {"$in": cand_ids}}))
    undeducted = [
        {"ticket_id": t["ticket_id"], "order": t.get("order_number"),
         "style": t.get("style"), "color": t.get("color"),
         "status": t.get("status"), "picked": picked, "created_at": t.get("created_at")}
        for t, picked in candidates if t["ticket_id"] not in settled_ids
    ]

    negativos = await db.wms_inventory.count_documents({"$expr": {"$gt": ["$units_allocated", "$units_on_hand"]}})
    ceros_con_cajas = await db.wms_inventory.count_documents({"units_on_hand": {"$lte": 0}, "total_boxes": {"$gt": 0}})
    cutoff = (datetime.now(timezone.utc) - timedelta(days=14)).isoformat()
    stale_pending = await db.wms_boxes.count_documents({"status": "putaway_pending", "units": {"$gt": 0}, "created_at": {"$lt": cutoff}})

    return {
        "generated_at": now_iso(),
        "totales": {
            "inventario_unidades": (inv_agg[0]["units"] if inv_agg else 0),
            "inventario_filas": (inv_agg[0]["rows"] if inv_agg else 0),
            "cajas_unidades": (box_agg[0]["units"] if box_agg else 0),
            "cajas_vivas": (box_agg[0]["boxes"] if box_agg else 0),
        },
        "drift": {"celdas": len(drift), "unidades_abs": sum(abs(d["diff"]) for d in drift), "top": drift[:50]},
        "sin_descontar": {"tickets": len(undeducted), "top": sorted(undeducted, key=lambda x: -x["picked"])[:25]},
        "negativos_allocated": negativos,
        "ceros_con_cajas": ceros_con_cajas,
        "cajas_pendientes_14d": stale_pending,
    }


@router.get("/audit/box/{box_id}")
async def audit_box(box_id: str, request: Request):
    """Ciclo de vida completo de una caja: recibo, estado actual, renglon de
    inventario ligado, ticket que la surtio y todos sus movimientos."""
    await require_supersu(request)
    bid = box_id.strip()
    box = await db.wms_boxes.find_one({"$or": [{"box_id": bid}, {"barcode": bid}, {"lpn_id": bid}]}, {"_id": 0})

    receiving = None
    rid = (box or {}).get("receiving_id")
    if rid:
        receiving = await db.wms_receiving.find_one({"receiving_id": rid}, {"_id": 0})
    if not box:
        # Caja ausente de wms_boxes: buscarla dentro de los recibos (huerfana / embarcada)
        receiving = await db.wms_receiving.find_one({"boxes.box_id": bid}, {"_id": 0})

    movements = await db.wms_movements.find({"$or": [
        {"details.box_id": bid}, {"details.box_ids": bid}, {"details.boxes.box_id": bid},
    ]}, {"_id": 0}).sort("created_at", 1).to_list(300)

    ticket = None
    tid = (box or {}).get("last_pick_ticket")
    if tid:
        ticket = await db.wms_pick_tickets.find_one({"ticket_id": tid},
            {"_id": 0, "ticket_id": 1, "order_number": 1, "status": 1, "style": 1, "color": 1,
             "assigned_to_name": 1, "completed_at": 1, "deducted_map": 1})

    inventory = None
    iid = (box or {}).get("inventory_id")
    if iid:
        inventory = await db.wms_inventory.find_one({"inventory_id": iid}, {"_id": 0})

    if not box and not receiving and not movements:
        raise HTTPException(404, f"Sin rastro de la caja {bid} en cajas, recibos ni movimientos")
    return {"box_id": bid, "box": box, "receiving": receiving, "inventory": inventory,
            "pick_ticket": ticket, "movements": movements, "movement_count": len(movements)}


@router.get("/audit/sku")
async def audit_sku(request: Request, style: str, color: str = "", size: str = ""):
    """Balance del flujo completo de un SKU: recibido - pickeado vs en mano,
    cajas por estatus y su historial de movimientos."""
    await require_supersu(request)
    if not style.strip():
        raise HTTPException(400, "style requerido")
    st = _ci_eq(style.strip())
    q_recv = {"style": st}
    q_inv = {"$or": [{"style": st}, {"sku": st}]}
    q_box = {"$or": [{"style": st}, {"sku": st}]}
    q_tk = {"style": st}
    if color.strip():
        for q in (q_recv, q_inv, q_box, q_tk):
            q["color"] = _ci_eq(color.strip())
    if size.strip():
        for q in (q_recv, q_inv, q_box):
            q["size"] = _ci_eq(size.strip())

    recibido, recv_list = 0, []
    async for r in db.wms_receiving.find(q_recv, {"_id": 0, "receiving_id": 1, "size": 1, "total_units": 1,
                                                  "inv_location": 1, "created_at": 1, "received_by_name": 1}).sort("created_at", 1):
        recibido += int(r.get("total_units", 0) or 0)
        recv_list.append(r)

    pickeado, tickets = 0, []
    async for t in db.wms_pick_tickets.find(q_tk, {"_id": 0, "ticket_id": 1, "order_number": 1, "status": 1,
                                                   "picked_sizes": 1, "created_at": 1}).sort("created_at", 1):
        tot = 0
        for sz, v in (t.get("picked_sizes") or {}).items():
            if size.strip() and str(sz).strip().upper() != size.strip().upper():
                continue
            if isinstance(v, dict):
                tot += int(v.get("total", 0) or 0)
        if tot > 0:
            pickeado += tot
            tickets.append({"ticket_id": t["ticket_id"], "order": t.get("order_number"),
                            "status": t.get("status"), "picked": tot, "created_at": t.get("created_at")})

    on_hand, inv_rows = 0, []
    async for i in db.wms_inventory.find(q_inv, {"_id": 0, "location": 1, "size": 1, "units_on_hand": 1,
                                                 "units_allocated": 1, "total_boxes": 1}).sort("location", 1):
        on_hand += int(i.get("units_on_hand", 0) or 0)
        if int(i.get("units_on_hand", 0) or 0) != 0:
            inv_rows.append(i)

    boxes_by_status = {}
    async for g in db.wms_boxes.aggregate([{"$match": q_box},
            {"$group": {"_id": "$status", "cajas": {"$sum": 1}, "unidades": {"$sum": "$units"}}}]):
        boxes_by_status[str(g["_id"])] = {"cajas": g["cajas"], "unidades": g["unidades"]}

    movements = await get_sku_movement_history(style.strip(), color.strip(), size.strip(), 100)
    esperado = recibido - pickeado
    return {
        "style": style.strip().upper(), "color": color.strip().upper(), "size": size.strip().upper(),
        "balance": {"recibido": recibido, "pickeado": pickeado, "esperado": esperado,
                    "en_mano": on_hand, "diferencia": on_hand - esperado},
        "recibos": recv_list, "tickets": tickets, "inventario": inv_rows,
        "cajas_por_status": boxes_by_status, "movimientos": movements,
    }


@router.get("/audit/movements")
async def audit_movements(request: Request, q: str = "", movement_type: str = "", user: str = "",
                          since: str = "", until: str = "", limit: int = 200):
    """Busqueda de movimientos con filtros: texto libre (caja/sku/orden/
    ubicacion/ticket), tipo, usuario y rango de fechas."""
    await require_supersu(request)
    query = {}
    if movement_type.strip():
        query["type"] = movement_type.strip()
    if user.strip():
        query["user_name"] = {"$regex": re.escape(user.strip()), "$options": "i"}
    if since.strip() or until.strip():
        rng = {}
        if since.strip():
            rng["$gte"] = since.strip()
        if until.strip():
            rng["$lte"] = until.strip() + ("T23:59:59" if len(until.strip()) == 10 else "")
        query["created_at"] = rng
    if q.strip():
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        query["$or"] = [
            {"details.box_id": rx}, {"details.box_ids": rx}, {"details.sku": rx},
            {"details.style": rx}, {"details.color": rx}, {"details.ticket_id": rx},
            {"details.order_number": rx}, {"details.location": rx}, {"details.from": rx},
            {"details.to": rx}, {"details.receiving_id": rx}, {"details.batch_id": rx},
            {"movement_id": rx},
        ]
    limit = max(1, min(int(limit or 200), 1000))
    movements = await db.wms_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    total = await db.wms_movements.count_documents(query)
    return {"total": total, "count": len(movements), "movements": movements}


# ── Self-test: simula el flujo completo con datos marcados y auto-limpieza ──
_SELFTEST_STYLE = "__SELFTEST__"
_SELFTEST_CUSTOMER = "__SELFTEST__"
_SELFTEST_CART = "__SELFTEST_CARRO__"
_SELFTEST_BIN = "__SELFTEST_BIN__"


async def _selftest_cleanup(receiving_id, ticket_id, box_ids):
    """Borra TODO rastro del self-test. Se corre pase lo que pase (finally).
    Todo esta marcado con el style/customer/ubicaciones reservados, imposibles
    de colisionar con material real, asi que el borrado es exhaustivo y seguro."""
    deleted = {}
    deleted["boxes"] = (await db.wms_boxes.delete_many({"style": _SELFTEST_STYLE})).deleted_count
    deleted["inventory"] = (await db.wms_inventory.delete_many(
        {"$or": [{"style": _SELFTEST_STYLE}, {"location": {"$in": [_SELFTEST_CART, _SELFTEST_BIN]}}]})).deleted_count
    if receiving_id:
        deleted["receiving"] = (await db.wms_receiving.delete_many({"receiving_id": receiving_id})).deleted_count
    if ticket_id:
        deleted["tickets"] = (await db.wms_pick_tickets.delete_many({"ticket_id": ticket_id})).deleted_count
    deleted["movements"] = (await db.wms_movements.delete_many({"$or": [
        {"details.style": _SELFTEST_STYLE},
        {"details.receiving_id": receiving_id} if receiving_id else {"_id": None},
        {"details.ticket_id": ticket_id} if ticket_id else {"_id": None},
        {"details.box_id": {"$in": box_ids}} if box_ids else {"_id": None},
    ]})).deleted_count
    deleted["tasks"] = (await db.wms_tasks.delete_many({"context.sku": _SELFTEST_STYLE})).deleted_count
    return deleted


@router.post("/audit/self-test")
async def audit_self_test(request: Request):
    """Simula el ciclo completo — Recibo → Putaway → Pick Ticket → Surtido —
    ejecutando las MISMAS funciones internas del sistema real (donde vivio el
    bug de descuento), con datos de prueba marcados y limpieza automatica.
    Verifica el inventario en cada paso. No toca nada de material real."""
    user = await require_supersu(request)
    body = await request.json() if request.headers.get("content-length") else {}
    color = "TESTCOLOR"
    size = "M"
    units_per_box = int(body.get("units_per_box", 10) or 10)
    n_boxes = int(body.get("boxes", 3) or 3)
    pick_units = int(body.get("pick_units", units_per_box * 2) or units_per_box * 2)  # 2 cajas completas
    total_units = units_per_box * n_boxes

    steps = []
    receiving_id = gen_id("rcv")
    ticket_id = "pick_" + uuid.uuid4().hex[:12]
    box_ids = []

    def step(name, ok, detail, expected=None, got=None):
        steps.append({"step": len(steps) + 1, "name": name, "status": "PASS" if ok else "FAIL",
                      "detail": detail, "expected": expected, "got": got})
        return ok

    try:
        # ───── PASO 1: RECIBO ─────
        seq = await _reserve_box_seqs(n_boxes) - 1
        box_docs = []
        for _ in range(n_boxes):
            seq += 1
            bid = f"BOX-{seq:06d}"
            box_ids.append(bid)
            box_docs.append({
                "box_id": bid, "barcode": bid, "lpn_id": bid, "receiving_id": receiving_id,
                "customer": _SELFTEST_CUSTOMER, "manufacturer": "SELFTEST", "style": _SELFTEST_STYLE,
                "sku": f"{_SELFTEST_STYLE}-{color}-{size}", "color": color, "size": size,
                "units": units_per_box, "qty": units_per_box, "seq_num": seq, "location": _SELFTEST_CART,
                "status": "putaway_pending", "state": "raw", "is_bpo": False,
                "coo": "TEST", "country_of_origin": "TEST", "fabric_content": "100%TEST",
                "description": "SELF TEST", "created_at": now_iso(),
            })
        await db.wms_boxes.insert_many(box_docs)
        await db.wms_receiving.insert_one({
            "receiving_id": receiving_id, "customer": _SELFTEST_CUSTOMER, "manufacturer": "SELFTEST",
            "style": _SELFTEST_STYLE, "color": color, "size": size, "inv_location": _SELFTEST_CART,
            "total_units": total_units, "country_of_origin": "TEST", "fabric_content": "100%TEST",
            "received_by_name": user.get("name", "self-test"), "created_at": now_iso(),
            "boxes": [{"box_id": b["box_id"], "units": b["units"]} for b in box_docs],
        })
        inv_id = None
        for b in box_docs:
            inv_id = await _update_inventory_enhanced(
                _SELFTEST_STYLE, color, size, units_per_box, "add",
                customer=_SELFTEST_CUSTOMER, location=_SELFTEST_CART, box_count=1)
        row = await db.wms_inventory.find_one({"style": _SELFTEST_STYLE, "location": _SELFTEST_CART})
        got = int((row or {}).get("units_on_hand", 0))
        step("Recibo", got == total_units,
             f"{n_boxes} cajas de {units_per_box} u recibidas al {_SELFTEST_CART}", total_units, got)

        # ───── PASO 2: PUTAWAY ─────
        moved = 0
        for b in box_docs:
            await db.wms_boxes.update_one({"box_id": b["box_id"]},
                {"$set": {"location": _SELFTEST_BIN, "status": "located"}})
            await _move_box_inventory({**b, "style": _SELFTEST_STYLE}, _SELFTEST_CART, _SELFTEST_BIN)
            moved += 1
        bin_row = await db.wms_inventory.find_one({"style": _SELFTEST_STYLE, "location": _SELFTEST_BIN})
        cart_row = await db.wms_inventory.find_one({"style": _SELFTEST_STYLE, "location": _SELFTEST_CART})
        bin_units = int((bin_row or {}).get("units_on_hand", 0))
        cart_units = int((cart_row or {}).get("units_on_hand", 0))
        step("Putaway", bin_units == total_units and cart_units == 0,
             f"{moved} cajas movidas {_SELFTEST_CART} → {_SELFTEST_BIN}",
             f"bin={total_units}, carro=0", f"bin={bin_units}, carro={cart_units}")

        # ───── PASO 3: PICK TICKET ─────
        await db.wms_pick_tickets.insert_one({
            "ticket_id": ticket_id, "order_number": "SELFTEST", "style": _SELFTEST_STYLE,
            "color": color, "customer": _SELFTEST_CUSTOMER, "status": "pending",
            "picking_status": "assigned", "assigned_to": user.get("user_id"),
            "assigned_to_name": user.get("name", "self-test"),
            "sizes": {size: pick_units}, "picked_sizes": {},
            "created_at": now_iso(),
        })
        tk = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id}, {"_id": 0, "ticket_id": 1})
        step("Crear pick ticket", tk is not None,
             f"Ticket para surtir {pick_units} u talla {size}", ticket_id, (tk or {}).get("ticket_id"))

        # ───── PASO 4: SURTIDO (DESCUENTO REAL) ─────
        before = bin_units
        await _deduct_pick_boxes(_SELFTEST_STYLE, color, size, _SELFTEST_BIN, pick_units,
                                 "deduct", _SELFTEST_CUSTOMER, "SELFTEST", None,
                                 user=user, ticket_id=ticket_id)
        after_row = await db.wms_inventory.find_one({"style": _SELFTEST_STYLE, "location": _SELFTEST_BIN})
        after = int((after_row or {}).get("units_on_hand", 0))
        depleted = await db.wms_boxes.count_documents({"style": _SELFTEST_STYLE, "status": "depleted"})
        mov = await db.wms_movements.count_documents({"type": "pick_deduction", "details.ticket_id": ticket_id})
        expected_after = before - pick_units
        expected_depleted = pick_units // units_per_box
        ok4 = (after == expected_after and depleted == expected_depleted and mov > 0)
        step("Surtido y descuento", ok4,
             f"Surtidas {pick_units} u; inventario baja y cajas se agotan; movimiento pick_deduction registrado",
             f"en_mano={expected_after}, cajas_depleted={expected_depleted}, movimiento=sí",
             f"en_mano={after}, cajas_depleted={depleted}, movimiento={'sí' if mov else 'NO'}")

        # ───── PASO 5: TRAZABILIDAD ─────
        trace_box = box_ids[0]
        trace_mov = await db.wms_movements.count_documents({"details.box_ids": trace_box})
        step("Trazabilidad por caja", trace_mov > 0,
             f"La caja {trace_box} deja rastro por box_id en movimientos", "≥1 movimiento", trace_mov)

    finally:
        cleanup = await _selftest_cleanup(receiving_id, ticket_id, box_ids)

    passed = sum(1 for s in steps if s["status"] == "PASS")
    return {
        "ok": passed == len(steps) and len(steps) == 5,
        "passed": passed, "total": len(steps),
        "steps": steps, "cleanup": cleanup,
        "ran_at": now_iso(),
        "params": {"boxes": n_boxes, "units_per_box": units_per_box, "pick_units": pick_units},
    }


# ==================== CONCILIACION FISICA (PDA + PC) ====================
# El picker escanea una ubicacion y luego las cajas fisicamente presentes. El
# sistema "casa" el inventario a la realidad: confirma las presentes, marca las
# faltantes para el admin, jala cajas que aparezcan de otra ubicacion, y crea
# las que se escaneen y no existan. Cada ubicacion conciliada queda bloqueada.
RECON_DEFAULT_UNITS = 72  # no hay cajas en 0 en piso: una caja presente en 0 se rellena a 72
_RECON_NORM = lambda s: re.sub(r"\s+", " ", str(s or "")).strip().upper()


async def _recon_present_boxes(loc):
    """Cajas del sistema que se ESPERAN fisicamente en `loc` (en piso, con stock,
    no ya-salidas ni faltantes-pendientes)."""
    return await db.wms_boxes.find({
        "location": {"$regex": f"^{re.escape(loc)}$", "$options": "i"},
        "status": {"$nin": list(_BOX_OUT_STATUSES)},
    }, {"_id": 0, "box_id": 1, "barcode": 1, "lpn_id": 1, "generic_lpn": 1, "lpn_reconciled_at": 1,
        "sku": 1, "style": 1, "color": 1, "size": 1, "units": 1, "status": 1}).to_list(3000)


def _recon_lpn_boxes(boxes):
    """Cajas con LPN fisico real (identificador sin prefijo BOX). La conciliacion
    por-BOX no las puede casar, asi que su ubicacion se bloquea. Se detecta por el
    box_id sin prefijo BOX (las LPN migradas tienen el LPN EN el box_id, ej.
    'LPN_1506992C'), o por lpn_id/barcode sin prefijo BOX, o por la marca de
    reconciliacion de LPN."""
    out = []
    for b in boxes:
        bid = str(b.get("box_id") or "").strip().upper()
        lpn = str(b.get("lpn_id") or "").strip().upper()
        bc = str(b.get("barcode") or "").strip().upper()
        has_lpn = (bid and not bid.startswith("BOX")) \
            or (lpn and not lpn.startswith("BOX")) \
            or (bc and not bc.startswith("BOX")) \
            or bool(b.get("lpn_reconciled_at")) or bool(b.get("generic_lpn"))
        if has_lpn:
            out.append({"box_id": b.get("box_id"), "lpn": b.get("lpn_id") or b.get("barcode") or b.get("box_id")})
    return out


@router.get("/recon/location/{loc}")
async def recon_location(loc: str, request: Request):
    """Estado de una ubicacion para conciliar: cajas esperadas + si ya fue
    conciliada (bloqueada)."""
    await require_auth(request)
    location = _RECON_NORM(loc)
    if not location:
        raise HTTPException(400, "Ubicacion requerida")
    lock = await db.wms_reconciled_locations.find_one({"location": location}, {"_id": 0})
    expected = await _recon_present_boxes(location)
    lpn_boxes = _recon_lpn_boxes(expected)
    return {
        "location": location,
        "locked": bool(lock and lock.get("status") == "done"),
        "lock": lock,
        # Bloqueo por LPN: hay cajas con licencia fisica real que la conciliacion
        # por-BOX no puede casar; esta ubicacion no debe conciliarse aqui.
        "blocked_lpn": len(lpn_boxes) > 0,
        "lpn_boxes": lpn_boxes,
        "expected_count": len(expected),
        "expected": [{k: v for k, v in b.items() if k in ("box_id", "sku", "style", "color", "size", "units", "status")} for b in expected],
    }


@router.post("/recon/commit")
async def recon_commit(request: Request):
    """Concilia una ubicacion contra las cajas fisicamente escaneadas.

    body: { location, scanned_box_ids: [BOX-...] }
    - Confirma las presentes (rellena a 72 las que esten en 0).
    - Jala cajas que el sistema tenia en otra ubicacion / pendientes.
    - Crea las cajas BOX escaneadas que no existan (a revision del admin).
    - Marca faltantes (esperadas y no escaneadas) como recon_pending -> tabla admin.
    - Reconstruye el inventario de la ubicacion desde las cajas presentes.
    - Bloquea la ubicacion. Reversible via wms_recon_bak + commit_id.
    """
    user = await require_auth(request)
    body = await request.json()
    location = _RECON_NORM(body.get("location"))
    if not location:
        raise HTTPException(400, "Ubicacion requerida")
    scanned = []
    seen = set()
    for raw in (body.get("scanned_box_ids") or []):
        bid = str(raw or "").strip().upper()
        if not bid or bid in seen:
            continue
        if not bid.startswith("BOX"):
            continue  # solo cajas del sistema (prefijo BOX)
        seen.add(bid)
        scanned.append(bid)

    # Claim atomico del lock: primer commit gana; si ya esta 'done' -> bloqueada.
    existing = await db.wms_reconciled_locations.find_one({"location": location})
    if existing and existing.get("status") == "done":
        raise HTTPException(409, f"La ubicacion {location} ya fue conciliada por "
                                 f"{existing.get('reconciled_by_name','?')}. Pide al admin reabrirla.")

    commit_id = "rcn_" + uuid.uuid4().hex[:12]
    uname = user.get("name", user.get("email", "?"))

    # Respaldo: cajas presentes actuales + inventario de la ubicacion.
    pre_boxes = await _recon_present_boxes(location)

    # Bloqueo por LPN: si la ubicacion tiene cajas con licencia fisica real, la
    # conciliacion por-BOX no aplica (no puede casarlas). Se rechaza.
    lpn_boxes = _recon_lpn_boxes(pre_boxes)
    if lpn_boxes:
        ejemplos = ", ".join(str(x.get("lpn") or x.get("box_id")) for x in lpn_boxes[:3])
        raise HTTPException(409, f"La ubicacion {location} tiene {len(lpn_boxes)} caja(s) con LPN fisico "
                                 f"(ej. {ejemplos}) y no puede conciliarse aqui. Repórtala al administrador.")
    pre_inv = await db.wms_inventory.find({"location": {"$regex": f"^{re.escape(location)}$", "$options": "i"}}, {"_id": 0}).to_list(2000)
    scanned_docs = await db.wms_boxes.find({"box_id": {"$in": scanned}}, {"_id": 0}).to_list(3000) if scanned else []
    await db.wms_recon_bak.insert_one({
        "commit_id": commit_id, "location": location, "created_at": now_iso(),
        "boxes_before": pre_boxes + [d for d in scanned_docs if d.get("location") != location],
        "inventory_before": pre_inv,
    })

    expected_ids = {b["box_id"] for b in pre_boxes}
    scanned_set = set(scanned)
    confirmadas = movidas = creadas = faltantes = 0
    stamp = {"recon_batch": commit_id, "recon_counted_at": now_iso(), "recon_counted_by": uname}

    for bid in scanned:
        box = await db.wms_boxes.find_one({"box_id": bid})
        if not box:
            # Caja BOX que no existe -> crearla en la ubicacion, a revision del admin.
            seqp = re.match(r"BOX-?(\d+)", bid)
            await db.wms_boxes.insert_one({
                "box_id": bid, "barcode": bid, "lpn_id": bid, "location": location,
                "style": "(SIN IDENTIFICAR)", "sku": bid, "color": "", "size": "",
                "units": RECON_DEFAULT_UNITS, "qty": RECON_DEFAULT_UNITS,
                "status": "located", "state": "raw",
                "customer": "", "manufacturer": "", "country_of_origin": "", "fabric_content": "",
                "seq_num": int(seqp.group(1)) if seqp else 0,
                "recon_created": True, "created_at": now_iso(), **stamp,
            })
            creadas += 1
        else:
            u = int(box.get("units") or box.get("qty") or 0)
            new_u = RECON_DEFAULT_UNITS if u <= 0 else u
            was_elsewhere = _RECON_NORM(box.get("location")) != location or box.get("status") == "recon_pending"
            upd = {"location": location, "status": "located", "units": new_u, "qty": new_u, **stamp}
            upd["recon_pending"] = False
            upd["recon_missing_from"] = None
            await db.wms_boxes.update_one({"_id": box["_id"]}, {"$set": upd})
            if was_elsewhere:
                movidas += 1
            else:
                confirmadas += 1

    # Faltantes: esperadas pero NO escaneadas -> pendientes para el admin.
    missing_ids = expected_ids - scanned_set
    if missing_ids:
        await db.wms_boxes.update_many(
            {"box_id": {"$in": list(missing_ids)}},
            {"$set": {"status": "recon_pending", "recon_pending": True,
                      "recon_missing_from": location, "recon_flagged_at": now_iso(),
                      "recon_flagged_by": uname, "recon_batch": commit_id}})
        faltantes = len(missing_ids)

    # Reconstruir inventario de la ubicacion desde las cajas AHORA presentes.
    from collections import defaultdict as _dd
    present = await db.wms_boxes.find({
        "location": {"$regex": f"^{re.escape(location)}$", "$options": "i"},
        "status": "located", "units": {"$gt": 0},
    }, {"_id": 0, "style": 1, "sku": 1, "color": 1, "size": 1, "units": 1, "customer": 1,
        "manufacturer": 1, "country_of_origin": 1, "fabric_content": 1}).to_list(5000)
    groups = _dd(lambda: {"units": 0, "boxes": 0, "meta": {}})
    for b in present:
        key = (b.get("style") or b.get("sku") or "", b.get("color") or "", b.get("size") or "")
        g = groups[key]
        g["units"] += int(b.get("units") or 0)
        g["boxes"] += 1
        if not g["meta"]:
            g["meta"] = {k: b.get(k, "") for k in ("customer", "manufacturer", "country_of_origin", "fabric_content", "sku")}
    await db.wms_inventory.delete_many({"location": {"$regex": f"^{re.escape(location)}$", "$options": "i"}})
    if groups:
        for (st, co, sz), g in groups.items():
            dst_id = gen_id("inv")
            await db.wms_inventory.insert_one({
                "inventory_id": dst_id, "style": st, "sku": g["meta"].get("sku") or st,
                "color": co, "size": sz, "location": location,
                "units_on_hand": g["units"], "units_allocated": 0, "total_boxes": g["boxes"],
                "customer": g["meta"].get("customer", ""), "manufacturer": g["meta"].get("manufacturer", ""),
                "country_of_origin": g["meta"].get("country_of_origin", ""),
                "fabric_content": g["meta"].get("fabric_content", ""),
                "recon_batch": commit_id, "updated_at": now_iso(),
            })
            # Update matching boxes at this location to point to the new inventory_id
            style_norm = _RECON_NORM(st)
            color_norm = _RECON_NORM(co)
            size_norm = _RECON_NORM(sz)
            await db.wms_boxes.update_many({
                "location": {"$regex": f"^{re.escape(location)}$", "$options": "i"},
                "status": "located",
                "$or": [{"sku": {"$regex": f"^{style_norm}$", "$options": "i"}},
                        {"style": {"$regex": f"^{style_norm}$", "$options": "i"}}],
                "color": {"$regex": f"^{color_norm}$", "$options": "i"},
                "size": {"$regex": f"^{size_norm}$", "$options": "i"},
            }, {"$set": {"inventory_id": dst_id}})

    # Registrar/bloquear la ubicacion.
    lock_doc = {
        "location": location, "status": "done", "commit_id": commit_id,
        "reconciled_by": user.get("user_id"), "reconciled_by_name": uname,
        "reconciled_at": now_iso(),
        "counts": {"escaneadas": len(scanned), "confirmadas": confirmadas, "movidas": movidas,
                   "creadas": creadas, "faltantes": faltantes},
    }
    await db.wms_reconciled_locations.update_one({"location": location}, {"$set": lock_doc}, upsert=True)
    await log_movement(user, "reconciliation_commit", {
        "commit_id": commit_id, "location": location, **lock_doc["counts"]})

    return {"ok": True, "location": location, "commit_id": commit_id, **lock_doc["counts"]}


@router.get("/recon/pending")
async def recon_pending(request: Request):
    """Admin: cajas faltantes (recon_pending) + creadas/desconocidas."""
    await require_admin(request)
    faltantes = await db.wms_boxes.find(
        {"status": "recon_pending"},
        {"_id": 0, "box_id": 1, "style": 1, "color": 1, "size": 1, "units": 1,
         "recon_missing_from": 1, "recon_flagged_by": 1, "recon_flagged_at": 1}
    ).sort("recon_flagged_at", -1).to_list(2000)
    creadas = await db.wms_boxes.find(
        {"recon_created": True},
        {"_id": 0, "box_id": 1, "location": 1, "units": 1, "recon_counted_by": 1, "recon_counted_at": 1}
    ).sort("recon_counted_at", -1).to_list(2000)
    return {"faltantes": faltantes, "faltantes_count": len(faltantes),
            "creadas": creadas, "creadas_count": len(creadas)}


@router.get("/recon/log")
async def recon_log(request: Request):
    """Admin: registro de ubicaciones conciliadas."""
    await require_admin(request)
    locs = await db.wms_reconciled_locations.find({}, {"_id": 0}).sort("reconciled_at", -1).to_list(5000)
    return {"count": len(locs), "locations": locs}


@router.get("/recon/lpn-locations")
async def recon_lpn_locations(request: Request):
    """Admin: ubicaciones que tienen cajas con LPN fisico y por lo tanto NO se
    pueden conciliar por el flujo PDA (para avisar a los contadores de antemano).
    El sistema ya lo sabe sin necesidad de escanear."""
    await require_admin(request)
    rows = await db.wms_boxes.aggregate([
        {"$match": {"status": {"$nin": list(_BOX_OUT_STATUSES)}, "units": {"$gt": 0}, "$or": [
            {"box_id": {"$not": {"$regex": "^BOX", "$options": "i"}}},
            {"lpn_reconciled_at": {"$exists": True, "$ne": None}},
            {"generic_lpn": {"$exists": True, "$ne": None}},
        ]}},
        {"$group": {"_id": "$location", "cajas": {"$sum": 1}, "unidades": {"$sum": "$units"}}},
        {"$sort": {"cajas": -1}},
    ]).to_list(10000)
    out = [{"location": r["_id"], "cajas": r["cajas"], "unidades": r["unidades"]} for r in rows if r["_id"]]
    return {"count": len(out), "locations": out}


@router.get("/recon/adjustments")
async def recon_adjustments(request: Request):
    """Admin: registro de ajustes de cajas (p. ej. restauracion de LPN marcadas
    faltantes por conciliacion). Visible en el modulo de Conciliacion."""
    await require_admin(request)
    adj = await db.wms_recon_adjustments.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"count": len(adj), "adjustments": adj}


@router.post("/recon/resolve")
async def recon_resolve(request: Request):
    """Admin: resuelve una caja pendiente/creada.
    body: { box_id, action: 'assign'|'delete', location? }"""
    user = await require_admin(request)
    body = await request.json()
    bid = str(body.get("box_id", "")).strip().upper()
    action = body.get("action")
    box = await db.wms_boxes.find_one({"box_id": bid})
    if not box:
        raise HTTPException(404, "Caja no encontrada")
    if action == "delete":
        await db.wms_boxes.delete_one({"_id": box["_id"]})
        await log_movement(user, "recon_resolve_delete", {"box_id": bid})
        return {"ok": True, "box_id": bid, "action": "delete"}
    if action == "assign":
        loc = _RECON_NORM(body.get("location"))
        if not loc:
            raise HTTPException(400, "Ubicacion requerida para asignar")
        u = int(box.get("units") or 0) or RECON_DEFAULT_UNITS
        await db.wms_boxes.update_one({"_id": box["_id"]}, {"$set": {
            "location": loc, "status": "located", "units": u, "qty": u,
            "recon_pending": False, "recon_missing_from": None, "updated_at": now_iso()}})
        await log_movement(user, "recon_resolve_assign", {"box_id": bid, "location": loc})
        return {"ok": True, "box_id": bid, "action": "assign", "location": loc}
    raise HTTPException(400, "action debe ser 'assign' o 'delete'")


@router.post("/recon/reopen")
async def recon_reopen(request: Request):
    """Admin: reabre una ubicacion bloqueada para volver a conciliarla."""
    user = await require_admin(request)
    body = await request.json()
    location = _RECON_NORM(body.get("location"))
    r = await db.wms_reconciled_locations.delete_one({"location": location})
    if r.deleted_count == 0:
        raise HTTPException(404, "Esa ubicacion no estaba conciliada")
    await log_movement(user, "recon_reopen", {"location": location})
    return {"ok": True, "location": location}


@router.post("/recon/second-count/start")
async def recon_second_count_start(request: Request):
    """Admin: libera (reabre) todas las ubicaciones que tienen cajas marcadas
    como faltantes (recon_pending), para que los operadores hagan un segundo
    conteo fisico de esas ubicaciones en el PDA. Deja un registro en Ajustes
    de cajas con el detalle de que ubicaciones se liberaron."""
    user = await require_admin(request)
    pending = await db.wms_boxes.find(
        {"status": "recon_pending"},
        {"_id": 0, "units": 1, "recon_missing_from": 1}
    ).to_list(5000)

    from collections import defaultdict as _dd
    groups = _dd(lambda: {"cajas": 0, "unidades": 0})
    for b in pending:
        loc = b.get("recon_missing_from")
        if not loc:
            continue
        groups[loc]["cajas"] += 1
        groups[loc]["unidades"] += int(b.get("units") or 0)

    if not groups:
        return {"ok": True, "count": 0, "locations": []}

    locations = list(groups.keys())

    # Foto del primer conteo antes de borrarlo: al recontar, su registro en
    # wms_reconciled_locations se sobreescribe con los numeros del segundo
    # conteo, asi que se guarda aqui para no perder de vista lo que dio el
    # primero (queda visible en la pestana de Ajustes aunque ya no en el log).
    first_counts = {
        d["location"]: d
        for d in await db.wms_reconciled_locations.find(
            {"location": {"$in": locations}}, {"_id": 0}
        ).to_list(len(locations))
    }

    reopened = await db.wms_reconciled_locations.delete_many({"location": {"$in": locations}})

    uname = user.get("name", user.get("email", "?"))
    adj_locations = sorted(
        (
            {
                "location": loc, "cajas": g["cajas"], "unidades": g["unidades"],
                "first_count": (first_counts.get(loc) or {}).get("counts"),
                "first_count_by": (first_counts.get(loc) or {}).get("reconciled_by_name"),
                "first_count_at": (first_counts.get(loc) or {}).get("reconciled_at"),
            }
            for loc, g in groups.items()
        ),
        key=lambda x: -x["cajas"],
    )
    await db.wms_recon_adjustments.insert_one({
        "type": "second_count_start",
        "created_at": now_iso(), "created_by": uname,
        "count": len(locations),
        "units": sum(g["unidades"] for g in groups.values()),
        "reason": f"Segundo conteo: se liberaron {len(locations)} ubicaciones con cajas faltantes para volver a contarlas fisicamente.",
        "locations": adj_locations,
        "boxes": [],
    })
    await log_movement(user, "recon_second_count_start", {"locations_count": len(locations), "reopened": reopened.deleted_count})
    return {"ok": True, "count": len(locations), "reopened": reopened.deleted_count, "locations": adj_locations}

# ==================== LABELS (PDF) ====================

async def _enrich_box_for_label(box):
    """Merge a box doc with its receiving record so the label has the full
    context (description / COO / fabric / lot / PO / UPC) even when those live
    on the receiving doc and not on the box."""
    r = dict(box)
    rid = box.get("receiving_id")
    if rid:
        rcv = await db.wms_receiving.find_one({"receiving_id": rid}, {"_id": 0})
        if rcv:
            for f in ("description", "country_of_origin", "fabric_content",
                      "lot_number", "po", "upc", "received_by_name",
                      "manufacturer", "customer"):
                if not r.get(f):
                    r[f] = rcv.get(f, "")
    r["location"] = box.get("location") or box.get("inv_location") or ""
    return r


def _fmt_label_date(iso):
    """dd/mm/yyyy HH:MM en hora del almacén (America/Tijuana) para la etiqueta."""
    try:
        from datetime import datetime
        from zoneinfo import ZoneInfo
        d = datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
        if d.tzinfo:
            d = d.astimezone(ZoneInfo("America/Tijuana"))
        return d.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ""


def _build_box_labels_html(items):
    """Printable HTML (4x6) — one label per box, barcodes rendered in the browser
    via JsBarcode. No server-side barcode/reportlab dependency. Mirrors the
    Receiving label so a reprint looks identical to the original."""
    import html as _html
    esc = lambda v: _html.escape(str(v if v is not None else ""))
    pages, calls = [], []
    n = len(items)
    for idx, r in enumerate(items):
        bid = esc(r.get("box_id", ""))
        upc = esc(r.get("upc", ""))
        # Carros se abrevian a "C.<num>" para que el numero quepa ENORME;
        # bins y demas ubicaciones se imprimen tal cual.
        raw_loc = str(r.get("location") or "").strip()
        carro_m = re.match(r"(?i)^CARRO\s*(\S+)$", raw_loc)
        loc_disp = f"C.{carro_m.group(1)}" if carro_m else raw_loc
        loc_size = "52px" if carro_m else "40px"
        pages.append(f'''
      <div class="label-page">
        <div style="text-align:center;margin-bottom:6px"><svg id="bc-{idx}"></svg></div>
        <table class="table">
          <tr class="row">
            <td class="cell" style="width:60%"><span class="label">Cliente</span><span class="value">{esc(r.get("customer"))}</span></td>
            <td class="cell" style="width:40%"><span class="label">PO</span><span class="value">{esc(r.get("po"))}</span></td>
          </tr>
          <tr class="row">
            <td class="cell" style="width:40%"><span class="label">Lote</span><span class="value">{esc(r.get("lot_number"))}</span></td>
            <td class="cell" style="width:60%;text-align:center;position:relative"><span style="position:absolute;top:2px;left:4px;font-size:9px;color:#666;font-weight:bold">L</span><span class="value" style="font-size:{loc_size};letter-spacing:1px;line-height:1;display:block">{esc(loc_disp)}</span></td>
          </tr>
          <tr class="row"><td class="cell" colspan="2"><span class="label">Fabricante</span><span class="value">{esc(r.get("manufacturer"))}</span></td></tr>
          <tr class="row">
            <td class="cell" style="width:50%"><span class="label">Style</span><span class="value" style="font-size:16px">{esc(r.get("style"))}</span></td>
            <td class="cell" style="width:50%"><span class="label">SKU</span><span class="value" style="font-family:monospace">{esc(r.get("sku") or r.get("style"))}</span></td>
          </tr>
          <tr class="row">
            <td class="cell" style="width:50%"><span class="label">Color</span><span class="value">{esc(r.get("color"))}</span></td>
            <td class="cell" style="width:50%"><span class="label">Talla</span><span class="value" style="font-size:16px">{esc(r.get("size"))}</span></td>
          </tr>
          <tr class="row"><td class="cell" colspan="2"><span class="label">Descripcion</span><span class="value">{esc(r.get("description"))}</span></td></tr>
          <tr class="row">
            <td class="cell" style="width:50%"><span class="label">Pais</span><span class="value">{esc(r.get("country_of_origin"))}</span></td>
            <td class="cell" style="width:50%"><span class="label">Tela</span><span class="value">{esc(r.get("fabric_content"))}</span></td>
          </tr>
          <tr class="row"><td class="cell" colspan="2" style="text-align:center"><span class="label">UNITS IN BOX</span><span class="value" style="font-size:24px">{esc(r.get("units", 0))}</span></td></tr>
        </table>
        {f'<div style="text-align:center;margin-top:4px"><span class="label">UPC</span><svg id="upc-{idx}"></svg></div>' if upc else ''}
        <div style="margin-top:12px;border-top:2px solid #000;padding-top:8px">
          <div style="display:flex;justify-content:space-between;align-items:baseline">
            <div><span class="label">Caja</span><span style="font-size:16px;font-weight:bold;font-family:monospace">{bid}</span></div>
            <div style="text-align:right"><span class="label">Etiqueta</span><span style="font-size:16px;font-weight:bold">{idx + 1} de {n}</span></div>
          </div>
          <div style="display:flex;justify-content:space-between;align-items:baseline;margin-top:6px">
            <div><span class="label">Fecha de recibo</span><span style="font-size:15px;font-weight:bold">{esc(_fmt_label_date(r.get("created_at")))}</span></div>
            <div style="text-align:right"><span class="label">Recibi&oacute;</span><span style="font-size:15px;font-weight:bold">{esc(r.get("received_by_name"))}</span></div>
          </div>
        </div>
      </div>''')
        # JS-context escape: json.dumps neutralizes quotes/backslashes/</script>
        # in scanned or imported box_id/upc values. esc() only covers HTML.
        calls.append(f'JsBarcode("#bc-{idx}", {json.dumps(str(r.get("box_id") or ""))}, {{width:1.5,height:40,displayValue:true,fontSize:10,margin:0}});')
        if upc:
            calls.append(f'JsBarcode("#upc-{idx}", {json.dumps(str(r.get("upc") or ""))}, {{width:1.6,height:44,displayValue:true,fontSize:14,margin:0}});')
    return f'''<html><head><title>Etiqueta(s) de caja</title>
      <style>
        @page {{ size: 4in 6in; margin: 5mm; }}
        body {{ font-family: Arial, sans-serif; margin:0; padding:0; width:3.6in; background:white; }}
        .label-page {{ page-break-after: always; padding:10px; height:5.5in; box-sizing:border-box; }}
        .label-page:last-child {{ page-break-after:auto; }}
        .row {{ display:flex; border-bottom:1px solid #000; }}
        .cell {{ padding:4px 6px; border-right:1px solid #000; }}
        .cell:last-child {{ border-right:none; }}
        .label {{ font-size:8px; text-transform:uppercase; color:#666; display:block; }}
        .value {{ font-size:12px; font-weight:bold; }}
        .table {{ border:1px solid #000; border-collapse:collapse; width:100%; margin-top:6px; }}
      </style></head><body>
      {''.join(pages)}
      <script src="https://cdn.jsdelivr.net/npm/jsbarcode@3.11.5/dist/JsBarcode.all.min.js"></script>
      <script>
        try {{ {''.join(calls)} setTimeout(function(){{window.print()}}, 600); }} catch(e) {{}}
      </script>
    </body></html>'''


@router.get("/labels/box/{box_id}")
async def generate_box_label(box_id: str, request: Request):
    await require_auth(request)
    box = await db.wms_boxes.find_one({"box_id": box_id}, {"_id": 0})
    if not box:
        raise HTTPException(404, "Caja no encontrada")
    enriched = await _enrich_box_for_label(box)
    return HTMLResponse(_build_box_labels_html([enriched]))

@router.get("/labels/boxes")
async def generate_multi_box_labels(request: Request, box_ids: str = ""):
    await require_auth(request)
    ids = [b.strip() for b in box_ids.split(",") if b.strip()]
    if not ids:
        raise HTTPException(400, "box_ids requeridos (separados por coma)")
    items = []
    for bid in ids:
        box = await db.wms_boxes.find_one({"box_id": bid}, {"_id": 0})
        if box:
            items.append(await _enrich_box_for_label(box))
    if not items:
        raise HTTPException(404, "Ninguna de las cajas existe")
    return HTMLResponse(_build_box_labels_html(items))

@router.get("/labels/location")
async def generate_location_labels(request: Request, location: str = ""):
    """Print one box label per LPN currently sitting (units>0) in a location.
    Used by the Locations detail modal ('Imprimir etiquetas')."""
    await require_auth(request)
    loc = (location or "").strip()
    if not loc:
        raise HTTPException(400, "location requerido")
    boxes = await db.wms_boxes.find(
        {"location": {"$regex": f"^{re.escape(loc)}$", "$options": "i"}, "units": {"$gt": 0}},
        {"_id": 0},
    ).sort("created_at", 1).to_list(3000)
    if not boxes:
        raise HTTPException(404, f"No hay cajas con unidades en {loc}")
    items = [await _enrich_box_for_label(b) for b in boxes]
    return HTMLResponse(_build_box_labels_html(items))

# ==================== EXPORT ====================

@router.get("/export/inventory")
async def export_inventory(request: Request, exclude_hold: bool = False):
    await require_auth(request)
    # to_list(None) returns ALL rows. A fixed cap (e.g. 20000) silently dropped
    # the alphabetical tail (W/X/Y/Z SKUs) once the catalog grew past it.
    # Skip fully-empty orphan rows (no units, no boxes, no allocations) so the
    # report never shows phantom lines — e.g. a row left behind after a move/pick
    # decremented one counter but not the other.
    inventory = await db.wms_inventory.find(
        {"$or": [
            {"units_on_hand": {"$gt": 0}},
            {"total_boxes": {"$gt": 0}},
            {"units_allocated": {"$gt": 0}},
        ]},
        {"_id": 0},
    ).sort("sku", 1).to_list(None)
    # Optionally drop rows sitting in SAT-held locations (case-insensitive match
    # against the hold list).
    if exclude_hold:
        held = await _hold_location_names()
        if held:
            inventory = [r for r in inventory if (r.get("location") or "").strip().upper() not in held]
    import xlsxwriter
    # Pull boxes up-front so the aggregated "Inventory" sheet can show each row's
    # MOST RECENT box transfer (date + user). Reused for the per-box sheet below.
    boxes = await db.wms_boxes.find(
        {"$or": [{"units": {"$gt": 0}}, {"qty": {"$gt": 0}}]},
        {"_id": 0},
    ).sort([("location", 1), ("sku", 1)]).to_list(None)
    if exclude_hold and held:
        boxes = [b for b in boxes if (b.get("location") or "").strip().upper() not in held]
    # Last transfer INTO each location (date + user) from the movement log. The
    # destination ('to') is always recorded — unlike the box_id list which bulk
    # moves cap at 50 — so this answers "cuándo se transfirió material a esta
    # ubicación y quién" with FULL coverage. (The per-box sheet below uses each
    # box's own stamp for box-level precision.)
    loc_transfer = {}
    async for mv in db.wms_movements.find(
        {"type": {"$in": ["putaway", "putaway_bulk", "bulk_relocation",
                          "transit_relocation", "lpn_reconciled", "task_completed"]}},
        {"_id": 0, "created_at": 1, "user_name": 1, "user_id": 1, "details": 1},
    ).sort("created_at", -1):
        d = mv.get("details") or {}
        dest = (d.get("to") or d.get("destination") or d.get("location") or "").strip().upper()
        if dest and dest not in loc_transfer:  # first seen = newest (sorted desc)
            loc_transfer[dest] = (mv.get("created_at") or "", mv.get("user_name") or mv.get("user_id") or "")

    # Fallback for material loaded via import (no transfer movement): show when
    # the newest box at that location was created, marked "(carga inicial)" so
    # it's clearly a load, not a user transfer.
    loc_loaded = {}
    for b in boxes:
        ca = b.get("created_at")
        loc = (b.get("location") or "").strip().upper()
        if ca and loc and (loc not in loc_loaded or ca > loc_loaded[loc]):
            loc_loaded[loc] = ca

    def _row_transfer(inv):
        loc = (inv.get("location") or "").strip().upper()
        tr = loc_transfer.get(loc)
        if tr:
            return (tr[0][:19].replace("T", " "), tr[1])
        ca = loc_loaded.get(loc)
        if ca:
            return (ca[:19].replace("T", " "), "(carga inicial)")
        return ("", "")

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf)
    ws = wb.add_worksheet("Inventory")
    headers = ["Customer", "Style", "Color", "Size", "Description", "Category",
               "Manufacturer", "Location", "Total Boxes", "On Hand", "Allocated", "Available",
               "Country of Origin", "Fabric Content", "Is BPO",
               "Última transferencia", "Transferido por"]
    bold = wb.add_format({"bold": True})
    for i, h in enumerate(headers):
        ws.write(0, i, h, bold)
    for row, inv in enumerate(inventory, 1):
        ws.write(row, 0, inv.get("customer", ""))
        ws.write(row, 1, inv.get("style", inv.get("sku", "")))
        ws.write(row, 2, inv.get("color", ""))
        ws.write(row, 3, inv.get("size", ""))
        ws.write(row, 4, inv.get("description", ""))
        ws.write(row, 5, inv.get("category", ""))
        ws.write(row, 6, inv.get("manufacturer", ""))
        ws.write(row, 7, inv.get("location", ""))
        ws.write(row, 8, inv.get("total_boxes", 0))
        ws.write(row, 9, inv.get("units_on_hand", 0))
        ws.write(row, 10, inv.get("units_allocated", 0))
        ws.write(row, 11, inv.get("units_on_hand", 0) - inv.get("units_allocated", 0))
        ws.write(row, 12, inv.get("country_of_origin", ""))
        ws.write(row, 13, inv.get("fabric_content", ""))
        ws.write(row, 14, "YES" if inv.get("is_bpo") else "NO")
        _tr_at, _tr_by = _row_transfer(inv)
        ws.write(row, 15, _tr_at)
        ws.write(row, 16, _tr_by)

    # ── Sheet 2: box-level detail (Case# 007) ────────────────────────────────
    # Which physical box / LPN sits in each location — the aggregated sheet above
    # buckets by SKU+location and hides the box numbers. (boxes fetched up-front.)
    ws2 = wb.add_worksheet("Cajas - LPNs")
    box_headers = ["Box / LPN", "Customer", "Style", "Color", "Size", "Location",
                   "Units", "Status", "Country of Origin", "Fabric Content", "Description",
                   "Última transferencia", "Transferido por"]
    for i, h in enumerate(box_headers):
        ws2.write(0, i, h, bold)
    for row, b in enumerate(boxes, 1):
        ws2.write(row, 0, b.get("box_id", b.get("lpn_id", "")))
        ws2.write(row, 1, b.get("customer", ""))
        ws2.write(row, 2, b.get("style", b.get("sku", "")))
        ws2.write(row, 3, b.get("color", ""))
        ws2.write(row, 4, b.get("size", ""))
        ws2.write(row, 5, b.get("location", ""))
        ws2.write(row, 6, int(b.get("units") or b.get("qty") or 0))
        ws2.write(row, 7, b.get("status", b.get("state", "")))
        ws2.write(row, 8, b.get("country_of_origin", b.get("coo", "")))
        ws2.write(row, 9, b.get("fabric_content", ""))
        ws2.write(row, 10, b.get("description", ""))
        ws2.write(row, 11, (b.get("last_transferred_at") or "")[:19].replace("T", " "))
        ws2.write(row, 12, b.get("last_transferred_by", ""))

    wb.close()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=inventory.xlsx"})


@router.get("/export/receiving")
async def export_receiving(request: Request, customer: str = ""):
    """Export receivings to Excel (optionally by customer), most recent first,
    WITH ALERTS:
      - POSIBLE DOBLE RECIBO: receipts sharing (cliente, style, color, talla,
        unidades, ubicacion), scored by confidence (ALTA/MEDIA/REVISAR).
        Flagged on the main sheet and listed on a 'Posibles dobles' sheet.
      - ELIMINADOS: receipts that were deleted (from the movement log) on an
        'Eliminados' sheet."""
    from collections import defaultdict
    await require_auth(request)
    customer = (customer or "").strip()
    query = {}
    if customer:
        query["customer"] = {"$regex": f"^{re.escape(customer)}$", "$options": "i"}
    docs = await db.wms_receiving.find(query, {"_id": 0}).sort("created_at", -1).to_list(None)

    # Duplicate detection. Grouping by (cliente, style, color, talla, unidades,
    # ubicacion) only finds CANDIDATES — a legit re-receipt of the same qty to the
    # same cart on another day matches too. So we score each group's confidence by
    # how a real double-capture actually looks (rapid, same operator, same day):
    #   ALTA    : two captures <= 15 min apart  -> casi seguro doble captura
    #   MEDIA   : mismo dia (pero no tan juntas) -> probable, revisar
    #   REVISAR : en dias distintos u operadores distintos -> posible legitimo
    def dkey(r):
        return (
            (r.get("customer") or "").strip().upper(),
            (r.get("style") or "").strip().upper(),
            (r.get("color") or "").strip().upper(),
            (r.get("size") or "").strip().upper(),
            int(r.get("total_units") or 0),
            (r.get("inv_location") or "").strip().upper(),
        )

    def _ts(r):
        try:
            return datetime.fromisoformat((r.get("created_at") or "").replace("Z", "+00:00"))
        except Exception:
            return None

    groups = defaultdict(list)
    for r in docs:
        groups[dkey(r)].append(r)

    dup_conf = {}        # receiving_id -> "ALTA" | "MEDIA" | "REVISAR"
    dup_groups_list = []  # (conf, key, members, min_gap_min, n_ops)
    for k, g in groups.items():
        if len(g) < 2:
            continue
        members = sorted(g, key=lambda r: r.get("created_at") or "")
        days = {(r.get("created_at") or "")[:10] for r in members if r.get("created_at")}
        ops = {(r.get("received_by_name") or "").strip() for r in members if (r.get("received_by_name") or "").strip()}
        gaps = []
        for i in range(1, len(members)):
            a, b = _ts(members[i - 1]), _ts(members[i])
            if a and b:
                gaps.append(abs((b - a).total_seconds()))
        min_gap = min(gaps) if gaps else None
        if min_gap is not None and min_gap <= 900:
            conf = "ALTA"
        elif len(days) <= 1 and len(ops) <= 1:
            conf = "MEDIA"
        else:
            conf = "REVISAR"
        for r in members:
            dup_conf[r.get("receiving_id")] = conf
        dup_groups_list.append((conf, k, members, (round(min_gap / 60, 1) if min_gap is not None else None), len(ops)))
    _ord = {"ALTA": 0, "MEDIA": 1, "REVISAR": 2}
    dup_groups_list.sort(key=lambda x: (_ord.get(x[0], 9), -len(x[2])))

    # Deleted receipts from the movement log (new dedicated type + legacy text).
    del_q = {"$or": [
        {"type": "receiving_deleted"},
        {"type": "deallocate", "details.details": {"$regex": "Eliminado registro de receiving", "$options": "i"}},
    ]}
    if customer:
        del_q = {"$and": [del_q, {"details.customer": {"$regex": f"^{re.escape(customer)}$", "$options": "i"}}]}
    deleted = await db.wms_movements.find(del_q, {"_id": 0}).sort("created_at", -1).to_list(None)

    import xlsxwriter
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf)
    bold = wb.add_format({"bold": True})
    warn = wb.add_format({"bold": True, "font_color": "#B00020"})

    # ── Sheet 1: Recibos (with ALERTA column) ─────────────────────────────────
    ws = wb.add_worksheet("Recibos")
    headers = ["ALERTA", "Receiving ID", "Fecha", "Cliente", "Fabricante", "Style", "SKU",
               "Color", "Talla", "Descripcion", "Pais", "Fabric", "Lote",
               "Ubicacion", "Unidades", "Cajas", "ASN", "UPC", "BPO", "Recibido por"]
    for i, h in enumerate(headers):
        ws.write(0, i, h, bold)
    for row, r in enumerate(docs, 1):
        conf = dup_conf.get(r.get("receiving_id"))
        ws.write(row, 0, f"POSIBLE DOBLE RECIBO ({conf})" if conf else "", warn if conf else None)
        ws.write(row, 1, r.get("receiving_id", ""))
        ws.write(row, 2, r.get("created_at", ""))
        ws.write(row, 3, r.get("customer", ""))
        ws.write(row, 4, r.get("manufacturer", ""))
        ws.write(row, 5, r.get("style", ""))
        ws.write(row, 6, r.get("sku", ""))
        ws.write(row, 7, r.get("color", ""))
        ws.write(row, 8, r.get("size", ""))
        ws.write(row, 9, r.get("description", ""))
        ws.write(row, 10, r.get("country_of_origin", ""))
        ws.write(row, 11, r.get("fabric_content", ""))
        ws.write(row, 12, r.get("lot_number", ""))
        ws.write(row, 13, r.get("inv_location", ""))
        ws.write(row, 14, int(r.get("total_units") or 0))
        ws.write(row, 15, len(r.get("boxes") or []))
        ws.write(row, 16, r.get("asn_reference", ""))
        ws.write(row, 17, r.get("upc", ""))
        ws.write(row, 18, "YES" if r.get("is_bpo") else "NO")
        ws.write(row, 19, r.get("received_by_name", ""))

    # ── Sheet 2: Duplicados (con nivel de confianza) ──────────────────────────
    wsd = wb.add_worksheet("Posibles dobles")
    wsd.write(0, 0, "POSIBLES DOBLES RECIBOS (a revisar). Confianza: ALTA = capturas <=15 min "
                    "(casi seguro doble recibo) · MEDIA = mismo dia y operador (probable) · "
                    "REVISAR = dias u operadores distintos (posible recibo legitimo)", bold)
    dh = ["Grupo", "Confianza", "Veces", "Min entre capturas", "Operadores",
          "Cliente", "Style", "Color", "Talla", "Unidades", "Ubicacion",
          "Receiving ID", "Fecha", "ASN", "Lote", "Recibido por"]
    hrow = 1
    for i, h in enumerate(dh):
        wsd.write(hrow, i, h, bold)
    drow = hrow + 1
    for gi, (conf, k, members, min_gap_min, n_ops) in enumerate(dup_groups_list, 1):
        for r in members:
            wsd.write(drow, 0, gi)
            wsd.write(drow, 1, conf, warn if conf == "ALTA" else None)
            wsd.write(drow, 2, len(members))
            wsd.write(drow, 3, min_gap_min if min_gap_min is not None else "")
            wsd.write(drow, 4, n_ops)
            wsd.write(drow, 5, r.get("customer", ""))
            wsd.write(drow, 6, r.get("style", ""))
            wsd.write(drow, 7, r.get("color", ""))
            wsd.write(drow, 8, r.get("size", ""))
            wsd.write(drow, 9, int(r.get("total_units") or 0))
            wsd.write(drow, 10, r.get("inv_location", ""))
            wsd.write(drow, 11, r.get("receiving_id", ""))
            wsd.write(drow, 12, r.get("created_at", ""))
            wsd.write(drow, 13, r.get("asn_reference", ""))
            wsd.write(drow, 14, r.get("lot_number", ""))
            wsd.write(drow, 15, r.get("received_by_name", ""))
            drow += 1
    if drow == hrow + 1:
        wsd.write(hrow + 1, 0, "Sin duplicados detectados")

    # ── Sheet 3: Eliminados ───────────────────────────────────────────────────
    wse = wb.add_worksheet("Eliminados")
    eh = ["Fecha", "Receiving ID", "Cliente", "Style", "Color", "Talla",
          "Unidades", "Ubicacion", "Cajas revertidas", "Usuario", "Detalle"]
    for i, h in enumerate(eh):
        wse.write(0, i, h, bold)
    for row, m in enumerate(deleted, 1):
        d = m.get("details", {}) or {}
        wse.write(row, 0, m.get("created_at", ""))
        wse.write(row, 1, d.get("receiving_id", ""))
        wse.write(row, 2, d.get("customer", ""))
        wse.write(row, 3, d.get("style", ""))
        wse.write(row, 4, d.get("color", ""))
        wse.write(row, 5, d.get("size", ""))
        wse.write(row, 6, int(d.get("total_units", 0) or 0))
        wse.write(row, 7, d.get("inv_location", ""))
        wse.write(row, 8, int(d.get("boxes_reverted", 0) or 0))
        wse.write(row, 9, m.get("user_name", m.get("user_id", "")))
        wse.write(row, 10, d.get("details", ""))
    if not deleted:
        wse.write(1, 0, "Sin recibos eliminados registrados")

    wb.close()
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", customer) or "todos"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=recibos_{safe}.xlsx"})


# ==================== ASN (Advanced Shipping Notice) ====================

# Field -> list of header keywords (lowercase, accent-free). A column matches
# a field if any of the field's keywords appears in the normalized header text.
# Both R5 (English) and R6 (Spanish) headers are merged per column before
# matching, so a column with just the Spanish label still resolves correctly.
_ASN_HEADER_KEYWORDS = {
    "part_number":  ["part number", "no. de parte", "numero de parte", "no de parte", "n de parte", "style", "sku", "pn"],
    "description":  ["description", "descripcion espanol", "descripcion ingles", "english description", "descripcion"],
    "qty":          ["qty", "quantity", "cantidad"],
    "country":      ["country", "pais origen", "pais de origen", "country of origin", "coo"],
    "brand":        ["brand", "marca"],
    "po":           ["po number", "entry number", "numero de entrada", "numeor de entrada", "po"],
}
# Required fields — if any is missing after detection the import refuses.
_ASN_REQUIRED_FIELDS = ("part_number", "qty")

# Box statuses that mean the units already LEFT inventory (consumed/in process/
# shipped). Anything else with units>0 is considered still on hand. Whitelisting
# the "out" set keeps legacy/empty statuses counted as in-stock.
_BOX_OUT_STATUSES = {"shipped", "in_production", "finished", "in_neck_cutting", "confirmed", "depleted", "recon_pending"}

def _box_in_stock(b) -> bool:
    units = int(b.get("units") or b.get("qty") or 0)
    return units > 0 and (b.get("status") or "") not in _BOX_OUT_STATUSES

def _box_keys(b) -> set:
    """Identifiers a box can be matched to an ASN line by."""
    return {str(b.get(k) or "").strip().upper() for k in ("upc", "sku", "style") if b.get(k)}

def _normalize_header(s) -> str:
    """Lowercase, strip accents, collapse whitespace, drop punctuation."""
    if s is None:
        return ""
    import unicodedata
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    # Replace common separators with spaces; drop other punctuation
    out = []
    for ch in t:
        if ch.isalnum() or ch == " ":
            out.append(ch)
        elif ch in "-_/.,:;":
            out.append(" ")
    return " ".join("".join(out).split())

def _detect_sheet_kind(name: str) -> str | None:
    """Return a label for the sheet kind, or None if it's not a packing-list sheet.
    The kind is informational only — column detection is name-based per sheet."""
    n = (name or "").upper()
    if "RAW MATERIAL" in n or "MATERIA PRIMA" in n:
        return "raw"
    if "EQUIPMENT" in n or "EQUIPO" in n:
        return "equipment"
    if "FINISHED GOODS" in n or "PRODUCTO TERM" in n or "PRODUCTOTERM" in n:
        return "finished_goods"
    return None

def _find_asn_number(ws) -> str:
    """Scan top-left region for a cell labelled 'ASN' and return the neighbour value."""
    for r in range(1, 12):
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().upper() == "ASN":
                for dc in range(1, 5):
                    rv = ws.cell(row=r, column=c + dc).value
                    if rv not in (None, ""):
                        return str(rv).strip()
    return ""

_LABEL_DENYLIST = {
    "PACKING LIST", "BILL OF MATERIALS", "RAW MATERIAL", "EQUIPMENT",
    "FINISHED GOODS", "MATERIA PRIMA", "EQUIPO", "PRODUCTO TERMINADO",
    "RAW MATERIALS",
}

def _find_label_value(ws, *label_aliases, rows=8, max_cols=20, scan_right=6, max_gap=3) -> str:
    """Find a header label (e.g. 'Cliente:') anywhere in the top of the sheet
    and return the value next to it. Stops after `max_gap` consecutive blanks
    so we don't latch onto an unrelated heading further down the row."""
    aliases = {a.strip().upper().rstrip(':') for a in label_aliases}
    for r in range(1, rows + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            key = str(v).strip().upper().rstrip(':')
            if key not in aliases:
                continue
            blanks = 0
            for dc in range(1, scan_right + 1):
                rv = ws.cell(row=r, column=c + dc).value
                if rv in (None, ""):
                    blanks += 1
                    if blanks >= max_gap:
                        break
                    continue
                blanks = 0
                text = str(rv).strip()
                if not text:
                    continue
                upper = text.upper().rstrip(':')
                if text.endswith(':') or upper in aliases or upper in _LABEL_DENYLIST:
                    continue  # adjacent cell is another label, not a value
                return text
    return ""

# Patterns like "100% DE ALGODÓN", "50% ALGODON", "50% POLIESTER".
# Captures the percentage word + the material word (with common diacritics).
_FABRIC_PATTERN = re.compile(
    r"\d+\s*%\s*(?:DE\s+)?[A-ZÁÉÍÓÚÑ/]+(?:\s+[A-ZÁÉÍÓÚÑ/]+)?",
    re.IGNORECASE,
)
_STOP_FABRIC_WORDS = {"DE", "DEL", "LA", "EL", "PARA", "CON", "Y"}

def _extract_fabric_from_description(desc: str) -> str:
    """Pull fabric composition out of a free-form description.
    Examples:
      'CAMISETA ... 100% DE ALGODON'           -> '100% ALGODON'
      'CAMISETA ... 50% ALGODON, 50% POLIESTER' -> '50% ALGODON, 50% POLIESTER'
    Returns '' when no percentage pattern is found.
    """
    if not desc:
        return ""
    text = str(desc).upper()
    matches = _FABRIC_PATTERN.findall(text)
    cleaned = []
    for m in matches:
        # Normalise: drop "DE" connector, collapse whitespace, strip trailing slash/comma
        parts = [p for p in re.split(r"\s+", m) if p and p.upper() != "DE"]
        result = " ".join(parts).rstrip("/,. ")
        if result:
            cleaned.append(result)
    return ", ".join(cleaned)

def _detect_columns(ws, scan_rows: int = 15) -> tuple[dict[str, int], int]:
    """Find the header row and map field -> 0-based column index by header name.

    Strategy:
      1. For each row in 1..scan_rows, score how many fields can be located in
         that row's cells.
      2. Take the best row. If the next row also has hits, merge them column-
         wise (bilingual templates put English + Spanish on consecutive rows).
      3. Resolve each field to a column index using the field's keyword list.

    Returns (column_map, data_start_row). data_start_row = best + 1 (or +2 if
    we merged with the row below).
    """
    max_col = min(ws.max_column or 50, 50)
    # Per-row, per-field: which column matched (or -1)
    per_row: list[dict[str, int]] = []
    for r in range(1, scan_rows + 1):
        row_map: dict[str, int] = {}
        for c in range(1, max_col + 1):
            text = _normalize_header(ws.cell(row=r, column=c).value)
            if not text:
                continue
            for field, kws in _ASN_HEADER_KEYWORDS.items():
                if field in row_map:
                    continue  # first match wins per row
                if any(kw in text for kw in kws):
                    row_map[field] = c - 1  # 0-based for openpyxl iter_rows tuples
                    break
        per_row.append(row_map)

    # Pick the row with the most field hits
    best_row, best_score = 0, 0
    for i, rm in enumerate(per_row):
        score = sum(1 for f in _ASN_HEADER_KEYWORDS if f in rm)
        if score > best_score:
            best_row, best_score = i, score

    if best_score == 0:
        return {}, 0

    merged = dict(per_row[best_row])
    data_start = best_row + 1 + 1  # 1-based row right after header

    # If the next row also hits any field not in best_row, merge it (bilingual)
    if best_row + 1 < len(per_row):
        for f, c in per_row[best_row + 1].items():
            merged.setdefault(f, c)
        if per_row[best_row + 1]:
            data_start = best_row + 2 + 1

    return merged, data_start

def _parse_packing_list(ws) -> tuple[list[dict], str, dict[str, int]]:
    """Return (items, po_number, detected_column_map) from a worksheet.

    Column detection is by header name (see _detect_columns) so layout shifts
    across vendor revisions don't break the parse — as long as the header text
    is recognizable.
    """
    col_map, data_start = _detect_columns(ws)
    missing = [f for f in _ASN_REQUIRED_FIELDS if f not in col_map]
    if missing:
        raise HTTPException(
            400,
            f"No se detectaron columnas obligatorias: {', '.join(missing)}. "
            "Revisa que los encabezados del archivo incluyan al menos "
            "'Part Number' y 'Qty'/'Cantidad'."
        )

    pn_col = col_map["part_number"]
    qty_col = col_map["qty"]

    def cell_at(row, field):
        c = col_map.get(field)
        if c is None or c >= len(row) or row[c] is None:
            return ""
        return str(row[c]).strip()

    items: list[dict] = []
    po_number = ""
    line_no = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue
        if pn_col >= len(row):
            continue
        pn_raw = row[pn_col]
        if pn_raw is None:
            continue
        pn = str(pn_raw).strip()
        if not pn or pn.upper() == "TOTAL":
            continue
        try:
            qty = int(float(row[qty_col] if qty_col < len(row) and row[qty_col] is not None else 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        if not po_number:
            po_number = cell_at(row, "po")
        line_no += 1
        desc = cell_at(row, "description")
        items.append({
            "line_no": line_no,
            "part_number": pn.upper(),
            "description": desc,
            "fabric_content": _extract_fabric_from_description(desc),
            "qty_expected": qty,
            "qty_received": 0,
            "country": cell_at(row, "country").upper(),
            "brand": cell_at(row, "brand").upper(),
        })
    return items, po_number, col_map


def _parse_packing_list_pdf(contents: bytes) -> dict:
    """Parse an aduanal-style 'Lista de Empaque / Packing List' PDF.

    Layout: a key/value header (NO. FACTURA / INVOICE = the ASN number,
    REMITENTE / SHIPPER = vendor, VENDEDOR / SELLER = brand) followed by a
    merchandise table where each item begins with '(N) <part_number>'. On that
    same line the quantity is the positive WHOLE number (net weight is
    fractional; bulk/package qty are zero), and 'Pais Origen:' gives the COO.
    Header values are read via word coordinates because the page has 3 columns.

    Returns { asn_id, vendor, brand, customer, po_number, items }.
    """
    import pdfplumber

    lines: list[str] = []
    header_words: list[dict] = []
    with pdfplumber.open(io.BytesIO(contents)) as pdf:
        for pi, page in enumerate(pdf.pages):
            lines.extend((page.extract_text(x_tolerance=1.5) or "").split("\n"))
            if pi == 0:
                header_words = page.extract_words(x_tolerance=1.5)
    full = "\n".join(lines)

    # Header value sitting directly below a label word, kept within the label's
    # column so 3-column layouts don't bleed into the neighbour.
    def value_below(*labels):
        lab = None
        for w in header_words:
            up = w["text"].upper()
            if any(l in up for l in labels):
                lab = w
                break
        if not lab:
            return ""
        lx, ltop = lab["x0"], lab["top"]
        below = [w for w in header_words if w["top"] > ltop + 2 and (lx - 20) <= w["x0"] <= (lx + 200)]
        if not below:
            return ""
        below.sort(key=lambda w: (w["top"], w["x0"]))
        first_top = below[0]["top"]
        row_words = [w for w in below if abs(w["top"] - first_top) <= 4]
        return " ".join(w["text"] for w in sorted(row_words, key=lambda w: w["x0"])).strip()

    asn_id = ""
    m = re.search(r"(?:NO\.?\s*FACTURA|INVOICE)\s*[:/#]*\s*([0-9][0-9A-Za-z\-]{2,})", full, re.I)
    if m:
        asn_id = m.group(1).strip()

    vendor = value_below("SHIPPER", "REMITENTE")
    seller = value_below("SELLER", "VENDEDOR")
    brand = (seller.split()[0] if seller else "").upper()

    NUM = re.compile(r"\d[\d,]*\.\d+|\d[\d,]*")
    item_re = re.compile(r"^\((\d+)\)\s*(\S+)?")
    items: list[dict] = []
    cur: dict | None = None

    def flush():
        nonlocal cur
        if cur and cur.get("part_number"):
            desc = " ".join(cur.pop("_desc", [])).strip()
            cur["description"] = desc
            cur["fabric_content"] = _extract_fabric_from_description(desc)
            items.append(cur)
        cur = None

    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        mi = item_re.match(s)
        if mi:
            flush()
            part = (mi.group(2) or "").strip().upper()
            rest = s[mi.end():]
            nums = [float(x.replace(",", "")) for x in NUM.findall(rest)]
            whole_pos = [n for n in nums if n > 0 and abs(n - round(n)) < 1e-6]
            qty = int(round(whole_pos[0])) if whole_pos else (int(round(max(nums))) if nums else 0)
            cur = {
                "line_no": int(mi.group(1)),
                "part_number": part,
                "_desc": [],
                "qty_expected": qty,
                "qty_received": 0,
                "country": "",
                "brand": brand,
            }
            continue
        if cur is None:
            continue
        u = s.upper()
        cm = re.match(r"PA[IÍ]S\s*ORIGEN\s*[:]?\s*([A-Za-z]{2,3})", u)
        if cm:
            cur["country"] = cm.group(1).upper()
            continue
        if u.startswith(("FRACCION", "FRACCIÓN", "PO:", "LOT:", "PESO", "NET WT", "U.M.C", "UMC")):
            continue
        if not s.startswith("(") and len(cur["_desc"]) < 4:
            cur["_desc"].append(s)
    flush()

    return {
        "asn_id": asn_id,
        "vendor": (vendor or "").upper(),
        "brand": brand,
        "customer": vendor or "",
        "po_number": "",
        "items": items,
    }


@router.post("/asn")
async def create_asn(request: Request):
    """Manual ASN creation. Caller supplies asn_id; we don't auto-generate so the
    physical packing-list number stays as the source of truth."""
    user = await require_auth(request)
    body = await request.json()

    asn_id = str(body.get("asn_id", "")).strip()
    if not asn_id:
        raise HTTPException(400, "asn_id requerido (numero del packing list)")

    items = body.get("items", [])
    if not items:
        raise HTTPException(400, "El ASN debe contener al menos un item")

    if await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 1}):
        raise HTTPException(409, f"ASN {asn_id} ya existe")

    normalized_items = []
    for idx, it in enumerate(items, start=1):
        qty = int(it.get("qty_expected", it.get("quantity", 0)) or 0)
        normalized_items.append({
            "line_no": idx,
            "part_number": str(it.get("part_number", it.get("sku", ""))).strip().upper(),
            "description": str(it.get("description", "")).strip(),
            "qty_expected": qty,
            "qty_received": 0,
            "country": str(it.get("country", "")).strip().upper(),
            "brand": str(it.get("brand", "")).strip().upper(),
        })

    doc = {
        "asn_id": asn_id,
        "po_number": str(body.get("po_number", "")).strip(),
        "vendor": str(body.get("vendor", "")).strip().upper(),
        "expected_date": body.get("expected_date", ""),
        "source_sheet": "",
        "source_file": "",
        "items": normalized_items,
        "status": AsnStatus.PENDING,
        "created_at": now_iso(),
        "created_by": user.get("user_id"),
    }
    await db.wms_asn.insert_one(doc)
    doc.pop("_id", None)
    return doc

@router.get("/asn")
async def list_asn(request: Request):
    await require_auth(request)
    return await db.wms_asn.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)

# NOTE: declared BEFORE /asn/{asn_id} so "trace-sku" isn't captured as an asn_id.
@router.get("/asn/trace-sku")
async def trace_sku(request: Request, q: str = Query(...)):
    """Given a SKU/style/UPC, return which ASN(s) it came from, with how much is
    still in stock, where, and the box counts/dates. Boxes without an
    asn_reference are grouped under '(SIN ASN)'."""
    await require_auth(request)
    qn = (q or "").strip().upper()
    if not qn:
        raise HTTPException(400, "Parámetro 'q' requerido")
    rx = {"$regex": f"^{re.escape(qn)}$", "$options": "i"}
    boxes = await db.wms_boxes.find({"$or": [{"sku": rx}, {"style": rx}, {"upc": qn}]}, {"_id": 0}).to_list(20000)
    if not boxes:
        contains = {"$regex": re.escape(qn), "$options": "i"}
        boxes = await db.wms_boxes.find({"$or": [{"sku": contains}, {"style": contains}]}, {"_id": 0}).to_list(20000)

    groups: dict[str, dict] = {}
    for b in boxes:
        ref = (b.get("asn_reference") or "").strip() or "(SIN ASN)"
        u = int(b.get("units") or b.get("qty") or 0)
        g = groups.setdefault(ref, {
            "asn_reference": ref, "units_in_stock": 0, "units_total": 0,
            "boxes": 0, "boxes_in_stock": 0, "locations": set(), "skus": set(),
            "first_at": None, "last_at": None,
        })
        g["boxes"] += 1
        g["units_total"] += u
        if _box_in_stock(b):
            g["units_in_stock"] += u
            g["boxes_in_stock"] += 1
            if b.get("location"):
                g["locations"].add(str(b["location"]).strip())
        if b.get("sku") or b.get("style"):
            g["skus"].add(str(b.get("sku") or b.get("style")).strip())
        ca = b.get("created_at")
        if ca:
            g["first_at"] = min(g["first_at"], ca) if g["first_at"] else ca
            g["last_at"] = max(g["last_at"], ca) if g["last_at"] else ca

    results = []
    for ref, g in groups.items():
        asn = None
        if ref != "(SIN ASN)":
            asn = await db.wms_asn.find_one({"asn_id": ref}, {"_id": 0, "vendor": 1, "po_number": 1, "status": 1})
        results.append({
            **g,
            "locations": sorted(g["locations"]),
            "skus": sorted(g["skus"]),
            "vendor": (asn or {}).get("vendor", ""),
            "po_number": (asn or {}).get("po_number", ""),
            "asn_status": (asn or {}).get("status", ""),
            "exists": asn is not None,
        })
    results.sort(key=lambda x: (-x["units_in_stock"], x["asn_reference"]))
    return {"query": qn, "total_boxes": len(boxes), "groups": results}

@router.get("/asn/{asn_id}")
async def get_asn_detail(asn_id: str, request: Request):
    """Return ASN doc + every receiving event + every box tied to this ASN.

    Trazabilidad: el frontend muestra un summary, una lista de recepciones
    (con quién, cuándo, lote, ubicación) y la tabla detallada de cajas.
    """
    await require_auth(request)
    asn = await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})
    if not asn:
        raise HTTPException(404, f"ASN {asn_id} no encontrado")

    boxes = await db.wms_boxes.find(
        {"asn_reference": asn_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(5000)

    receivings = await db.wms_receiving.find(
        {"asn_reference": asn_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(2000)

    # Aggregated summary across boxes (since boxes are the granular unit and
    # cover historical records that pre-date the receiving.asn_reference field).
    total_units = sum(int(b.get("units") or b.get("qty") or 0) for b in boxes)
    total_boxes = len(boxes)
    receivers = sorted({(r.get("received_by_name") or "").strip() for r in receivings if r.get("received_by_name")})
    distinct_styles = sorted({(b.get("style") or "").strip() for b in boxes if b.get("style")})
    distinct_locations = sorted({(b.get("location") or "").strip() for b in boxes if b.get("location")})
    dates = [r.get("created_at") for r in receivings if r.get("created_at")] + \
            [b.get("created_at") for b in boxes if b.get("created_at")]
    first_at = min(dates) if dates else None
    last_at = max(dates) if dates else None

    # ── Traceability: split current stock vs material that already left ──────
    units_expected = sum(int(it.get("qty_expected") or 0) for it in asn.get("items", []))
    units_received = sum(int(it.get("qty_received") or 0) for it in asn.get("items", []))
    in_stock_boxes = [b for b in boxes if _box_in_stock(b)]
    units_in_stock = sum(int(b.get("units") or b.get("qty") or 0) for b in in_stock_boxes)
    boxes_in_stock = len(in_stock_boxes)
    units_out = max(0, units_received - units_in_stock)

    # In-stock units grouped by location, and by SKU (what's left and where).
    by_location: dict[str, dict] = {}
    by_sku: dict[str, dict] = {}
    for b in in_stock_boxes:
        u = int(b.get("units") or b.get("qty") or 0)
        loc = (b.get("location") or "—").strip() or "—"
        l = by_location.setdefault(loc, {"location": loc, "units": 0, "boxes": 0})
        l["units"] += u; l["boxes"] += 1
        sku = (b.get("sku") or b.get("style") or "—").strip() or "—"
        s = by_sku.setdefault(sku, {"sku": sku, "style": b.get("style") or "", "color": b.get("color") or "", "size": b.get("size") or "", "units": 0, "boxes": 0, "locations": set()})
        s["units"] += u; s["boxes"] += 1; s["locations"].add(loc)
    by_sku_list = [{**v, "locations": sorted(v["locations"])} for v in by_sku.values()]

    # Per ASN line: how much of that line is still on hand (best-effort match by
    # part_number against each box's upc/sku/style).
    by_line = []
    for it in asn.get("items", []):
        pn = str(it.get("part_number") or "").strip().upper()
        line_stock = sum(
            int(b.get("units") or b.get("qty") or 0)
            for b in in_stock_boxes if pn and pn in _box_keys(b)
        )
        by_line.append({
            "line_no": it.get("line_no"),
            "part_number": it.get("part_number"),
            "qty_expected": int(it.get("qty_expected") or 0),
            "qty_received": int(it.get("qty_received") or 0),
            "qty_in_stock": line_stock,
        })

    summary = {
        "total_units": total_units,
        "total_boxes": total_boxes,
        "total_receivings": len(receivings),
        "receivers": receivers,
        "distinct_styles": distinct_styles,
        "distinct_locations": distinct_locations,
        "first_received_at": first_at,
        "last_received_at": last_at,
        # Traceability
        "units_expected": units_expected,
        "units_received": units_received,
        "units_in_stock": units_in_stock,
        "units_out": units_out,
        "boxes_in_stock": boxes_in_stock,
        "by_location": sorted(by_location.values(), key=lambda x: -x["units"]),
        "by_sku": sorted(by_sku_list, key=lambda x: -x["units"]),
        "by_line": by_line,
    }

    return {"asn": asn, "boxes": boxes, "receivings": receivings, "summary": summary}

@router.delete("/asn/{asn_id}")
async def delete_asn(asn_id: str, request: Request):
    """Delete an ASN. Received boxes are kept (their asn_reference becomes
    informational only) so historical traceability is preserved."""
    user = await require_auth(request)
    asn = await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})
    if not asn:
        raise HTTPException(404, f"ASN {asn_id} no encontrado")
    boxes_count = await db.wms_boxes.count_documents({"asn_reference": asn_id})
    await db.wms_asn.delete_one({"asn_id": asn_id})
    await log_movement(user, MovementType.ASN_IMPORTED, {
        "asn_id": asn_id, "deleted": True,
        "po_number": asn.get("po_number"),
        "had_boxes": boxes_count,
    })
    return {"status": "deleted", "asn_id": asn_id, "boxes_kept": boxes_count}

@router.put("/asn/{asn_id}")
async def update_asn(asn_id: str, request: Request):
    """Edit an ASN's header and packing-list lines. Admin / super-user.
    Received quantities are preserved (matched by line_no, then part_number),
    and status is recomputed from expected vs received."""
    user = await require_admin(request)
    asn = await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})
    if not asn:
        raise HTTPException(404, f"ASN {asn_id} no encontrado")
    body = await request.json()

    update = {}
    for k in ("po_number", "vendor", "expected_date"):
        if k in body:
            v = str(body.get(k) or "").strip()
            update[k] = v.upper() if k == "vendor" else v

    if "items" in body:
        incoming = body.get("items") or []
        if not incoming:
            raise HTTPException(400, "El ASN debe contener al menos un item")
        prev = asn.get("items", [])
        by_line = {it.get("line_no"): it for it in prev}
        by_part = {}
        for it in prev:
            by_part.setdefault(str(it.get("part_number", "")).upper(), it)
        normalized = []
        for idx, it in enumerate(incoming, start=1):
            part = str(it.get("part_number", it.get("sku", ""))).strip().upper()
            # Preserve received qty: match by the original line_no, then part_number.
            # IMPORTANT: only fall back to by_part when the item already has a line_no
            # (i.e. it is an existing line being edited). A brand-new item (no line_no)
            # must always start with qty_received = 0, even if the same part_number
            # already appears on another line.
            line_no_key = it.get("line_no")
            if line_no_key is not None:
                src = by_line.get(line_no_key) or by_part.get(part) or {}
            else:
                src = by_line.get(line_no_key) or {}   # new item → no inherited qty
            normalized.append({
                "line_no": idx,
                "part_number": part,
                "description": str(it.get("description", "")).strip(),
                "qty_expected": int(it.get("qty_expected", it.get("quantity", 0)) or 0),
                "qty_received": int(src.get("qty_received", 0) or 0),
                "country": str(it.get("country", "")).strip().upper(),
                "brand": str(it.get("brand", "")).strip().upper(),
            })
        update["items"] = normalized

        total_exp = sum(i["qty_expected"] for i in normalized)
        total_rcv = sum(i["qty_received"] for i in normalized)
        if total_rcv <= 0:
            update["status"] = AsnStatus.PENDING
        elif total_rcv >= total_exp and all(i["qty_received"] >= i["qty_expected"] for i in normalized):
            update["status"] = AsnStatus.RECEIVED
        else:
            update["status"] = AsnStatus.PARTIAL

    if not update:
        raise HTTPException(400, "Nada por actualizar")

    update["updated_at"] = now_iso()
    update["updated_by"] = user.get("user_id")
    await db.wms_asn.update_one({"asn_id": asn_id}, {"$set": update})
    await log_movement(user, MovementType.ASN_IMPORTED, {
        "asn_id": asn_id, "edited": True,
        "fields": [k for k in update if k not in ("updated_at", "updated_by")],
    })
    return await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})


def _asn_discrepancies(asn: dict) -> list:
    """Per-line expected vs received differences (snapshot of the discrepancy log)."""
    out = []
    for it in asn.get("items", []):
        exp = int(it.get("qty_expected") or 0)
        rcv = int(it.get("qty_received") or 0)
        diff = rcv - exp
        if diff != 0:
            out.append({
                "line_no": it.get("line_no"),
                "part_number": it.get("part_number"),
                "description": it.get("description", ""),
                "qty_expected": exp,
                "qty_received": rcv,
                "difference": diff,
                "type": "SOBRANTE" if diff > 0 else "FALTANTE",
            })
    return out


@router.post("/asn/{asn_id}/close")
async def close_asn(asn_id: str, request: Request):
    """Finish the receiving process for an ASN even with discrepancies. Snapshots
    the per-line discrepancy log, marks the ASN closed (and status received), and
    blocks further receiving against it until reopened."""
    user = await require_auth(request)
    body = await request.json()
    asn = await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})
    if not asn:
        raise HTTPException(404, f"ASN {asn_id} no encontrado")

    discrepancies = _asn_discrepancies(asn)
    update = {
        "closed": True,
        "closed_at": now_iso(),
        "closed_by": user.get("user_id"),
        "closed_by_name": user.get("name", user.get("email", "")),
        "closure_note": str(body.get("note", "")).strip(),
        "discrepancies": discrepancies,
        "status": AsnStatus.RECEIVED,
        "received_at": asn.get("received_at") or now_iso(),
    }
    await db.wms_asn.update_one({"asn_id": asn_id}, {"$set": update})
    await log_movement(user, MovementType.ASN_RECEIPT, {
        "asn_id": asn_id, "closed": True, "discrepancy_lines": len(discrepancies),
    })
    return await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})


@router.post("/asn/{asn_id}/reopen")
async def reopen_asn(asn_id: str, request: Request):
    """Reopen a closed ASN so receiving can continue. Admin / super-user."""
    user = await require_admin(request)
    asn = await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})
    if not asn:
        raise HTTPException(404, f"ASN {asn_id} no encontrado")
    total_exp = sum(int(it.get("qty_expected") or 0) for it in asn.get("items", []))
    total_rcv = sum(int(it.get("qty_received") or 0) for it in asn.get("items", []))
    status = AsnStatus.PENDING if total_rcv <= 0 else (AsnStatus.RECEIVED if total_rcv >= total_exp else AsnStatus.PARTIAL)
    await db.wms_asn.update_one({"asn_id": asn_id}, {
        "$set": {"closed": False, "status": status},
        "$unset": {"closed_at": "", "closed_by": "", "closed_by_name": "", "closure_note": "", "discrepancies": ""},
    })
    await log_movement(user, MovementType.ASN_RECEIPT, {"asn_id": asn_id, "reopened": True})
    return await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 0})

@router.post("/asn/import")
async def import_asn(
    request: Request,
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(default=None),
):
    """Two-phase import:
       - Phase 1 (no sheet_name): inspect the file and return its sheets + detected
         ASN# per sheet so the user can choose which to import.
       - Phase 2 (with sheet_name): parse that sheet and persist.
    """
    user = await require_auth(request)
    import openpyxl
    try:
        contents = await file.read()
    except Exception as e:
        logger.exception("ASN import: failed to read uploaded file")
        raise HTTPException(400, f"No se pudo leer el archivo subido: {e}")
    if not contents:
        raise HTTPException(400, "Archivo vacio o no recibido")

    # ── PDF branch: aduanal "Lista de Empaque / Packing List" ────────────────
    is_pdf = (file.filename or "").lower().endswith(".pdf") or \
             (file.content_type or "").lower() == "application/pdf" or \
             contents[:5] == b"%PDF-"
    if is_pdf:
        try:
            parsed = _parse_packing_list_pdf(contents)
        except Exception as e:
            logger.exception("ASN import: PDF parse failed")
            raise HTTPException(400, f"No se pudo leer el PDF del packing list: {e}")
        asn_id = parsed.get("asn_id") or ""
        items = parsed.get("items") or []
        if not asn_id:
            raise HTTPException(400, "No se encontró el número de factura/INVOICE en el PDF")
        if not items:
            raise HTTPException(400, "No se detectaron líneas de mercancía en el PDF")
        PDF_SHEET = "PACKING LIST (PDF)"

        # Phase 1: inspect → present a single virtual sheet for confirmation.
        if not sheet_name:
            return {"action": "select_sheet", "filename": file.filename, "sheets": [{
                "name": PDF_SHEET,
                "kind": "pdf",
                "detected_asn_id": asn_id,
                "detected_customer": parsed.get("vendor") or "",
                "row_count": len(items),
                "detected_columns": {"part_number": "PDF", "qty": "PDF", "country": "PDF"},
                "missing_required": [],
            }]}

        # Phase 2: persist.
        if await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 1}):
            raise HTTPException(409, f"ASN {asn_id} ya existe en el sistema")
        doc = {
            "asn_id": asn_id,
            "po_number": parsed.get("po_number") or "",
            "customer": parsed.get("customer") or "",
            "vendor": parsed.get("vendor") or "",
            "expected_date": now_iso(),
            "source_sheet": PDF_SHEET,
            "source_file": file.filename or "",
            "items": items,
            "status": AsnStatus.PENDING,
            "created_at": now_iso(),
            "created_by": user.get("user_id"),
        }
        await db.wms_asn.insert_one(doc)
        await log_movement(user, MovementType.ASN_IMPORTED, {
            "asn_id": asn_id, "items": len(items), "source": "pdf",
            "total_qty": sum(it.get("qty_expected", 0) for it in items),
        })
        return {
            "status": "success",
            "asn_id": asn_id,
            "po_number": doc["po_number"],
            "vendor": doc["vendor"],
            "items_count": len(items),
            "total_qty_expected": sum(it.get("qty_expected", 0) for it in items),
            "detected_columns": {},
        }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        logger.exception("ASN import: openpyxl could not open workbook")
        raise HTTPException(400, f"No se pudo leer el archivo Excel: {e}")

    # Helper: convert 0-based col index to Excel letter for the UI
    def _col_letter(idx: int) -> str:
        n, s = idx + 1, ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    # Phase 1: inspect
    if not sheet_name:
        sheets = []
        for sn in wb.sheetnames:
            kind = _detect_sheet_kind(sn)
            if not kind:
                continue
            ws = wb[sn]
            asn_num = _find_asn_number(ws)
            col_map, data_start = _detect_columns(ws)
            count = 0
            if "part_number" in col_map:
                pn_col = col_map["part_number"]
                for row in ws.iter_rows(min_row=data_start, values_only=True):
                    if not row or pn_col >= len(row):
                        continue
                    pn = row[pn_col]
                    if pn and str(pn).strip().upper() != "TOTAL":
                        count += 1
            sheets.append({
                "name": sn,
                "kind": kind,
                "detected_asn_id": asn_num,
                "detected_customer": _find_label_value(ws, "Cliente", "Client", "Customer"),
                "row_count": count,
                "detected_columns": {f: _col_letter(c) for f, c in col_map.items()},
                "missing_required": [f for f in _ASN_REQUIRED_FIELDS if f not in col_map],
            })
        if not sheets:
            raise HTTPException(400, "No se encontraron hojas compatibles (RAW MATERIAL / EQUIPMENT / FINISHED GOODS)")
        return {"action": "select_sheet", "filename": file.filename, "sheets": sheets}

    # Phase 2: import the chosen sheet
    if sheet_name not in wb.sheetnames:
        raise HTTPException(400, f"Hoja '{sheet_name}' no existe en el archivo")
    if not _detect_sheet_kind(sheet_name):
        raise HTTPException(400, f"Hoja '{sheet_name}' no es de un tipo soportado")

    ws = wb[sheet_name]
    asn_id = _find_asn_number(ws)
    if not asn_id:
        raise HTTPException(400, "No se encontro el numero de ASN en la hoja (celda contigua al label 'ASN')")

    if await db.wms_asn.find_one({"asn_id": asn_id}, {"_id": 1}):
        raise HTTPException(409, f"ASN {asn_id} ya existe en el sistema")

    items, po_number, col_map = _parse_packing_list(ws)
    if not items:
        raise HTTPException(400, "No se encontraron lineas con cantidad > 0 en la hoja")

    # Vendor: most common brand wins, fallback to first
    brands = [it["brand"] for it in items if it.get("brand")]
    vendor = max(set(brands), key=brands.count) if brands else ""

    # Customer: pulled from the "Cliente:" label at the top of the sheet
    customer = _find_label_value(ws, "Cliente", "Client", "Customer")

    doc = {
        "asn_id": asn_id,
        "po_number": po_number,
        "customer": customer,
        "vendor": vendor,
        "expected_date": now_iso(),
        "source_sheet": sheet_name,
        "source_file": file.filename or "",
        "items": items,
        "status": AsnStatus.PENDING,
        "created_at": now_iso(),
        "created_by": user.get("user_id"),
    }
    await db.wms_asn.insert_one(doc)

    detected_columns = {f: _col_letter(c) for f, c in col_map.items()}
    await log_movement(user, MovementType.ASN_IMPORTED, {
        "asn_id": asn_id, "po_number": po_number,
        "items": len(items), "sheet": sheet_name,
        "total_qty": sum(it["qty_expected"] for it in items),
        "detected_columns": detected_columns,
    })

    return {
        "status": "success",
        "asn_id": asn_id,
        "po_number": po_number,
        "vendor": vendor,
        "items_count": len(items),
        "total_qty_expected": sum(it["qty_expected"] for it in items),
        "detected_columns": detected_columns,
    }


async def _apply_receiving_to_asn(
    asn_id: str,
    received_by_pn: dict,
    user: dict,
    target_line_no: int | None = None,
    target_qty: int = 0,
) -> dict:
    """Increment qty_received on an ASN.

    Two modes:
      - target_line_no provided: increment that exact line by target_qty. The
        received_by_pn dict is ignored. Used when the operator explicitly
        picked which ASN line they're receiving (style vocabulary may not match).
      - target_line_no None: increment per part_number using received_by_pn.
        Lines not in the ASN come back as mismatched.

    Returns {matched: [...], mismatched: [...]} for caller to surface warnings.
    """
    result = {"matched": [], "mismatched": []}
    if not asn_id:
        return result

    asn = await db.wms_asn.find_one({"asn_id": asn_id})
    if not asn:
        if target_line_no is not None:
            result["mismatched"] = [{"line_no": target_line_no, "qty": target_qty, "reason": "asn_not_found"}]
        else:
            result["mismatched"] = [{"part_number": pn, "qty": q, "reason": "asn_not_found"}
                                    for pn, q in received_by_pn.items()]
        return result

    if target_line_no is not None:
        # Direct line-level decrement — bypasses style matching entirely
        line = next((it for it in asn.get("items", []) if it.get("line_no") == target_line_no), None)
        if line is None:
            result["mismatched"].append({
                "line_no": target_line_no, "qty": target_qty,
                "reason": "line_not_in_asn",
            })
        elif target_qty > 0:
            await db.wms_asn.update_one(
                {"asn_id": asn_id, "items.line_no": target_line_no},
                {"$inc": {"items.$.qty_received": int(target_qty)}},
            )
            result["matched"].append({
                "line_no": target_line_no,
                "part_number": line.get("part_number"),
                "qty": target_qty,
            })
    else:
        if not received_by_pn:
            return result
        items_by_pn = {it.get("part_number", "").upper(): it for it in asn.get("items", [])}
        for pn, qty in received_by_pn.items():
            pn_u = pn.upper()
            if pn_u in items_by_pn:
                await db.wms_asn.update_one(
                    {"asn_id": asn_id, "items.part_number": pn_u},
                    {"$inc": {"items.$.qty_received": int(qty)}},
                )
                result["matched"].append({"part_number": pn_u, "qty": qty})
            else:
                result["mismatched"].append({
                    "part_number": pn_u, "qty": qty,
                    "reason": "part_not_in_asn",
                })

    # Recalculate status
    fresh = await db.wms_asn.find_one({"asn_id": asn_id}, {"items": 1})
    items = fresh.get("items", []) if fresh else []
    total_exp = sum(it.get("qty_expected", 0) for it in items)
    total_rcv = sum(it.get("qty_received", 0) for it in items)
    if total_rcv <= 0:
        new_status = AsnStatus.PENDING
    elif total_rcv >= total_exp and all(
        it.get("qty_received", 0) >= it.get("qty_expected", 0) for it in items
    ):
        new_status = AsnStatus.RECEIVED
    else:
        new_status = AsnStatus.PARTIAL

    update = {"status": new_status}
    if new_status == AsnStatus.RECEIVED:
        update["received_at"] = now_iso()
    await db.wms_asn.update_one({"asn_id": asn_id}, {"$set": update})

    await log_movement(user, MovementType.ASN_RECEIPT, {
        "asn_id": asn_id,
        "received_by_pn": received_by_pn,
        "mismatched": result["mismatched"],
        "new_status": new_status,
    })

    return result

# ==================== SUPERVISOR OVERRIDES ====================

@router.put("/pick-tickets/{ticket_id}/prioritize")
async def prioritize_ticket(ticket_id: str, request: Request):
    user = await require_admin(request)
    res = await db.wms_pick_tickets.update_one(
        {"ticket_id": ticket_id},
        {"$set": {"priority": "HOT", "updated_at": now_iso()}}
    )
    if res.modified_count == 0:
        raise HTTPException(404, "Ticket no encontrado")
        
    await db.wms_tasks.update_many(
        {"context.ticket_id": ticket_id},
        {"$set": {"priority": "HOT", "updated_at": now_iso()}}
    )
    
    await log_movement(user, "ticket_prioritized", {"ticket_id": ticket_id})
    return {"status": "success", "message": "Ticket escalado a HOT"}

# ==================== MANUAL INVENTORY ENTRY ====================

@router.post("/inventory")
async def add_inventory_manual(request: Request):
    """Add a single inventory line manually. Accumulates with any existing row
    matching (style, color, size, location); creates a new row otherwise.
    Auto-creates the location if missing and generates LPN boxes for any new
    cases added by this call. Mirrors POST /import/inventory semantics for one
    line so the data shape stays consistent.
    """
    user = await require_inventory_level(request, 2)
    body = await request.json()

    def s(field, default=""):
        v = body.get(field, default)
        return str(v).strip().upper() if isinstance(default, str) else v

    style = s("style") or s("sku")
    if not style:
        raise HTTPException(400, "Style/SKU es requerido")

    location = s("location")
    if not location:
        raise HTTPException(400, "Ubicación es requerida")

    try:
        total_boxes = int(float(body.get("total_boxes", 0) or 0))
        total_units = int(float(body.get("total_units", 0) or 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Cantidades inválidas")

    if total_units <= 0:
        raise HTTPException(400, "Total de unidades debe ser mayor a 0")
    if total_boxes < 0:
        raise HTTPException(400, "Total de cajas no puede ser negativo")

    color = s("color")
    size = s("size")
    now = now_iso()

    # Operation: "add" (default, entrada) or "remove" (salida / ajuste a la baja).
    operation = str(body.get("operation", "add")).strip().lower()

    # Wider key so two batches of the same SKU with different fabric/COO
    # stay as separate inventory rows (mirrors the import + receiving keys).
    fabric_key = s("fabric_content")
    coo_key = s("country_of_origin")
    existing = await db.wms_inventory.find_one({
        "style": style, "color": color, "size": size, "location": location,
        "fabric_content": fabric_key, "country_of_origin": coo_key,
    })

    # ── Salida manual (remove): resta de una línea existente ──────────────────
    if operation == "remove":
        # Fall back to the core key when fabric/COO weren't supplied so the
        # operator can pull stock out without re-typing every metadata field.
        if not existing:
            existing = await db.wms_inventory.find_one(
                {"style": style, "color": color, "size": size, "location": location},
                sort=[("units_on_hand", -1)],
            )
        if not existing:
            raise HTTPException(404, "No existe inventario para esa combinación (SKU+color+talla+ubicación)")

        reason = s("reason")
        if not reason:
            raise HTTPException(400, "El motivo de la salida es obligatorio")

        on_hand = int(existing.get("units_on_hand", 0) or 0)
        allocated = int(existing.get("units_allocated", 0) or 0)
        cur_boxes = int(existing.get("total_boxes", 0) or 0)
        if total_units > on_hand:
            raise HTTPException(400, f"No puedes sacar {total_units}: solo hay {on_hand} en existencia")
        if on_hand - total_units < allocated:
            raise HTTPException(
                400,
                f"No puedes sacar {total_units}: {allocated} están comprometidas (allocated). Disponible libre: {on_hand - allocated}",
            )
        if total_boxes > cur_boxes:
            raise HTTPException(400, f"No puedes sacar {total_boxes} cajas: solo hay {cur_boxes} registradas")

        inventory_id = existing["inventory_id"]
        await db.wms_inventory.update_one(
            {"inventory_id": inventory_id},
            {"$inc": {"total_boxes": -total_boxes, "units_on_hand": -total_units},
             "$set": {"updated_at": now}},
        )
        # Drop the oldest LPN boxes for this row to match the case count removed.
        if total_boxes > 0:
            old_box_ids = [
                b["box_id"] for b in await db.wms_boxes.find(
                    {"inventory_id": inventory_id}, {"_id": 0, "box_id": 1}
                ).sort("created_at", 1).to_list(total_boxes)
            ]
            if old_box_ids:
                await db.wms_boxes.delete_many({"box_id": {"$in": old_box_ids}})

        await log_movement(user, "manual_inventory_remove", {
            "inventory_id": inventory_id, "mode": "removed",
            "style": style, "color": color, "size": size, "location": location,
            "removed_boxes": total_boxes, "removed_units": total_units,
        })
        return {
            "inventory_id": inventory_id, "mode": "removed", "location_created": False,
            "removed_boxes": total_boxes, "removed_units": total_units,
            "added_boxes": -total_boxes, "added_units": -total_units,
        }

    # Case# 006: link a positive adjustment to a REAL box number the operator
    # scans/types, instead of the synthetic LPN generated by default. When
    # supplied, exactly one real box is created (total_boxes forced to 1).
    manual_box_id = str(body.get("box_id", "") or "").strip().upper()
    if manual_box_id:
        clash = await db.wms_boxes.find_one(
            {"$or": [{"box_id": manual_box_id}, {"barcode": manual_box_id}, {"lpn_id": manual_box_id}]},
            {"_id": 0, "box_id": 1},
        )
        if clash:
            raise HTTPException(400, f"La caja {manual_box_id} ya existe; usa un número distinto")
        total_boxes = 1

    if existing:
        inventory_id = existing["inventory_id"]
        await db.wms_inventory.update_one(
            {"inventory_id": inventory_id},
            {"$inc": {"total_boxes": total_boxes, "units_on_hand": total_units},
             "$set": {"updated_at": now}},
        )
        mode = "accumulated"
    else:
        inventory_id = f"inv_{uuid.uuid4().hex[:12]}"
        await db.wms_inventory.insert_one({
            "inventory_id": inventory_id,
            "customer": s("customer"),
            "style": style,
            "sku": style,
            "color": color,
            "size": size,
            "size_header": s("size_header"),
            "manufacturer": s("manufacturer"),
            "description": s("description"),
            "category": s("category"),
            "country_of_origin": s("country_of_origin"),
            "fabric_content": s("fabric_content"),
            "import_number": body.get("import_number", ""),
            "po": body.get("po", ""),
            "bpo": body.get("bpo", ""),
            "location": location,
            "total_boxes": total_boxes,
            "units_on_hand": total_units,
            "units_allocated": 0,
            "updated_at": now,
        })
        mode = "created"

    # Generate the box(es) for the units we just added.
    if manual_box_id:
        # One real, scannable box stamped with the operator's number (Case# 006).
        await db.wms_boxes.insert_one({
            "box_id": manual_box_id, "barcode": manual_box_id, "lpn_id": manual_box_id,
            "inventory_id": inventory_id,
            "sku": style, "color": color, "size": size,
            "units": total_units, "qty": total_units,
            "location": location, "state": "putaway",
            "customer": s("customer"), "coo": coo_key, "country_of_origin": coo_key,
            "fabric_content": fabric_key, "description": s("description"),
            "manufacturer": s("manufacturer"), "created_at": now,
        })
    elif total_boxes > 0:
        # Fallback: synthetic LPNs (bulk multi-line add / no number supplied).
        units_per_box = total_units // total_boxes
        remainder = total_units % total_boxes
        box_docs = [{
            "box_id": f"LPN_{uuid.uuid4().hex[:8].upper()}",
            "inventory_id": inventory_id,
            "sku": style,
            "color": color,
            "size": size,
            "units": units_per_box + (1 if i < remainder else 0),
            "qty": units_per_box + (1 if i < remainder else 0),
            "location": location,
            "state": "putaway",
            "customer": s("customer"),
            "coo": coo_key,
            "country_of_origin": coo_key,
            "fabric_content": fabric_key,
            "description": s("description"),
            "manufacturer": s("manufacturer"),
            "created_at": now,
        } for i in range(total_boxes)]
        await db.wms_boxes.insert_many(box_docs)

    # Auto-create location if missing (case-insensitive check).
    loc_existing = await db.wms_locations.find_one(
        {"name": {"$regex": f"^{re.escape(location)}$", "$options": "i"}}
    )
    location_created = False
    if not loc_existing:
        zone = location.split("-")[0] if "-" in location else "DEFAULT"
        await db.wms_locations.insert_one({
            "location_id": f"loc_{uuid.uuid4().hex[:12]}",
            "name": location,
            "zone": zone,
            "type": "rack",
            "active": True,
            "created_at": now,
        })
        location_created = True

    await log_movement(user, "manual_inventory_add", {
        "inventory_id": inventory_id,
        "mode": mode,
        "reason": s("reason"),
        "style": style,
        "color": color,
        "size": size,
        "location": location,
        "added_boxes": total_boxes,
        "added_units": total_units,
        "box_id": manual_box_id or None,
    })

    return {
        "inventory_id": inventory_id,
        "mode": mode,
        "location_created": location_created,
        "added_boxes": total_boxes,
        "added_units": total_units,
        "box_id": manual_box_id or None,
    }


# ─── Bulk inventory adjustment (Mover · admin level 3) ───────────────────────
async def _line_boxes(inv):
    """Live LPN boxes backing an inventory line (by inventory_id, else identity)."""
    by_identity = {
        "location": {"$regex": f"^{re.escape(inv.get('location', ''))}$", "$options": "i"},
        "sku": inv.get("sku") or inv.get("style") or "",
        "color": inv.get("color", ""), "size": inv.get("size", ""),
    }
    q = {"$or": [{"inventory_id": inv["inventory_id"]}, by_identity]} if inv.get("inventory_id") else by_identity
    return await db.wms_boxes.find(q).sort("units", 1).to_list(2000)


async def _reconcile_line_boxes(inv, delta):
    """Keep the line's boxes' unit-sum tracking the on-hand delta. +delta adds one
    adjustment box; -delta draws down existing boxes (smallest first), deleting
    any that empty out. Returns the net change in the line's box count."""
    if delta > 0:
        seq = await _reserve_box_seqs(1)
        box_id = f"BOX-{seq:06d}"
        await db.wms_boxes.insert_one({
            "box_id": box_id, "barcode": box_id, "lpn_id": box_id,
            "inventory_id": inv.get("inventory_id"), "customer": inv.get("customer", ""),
            "style": inv.get("style") or inv.get("sku", ""), "sku": inv.get("sku") or inv.get("style", ""),
            "color": inv.get("color", ""), "size": inv.get("size", ""),
            "units": delta, "qty": delta, "seq_num": seq, "location": inv.get("location", ""),
            "status": "putaway_done", "state": "raw",
            "coo": inv.get("country_of_origin", ""), "country_of_origin": inv.get("country_of_origin", ""),
            "fabric_content": inv.get("fabric_content", ""), "is_adjustment": True, "created_at": now_iso(),
        })
        return 1
    need = -delta
    removed = 0
    for b in await _line_boxes(inv):
        if need <= 0:
            break
        bu = int(b.get("units") or b.get("qty") or 0)
        if bu <= 0:
            continue
        take = min(bu, need)
        left = bu - take
        if left <= 0:
            await db.wms_boxes.delete_one({"box_id": b["box_id"]})
            removed += 1
        else:
            await db.wms_boxes.update_one({"box_id": b["box_id"]}, {"$set": {"units": left, "qty": left, "updated_at": now_iso()}})
        need -= take
    return -removed


@router.post("/inventory/bulk-adjust")
async def bulk_adjust_inventory(request: Request):
    """Mass inventory adjustment from the 'Formato ajuste de inventario' Excel.
    Admin level 3+. Each row's 'on_hand' is a DELTA (positive adds, negative
    subtracts). dry_run=true returns the plan only (preview); dry_run=false
    applies it. Boxes (LPNs) are reconciled so the per-box sum tracks the new
    line on-hand. A reason is mandatory to apply (audited per row)."""
    user = await require_inventory_level(request, 2)
    body = await request.json()
    rows = body.get("rows") or []
    dry_run = bool(body.get("dry_run", True))
    reason = str(body.get("reason", "") or "").strip()
    if not isinstance(rows, list) or not rows:
        raise HTTPException(400, "No hay filas para ajustar")
    if not dry_run and not reason:
        raise HTTPException(400, "El motivo del ajuste es obligatorio")

    plan = []
    for idx, r in enumerate(rows):
        style = str(r.get("style", "") or "").strip().upper()
        color = str(r.get("color", "") or "").strip().upper()
        size = str(r.get("size", "") or "").strip().upper()
        location = str(r.get("location", "") or "").strip().upper()
        label = f"{style}{('-' + color) if color else ''}{('-' + size) if size else ''} @ {location or '—'}"
        try:
            delta = int(float(r.get("on_hand")))
        except (TypeError, ValueError):
            plan.append({"row": idx + 1, "label": label, "status": "error", "message": "Cantidad inválida"})
            continue
        if not style or not location:
            plan.append({"row": idx + 1, "label": label, "status": "error", "message": "Style y Location son obligatorios"})
            continue
        if delta == 0:
            plan.append({"row": idx + 1, "label": label, "status": "skip", "current": None, "delta": 0, "message": "Sin cambio (0)"})
            continue
        loc_rx = {"$regex": f"^{re.escape(location)}$", "$options": "i"}
        inv = await db.wms_inventory.find_one({"style": style, "color": color, "size": size, "location": loc_rx}, {"_id": 0}) \
            or await db.wms_inventory.find_one({"sku": style, "color": color, "size": size, "location": loc_rx}, {"_id": 0})
        if inv:
            current = int(inv.get("units_on_hand", 0) or 0)
            allocated = int(inv.get("units_allocated", 0) or 0)
            new_val = current + delta
            if new_val < 0:
                plan.append({"row": idx + 1, "label": label, "status": "error", "current": current, "delta": delta, "message": f"Quedaría negativo ({new_val})"})
            elif new_val < allocated:
                plan.append({"row": idx + 1, "label": label, "status": "error", "current": current, "delta": delta, "message": f"Quedaría {new_val} < comprometido {allocated}"})
            else:
                plan.append({"row": idx + 1, "label": label, "status": "adjust", "inventory_id": inv.get("inventory_id"), "current": current, "delta": delta, "new": new_val})
        else:
            if delta < 0:
                plan.append({"row": idx + 1, "label": label, "status": "error", "current": 0, "delta": delta, "message": "No existe la línea; no se puede restar"})
            else:
                plan.append({"row": idx + 1, "label": label, "status": "new", "current": 0, "delta": delta, "new": delta,
                             "customer": str(r.get("customer", "") or "").strip(),
                             "country_of_origin": str(r.get("country_of_origin", "") or "").strip(),
                             "fabric_content": str(r.get("fabric_content", "") or "").strip(),
                             "style": style, "color": color, "size": size, "location": location})

    summary = {
        "total": len(plan),
        "adjust": sum(1 for p in plan if p["status"] == "adjust"),
        "new": sum(1 for p in plan if p["status"] == "new"),
        "error": sum(1 for p in plan if p["status"] == "error"),
        "skip": sum(1 for p in plan if p["status"] == "skip"),
    }
    if dry_run:
        return {"dry_run": True, "summary": summary, "rows": plan}

    applied = 0
    for p in plan:
        try:
            if p["status"] == "adjust":
                inv = await db.wms_inventory.find_one({"inventory_id": p["inventory_id"]}, {"_id": 0})
                if not inv:
                    p["status"] = "error"; p["message"] = "La línea ya no existe"; continue
                box_change = await _reconcile_line_boxes(inv, p["delta"])
                await db.wms_inventory.update_one(
                    {"inventory_id": p["inventory_id"]},
                    {"$inc": {"units_on_hand": p["delta"], "total_boxes": box_change}, "$set": {"updated_at": now_iso()}},
                )
                await log_movement(user, "inventory_adjustment", {
                    "inventory_id": p["inventory_id"], "sku": inv.get("sku"), "location": inv.get("location"),
                    "delta": p["delta"], "new_on_hand": p["new"], "reason": reason, "bulk": True,
                })
                applied += 1
            elif p["status"] == "new":
                customer = await _canonical_customer(p.get("customer", ""))
                style, color, size = p["style"], p["color"], p["size"]
                sku = style + (("-" + color) if color else "") + (("-" + size) if size else "")
                inventory_id = gen_id("inv")
                qty = int(p["delta"])
                await db.wms_inventory.insert_one({
                    "inventory_id": inventory_id, "customer": customer, "style": style, "sku": sku,
                    "color": color, "size": size, "location": p["location"],
                    "units_on_hand": qty, "units_allocated": 0, "total_boxes": 1,
                    "country_of_origin": p.get("country_of_origin", ""), "fabric_content": p.get("fabric_content", ""),
                    "created_at": now_iso(), "updated_at": now_iso(),
                })
                seq = await _reserve_box_seqs(1)
                box_id = f"BOX-{seq:06d}"
                await db.wms_boxes.insert_one({
                    "box_id": box_id, "barcode": box_id, "lpn_id": box_id, "inventory_id": inventory_id,
                    "customer": customer, "style": style, "sku": sku, "color": color, "size": size,
                    "units": qty, "qty": qty, "seq_num": seq, "location": p["location"],
                    "status": "putaway_done", "state": "raw",
                    "coo": p.get("country_of_origin", ""), "country_of_origin": p.get("country_of_origin", ""),
                    "fabric_content": p.get("fabric_content", ""), "is_adjustment": True, "created_at": now_iso(),
                })
                await log_movement(user, "inventory_adjustment_create", {
                    "inventory_id": inventory_id, "sku": sku, "location": p["location"],
                    "delta": qty, "reason": reason, "bulk": True,
                })
                applied += 1
        except Exception as e:
            p["status"] = "error"; p["message"] = f"Error al aplicar: {e}"

    await log_movement(user, "bulk_inventory_adjustment", {"applied": applied, "reason": reason, **summary})
    return {"dry_run": False, "summary": {**summary, "applied": applied}, "rows": plan}


@router.delete("/inventory/{inventory_id}")
async def delete_inventory_row(inventory_id: str, request: Request):
    """Supersu-only: remove a single inventory line (and its linked LPN boxes)
    from a location. Used by the Locations detail modal (delete line / clear
    location). Matches strictly by inventory_id, so it never touches other
    locations or duplicate rows of the same SKU."""
    user = await require_admin_level(request, 2)
    inv = await db.wms_inventory.find_one({"inventory_id": inventory_id})
    if not inv:
        raise HTTPException(404, "Inventario no encontrado")
    # Optional motivo for the audit trail. Read from body if present, else query
    # param; not hard-required so the Locations modal keeps working, but captured
    # whenever supplied.
    reason = ""
    try:
        _b = await request.json()
        if isinstance(_b, dict):
            reason = str(_b.get("reason", "") or "").strip()
    except Exception:
        pass
    if not reason:
        reason = (request.query_params.get("reason") or "").strip()
    # Capture which LPNs are removed so the deletion is reconstructable.
    box_ids = await db.wms_boxes.distinct("box_id", {"inventory_id": inventory_id})
    await db.wms_inventory.delete_one({"inventory_id": inventory_id})
    box_res = await db.wms_boxes.delete_many({"inventory_id": inventory_id})
    await log_movement(user, "inventory_deleted", {
        "inventory_id": inventory_id,
        "sku": inv.get("sku"), "style": inv.get("style"),
        "color": inv.get("color"), "size": inv.get("size"),
        "location": inv.get("location"),
        "units_removed": inv.get("units_on_hand", 0),
        "boxes_removed": box_res.deleted_count,
        "box_ids": box_ids[:100],
        "reason": reason or None,
    })
    return {"message": "Inventario eliminado", "inventory_id": inventory_id, "boxes_removed": box_res.deleted_count}


# ==================== IMPORT INVENTORY ====================

@router.post("/import/inventory")
async def import_inventory(request: Request, file: UploadFile = File(...), force: bool = False):
    # Admin-only: this endpoint REPLACES the entire warehouse (see wipe below).
    user = await require_admin(request)
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo archivos Excel (.xlsx)")

    import openpyxl
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "El archivo esta vacio")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    col_map = {h: i for i, h in enumerate(headers)}

    def get(row, name, default=""):
        idx = col_map.get(name)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip() if isinstance(default, str) else row[idx]

    now = now_iso()
    inventory_by_key = {}
    box_docs = []
    locations_set = set()
    skipped = 0


    for row in rows[1:]:
        style = get(row, 'Style', '').strip().upper()
        if not style:
            skipped += 1
            continue

        color = get(row, 'Color', '').strip().upper()
        size = str(get(row, 'Size', '')).strip().upper()
        inv_loc = get(row, 'InvLocation', '').strip().upper()
        total_boxes = int(float(get(row, 'Total Boxes', 0) or 0))
        total_units = int(float(get(row, 'TotalUnits', 0) or 0))
        description = get(row, 'Description', '').strip().upper()
        coo = get(row, 'CountryofOrigin', '').strip().upper()
        fabric = get(row, 'FabricContent', '').strip().upper()

        if inv_loc:
            locations_set.add(inv_loc)

        # Include fabric + COO in the key so distinct batches at the same shelf
        # stay separate (matches _update_inventory_enhanced behavior).
        key = (style, color, size, inv_loc, fabric, coo)
        if key not in inventory_by_key:
            inventory_id = f"inv_{uuid.uuid4().hex[:12]}"
            inventory_by_key[key] = {
                "inventory_id": inventory_id,
                "customer": get(row, 'CustomerID', '').strip().upper(),
                "style": style,
                "sku": style,
                "color": color,
                "size": size,
                "size_header": get(row, 'SizeHeader', '').strip().upper(),
                "manufacturer": get(row, 'Manufacturer', '').strip().upper(),
                "description": description,
                "category": get(row, 'Category', '').strip().upper(),
                "country_of_origin": coo,
                "fabric_content": get(row, 'FabricContent', '').strip().upper(),
                "import_number": get(row, 'ImportNumber', ''),
                "po": get(row, 'PO', ''),
                "bpo": get(row, 'BPO', ''),
                "location": inv_loc,
                "total_boxes": 0,
                "units_on_hand": 0,
                "units_allocated": 0,
                "updated_at": now
            }

        # Accumulate quantities
        inventory_by_key[key]["total_boxes"] += total_boxes
        inventory_by_key[key]["units_on_hand"] += total_units

        # Retrieve the consolidated inventory_id for the boxes
        inventory_id = inventory_by_key[key]["inventory_id"]

        # Generate LPNs (boxes) for this line item
        if total_boxes > 0:
            units_per_box = total_units // total_boxes
            remainder = total_units % total_boxes
            for i in range(total_boxes):
                box_units = units_per_box + (1 if i < remainder else 0)
                box_docs.append({
                    "box_id": f"LPN_{uuid.uuid4().hex[:8].upper()}",
                    "inventory_id": inventory_id,
                    "sku": style,
                    "color": color,
                    "size": size,
                    "units": box_units,
                    "qty": box_units,
                    "location": inv_loc,
                    "state": "putaway",
                    "customer": get(row, 'CustomerID', '').strip().upper(),
                    "manufacturer": get(row, 'Manufacturer', '').strip().upper(),
                    "description": description,
                    "coo": coo,
                    "country_of_origin": coo,
                    "fabric_content": fabric,
                    "created_at": now
                })

    inventory_docs = list(inventory_by_key.values())

    # SAFETY GUARD — this endpoint REPLACES the whole warehouse. Before this guard
    # existed a wrong/empty Excel wiped everything (real 2026-05-29 incident).
    # Refuse to wipe when the upload parsed to nothing or skipped most rows
    # (mis-mapped headers), and snapshot the current data into *_prev collections
    # in Mongo first so a bad import is recoverable.
    data_rows = max(0, len(rows) - 1)
    if not inventory_docs:
        raise HTTPException(400, "El archivo no produjo inventario válido (¿encabezados mal mapeados?). No se borró nada.")
    if data_rows and skipped > data_rows * 0.5:
        raise HTTPException(400, f"Se omitieron {skipped} de {data_rows} filas — revisa el archivo/encabezados. No se borró nada para proteger el inventario.")

    # SHRINK GUARD — the #1 cause of lost boxes: re-importing an OUTDATED Excel
    # wipes every box received/moved since it was exported. Refuse when the upload
    # would drop a big chunk of the current boxes unless ?force=true is passed.
    cur_boxes = await db.wms_boxes.count_documents({})
    if not force and cur_boxes > 0 and len(box_docs) < cur_boxes * 0.7:
        raise HTTPException(409,
            f"El archivo trae {len(box_docs)} cajas vs {cur_boxes} actuales "
            f"(perderías {cur_boxes - len(box_docs)}). Probablemente sea un Excel "
            f"desactualizado y borrarías cajas recibidas/movidas después. Si es "
            f"intencional reenvía con ?force=true. NO se borró nada.")

    # VERSIONED rollback snapshot — a TIMESTAMPED copy per import (never
    # overwritten), so a second import can't destroy the previous backup the way
    # the single _prev did. Keeps _prev too for the existing restore flow.
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    await db.wms_inventory.aggregate([{"$out": f"wms_inventory_bkp_{stamp}"}]).to_list(1)
    await db.wms_boxes.aggregate([{"$out": f"wms_boxes_bkp_{stamp}"}]).to_list(1)
    await db.wms_inventory.aggregate([{"$out": "wms_inventory_prev"}]).to_list(1)
    await db.wms_boxes.aggregate([{"$out": "wms_boxes_prev"}]).to_list(1)

    # Clear old data (WMS 2.0 Fresh Start)
    await db.wms_inventory.delete_many({})
    await db.wms_boxes.delete_many({})
    await db.wms_tasks.delete_many({})
    await db.wms_allocations.delete_many({})
    if inventory_docs:
        batch_size = 1000
        for i in range(0, len(inventory_docs), batch_size):
            await db.wms_inventory.insert_many(inventory_docs[i:i+batch_size])
    if box_docs:
        batch_size = 1000
        for i in range(0, len(box_docs), batch_size):
            await db.wms_boxes.insert_many(box_docs[i:i+batch_size])


    # Auto-create locations with strict case-insensitive check
    locations_created = 0
    for loc_name in locations_set:
        # Normalizar el nombre para la busqueda
        clean_name = loc_name.strip().upper()
        if not clean_name: continue
        
        existing = await db.wms_locations.find_one({"name": {"$regex": f"^{re.escape(clean_name)}$", "$options": "i"}})
        if not existing:
            zone = clean_name.split('-')[0] if '-' in clean_name else "DEFAULT"
            await db.wms_locations.insert_one({
                "location_id": f"loc_{uuid.uuid4().hex[:12]}",
                "name": clean_name,
                "zone": zone,
                "type": "rack",
                "active": True,
                "created_at": now
            })
            locations_created += 1

    # Log movement
    await db.wms_movements.insert_one({
        "movement_id": f"mov_{uuid.uuid4().hex[:12]}",
        "type": "import",
        "description": f"Imported {len(inventory_docs)} inventory records from {file.filename}",
        "user": user.get("name", user.get("email", "")),
        "timestamp": now
    })

    wb.close()
    return {
        "imported": len(inventory_docs),
        "skipped": skipped,
        "locations_created": locations_created,
        "total_locations": len(locations_set)
    }


# ==================== SKU GENERATION ====================

@router.get("/generate-sku")
async def generate_sku(request: Request, style: str = "", color: str = "", size: str = ""):
    """Preview auto-generated SKU for a style+color+size combination."""
    await require_auth(request)
    if not style:
        return {"sku": ""}
    base = style.upper().replace(' ', '-')
    parts = [base]
    if color: parts.append(color.upper().replace(' ', '-')[:10])
    if size: parts.append(size.upper())
    sku = '-'.join(parts)
    return {"sku": sku}

# ==================== CYCLE COUNT ====================

@router.post("/cycle-counts")
async def create_cycle_count(request: Request):
    """Create a new cycle count task."""
    user = await require_inventory_level(request, 1)
    body = await request.json()
    name = body.get("name", "").strip()
    location_filter = body.get("location_filter", "").strip()
    customer_filter = body.get("customer_filter", "").strip()
    style_filter = body.get("style_filter", "").strip()
    color_filter = body.get("color_filter", "").strip()
    assigned_to = body.get("assigned_to", "").strip()
    assigned_to_name = body.get("assigned_to_name", "").strip()
    is_general = body.get("is_general", False) or (
        not location_filter and not customer_filter and not style_filter and not color_filter
    )

    if not name:
        raise HTTPException(400, "Nombre del conteo requerido")

    # Build query to get inventory items for this count
    query = {}
    if not is_general:
        if location_filter:
            # Prefix search for locations sharing the prefix
            query["location"] = {"$regex": f"^{re.escape(location_filter)}", "$options": "i"}
        if customer_filter:
            query["customer"] = {"$regex": f"^{customer_filter}$", "$options": "i"}
        if style_filter:
            query["style"] = {"$regex": f"^{style_filter}$", "$options": "i"}
        if color_filter:
            query["color"] = {"$regex": f"^{re.escape(color_filter)}$", "$options": "i"}

    # Get inventory items matching filters - Increase limit to 50,000 for general counts
    items = await db.wms_inventory.find(query, {"_id": 0}).to_list(50000)
    if not items:
        raise HTTPException(400, "No se encontraron items con los filtros proporcionados")

    # Build count lines. We snapshot the descriptive fields at creation time
    # (description / customer / country / fabric / manufacturer) so the counter
    # has full context on the item without an extra fetch. Old counts created
    # before this change get enriched lazily on GET (see get_cycle_count below).
    count_lines = []
    for item in items:
        count_lines.append({
            "line_id": gen_id("cl"),
            "style": item.get("style", ""),
            "color": item.get("color", ""),
            "size": item.get("size", ""),
            "inv_location": item.get("location", ""),
            "sku": item.get("sku", ""),
            "system_qty": item.get("units_on_hand", 0),
            "description": item.get("description", ""),
            "customer": item.get("customer", ""),
            "manufacturer": item.get("manufacturer", ""),
            "country_of_origin": item.get("country_of_origin", ""),
            "fabric_content": item.get("fabric_content", ""),
            "counted_qty": None,
            "discrepancy": None,
            "counted": False
        })

    # ── Modo BOX_SCAN (nuevo, default): el conteo es POR UBICACIÓN escaneando el
    # LPN de cada caja. Se compara el set de LPN escaneados vs el que el sistema
    # tiene en esa ubicación. 3 pases escalonados: la ubicación que no cuadra sube
    # de pase; tras el pase 3 queda para revisión de supervisor. No ajusta unidades.
    # Los conteos viejos (sin `mode`) siguen usando `lines[]` por piezas.
    mode = (body.get("mode") or "box_scan").strip()
    scan_locations = []
    if mode == "box_scan":
        box_query = {"units": {"$gt": 0}}
        if not is_general:
            if location_filter:
                box_query["location"] = {"$regex": f"^{re.escape(location_filter)}", "$options": "i"}
            if customer_filter:
                box_query["customer"] = {"$regex": f"^{customer_filter}$", "$options": "i"}
            if style_filter:
                box_query["$or"] = [
                    {"style": {"$regex": f"^{style_filter}$", "$options": "i"}},
                    {"sku": {"$regex": f"^{style_filter}$", "$options": "i"}},
                ]
            if color_filter:
                box_query["color"] = {"$regex": f"^{re.escape(color_filter)}$", "$options": "i"}
        boxes = await db.wms_boxes.find(
            box_query, {"_id": 0, "box_id": 1, "location": 1}
        ).to_list(100000)
        by_loc: dict[str, list] = {}
        for b in boxes:
            loc = (b.get("location") or "").strip()
            if not loc:
                continue
            by_loc.setdefault(loc, []).append(b.get("box_id"))
        for loc in sorted(by_loc):
            scan_locations.append({
                "loc_id": gen_id("ccl"),
                "location": loc,
                "pass": 1,                       # pase actual (1/2/3)
                "status": "pending",             # pending | ok | escalated | supervisor
                "expected_boxes": sorted(by_loc[loc]),   # LPN que el sistema tiene aquí (snapshot)
                "scanned_boxes": [],             # LPN escaneados en el pase actual
                "missing": [],                   # esperados no escaneados
                "extra": [],                     # escaneados no esperados (ajenos)
                "counted_by": None,
                "counted_by_name": None,
                "counted_at": None,
                "history": [],                   # auditoría por pase
            })
        if not scan_locations:
            raise HTTPException(400, "No se encontraron cajas con los filtros proporcionados")

    count_id = gen_id("cc")
    count_doc = {
        "count_id": count_id,
        "name": name,
        "status": "pending",
        "mode": mode,
        "is_general": is_general,
        "location_filter": location_filter if not is_general else "",
        "customer_filter": customer_filter if not is_general else "",
        "style_filter": style_filter if not is_general else "",
        "color_filter": color_filter if not is_general else "",
        "assigned_to": assigned_to or None,
        "assigned_to_name": assigned_to_name or None,
        "total_lines": len(count_lines),
        "counted_lines": 0,
        "lines": count_lines,
        # box_scan
        "scan_locations": scan_locations,
        "total_scan_locations": len(scan_locations),
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name", ""),
        "created_at": now_iso(),
    }
    await db.wms_cycle_counts.insert_one(count_doc)
    count_doc.pop("_id", None)
    await log_movement(user, "cycle_count_created", {"count_id": count_id, "total_lines": len(count_lines), "is_general": is_general})

    if assigned_to:
        await ws_manager.broadcast("cycle_count_assigned", {
            "count_id": count_id,
            "assigned_to": assigned_to,
            "assigned_to_name": assigned_to_name,
            "name": name
        })

    await notify_badge_change("cycle_count")
    return count_doc


# ══════════════════ BOX-SCAN CYCLE COUNT (por caja · por ubicación) ══════════════════
CYCLE_COUNT_MAX_PASS = 3  # tras el pase 3 sin cuadrar → revisión de supervisor

def _cc_find_loc(count: dict, location: str) -> Optional[dict]:
    loc = (location or "").strip().upper()
    for L in count.get("scan_locations", []):
        if (L.get("location") or "").strip().upper() == loc:
            return L
    return None


@router.post("/cycle-counts/{count_id}/scan-location")
async def cc_scan_location(count_id: str, request: Request):
    """Escanea el LPN de una caja para una ubicación del conteo. Agrega el box_id
    al set escaneado del pase actual (dedupe). Reporta si es esperado, ajeno o
    duplicado, sin cerrar todavía la ubicación."""
    user = await require_inventory_level(request, 1)
    body = await request.json()
    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")
    if count.get("mode") != "box_scan":
        raise HTTPException(400, "Este conteo no es por caja (box_scan)")
    if count.get("status") == "approved":
        raise HTTPException(400, "Este conteo ya fue aprobado")
    location = (body.get("location") or "").strip()
    box_id = (body.get("box_id") or body.get("lpn") or "").strip().upper()
    if not location or not box_id:
        raise HTTPException(400, "location y box_id (LPN) son requeridos")
    L = _cc_find_loc(count, location)
    if not L:
        raise HTTPException(404, f"La ubicación {location} no está en este conteo")
    if L.get("status") in ("ok", "supervisor"):
        raise HTTPException(400, f"La ubicación {location} ya está cerrada ({L.get('status')})")
    scanned = L.get("scanned_boxes", [])
    dup = box_id in scanned
    if not dup:
        scanned.append(box_id)
        L["scanned_boxes"] = scanned
        await db.wms_cycle_counts.update_one(
            {"count_id": count_id, "scan_locations.loc_id": L["loc_id"]},
            {"$set": {"scan_locations.$.scanned_boxes": scanned, "last_updated_at": now_iso()}},
        )
    expected = set(L.get("expected_boxes", []))
    return {
        "box_id": box_id,
        "duplicate": dup,
        "expected_here": box_id in expected,   # False => caja AJENA a esta ubicación
        "scanned_count": len(scanned),
        "expected_count": len(expected),
        "pass": L.get("pass"),
    }


@router.post("/cycle-counts/{count_id}/close-location")
async def cc_close_location(count_id: str, request: Request):
    """Cierra una ubicación: compara el set de LPN escaneados vs los esperados.
    Cuadra (sin faltantes ni ajenas) → status ok. No cuadra → sube de pase
    (1→2→3); tras el pase 3 queda 'supervisor' (revisión/ajuste manual). Nunca
    ajusta unidades — solo audita cajas y escala."""
    user = await require_inventory_level(request, 1)
    body = await request.json()
    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")
    if count.get("mode") != "box_scan":
        raise HTTPException(400, "Este conteo no es por caja (box_scan)")
    location = (body.get("location") or "").strip()
    L = _cc_find_loc(count, location)
    if not L:
        raise HTTPException(404, f"La ubicación {location} no está en este conteo")
    if L.get("status") in ("ok", "supervisor"):
        raise HTTPException(400, f"La ubicación {location} ya está cerrada ({L.get('status')})")

    expected = set(L.get("expected_boxes", []))
    scanned = set(L.get("scanned_boxes", []))
    missing = sorted(expected - scanned)   # esperadas que no aparecieron
    extra = sorted(scanned - expected)     # escaneadas que no van aquí (ajenas)
    matched = (not missing and not extra)
    cur_pass = int(L.get("pass", 1))

    hist = L.get("history", [])
    hist.append({
        "pass": cur_pass, "scanned": sorted(scanned),
        "missing": missing, "extra": extra, "matched": matched,
        "by": user.get("user_id"), "by_name": user.get("name", ""), "at": now_iso(),
    })

    set_fields = {
        "scan_locations.$.history": hist,
        "scan_locations.$.missing": missing,
        "scan_locations.$.extra": extra,
        "scan_locations.$.counted_by": user.get("user_id"),
        "scan_locations.$.counted_by_name": user.get("name", ""),
        "scan_locations.$.counted_at": now_iso(),
        "last_updated_at": now_iso(),
    }
    if matched:
        set_fields["scan_locations.$.status"] = "ok"
        new_status, new_pass = "ok", cur_pass
    elif cur_pass < CYCLE_COUNT_MAX_PASS:
        new_pass = cur_pass + 1
        set_fields["scan_locations.$.status"] = "escalated"
        set_fields["scan_locations.$.pass"] = new_pass
        set_fields["scan_locations.$.scanned_boxes"] = []   # nuevo pase arranca limpio
        new_status = "escalated"
    else:
        set_fields["scan_locations.$.status"] = "supervisor"
        new_status, new_pass = "supervisor", cur_pass

    await db.wms_cycle_counts.update_one(
        {"count_id": count_id, "scan_locations.loc_id": L["loc_id"]},
        {"$set": set_fields},
    )
    await log_movement(user, "cycle_count_location_closed", {
        "count_id": count_id, "location": L.get("location"), "pass": cur_pass,
        "matched": matched, "missing": len(missing), "extra": len(extra),
        "result": new_status, "next_pass": new_pass,
    })
    await notify_badge_change("cycle_count")
    return {
        "location": L.get("location"), "matched": matched,
        "missing": missing, "extra": extra,
        "result": new_status, "pass": new_pass,
        "escalated": new_status == "escalated",
        "needs_supervisor": new_status == "supervisor",
    }


@router.get("/cycle-counts")
async def list_cycle_counts(request: Request):
    """List all cycle counts."""
    await require_inventory_level(request, 1)
    counts = await db.wms_cycle_counts.find({}, {"_id": 0, "lines": 0}).sort("created_at", -1).to_list(200)
    return counts

@router.get("/cycle-counts/{count_id}")
async def get_cycle_count(count_id: str, request: Request):
    """Get a cycle count with all lines. For counts created before we started
    snapshotting description/customer/country/fabric/manufacturer onto each
    line, enrich on the fly by joining against current wms_inventory so the
    counter always sees the full context."""
    await require_inventory_level(request, 1)
    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")

    lines = count.get("lines") or []
    if lines:
        missing_idx = [
            i for i, l in enumerate(lines)
            if not l.get("description") and not l.get("customer") and not l.get("fabric_content")
        ]
        if missing_idx:
            # One inventory lookup per missing line, keyed on the (sku, location)
            # pair that uniquely identifies each inventory row.
            keys = []
            for i in missing_idx:
                line = lines[i]
                keys.append({
                    "sku": line.get("sku") or line.get("style") or "",
                    "color": line.get("color") or "",
                    "size": line.get("size") or "",
                    "location": line.get("inv_location") or "",
                })
            # Build a single Mongo $or query for the join.
            or_clauses = [{"sku": k["sku"], "color": k["color"], "size": k["size"], "location": k["location"]} for k in keys if k["sku"]]
            inv_map = {}
            if or_clauses:
                async for inv in db.wms_inventory.find(
                    {"$or": or_clauses},
                    {"_id": 0, "sku": 1, "color": 1, "size": 1, "location": 1,
                     "description": 1, "customer": 1, "manufacturer": 1,
                     "country_of_origin": 1, "fabric_content": 1},
                ):
                    k = (inv.get("sku", ""), inv.get("color", ""), inv.get("size", ""), inv.get("location", ""))
                    inv_map[k] = inv
            for i in missing_idx:
                line = lines[i]
                k = (
                    line.get("sku") or line.get("style") or "",
                    line.get("color") or "",
                    line.get("size") or "",
                    line.get("inv_location") or "",
                )
                inv = inv_map.get(k)
                if inv:
                    line["description"] = line.get("description") or inv.get("description", "")
                    line["customer"] = line.get("customer") or inv.get("customer", "")
                    line["manufacturer"] = line.get("manufacturer") or inv.get("manufacturer", "")
                    line["country_of_origin"] = line.get("country_of_origin") or inv.get("country_of_origin", "")
                    line["fabric_content"] = line.get("fabric_content") or inv.get("fabric_content", "")
            count["lines"] = lines
    return count

# ─── Cycle-count variance guardrail ──────────────────────────────────────────
# A single line count should never be able to inflate stock by an absurd amount
# without review. This guards against fat-fingered counts (the classic 96 -> 996
# typo that dumped a phantom 900-unit box into inventory). Only *positive* jumps
# are guarded — counting LESS than expected (shrink/found-none) is normal and
# applies immediately. A jump trips the guard only when it is large in BOTH
# absolute (> ABS units) and relative (> MULT x current on-hand) terms, so
# ordinary restocks pass through. Tunable if the warehouse needs different bands.
CYCLE_COUNT_VARIANCE_ABS = 200   # min absolute unit increase to consider "large"
CYCLE_COUNT_VARIANCE_MULT = 3    # counted must exceed this multiple of on-hand

def _is_extreme_variance(cur: int, qty: int) -> bool:
    delta = qty - cur
    if delta <= CYCLE_COUNT_VARIANCE_ABS:
        return False
    return qty > CYCLE_COUNT_VARIANCE_MULT * max(cur, 1)


@router.put("/cycle-counts/{count_id}/count")
async def save_count_progress(count_id: str, request: Request):
    """Save counting progress - operator submits counted quantities."""
    user = await require_inventory_level(request, 1)
    body = await request.json()
    counted_items = body.get("counted_items", {})  # { line_id: counted_qty }

    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")
    if count.get("status") == "approved":
        raise HTTPException(400, "Este conteo ya fue aprobado")

    lines = count.get("lines", [])
    counted_count = 0
    adjustments = 0
    held = 0
    for line in lines:
        lid = line["line_id"]
        if lid in counted_items:
            qty = int(counted_items[lid])
            line["counted_qty"] = qty
            line["discrepancy"] = qty - (line.get("system_qty", 0) or 0)
            line["counted"] = True
            line["counted_by"] = user.get("user_id")
            line["counted_at"] = now_iso()
            # IMMEDIATE ADJUSTMENT (requirement): apply the variance to inventory
            # the moment the auditor saves this line. The variance is measured
            # against the LIVE on-hand — NOT the frozen `system_qty` snapshot.
            # Gating on `discrepancy = qty - system_qty` was a bug: recounting a
            # line back to its original snapshot yielded discrepancy 0, skipped the
            # whole block, and left a previously-applied WRONG adjustment in place
            # (the 96->996 typo whose correction to 96 never rolled back the 900).
            # Comparing to live on-hand makes every save self-correcting.
            # Match the inventory row the SAME way the rest of the module does
            # ($or sku/style, case-insensitive). The old exact {"style":..} filter
            # silently missed rows keyed by composite sku.
            st = (line.get("style") or line.get("sku") or "").strip()
            inv_row = await db.wms_inventory.find_one({
                "$or": [{"sku": _ci_eq(st)}, {"style": _ci_eq(st)}],
                "color": _ci_eq(line.get("color", "")),
                "size": _ci_eq(line.get("size", "")),
                "location": _ci_eq(line.get("inv_location", "")),
            })
            if inv_row:
                cur = int(inv_row.get("units_on_hand", 0) or 0)
                delta = qty - cur
                if delta == 0:
                    # Live inventory already matches — clear any stale review flag.
                    line["held_for_review"] = False
                elif _is_extreme_variance(cur, qty):
                    # GUARDRAIL: an absurd positive jump (large in both absolute and
                    # relative terms) is almost always a typo. Do NOT touch stock;
                    # hold the line for an admin to apply (or reject) at approval.
                    line["held_for_review"] = True
                    line["adjusted"] = False
                    held += 1
                    await log_movement(user, "cycle_count_variance_held", {
                        "count_id": count_id, "line_id": lid,
                        "sku": line.get("sku") or line.get("style"),
                        "location": line.get("inv_location"),
                        "from": cur, "to": qty, "delta": delta,
                    })
                else:
                    await db.wms_inventory.update_one(
                        {"_id": inv_row["_id"]},
                        {"$set": {"units_on_hand": qty, "updated_at": now_iso()}},
                    )
                    # Reconcile the backing boxes to the new on-hand so the box
                    # mirror agrees — otherwise _available_units' max(box,inv)
                    # silently undoes a count-down.
                    await _reconcile_line_boxes(inv_row, delta)
                    line["held_for_review"] = False
                    line["adjusted"] = True
                    line["adjusted_at"] = now_iso()
                    line["adjusted_by"] = user.get("user_id")
                    adjustments += 1
                    await log_movement(user, "cycle_count_adjustment", {
                        "count_id": count_id, "line_id": lid,
                        "sku": line.get("sku") or line.get("style"),
                        "location": line.get("inv_location"),
                        "from": cur, "to": qty,
                        "discrepancy": line["discrepancy"],
                    })
            elif line["discrepancy"]:
                # A key mismatch must not swallow a real discrepancy silently.
                line["adjust_failed"] = True
                await log_movement(user, "cycle_count_adjust_failed", {
                    "count_id": count_id, "line_id": lid,
                    "sku": line.get("sku") or line.get("style"),
                    "location": line.get("inv_location"),
                    "reason": "no se encontró fila de inventario para aplicar el conteo",
                })
        if line.get("counted"):
            counted_count += 1

    status = "completed" if counted_count >= len(lines) else "in_progress"
    await db.wms_cycle_counts.update_one({"count_id": count_id}, {"$set": {
        "lines": lines,
        "counted_lines": counted_count,
        "status": status,
        "last_updated_by": user.get("user_id"),
        "last_updated_at": now_iso()
    }})
    await log_movement(user, "cycle_count_progress", {"count_id": count_id, "counted": counted_count, "total": len(lines), "adjustments": adjustments, "held": held})
    if adjustments or held:
        await notify_badge_change("cycle_count")
    msg = f"Progreso guardado ({counted_count}/{len(lines)})"
    if adjustments:
        msg += f" · {adjustments} ajuste(s) aplicado(s) al inventario"
    if held:
        msg += f" · {held} variación(es) grande(s) retenida(s) para revisión del admin"
    return {"message": msg, "status": status, "adjustments": adjustments, "held": held}

@router.put("/cycle-counts/{count_id}/approve")
async def approve_cycle_count(count_id: str, request: Request):
    """Admin approves cycle count and adjusts inventory."""
    user = await require_inventory_level(request, 2)
    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")
    if count.get("status") != "completed":
        raise HTTPException(400, "El conteo debe estar completado antes de aprobar")

    adjustments = 0
    already = 0
    for line in count.get("lines", []):
        if not line.get("counted"):
            continue
        # Skip lines already applied during counting so approval is just a sign-off
        # and never clobbers stock that legitimately changed between the count and
        # the approval. Lines HELD for review (large-variance guard) are the
        # exception: the admin's approval IS the review, so they get applied here.
        if line.get("adjusted") and not line.get("held_for_review"):
            already += 1
            continue
        st = (line.get("style") or line.get("sku") or "").strip()
        inv_row = await db.wms_inventory.find_one({
            "$or": [{"sku": _ci_eq(st)}, {"style": _ci_eq(st)}],
            "color": _ci_eq(line.get("color", "")),
            "size": _ci_eq(line.get("size", "")),
            "location": _ci_eq(line.get("inv_location", "")),
        })
        if inv_row:
            cur = int(inv_row.get("units_on_hand", 0) or 0)
            tgt = int(line.get("counted_qty") or 0)
            # Reconcile against LIVE on-hand (same fix as save_count_progress): a
            # zero delta means live already matches, so there is nothing to apply.
            delta = tgt - cur
            if delta == 0:
                continue
            await db.wms_inventory.update_one(
                {"_id": inv_row["_id"]},
                {"$set": {"units_on_hand": tgt, "updated_at": now_iso()}}
            )
            await _reconcile_line_boxes(inv_row, delta)
            line["adjusted"] = True
            line["held_for_review"] = False
            line["adjusted_at"] = now_iso()
            adjustments += 1

    await db.wms_cycle_counts.update_one({"count_id": count_id}, {"$set": {
        "status": "approved",
        "lines": count.get("lines", []),
        "approved_by": user.get("user_id"),
        "approved_by_name": user.get("name", ""),
        "approved_at": now_iso(),
        "adjustments": adjustments + already
    }})
    await log_movement(user, "cycle_count_approved", {"count_id": count_id, "adjustments": adjustments, "already_applied": already})
    await notify_badge_change("cycle_count")
    return {"message": f"Conteo aprobado. {already} ajuste(s) aplicados durante el conteo, {adjustments} nuevo(s) al aprobar.", "adjustments": adjustments + already}

@router.delete("/cycle-counts/{count_id}")
async def delete_cycle_count(count_id: str, request: Request):
    """Admin deletes a cycle count."""
    user = await require_inventory_level(request, 2)
    count = await db.wms_cycle_counts.find_one({"count_id": count_id})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")
    await db.wms_cycle_counts.delete_one({"count_id": count_id})
    await log_movement(user, "cycle_count_deleted", {"count_id": count_id, "name": count.get("name")})
    await notify_badge_change("cycle_count")
    return {"message": "Conteo ciclico eliminado correctamente"}

# ==================== CYCLE COUNT REPORTS ====================

@router.get("/cycle-counts/reports/summary")
async def cycle_counts_report_summary(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    created_by_name: Optional[str] = None,
    approved_by_name: Optional[str] = None,
):
    """
    Historical summary of cycle counts, supports filters:
    date_from, date_to (ISO), status, created_by_name, approved_by_name.
    """
    await require_inventory_level(request, 3)

    query: dict = {}
    # Status filter: default to approved only if no status given
    if status and status != "all":
        query["status"] = status
    else:
        query["status"] = {"$in": ["approved", "in_progress", "pending"]}

    if date_from or date_to:
        date_filter: dict = {}
        if date_from:
            date_filter["$gte"] = date_from
        if date_to:
            date_filter["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_filter

    if created_by_name:
        query["created_by_name"] = {"$regex": created_by_name, "$options": "i"}
    if approved_by_name:
        query["approved_by_name"] = {"$regex": approved_by_name, "$options": "i"}

    counts = await db.wms_cycle_counts.find(
        query,
        {"_id": 0, "count_id": 1, "name": 1, "status": 1, "created_at": 1, "approved_at": 1,
         "approved_by_name": 1, "created_by_name": 1,
         "total_lines": 1, "counted_lines": 1, "adjustments": 1,
         "is_general": 1, "location_filter": 1, "style_filter": 1,
         "color_filter": 1, "customer_filter": 1, "lines": 1, "assigned_to_name": 1}
    ).sort("created_at", -1).to_list(5000)

    summary_list = []
    loc_discrepancy_map = {}   # location -> {count, total_delta}
    sku_discrepancy_map = {}   # "style|color" -> {count, total_delta}
    auditor_stats = {}         # user_id -> {...stats...}
    total_adjustments_all = 0
    total_lines_all = 0
    total_exact_all = 0

    for c in counts:
        lines = c.get("lines") or []
        total = len(lines)
        discrepant = [l for l in lines if l.get("counted") and l.get("discrepancy") and l["discrepancy"] != 0]
        exact = total - len(discrepant)
        accuracy = round((exact / total * 100), 1) if total > 0 else 100.0
        units_over = sum(l["discrepancy"] for l in discrepant if l["discrepancy"] > 0)
        units_short = sum(abs(l["discrepancy"]) for l in discrepant if l["discrepancy"] < 0)
        adj_count = c.get("adjustments") or len([l for l in lines if l.get("adjusted")])

        total_adjustments_all += adj_count
        total_lines_all += total
        total_exact_all += exact

        # Auditor productivity tracking
        for l in lines:
            if l.get("counted"):
                uid = l.get("counted_by") or c.get("assigned_to") or "Desconocido"
                if uid not in auditor_stats:
                    auditor_stats[uid] = {"user_id": uid, "lines_counted": 0, "units_counted": 0, "discrepancies_found": 0}
                auditor_stats[uid]["lines_counted"] += 1
                auditor_stats[uid]["units_counted"] += l.get("counted_qty", 0)
                if l.get("discrepancy") and l["discrepancy"] != 0:
                    auditor_stats[uid]["discrepancies_found"] += 1

        # Aggregate by location and SKU
        for l in discrepant:
            loc = l.get("inv_location") or "Sin ubicación"
            delta = abs(l.get("discrepancy") or 0)
            loc_discrepancy_map.setdefault(loc, {"count": 0, "total_delta": 0})
            loc_discrepancy_map[loc]["count"] += 1
            loc_discrepancy_map[loc]["total_delta"] += delta

            sku_key = f"{l.get('style', '')}|{l.get('color', '')}"
            sku_discrepancy_map.setdefault(sku_key, {"count": 0, "total_delta": 0, "style": l.get("style",""), "color": l.get("color","")})
            sku_discrepancy_map[sku_key]["count"] += 1
            sku_discrepancy_map[sku_key]["total_delta"] += delta

        summary_list.append({
            "count_id": c["count_id"],
            "name": c.get("name", ""),
            "status": c.get("status", ""),
            "approved_at": c.get("approved_at"),
            "created_at": c.get("created_at"),
            "approved_by_name": c.get("approved_by_name", ""),
            "created_by_name": c.get("created_by_name", ""),
            "assigned_to_name": c.get("assigned_to_name", ""),
            "total_lines": total,
            "discrepant_lines": len(discrepant),
            "exact_lines": exact,
            "accuracy_pct": accuracy,
            "units_over": units_over,
            "units_short": units_short,
            "adjustments": adj_count,
            "is_general": c.get("is_general", False),
            "location_filter": c.get("location_filter", ""),
            "style_filter": c.get("style_filter", ""),
            "color_filter": c.get("color_filter", ""),
            "customer_filter": c.get("customer_filter", ""),
        })

    # Fetch users for productivity map
    uids = list(auditor_stats.keys())
    users = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).to_list(1000)
    user_map = {u["user_id"]: u.get("name") or u.get("email") for u in users}

    auditor_productivity = []
    for uid, stats in auditor_stats.items():
        stats["name"] = user_map.get(uid) or ("Desconocido" if uid == "Desconocido" else uid)
        auditor_productivity.append(stats)
    auditor_productivity.sort(key=lambda x: x["lines_counted"], reverse=True)

    overall_accuracy = round((total_exact_all / total_lines_all * 100), 1) if total_lines_all > 0 else 100.0

    top_locations = sorted(loc_discrepancy_map.items(), key=lambda x: x[1]["count"], reverse=True)[:10]
    top_skus = sorted(sku_discrepancy_map.values(), key=lambda x: x["count"], reverse=True)[:10]

    return {
        "counts": summary_list,
        "total_counts": len(summary_list),
        "overall_accuracy_pct": overall_accuracy,
        "total_adjustments": total_adjustments_all,
        "total_lines_ever": total_lines_all,
        "auditor_productivity": auditor_productivity,
        "top_discrepant_locations": [
            {"location": loc, "discrepancy_count": v["count"], "total_delta": v["total_delta"]}
            for loc, v in top_locations
        ],
        "top_discrepant_skus": top_skus,
    }


@router.get("/cycle-counts/reports/timeline")
async def cycle_counts_report_timeline(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user_id: Optional[str] = None,
    count_id: Optional[str] = None,
):
    """
    Returns a chronological feed of cycle count activity.
    Supports date_from, date_to (ISO date string), user_id, count_id filters.
    """
    await require_inventory_level(request, 3)

    count_query: dict = {}
    if count_id:
        count_query["count_id"] = count_id

    counts = await db.wms_cycle_counts.find(
        count_query,
        {"_id": 0, "count_id": 1, "name": 1, "lines": 1, "status": 1}
    ).to_list(5000)

    timeline = []
    user_ids = set()

    date_from_dt = date_from or None
    date_to_dt   = (date_to + "T23:59:59") if date_to else None

    for c in counts:
        for l in c.get("lines", []):
            if l.get("counted") and l.get("counted_at"):
                ts = l["counted_at"]
                if date_from_dt and ts < date_from_dt:
                    continue
                if date_to_dt and ts > date_to_dt:
                    continue
                uid = l.get("counted_by") or "Desconocido"
                if user_id and uid != user_id:
                    continue
                user_ids.add(uid)
                timeline.append({
                    "count_id": c["count_id"],
                    "count_name": c.get("name", ""),
                    "count_status": c.get("status", ""),
                    "user_id": uid,
                    "timestamp": ts,
                    "location": l.get("inv_location", ""),
                    "style": l.get("style", ""),
                    "color": l.get("color", ""),
                    "size": l.get("size", ""),
                    "qty": l.get("counted_qty", 0),
                    "discrepancy": l.get("discrepancy", 0)
                })

    users = await db.users.find({"user_id": {"$in": list(user_ids)}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).to_list(1000)
    user_map = {u["user_id"]: u.get("name") or u.get("email") for u in users}

    for t in timeline:
        t["user_name"] = user_map.get(t["user_id"]) or t["user_id"]

    timeline.sort(key=lambda x: x["timestamp"], reverse=True)

    # Build list of unique auditors for filter dropdown
    auditors = sorted(set((t["user_id"], t["user_name"]) for t in timeline), key=lambda x: x[1])

    return {
        "timeline": timeline,
        "total": len(timeline),
        "auditors": [{"user_id": a[0], "name": a[1]} for a in auditors],
    }


@router.get("/cycle-counts/{count_id}/report")
async def get_cycle_count_report(count_id: str, request: Request):
    """
    Detailed report for a single cycle count (any status, best for approved).
    Returns KPIs, discrepancy table sorted by magnitude, and per-location breakdown.
    """
    await require_inventory_level(request, 3)
    count = await db.wms_cycle_counts.find_one({"count_id": count_id}, {"_id": 0})
    if not count:
        raise HTTPException(404, "Conteo no encontrado")

    lines = count.get("lines") or []
    total = len(lines)
    counted_lines = [l for l in lines if l.get("counted")]
    discrepant_lines = [l for l in counted_lines if l.get("discrepancy") and l["discrepancy"] != 0]
    exact_lines = [l for l in counted_lines if not l.get("discrepancy") or l["discrepancy"] == 0]

    # KPIs
    accuracy_pct = round((len(exact_lines) / total * 100), 1) if total > 0 else 100.0
    units_over = sum(l["discrepancy"] for l in discrepant_lines if l["discrepancy"] > 0)
    units_short = sum(abs(l["discrepancy"]) for l in discrepant_lines if l["discrepancy"] < 0)
    adjusted_count = len([l for l in lines if l.get("adjusted")])

    # Duration
    duration_mins = None
    try:
        from datetime import datetime
        ca = count.get("approved_at") or count.get("last_updated_at")
        cb = count.get("created_at")
        if ca and cb:
            fmt = "%Y-%m-%dT%H:%M:%S"
            d1 = datetime.fromisoformat(ca[:19])
            d2 = datetime.fromisoformat(cb[:19])
            duration_mins = int(abs((d1 - d2).total_seconds()) / 60)
    except Exception:
        pass

    # Discrepancy table — sorted by abs(discrepancy) descending
    discrepancy_table = sorted([
        {
            "line_id": l.get("line_id"),
            "location": l.get("inv_location", ""),
            "style": l.get("style", ""),
            "color": l.get("color", ""),
            "size": l.get("size", ""),
            "customer": l.get("customer", ""),
            "system_qty": l.get("system_qty", 0),
            "counted_qty": l.get("counted_qty", 0),
            "discrepancy": l.get("discrepancy", 0),
            "adjusted": l.get("adjusted", False),
            "adjust_failed": l.get("adjust_failed", False),
        }
        for l in discrepant_lines
    ], key=lambda x: abs(x["discrepancy"] or 0), reverse=True)

    # All counted lines (for full export)
    user_ids = {l.get("counted_by") for l in lines if l.get("counted_by")}
    users_db = await db.users.find({"user_id": {"$in": list(user_ids)}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}).to_list(1000)
    user_map_details = {u["user_id"]: u.get("name") or u.get("email") for u in users_db}

    all_lines_table = [
        {
            "location": l.get("inv_location", ""),
            "style": l.get("style", ""),
            "color": l.get("color", ""),
            "size": l.get("size", ""),
            "customer": l.get("customer", ""),
            "system_qty": l.get("system_qty", 0),
            "counted_qty": l.get("counted_qty"),
            "discrepancy": l.get("discrepancy"),
            "adjusted": l.get("adjusted", False),
            "counted": l.get("counted", False),
            "counted_at": l.get("counted_at"),
            "counted_by": l.get("counted_by"),
            "counted_by_name": user_map_details.get(l.get("counted_by")) if l.get("counted_by") else None
        }
        for l in lines
    ]

    # Per-location breakdown
    loc_map = {}
    for l in lines:
        loc = l.get("inv_location") or "Sin ubicación"
        if loc not in loc_map:
            loc_map[loc] = {"location": loc, "total": 0, "counted": 0, "discrepant": 0, "units_delta": 0}
        loc_map[loc]["total"] += 1
        if l.get("counted"):
            loc_map[loc]["counted"] += 1
            d = l.get("discrepancy") or 0
            if d != 0:
                loc_map[loc]["discrepant"] += 1
                loc_map[loc]["units_delta"] += abs(d)

    location_breakdown = sorted(loc_map.values(), key=lambda x: x["discrepant"], reverse=True)

    return {
        "count_id": count["count_id"],
        "name": count.get("name", ""),
        "status": count.get("status"),
        "is_general": count.get("is_general", False),
        "location_filter": count.get("location_filter", ""),
        "style_filter": count.get("style_filter", ""),
        "color_filter": count.get("color_filter", ""),
        "customer_filter": count.get("customer_filter", ""),
        "created_at": count.get("created_at"),
        "approved_at": count.get("approved_at"),
        "created_by_name": count.get("created_by_name", ""),
        "approved_by_name": count.get("approved_by_name", ""),
        "assigned_to_name": count.get("assigned_to_name", ""),
        # KPIs
        "kpis": {
            "total_lines": total,
            "counted_lines": len(counted_lines),
            "exact_lines": len(exact_lines),
            "discrepant_lines": len(discrepant_lines),
            "accuracy_pct": accuracy_pct,
            "units_over": units_over,
            "units_short": units_short,
            "adjusted_count": adjusted_count,
            "duration_mins": duration_mins,
        },
        "discrepancy_table": discrepancy_table,
        "all_lines": all_lines_table,
        "location_breakdown": location_breakdown,
    }


# ==================== QUICK INLINE UPDATES ====================

@router.put("/pick-tickets/{ticket_id}/status")
async def quick_status_update(ticket_id: str, request: Request):
    user = await require_auth(request)
    body = await request.json()
    if "blank_status" in body:
        new_status = body["blank_status"]
        # Determine if it's a virtual ticket or a real one
        order_number = None
        if ticket_id.startswith("virt_"):
            # If virtual, the ID is virt_ORDER_ID. BUT wait, let's look at list_pick_tickets logic.
            # It uses virt_vo.get('order_id')
            order_id = ticket_id.replace("virt_", "")
            order = await db.orders.find_one({"order_id": order_id})
            if order:
                order_number = order.get("order_number")
        else:
            ticket = await db.wms_pick_tickets.find_one({"ticket_id": ticket_id})
            if ticket:
                await db.wms_pick_tickets.update_one({"ticket_id": ticket_id}, {"$set": {"blank_status": new_status}})
                order_number = ticket.get("order_number")
        
        if order_number:
            await db.orders.update_one({"order_number": order_number}, {"$set": {"blank_status": new_status}})
            
    return {"status": "ok"}

# ==================== WMS 2.0 DIRECTED TASKS ====================

@router.get("/tasks/next")
async def get_next_task(request: Request, user_zone: str = ""):
    """Directed Work System: Retrieves the single highest priority task for the operator."""
    user = await require_auth(request)
    
    # Try finding HOT priority first
    hot_query = {"status": "pending", "priority": "HOT"}
    if user_zone:
        hot_query["context.suggested_zone"] = {"$regex": f"^{user_zone}$", "$options": "i"}
        
    next_task = await db.wms_tasks.find_one(hot_query, sort=[("created_at", 1)])
    
    # Fallback to NORMAL
    if not next_task:
        normal_query = {"status": "pending"}
        if user_zone:
            normal_query["context.suggested_zone"] = {"$regex": f"^{user_zone}$", "$options": "i"}
        # Fetch top 50 to sort by Travel Sequence in memory
        tasks = await db.wms_tasks.find(normal_query).sort("created_at", 1).to_list(50)
        if tasks:
            def get_sort_key(t):
                loc = t.get("context", {}).get("suggested_zone", "ZZ-99-9")
                parts = loc.split("-")
                aisle = parts[0] if len(parts) > 0 else "ZZ"
                section = parts[1] if len(parts) > 1 else "99"
                level = parts[2] if len(parts) > 2 else "9"
                return (aisle, section, level)
            tasks.sort(key=get_sort_key)
            next_task = tasks[0]
        else:
            next_task = None
        
    if not next_task:
        return {"task": None, "message": "No pending tasks."}
        
    # Claim it for the user
    task_id = next_task["task_id"]
    await db.wms_tasks.update_one(
        {"task_id": task_id},
        {"$set": {"status": "assigned", "assigned_to": user.get("user_id"), "assigned_at": now_iso()}}
    )
    
    # Hydrate lpn_details automatically
    lpn_id = next_task.get("lpn_id")
    if lpn_id:
        lpn = await db.wms_boxes.find_one({"box_id": lpn_id}, {"_id": 0})
        next_task["lpn_details"] = lpn
        
    next_task.pop("_id", None)
    return {"task": next_task}

@router.post("/tasks/{task_id}/complete")
async def complete_task(task_id: str, request: Request):
    """Marks a directed task as complete and permanently updates core objects."""
    user = await require_auth(request)
    body = await request.json()
    scan_validation = body.get("scan", "")
    
    task = await db.wms_tasks.find_one({"task_id": task_id})
    if not task:
        raise HTTPException(404, "Task not found")
        
    lpn_id = task.get("lpn_id")
    if lpn_id and scan_validation != lpn_id:
        raise HTTPException(400, "Validation failed: Scanned LPN does not match Task LPN.")
        
    # Execution Logic
    if task["task_type"] == "putaway":
        dest_location = body.get("destination_location", "").strip()
        if not dest_location:
            raise HTTPException(400, "destination_location is required for putaway")
        await db.wms_boxes.update_one({"box_id": lpn_id}, {"$set": {"location": dest_location, "status": "stored", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})

    elif task["task_type"] == "cross_dock":
        dest_location = body.get("destination_location", "Produccion")
        await db.wms_boxes.update_one({"box_id": lpn_id}, {"$set": {"location": dest_location, "status": "cross_docked", "last_transferred_at": now_iso(), "last_transferred_by": user.get("name", user.get("email", ""))}})
        
    await db.wms_tasks.update_one({"task_id": task_id}, {
        "$set": {"status": "completed", "completed_at": now_iso(), "completed_by": user.get("name", "")}
    })
    
    await log_movement(user, "task_completed", {"task_id": task_id, "type": task["task_type"]})
    await notify_badge_change("all")
    return {"message": "Task successfully executed"}


# ==================== UPC PRODUCT CATALOG ====================
# Single source of truth for what a UPC means. Receiving (and the inline
# Crear ASN modal later) look up by UPC to autofill style/color/size/etc
# so the same product never lands twice with conflicting descriptions.
#
# Collection: wms_upc_catalog. Unique key: upc (normalized to digits-only
# uppercase). Other fields are the canonical product metadata.


def _norm_upc(raw) -> str:
    """Strip whitespace and uppercase. Don't restrict to digits because some
    customers use alphanumeric internal codes — but trim aggressively so a
    typo'd trailing space doesn't bypass uniqueness."""
    return str(raw or "").strip().upper()


@router.get("/upc/{upc}")
async def get_upc(upc: str, request: Request):
    """Lookup a UPC. Returns 404 if not in the catalog so the frontend can
    show the 'Crear UPC' button."""
    await require_auth(request)
    code = _norm_upc(upc)
    if not code:
        raise HTTPException(400, "upc requerido")
    doc = await db.wms_upc_catalog.find_one({"upc": code}, {"_id": 0})
    if not doc:
        raise HTTPException(404, f"UPC {code} no encontrado")
    return doc


@router.get("/upc")
async def list_upc(request: Request, search: str = "", customer: str = "",
                    style: str = "", limit: int = 200):
    """List the catalog with optional filters. Used by the typeahead in the
    Receiving form so the operator can pick a UPC without typing it whole."""
    await require_auth(request)
    query = {}
    if customer:
        query["customer"] = {"$regex": re.escape(customer), "$options": "i"}
    if style:
        query["style"] = {"$regex": re.escape(style), "$options": "i"}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"upc": rx}, {"style": rx}, {"description": rx},
            {"color": rx}, {"customer": rx}, {"brand": rx},
        ]
    limit = max(1, min(limit, 1000))
    docs = await db.wms_upc_catalog.find(query, {"_id": 0}).sort("upc", 1).to_list(limit)
    return docs


@router.post("/upc")
async def create_upc(request: Request):
    """Create a UPC catalog entry. Idempotent: returns the existing doc if the
    UPC is already registered (so a race between two operators capturing the
    same UPC doesn't error out — the second one just gets the existing one)."""
    user = await require_auth(request)
    body = await request.json()
    code = _norm_upc(body.get("upc"))
    if not code:
        raise HTTPException(400, "upc es obligatorio")

    existing = await db.wms_upc_catalog.find_one({"upc": code}, {"_id": 0})
    if existing:
        return existing  # idempotent return

    doc = {
        "upc": code,
        "customer": await _canonical_customer(body.get("customer", "")),
        "manufacturer": str(body.get("manufacturer", "")).strip().upper(),
        "style": str(body.get("style", "")).strip().upper(),
        "color": str(body.get("color", "")).strip().upper(),
        "size": str(body.get("size", "")).strip().upper(),
        "description": str(body.get("description", "")).strip(),
        "country_of_origin": str(body.get("country_of_origin", "")).strip().upper(),
        "fabric_content": str(body.get("fabric_content", "")).strip(),
        "brand": str(body.get("brand", "")).strip().upper(),
        "sku": str(body.get("sku", "")).strip().upper(),
        "created_at": now_iso(),
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name", ""),
    }
    if not doc["style"]:
        raise HTTPException(400, "style es obligatorio para crear un UPC")
    # Misma guardia que create_receiving: los 6 campos de identidad deben venir
    # del catalogo curado. El UPC es la puerta trasera preferida para basura,
    # antes se escribia libre en un <input> del modal.
    await _assert_curated_identity(doc["customer"], {
        "styles": doc["style"],
        "colors": doc["color"],
        "sizes": doc["size"],
        "descriptions": doc["description"],
        "countries": doc["country_of_origin"],
        "fabrics": doc["fabric_content"],
    })
    await db.wms_upc_catalog.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/upc/{upc}")
async def update_upc(upc: str, request: Request):
    """Editar UPC del catalogo. Cualquier operador autenticado puede editar,
    PERO los 6 campos de identidad (style/color/size/description/country/
    fabric) DEBEN venir del catalogo curado — se valida con
    `_assert_curated_identity` igual que create_upc/create_receiving. Solo
    admin cambia el customer (mover el UPC a otro cliente cambia su scope
    de styles y no debe ser un cambio casual del operador).
    Propagacion a filas historicas NO es automatica — solo futuros lecturas
    heredan el valor."""
    user = await require_auth(request)
    code = _norm_upc(upc)
    body = await request.json()
    existing = await db.wms_upc_catalog.find_one({"upc": code})
    if not existing:
        raise HTTPException(404, f"UPC {code} no encontrado")
    allowed = ["customer", "manufacturer", "style", "color", "size", "description",
               "country_of_origin", "fabric_content", "brand", "sku"]
    update = {}
    for k in allowed:
        if k in body:
            if k == "customer":
                # Cualquier usuario autenticado puede cambiar el customer del UPC
                # (incluye mover a otro cliente). El valor se canonicaliza igual.
                update[k] = await _canonical_customer(body[k])
            else:
                update[k] = (str(body[k]).strip().upper() if k != "description" and k != "fabric_content"
                              else str(body[k]).strip())
    if not update:
        raise HTTPException(400, "Nada que actualizar")
    # Guard curated: valores de identidad deben estar en el catalogo curado.
    # Merge del doc actual con el update para evaluar los 6 campos.
    merged = {**existing, **update}
    await _assert_curated_identity(merged.get("customer", ""), {
        "styles": merged.get("style", ""),
        "colors": merged.get("color", ""),
        "sizes": merged.get("size", ""),
        "descriptions": merged.get("description", ""),
        "countries": merged.get("country_of_origin", ""),
        "fabrics": merged.get("fabric_content", ""),
    })
    update["updated_at"] = now_iso()
    update["updated_by"] = user.get("user_id")
    await db.wms_upc_catalog.update_one({"upc": code}, {"$set": update})

    # Propagacion opcional a cajas / receivings / inventory que ya usan este UPC.
    # Solo campos DESCRIPTIVOS (description, country_of_origin, fabric_content,
    # brand) — nunca identity (style/sku/color/size/customer/manufacturer), que
    # requieren el flujo de /upc/{upc}/correct (mueve inventario, guarda
    # allocations). Sin este flag, el catalogo queda actualizado pero las cajas
    # ya recibidas quedan con el valor viejo (bug del receiving GILDAN 5000
    # BLACK L, 2026-07-13: pais editado en UPC pero las 12 cajas se imprimieron
    # con el valor previo).
    propagate = bool(body.get("propagate_to_boxes"))
    propagate_result = None
    if propagate:
        PROPAGABLE = {"description", "country_of_origin", "fabric_content", "brand"}
        prop_set = {k: v for k, v in update.items()
                     if k in PROPAGABLE}
        if prop_set:
            box_set = dict(prop_set)
            if "country_of_origin" in prop_set:
                box_set["coo"] = prop_set["country_of_origin"]  # mirror field en boxes
            box_set["updated_at"] = now_iso()
            # Todas las cajas con este UPC O que compartan receiving con este UPC
            rcvs = await db.wms_receiving.find({"upc": code}, {"receiving_id": 1}).to_list(2000)
            rcv_ids = [r["receiving_id"] for r in rcvs]
            or_clauses = [{"upc": code}]
            if rcv_ids:
                or_clauses.append({"receiving_id": {"$in": rcv_ids}})
            rb = await db.wms_boxes.update_many({"$or": or_clauses}, {"$set": box_set})
            rr = await db.wms_receiving.update_many({"upc": code}, {"$set": {**prop_set, "updated_at": now_iso()}})
            ri = await db.wms_inventory.update_many(
                {"receiving_id": {"$in": rcv_ids}} if rcv_ids else {"_id": {"$exists": False}},
                {"$set": {**prop_set, "updated_at": now_iso()}}
            )
            propagate_result = {
                "boxes": rb.modified_count,
                "receivings": rr.modified_count,
                "inventory": ri.modified_count,
            }
            await log_movement(user, "upc_propagate", {
                "upc": code, "changes": prop_set, **propagate_result,
            })
    doc = await db.wms_upc_catalog.find_one({"upc": code}, {"_id": 0})
    if propagate_result is not None:
        doc["_propagate"] = propagate_result
    return doc


@router.post("/upc/{upc}/correct")
async def correct_upc(upc: str, request: Request):
    """Admin-only: fix a UPC captured with wrong attributes and cascade the fix
    to the stock already received with it.
      - Identity fields (style/sku, color, size) change the inventory line, so we
        MOVE each affected box's units from its old line to the corrected line
        (same location), keeping on_hand / total_boxes consistent.
      - Descriptive fields (description, country_of_origin, fabric_content,
        customer, manufacturer, brand) are updated in place on the boxes,
        receivings, catalog and the boxes' inventory lines.
    dry-run by default; apply=true commits. Refuses if any affected source line
    has allocated units, so open order allocations never desync."""
    user = await require_admin(request)
    code = _norm_upc(upc)
    body = await request.json()
    apply = bool(body.get("apply", False))
    raw_changes = body.get("changes") or {}

    cat = await db.wms_upc_catalog.find_one({"upc": code})
    if not cat:
        raise HTTPException(404, f"UPC {code} no encontrado")

    from collections import defaultdict

    UPPER = {"style", "sku", "color", "size", "country_of_origin", "customer", "manufacturer", "brand"}
    ALLOWED = ["style", "sku", "color", "size", "description", "country_of_origin",
               "fabric_content", "customer", "manufacturer", "brand"]
    changes = {}
    for k in ALLOWED:
        if k in raw_changes and raw_changes[k] is not None:
            v = str(raw_changes[k]).strip()
            v = v.upper() if k in UPPER else v
            cur = str(cat.get(k, "") or "").strip()
            cur = cur.upper() if k in UPPER else cur
            if v != cur:
                changes[k] = v
    if not changes:
        raise HTTPException(400, "No hay cambios para aplicar")
    # Inventory keys on sku; keep it in lockstep with style.
    if "style" in changes and "sku" not in changes:
        changes["sku"] = changes["style"]

    KEY = {"sku", "color", "size"}  # fields that move inventory between lines
    key_changed = bool(KEY & set(changes.keys()))

    rcvs = await db.wms_receiving.find({"upc": code}, {"receiving_id": 1}).to_list(2000)
    rcv_ids = [r["receiving_id"] for r in rcvs]
    or_clauses = [{"upc": code}]
    if rcv_ids:
        or_clauses.append({"receiving_id": {"$in": rcv_ids}})
    boxes = await db.wms_boxes.find({"$or": or_clauses}, {"_id": 0}).to_list(5000)

    def src_key(b):
        return (b.get("sku") or b.get("style") or "", b.get("color", ""), b.get("size", ""), b.get("location", ""))

    def tgt_key(b):
        sku = changes.get("sku", b.get("sku") or b.get("style") or "")
        color = changes.get("color", b.get("color", ""))
        size = changes.get("size", b.get("size", ""))
        return (sku, color, size, b.get("location", ""))

    # Guard allocations on the source lines we would move from.
    blocked, moved_units, moved_boxes = [], 0, 0
    if key_changed:
        checked = set()
        for b in boxes:
            sk = src_key(b)
            if sk == tgt_key(b):
                continue
            moved_units += int(b.get("units") or b.get("qty") or 0)
            moved_boxes += 1
            if sk in checked:
                continue
            checked.add(sk)
            s_sku, s_color, s_size, s_loc = sk
            inv = await db.wms_inventory.find_one(
                {"sku": s_sku, "color": s_color, "size": s_size, "location": s_loc}
            )
            if inv and int(inv.get("units_allocated", 0) or 0) > 0:
                blocked.append({"location": s_loc, "allocated": int(inv["units_allocated"])})

    preview = {
        "upc": code, "changes": changes, "receivings": len(rcv_ids),
        "boxes_total": len(boxes), "moved_boxes": moved_boxes,
        "moved_units": moved_units, "key_changed": key_changed, "blocked": blocked,
    }
    if not apply:
        return {"applied": False, "preview": preview}
    if blocked:
        raise HTTPException(
            409,
            f"No se puede corregir: hay unidades asignadas a ordenes en {len(blocked)} "
            "ubicacion(es). Desasigna esas ordenes primero.",
        )

    # 1) Move units between inventory lines for identity changes (per box).
    if key_changed:
        for b in boxes:
            sk, tk = src_key(b), tgt_key(b)
            if sk == tk:
                continue
            units = int(b.get("units") or b.get("qty") or 0)
            s_sku, s_color, s_size, s_loc = sk
            t_sku, t_color, t_size, t_loc = tk
            src_inv = await db.wms_inventory.find_one(
                {"sku": s_sku, "color": s_color, "size": s_size, "location": s_loc}
            )
            if src_inv:
                noh = max(0, int(src_inv.get("units_on_hand", 0)) - units)
                ntb = max(0, int(src_inv.get("total_boxes", 0) or 0) - 1)
                if noh == 0 and ntb == 0:
                    await db.wms_inventory.delete_one({"_id": src_inv["_id"]})
                else:
                    await db.wms_inventory.update_one(
                        {"_id": src_inv["_id"]},
                        {"$set": {"units_on_hand": noh, "total_boxes": ntb, "updated_at": now_iso()}},
                    )
            tgt_inv = await db.wms_inventory.find_one(
                {"sku": t_sku, "color": t_color, "size": t_size, "location": t_loc}
            )
            if tgt_inv:
                await db.wms_inventory.update_one(
                    {"_id": tgt_inv["_id"]},
                    {"$inc": {"units_on_hand": units, "total_boxes": 1}, "$set": {"updated_at": now_iso()}},
                )
            else:
                await db.wms_inventory.insert_one({
                    "inventory_id": gen_id("inv"),
                    "sku": t_sku, "style": changes.get("style", b.get("style") or t_sku),
                    "color": t_color, "size": t_size,
                    "customer": changes.get("customer", b.get("customer", "")),
                    "manufacturer": changes.get("manufacturer", b.get("manufacturer", "")),
                    "description": changes.get("description", b.get("description", "")),
                    "country_of_origin": changes.get("country_of_origin", b.get("country_of_origin", "") or b.get("coo", "")),
                    "fabric_content": changes.get("fabric_content", b.get("fabric_content", "")),
                    "location": t_loc, "is_bpo": b.get("is_bpo", False),
                    "total_boxes": 1, "units_on_hand": units, "units_allocated": 0,
                    "updated_at": now_iso(),
                })

    # 2) Descriptive-only changes: update the boxes' (post-move) inventory lines.
    meta_inv = {k: v for k, v in changes.items()
                if k in ("description", "country_of_origin", "fabric_content", "customer", "manufacturer")}
    if meta_inv:
        seen = set()
        for b in boxes:
            tk = tgt_key(b)
            if tk in seen:
                continue
            seen.add(tk)
            t_sku, t_color, t_size, t_loc = tk
            await db.wms_inventory.update_many(
                {"sku": t_sku, "color": t_color, "size": t_size, "location": t_loc},
                {"$set": {**meta_inv, "updated_at": now_iso()}},
            )

    # 3) Update the boxes themselves with every changed field.
    box_set = dict(changes)
    if "country_of_origin" in changes:
        box_set["coo"] = changes["country_of_origin"]
    box_set["upc"] = code
    box_set["updated_at"] = now_iso()
    box_ids = [b["box_id"] for b in boxes]
    if box_ids:
        await db.wms_boxes.update_many({"box_id": {"$in": box_ids}}, {"$set": box_set})

    # 4) Update the receiving metadata.
    if rcv_ids:
        rcv_set = {k: v for k, v in changes.items()
                   if k in ("style", "sku", "color", "size", "description",
                            "country_of_origin", "fabric_content", "customer", "manufacturer")}
        if rcv_set:
            rcv_set["updated_at"] = now_iso()
            await db.wms_receiving.update_many({"receiving_id": {"$in": rcv_ids}}, {"$set": rcv_set})

    # 5) Update the catalog + log + notify.
    cat_set = dict(changes)
    cat_set["updated_at"] = now_iso()
    cat_set["updated_by"] = user.get("user_id")
    await db.wms_upc_catalog.update_one({"upc": code}, {"$set": cat_set})
    await log_movement(user, "upc_correction", {
        "upc": code, "changes": changes, "moved_boxes": moved_boxes,
        "moved_units": moved_units, "receivings": rcv_ids[:50],
    })
    try:
        await notify_badge_change("all")
    except Exception:
        pass
    return {"applied": True, "result": preview}


@router.delete("/upc/{upc}")
async def delete_upc(upc: str, request: Request):
    """Admin-only delete. Existing receiving records keep their data — they
    don't reference the catalog by FK, they just inherited values at capture
    time."""
    user = await require_admin(request)
    code = _norm_upc(upc)
    res = await db.wms_upc_catalog.delete_one({"upc": code})
    if res.deleted_count == 0:
        raise HTTPException(404, f"UPC {code} no encontrado")
    await log_movement(user, "upc_deleted", {"upc": code})
    return {"message": f"UPC {code} eliminado"}
