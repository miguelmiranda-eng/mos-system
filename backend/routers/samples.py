"""Módulo de Ejemplos (Samples) — calendarización de órdenes-ejemplo por día.

Capa de scheduling sobre `orders`: cada `sample_task` apunta a una orden real
(order_id) y sólo guarda lo propio del módulo (día programado, prioridad,
color de identificación, operador asignado, notas). Los datos de la orden
(cliente, style, cantidad, cancel_date, priority) se unen EN VIVO desde
`orders` — no se duplican.

Órdenes provisionales: si el usuario abre un ejemplo antes de que exista un
número de Printavo, se crea una orden real con `order_number = PROV-<seq>` y
`is_provisional = True`. Cuando Printavo trae el número definitivo, la UI
sugiere el match (auto por style+client+due_date) y el usuario confirma con
`POST /tasks/{id}/promote` — se fusiona el ejemplo a la orden real y se
elimina el PROV-.
"""
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from deps import db, require_auth, require_role, log_activity, DEFAULT_OPTIONS, logger
from datetime import datetime, timezone, timedelta
import uuid, re, io

router = APIRouter(prefix="/api/samples")

WRITE_ROLES = ["sample_lead"]          # + admin/supersu siempre (SUPER_ROLES)
HOT_PRIORITIES = {"RUSH", "SPECIAL RUSH", "OVERSOLD", "EVENT"}

# Estatus finales del ejemplo (los que quitan el ejemplo del "activo"). Reusan
# los mismos strings que production_statuses para no crear un vocabulario paralelo.
TERMINAL_STATUSES = {"EJEMPLO APROBADO", "CANCELLED"}

# ── AVANCE DEL EJEMPLO ──────────────────────────────────────────────────────
# Cuatro palomeos independientes. Son PROPIOS del ejemplo y NO escriben en los
# campos homónimos de la orden (art_sep_status, art_neck_status, trim_status,
# screens): el arte de un EJEMPLO no es el de la orden de producción, y
# mezclarlos ensuciaría los contadores del módulo de Arte, que miden
# separaciones hechas por día para producción.
SAMPLE_FLAGS = ("arte", "neck", "trim", "screens")

# Aprobación del cliente. Nace en RECIBIDO y sólo tiene un paso adelante.
APPROVAL_RECEIVED = "RECIBIDO"
APPROVAL_APPROVED = "APROBADO"
APPROVAL_VALUES = (APPROVAL_RECEIVED, APPROVAL_APPROVED)

# ── LAS TRES PESTAÑAS ───────────────────────────────────────────────────────
# Son EXCLUYENTES y se derivan, no se guardan: un estado guardado se
# desincroniza en cuanto alguien promueve una orden o cambia la aprobación por
# otra vía. La precedencia:
#
#   1. aprobado por el cliente      → APROBADOS     (gana, TENGA O NO PO)
#   2. tiene número de orden (PO)   → READY TO MOS
#   3. lo demás                     → EJEMPLOS      (donde se crean y programan)
#
# POR QUÉ LA APROBACIÓN GANA
# Primero se hizo al revés (PO sobre aprobación) y la orden 2722 —que ya tenía
# número real— se quedó en Ready to MOS al aprobarla, cuando quien la aprobó
# esperaba verla en Aprobados. Las dos reglas del negocio se cumplen igual con
# esta precedencia: un ejemplo recién creado CON número de orden todavía no está
# aprobado, así que sigue entrando directo a Ready to MOS.
STAGE_EJEMPLOS = "ejemplos"
STAGE_APROBADOS = "aprobados"
STAGE_READY = "ready_to_mos"
STAGES = (STAGE_EJEMPLOS, STAGE_APROBADOS, STAGE_READY)


def _default_flags():
    return {f: False for f in SAMPLE_FLAGS}


def _stage_of(task) -> str:
    """Pestaña a la que pertenece el ejemplo. Ver la precedencia arriba."""
    if (task.get("approval") or APPROVAL_RECEIVED) == APPROVAL_APPROVED:
        return STAGE_APROBADOS
    if not task.get("is_provisional"):
        return STAGE_READY
    return STAGE_EJEMPLOS

# Campos de la orden que se muestran en la tarjeta (se unen en vivo).
_ORDER_FIELDS = {
    "_id": 0, "order_id": 1, "order_number": 1, "client": 1, "branding": 1,
    "quantity": 1, "color": 1, "style": 1, "sizes": 1, "priority": 1,
    "due_date": 1, "cancel_date": 1, "board": 1, "production_status": 1,
    "sample": 1, "job_title_a": 1, "job_title_b": 1,
    "art_sep_status": 1, "art_neck_status": 1, "screens": 1,
    "is_provisional": 1,
}


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def today_iso():
    return datetime.now(timezone.utc).date().isoformat()


def gen_id(prefix="sample"):
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


async def require_sample_write(request: Request):
    return await require_role(request, WRITE_ROLES)


def _is_hot(order):
    return (order.get("priority") or "").upper() in HOT_PRIORITIES


async def _broadcast(order_id):
    try:
        from ws_manager import ws_manager
        await ws_manager.broadcast({"type": "SAMPLE_UPDATED", "order_id": order_id})
    except Exception as e:
        logger.error(f"[samples] broadcast failed: {e}")


# ── Blank availability (WMS) ────────────────────────────────────────────────
async def _blank_status(order):
    """¿Hay playeras en almacén para producir el ejemplo? Suma units_on_hand
    por style+color en wms_inventory y compara contra la cantidad de la orden.
    Devuelve dict con {ok, on_hand, needed}. Best-effort: si no hay style/color
    → ok = False y needed = 0.
    """
    style = (order.get("style") or "").strip().upper()
    color = (order.get("color") or "").strip().upper()
    needed = int(order.get("quantity") or 0)
    if not style:
        return {"ok": False, "on_hand": 0, "needed": needed, "reason": "sin style"}
    q = {"$or": [{"style": style}, {"sku": style}]}
    if color:
        q["color"] = color
    on_hand = 0
    async for row in db.wms_inventory.find(q, {"_id": 0, "units_on_hand": 1}):
        try:
            on_hand += int(row.get("units_on_hand") or 0)
        except (TypeError, ValueError):
            pass
    ok = on_hand >= max(1, needed)
    return {"ok": ok, "on_hand": on_hand, "needed": needed}


async def _paint_ready(order_id):
    """¿La tinta de esta orden está lista? Consulta paint_tasks."""
    pt = await db.paint_tasks.find_one(
        {"order_id": order_id}, {"_id": 0, "ink_status": 1})
    if not pt:
        return {"ok": None, "reason": "no_paint_task"}  # None = no aplica
    return {"ok": pt.get("ink_status") == "lista", "status": pt.get("ink_status")}


async def _resources(order):
    """Estado de todos los recursos del ejemplo (blank + arte + neck + screens
    + tinta). Se calcula en vivo para no persistir estado incorrecto.
    """
    blank = await _blank_status(order)
    paint = await _paint_ready(order["order_id"])
    art_ok = bool(order.get("art_sep_status"))
    neck_ok = bool(order.get("art_neck_status"))
    screens_ok = bool(order.get("screens"))
    return {
        "blank_ok": blank["ok"], "blank_on_hand": blank.get("on_hand", 0),
        "blank_needed": blank.get("needed", 0),
        "art_ok": art_ok, "neck_ok": neck_ok, "screens_ok": screens_ok,
        "paint_ok": paint["ok"], "paint_status": paint.get("status"),
    }


def _overdue(order, task):
    """True si el ejemplo ya pasó su cancel_date y no está en un estatus final."""
    cd = (order.get("cancel_date") or "").strip()
    if not cd or (task.get("status") or "").upper() in TERMINAL_STATUSES:
        return False
    try:
        cancel = datetime.strptime(cd[:10], "%Y-%m-%d").date()
    except ValueError:
        return False
    return datetime.now(timezone.utc).date() > cancel


async def _enrich(tasks):
    """Une cada task con su orden en vivo (`.order`) y sus recursos (`.resources`)."""
    ids = [t["order_id"] for t in tasks if t.get("order_id")]
    omap = {}
    if ids:
        async for o in db.orders.find({"order_id": {"$in": ids}}, _ORDER_FIELDS):
            omap[o["order_id"]] = o
    for t in tasks:
        o = omap.get(t.get("order_id"), {})
        t["order"] = o
        t["resources"] = await _resources(o) if o else {}
        t["overdue"] = _overdue(o, t) if o else False
        # Defaults de lectura para los ejemplos creados antes de esta versión.
        flags = t.get("flags") or {}
        t["flags"] = {f: bool(flags.get(f)) for f in SAMPLE_FLAGS}
        t["approval"] = t.get("approval") or APPROVAL_RECEIVED
        t["stage"] = _stage_of(t)
    return tasks


# ═══════════════ CALENDARIO ═══════════════════════════════════════════════════
@router.get("/calendar")
async def calendar(request: Request, start: str = "", end: str = ""):
    """Tasks en rango [start, end] (YYYY-MM-DD) — para vista día/semana/mes.
    Si no se pasan fechas, devuelve la semana actual (lun–dom)."""
    await require_auth(request)
    try:
        if start and end:
            s = datetime.strptime(start, "%Y-%m-%d").date()
            e = datetime.strptime(end, "%Y-%m-%d").date()
        else:
            base = datetime.now(timezone.utc).date()
            s = base - timedelta(days=base.weekday())
            e = s + timedelta(days=6)
    except ValueError:
        raise HTTPException(400, "start/end deben ser YYYY-MM-DD")

    days_iso = [(s + timedelta(days=i)).isoformat() for i in range((e - s).days + 1)]
    tasks = await db.sample_tasks.find(
        {"scheduled_date": {"$in": days_iso}}, {"_id": 0}
    ).sort([("is_hot", -1), ("priority_rank", 1)]).to_list(3000)
    await _enrich(tasks)

    by_day = {d: [] for d in days_iso}
    for t in tasks:
        by_day.setdefault(t.get("scheduled_date"), []).append(t)

    return {
        "start": s.isoformat(), "end": e.isoformat(),
        "days": [{"date": d, "tasks": by_day.get(d, [])} for d in days_iso],
        "hot_count": sum(1 for t in tasks if t.get("is_hot")),
        "overdue_count": sum(1 for t in tasks if t.get("overdue")),
        "total": len(tasks),
    }


# ── Backlog (sin programar) ──────────────────────────────────────────────────
@router.get("/backlog")
async def backlog(request: Request):
    await require_auth(request)
    tasks = await db.sample_tasks.find(
        {"$or": [{"scheduled_date": None}, {"scheduled_date": {"$exists": False}}]},
        {"_id": 0}
    ).sort([("is_hot", -1), ("priority_rank", 1)]).to_list(1000)
    await _enrich(tasks)
    return {"backlog": tasks}


# ── Buscar cualquier orden para agregar al calendario ────────────────────────
@router.get("/search")
async def search_orders(request: Request, q: str = ""):
    await require_auth(request)
    q = (q or "").strip()
    if not q:
        return {"results": []}
    rx = {"$regex": re.escape(q), "$options": "i"}
    orders = await db.orders.find(
        {"board": {"$ne": "PAPELERA DE RECICLAJE"},
         "$or": [{"order_number": rx}, {"client": rx}, {"style": rx}]},
        _ORDER_FIELDS
    ).sort("created_at", -1).limit(25).to_list(25)

    ids = [o["order_id"] for o in orders]
    in_samples = set()
    if ids:
        in_samples = {d["order_id"] for d in
                      await db.sample_tasks.find(
                          {"order_id": {"$in": ids}}, {"_id": 0, "order_id": 1}).to_list(2000)}
    for o in orders:
        o["in_samples"] = o["order_id"] in in_samples
    return {"results": orders}


# ═══════════════ CRUD ═══════════════════════════════════════════════════════
async def _next_prov_seq():
    """Secuencia atómica para PROV-<n>. Guardada en la colección counters."""
    doc = await db.counters.find_one_and_update(
        {"_id": "prov_order"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    # Motor retorna el documento actualizado (after update por default con return_document=True)
    return int((doc or {}).get("seq") or 1)


async def _create_provisional_order(user, hint, explicit_order_number=None):
    """Crea una orden mínima para poder colgar el ejemplo. Si el usuario dio
    un `explicit_order_number`, se usa tal cual y la orden NO es provisional
    (asume que es el número real de Printavo/MOS). Si no lo dio, se genera
    `PROV-<n>` autoincremental.
    """
    if explicit_order_number:
        order_number = str(explicit_order_number).strip()
        is_prov = False
        board = "SCHEDULING"
    else:
        seq = await _next_prov_seq()
        order_number = f"PROV-{seq:04d}"
        is_prov = True
        board = "EJEMPLOS"
    now = now_iso()
    doc = {
        "order_id": f"ord_{uuid.uuid4().hex[:12]}",
        "order_number": order_number,
        "client": (hint.get("client") or "").strip().upper(),
        "style": (hint.get("style") or "").strip().upper(),
        "color": (hint.get("color") or "").strip().upper(),
        "quantity": int(hint.get("quantity") or 0),
        "priority": hint.get("priority") or "PRIORITY 2",
        "cancel_date": hint.get("cancel_date") or None,
        "due_date": hint.get("due_date") or None,
        "board": board,
        "sample": "EJEMPLO PRIMERO",
        "production_status": "EN ESPERA",
        "is_provisional": is_prov,
        "source": "sample_manual" if not is_prov else "sample_provisional",
        "created_at": now, "updated_at": now,
        "created_by": user.get("user_id"),
    }
    await db.orders.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@router.post("/tasks")
async def add_task(request: Request):
    """Agrega una orden al calendario de ejemplos. Modos:
      - {"order_id": "ord_..."}          → task para una orden existente
      - {"order_number": "12345"}        → busca por número; si no existe,
                                            requiere `provisional=True` para crearla
      - {"provisional": True, "style":..., "client":..., ...} → crea PROV- + task
    """
    user = await require_sample_write(request)
    body = await request.json()
    oid = (body.get("order_id") or "").strip()
    onum = (body.get("order_number") or "").strip()
    provisional = bool(body.get("provisional"))

    order = None
    if oid:
        order = await db.orders.find_one({"order_id": oid}, _ORDER_FIELDS)
    if not order and onum:
        order = await db.orders.find_one(
            {"order_number": {"$regex": f"^{re.escape(onum)}$", "$options": "i"}},
            _ORDER_FIELDS)
    if not order:
        if not provisional:
            raise HTTPException(
                404,
                "Orden no encontrada. Usa provisional=True para crear un PROV-.",
            )
        # Si el usuario dio un número explícito → orden real con ese número.
        # Si no → PROV-<n> autogenerado.
        order = await _create_provisional_order(user, body, explicit_order_number=onum or None)

    if await db.sample_tasks.find_one({"order_id": order["order_id"]}, {"_id": 0}):
        raise HTTPException(409, "La orden ya está en el calendario de ejemplos")

    doc = {
        "sample_task_id": gen_id(),
        "order_id": order["order_id"],
        "order_number": order.get("order_number"),
        "is_provisional": bool(order.get("is_provisional")),
        "scheduled_date": body.get("scheduled_date") or None,
        "priority_rank": 0,
        "is_hot": _is_hot(order),
        "status": body.get("status") or order.get("sample") or "EJEMPLO PRIMERO",
        # Avance propio del ejemplo (ver SAMPLE_FLAGS).
        "flags": {f: bool((body.get("flags") or {}).get(f)) for f in SAMPLE_FLAGS},
        "flags_meta": {},
        "approval": (body.get("approval") or APPROVAL_RECEIVED).strip().upper()
                    if (body.get("approval") or APPROVAL_RECEIVED).strip().upper() in APPROVAL_VALUES
                    else APPROVAL_RECEIVED,
        "operator_id": body.get("operator_id") or None,
        "operator": body.get("operator") or None,
        "color_tag": body.get("color_tag") or None,   # hex opcional
        "notes": body.get("notes") or "",
        "created_at": now_iso(), "updated_at": now_iso(),
        "created_by": user.get("name"),
    }
    await db.sample_tasks.insert_one(doc)
    doc.pop("_id", None)
    await _enrich([doc])
    await log_activity(user, "sample_task_added",
                       {"order_number": doc["order_number"],
                        "provisional": doc["is_provisional"]})
    await _broadcast(order["order_id"])
    return doc


@router.put("/tasks/{task_id}")
async def update_task(task_id: str, request: Request):
    user = await require_sample_write(request)
    body = await request.json()
    allowed = ("scheduled_date", "priority_rank", "is_hot", "color_tag",
               "notes", "operator_id", "operator")
    upd = {k: body[k] for k in allowed if k in body}
    if not upd:
        raise HTTPException(400, "Nada que actualizar")
    upd["updated_at"] = now_iso()
    res = await db.sample_tasks.update_one({"sample_task_id": task_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Ejemplo no encontrado")
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    await _enrich([task])
    await _broadcast(task.get("order_id"))
    return task


@router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, request: Request):
    user = await require_sample_write(request)
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")
    await db.sample_tasks.delete_one({"sample_task_id": task_id})
    await log_activity(user, "sample_task_removed",
                       {"order_number": task.get("order_number")})
    await _broadcast(task.get("order_id"))
    return {"ok": True}


# ── Drag-and-drop: reorder / mover a otro día ────────────────────────────────
@router.post("/reorder")
async def reorder(request: Request):
    await require_sample_write(request)
    body = await request.json()
    date = body.get("date")            # None => backlog
    ordered = body.get("ordered_ids") or []
    for i, tid in enumerate(ordered):
        await db.sample_tasks.update_one(
            {"sample_task_id": tid},
            {"$set": {"scheduled_date": date, "priority_rank": i,
                      "updated_at": now_iso()}})
    return {"ok": True, "count": len(ordered)}


# ── Cambiar estatus (reusa production_statuses) ──────────────────────────────
@router.put("/tasks/{task_id}/status")
async def set_status(task_id: str, request: Request):
    user = await require_sample_write(request)
    body = await request.json()
    status = (body.get("status") or "").strip()
    # Los estatus válidos vienen de la config real (config_id="main"), con
    # fallback a los defaults estáticos si aún no se ha personalizado.
    cfg = await db.config_options.find_one({"config_id": "main"}, {"_id": 0}) or {}
    prod = cfg.get("production_statuses") or DEFAULT_OPTIONS["production_statuses"]
    samp = cfg.get("samples") or DEFAULT_OPTIONS["samples"]
    valid = set(prod) | set(samp)
    if status not in valid:
        raise HTTPException(400, f"Estatus '{status}' no está en el catálogo")
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")

    await db.sample_tasks.update_one(
        {"sample_task_id": task_id},
        {"$set": {"status": status, "updated_at": now_iso()}})

    # Espeja a la orden: mantiene consistencia con el board legacy `sample`.
    if status in DEFAULT_OPTIONS["samples"]:
        await db.orders.update_one(
            {"order_id": task["order_id"]}, {"$set": {"sample": status}})

    await log_activity(user, "sample_status",
                       {"order_number": task.get("order_number"), "status": status})
    await _broadcast(task["order_id"])
    task["status"] = status
    await _enrich([task])
    return task


# ═══════════════ ASIGNACIÓN DE OPERADOR ═════════════════════════════════════
@router.put("/tasks/{task_id}/assign")
async def assign_operator(task_id: str, request: Request):
    user = await require_sample_write(request)
    body = await request.json()
    op_id = (body.get("operator_id") or "").strip() or None
    op_name = None
    if op_id:
        op = await db.operators.find_one({"operator_id": op_id}, {"_id": 0})
        if not op:
            raise HTTPException(404, "Operador no encontrado")
        op_name = op.get("name")

    # `find_one_and_update` = mutación atómica: dos asignaciones concurrentes
    # no pueden dejar el task en un estado intermedio.
    task = await db.sample_tasks.find_one_and_update(
        {"sample_task_id": task_id},
        {"$set": {"operator_id": op_id, "operator": op_name,
                  "updated_at": now_iso()}},
        return_document=True, projection={"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")
    await log_activity(user, "sample_operator_assigned",
                       {"order_number": task.get("order_number"), "operator": op_name})
    await _broadcast(task.get("order_id"))
    await _enrich([task])
    return task


# ═══════════════ RECURSOS EN VIVO ═══════════════════════════════════════════
@router.get("/resources/{order_id}")
async def resources(order_id: str, request: Request):
    """Estado en vivo de todos los recursos que necesita el ejemplo."""
    await require_auth(request)
    order = await db.orders.find_one({"order_id": order_id}, _ORDER_FIELDS)
    if not order:
        raise HTTPException(404, "Orden no encontrada")
    return await _resources(order)


# ═══════════════ AUTO-SCHEDULER (cancel_date - N días) ══════════════════════
def _shift_to_weekday(d: "date") -> "date":
    """Si cae en sábado o domingo, retrocede al viernes anterior (más lead time)."""
    from datetime import date as _date, timedelta as _td
    if d.weekday() == 5:            # sábado
        return d - _td(days=1)
    if d.weekday() == 6:            # domingo
        return d - _td(days=2)
    return d


@router.post("/auto-schedule")
async def auto_schedule(request: Request):
    """Programa automáticamente los ejemplos del backlog con distribución
    inteligente y respeto a la capacidad diaria del equipo.

    Body:
      {
        "days_before": 14,       # ideal: cancel_date - N (default 14)
        "target_per_day": 7,     # promedio deseado (default 7)
        "max_per_day": 15,       # cap DURO (default 15)
        "skip_weekends": true,
        "filter": "all",         # all | with_style | no_style
        "apply": false
      }

    Algoritmo:
      1. Ordena tasks por cancel_date ASC (los más urgentes primero).
      2. Para cada uno calcula el día ideal = cancel - days_before (min: mañana).
      3. Intenta colocar respetando dos ventanas:
         a) Ventana ±5 días alrededor del ideal buscando un día con carga
            < target_per_day. Nunca después de cancel_date.
         b) Si no encuentra, amplía a ±10 días con carga < max_per_day.
         c) Si todo está lleno, coloca en el ideal marcado como `over_cap`.
      4. La carga inicial cuenta los sample_tasks ya programados en Mongo.
      5. is_hot = True si la holgura final < 3 días.
    """
    user = await require_sample_write(request)
    body = await request.json()
    days_before = int(body.get("days_before") or 14)
    target_per_day = int(body.get("target_per_day") or 7)
    max_per_day = int(body.get("max_per_day") or 15)
    skip_weekends = bool(body.get("skip_weekends", True))
    filter_mode = (body.get("filter") or "all").lower()
    apply_it = bool(body.get("apply", False))
    # `include_scheduled=True` → los ya programados entran al pool para
    # redistribuir. Útil cuando hay días saturados que hay que abrir.
    include_scheduled = bool(body.get("include_scheduled", False))

    if days_before < 0 or days_before > 90:
        raise HTTPException(400, "days_before debe estar entre 0 y 90")
    if not (1 <= target_per_day <= max_per_day <= 100):
        raise HTTPException(400, "target_per_day y max_per_day deben cumplir 1 ≤ target ≤ max ≤ 100")

    from datetime import date as _date, timedelta as _td
    today = datetime.now(timezone.utc).date()
    tomorrow = today + _td(days=1)

    # Carga ya existente por día (solo los que NO están en el pool)
    existing_load = {}
    if not include_scheduled:
        async for row in db.sample_tasks.aggregate([
            {"$match": {"scheduled_date": {"$ne": None, "$exists": True}}},
            {"$group": {"_id": "$scheduled_date", "n": {"$sum": 1}}},
        ]):
            if row["_id"]:
                existing_load[row["_id"]] = row["n"]

    # Pool de tasks a procesar
    if include_scheduled:
        q = {}   # TODO: idealmente excluir estatus terminales (EJEMPLO APROBADO, CANCELLED)
    else:
        q = {"$or": [{"scheduled_date": None}, {"scheduled_date": {"$exists": False}}]}
    tasks = await db.sample_tasks.find(q, {"_id": 0}).to_list(5000)
    # Excluir tasks terminales del rebalance para no molestar aprobados
    tasks = [t for t in tasks if (t.get("status") or "").upper() not in {"EJEMPLO APROBADO", "CANCELLED"}]
    await _enrich(tasks)

    # Filtro y precómputo del ideal por task; los que no tienen cancel_date
    # se descartan a `skipped` de una vez.
    prepared, skipped = [], []
    for t in tasks:
        o = t.get("order") or {}
        has_style = bool(o.get("style"))
        if filter_mode == "with_style" and not has_style:
            continue
        if filter_mode == "no_style" and has_style:
            continue
        cancel_raw = (o.get("cancel_date") or "").strip()
        if not cancel_raw:
            # Si ya tenía fecha manual y no tiene cancel, respetamos su fecha
            # actual: no la movemos (no hay criterio para reprogramarla).
            if include_scheduled and t.get("scheduled_date"):
                continue
            skipped.append({
                "sample_task_id": t["sample_task_id"],
                "order_number": t["order_number"],
                "reason": "sin cancel_date",
            })
            continue
        try:
            cancel_dt = datetime.strptime(cancel_raw[:10], "%Y-%m-%d").date()
        except ValueError:
            skipped.append({
                "sample_task_id": t["sample_task_id"],
                "order_number": t["order_number"],
                "reason": f"cancel_date inválido: {cancel_raw}",
            })
            continue
        ideal = cancel_dt - _td(days=days_before)
        if ideal < tomorrow:
            ideal = tomorrow
        if skip_weekends:
            ideal = _shift_to_weekday(ideal)
        prepared.append({
            "task": t, "order": o,
            "ideal": ideal, "cancel_dt": cancel_dt,
            "previous_date": t.get("scheduled_date"),
        })

    # Ordena por cancel_date ASC — los más urgentes primero (para que si hay
    # cuellos de botella los "cómodos" cedan su día ideal).
    prepared.sort(key=lambda p: (p["cancel_dt"], p["ideal"]))

    new_load = {}     # carga que este batch va sumando por día
    def load_on(d):
        k = d.isoformat()
        return existing_load.get(k, 0) + new_load.get(k, 0)

    def try_place(ideal, cancel_dt, cap):
        """Busca la primera fecha en ventana [ideal-N, ideal+M] con carga < cap.
        Nunca después de cancel_dt ni antes de mañana. Devuelve fecha o None."""
        # Explora en orden: 0, -1, +1, -2, +2, ... hasta ±10 días
        for offset in [0, -1, 1, -2, 2, -3, 3, -4, 4, -5, 5, -6, 6, -7, 7, -8, 8, -9, 9, -10, 10]:
            candidate = ideal + _td(days=offset)
            if candidate < tomorrow: continue
            if candidate > cancel_dt: continue
            if skip_weekends:
                candidate = _shift_to_weekday(candidate)
                if candidate < tomorrow or candidate > cancel_dt: continue
            if load_on(candidate) < cap:
                return candidate
        return None

    scheduled = []
    over_cap_count = 0
    shifted_count = 0
    moved_count = 0
    kept_count = 0

    # Sub-orden: si en modo rebalance, procesa PRIMERO los ya-programados
    # cuya fecha actual entra dentro de la ventana ideal, para consumir su
    # slot y no forzar movimientos innecesarios.
    if include_scheduled:
        def _keep_priority(p):
            prev = p.get("previous_date")
            if not prev: return (1, p["cancel_dt"])
            try:
                prev_dt = datetime.strptime(prev, "%Y-%m-%d").date()
            except ValueError:
                return (1, p["cancel_dt"])
            if abs((prev_dt - p["ideal"]).days) <= 5:
                return (0, p["cancel_dt"])   # su fecha actual es "razonable"
            return (1, p["cancel_dt"])
        prepared.sort(key=_keep_priority)

    for p in prepared:
        ideal = p["ideal"]
        cancel_dt = p["cancel_dt"]
        prev_iso = p.get("previous_date")

        # En rebalance: si ya tenía fecha razonable y su día no está saturado,
        # respetamos su fecha actual (evita churn de la programación manual).
        placed = None
        if include_scheduled and prev_iso:
            try:
                prev_dt = datetime.strptime(prev_iso, "%Y-%m-%d").date()
                if (tomorrow <= prev_dt <= cancel_dt
                        and abs((prev_dt - ideal).days) <= 5
                        and load_on(prev_dt) < target_per_day):
                    placed = prev_dt
            except ValueError:
                pass

        # Si no pudimos respetar el actual, aplicar la búsqueda normal.
        if placed is None:
            placed = try_place(ideal, cancel_dt, target_per_day)
        if placed is None:
            placed = try_place(ideal, cancel_dt, max_per_day)
        over_cap = placed is None
        if over_cap:
            placed = ideal
            over_cap_count += 1

        shift_days = (placed - ideal).days
        if shift_days != 0: shifted_count += 1
        placed_iso = placed.isoformat()
        moved = bool(prev_iso) and prev_iso != placed_iso
        if moved: moved_count += 1
        elif prev_iso: kept_count += 1

        holgura = (cancel_dt - placed).days
        is_hot = holgura < 3
        new_load[placed_iso] = new_load.get(placed_iso, 0) + 1

        rec = {
            "sample_task_id": p["task"]["sample_task_id"],
            "order_number":   p["task"]["order_number"],
            "style":  p["order"].get("style") or "",
            "client": p["order"].get("client") or "",
            "cancel_date":    cancel_dt.isoformat(),
            "ideal_date":     ideal.isoformat(),
            "previous_date":  prev_iso,
            "scheduled_date": placed_iso,
            "shift_days":     shift_days,
            "holgura_dias":   holgura,
            "is_hot":         is_hot,
            "over_cap":       over_cap,
            "moved":          moved,
            "kept":           bool(prev_iso) and not moved,
        }
        scheduled.append(rec)

        # Solo escribimos si realmente cambia (evita 100 updates inútiles).
        if apply_it and (prev_iso != placed_iso or (is_hot and not p["task"].get("is_hot"))):
            await db.sample_tasks.update_one(
                {"sample_task_id": p["task"]["sample_task_id"]},
                {"$set": {
                    "scheduled_date": placed_iso,
                    "is_hot": is_hot or p["task"].get("is_hot", False),
                    "updated_at": now_iso(),
                }},
            )

    if apply_it and scheduled:
        await log_activity(user, "samples_auto_scheduled", {
            "count": len(scheduled),
            "days_before": days_before,
            "target_per_day": target_per_day, "max_per_day": max_per_day,
            "over_cap": over_cap_count, "shifted": shifted_count,
            "filter": filter_mode, "skip_weekends": skip_weekends,
        })
        await _broadcast(scheduled[0]["sample_task_id"])

    scheduled.sort(key=lambda r: r["scheduled_date"])

    # Distribución final por día (incluye la carga existente para que el
    # usuario vea el total real, no solo lo que este batch agregó).
    all_days = set(existing_load) | set(new_load)
    by_day = []
    for d in sorted(all_days):
        by_day.append({
            "date": d,
            "existing": existing_load.get(d, 0),
            "added": new_load.get(d, 0),
            "total": existing_load.get(d, 0) + new_load.get(d, 0),
        })

    return {
        "applied": apply_it,
        "days_before": days_before,
        "target_per_day": target_per_day,
        "max_per_day": max_per_day,
        "skip_weekends": skip_weekends,
        "filter": filter_mode,
        "include_scheduled": include_scheduled,
        "total_backlog": len(tasks),
        "scheduled_count": len(scheduled),
        "skipped_count": len(skipped),
        "shifted_count": shifted_count,
        "over_cap_count": over_cap_count,
        "moved_count": moved_count,
        "kept_count": kept_count,
        "scheduled": scheduled,
        "skipped": skipped,
        "by_day": by_day,
    }


# ═══════════════ PROVISIONAL → REAL ═════════════════════════════════════════
async def _prov_match_candidates(order):
    """Dada una orden REAL recién creada, busca PROV- que probablemente sean
    la misma cosa (mismo style + mismo cliente + due_date cercano ±5 días).
    """
    style = (order.get("style") or "").strip().upper()
    client = (order.get("client") or "").strip().upper()
    if not style or not client:
        return []
    q = {"is_provisional": True, "style": style, "client": client,
         "board": {"$ne": "PAPELERA DE RECICLAJE"}}
    prov = await db.orders.find(q, _ORDER_FIELDS).to_list(20)
    due = (order.get("due_date") or "").strip()
    candidates = []
    for p in prov:
        score = 100
        pdue = (p.get("due_date") or "").strip()
        if due and pdue:
            try:
                d1 = datetime.strptime(due[:10], "%Y-%m-%d").date()
                d2 = datetime.strptime(pdue[:10], "%Y-%m-%d").date()
                delta = abs((d1 - d2).days)
                if delta > 5:
                    continue
                score -= delta * 5
            except ValueError:
                pass
        candidates.append({**p, "match_score": score})
    return sorted(candidates, key=lambda x: -x["match_score"])


@router.get("/prov-suggestions")
async def prov_suggestions(request: Request, order_id: str = ""):
    """Devuelve PROV- candidatos a fusionar con `order_id`. La UI muestra un
    banner "¿Esta orden real es este ejemplo PROV-?" con botón "Fusionar".
    """
    await require_auth(request)
    if not order_id:
        raise HTTPException(400, "order_id requerido")
    order = await db.orders.find_one({"order_id": order_id}, _ORDER_FIELDS)
    if not order:
        raise HTTPException(404, "Orden no encontrada")
    if order.get("is_provisional"):
        return {"suggestions": []}   # PROV- no sugiere PROV-
    cands = await _prov_match_candidates(order)
    return {"suggestions": cands}


@router.post("/tasks/{task_id}/promote")
async def promote(task_id: str, request: Request):
    """Fusiona un ejemplo PROV- con una orden real existente.

    Body: {"order_number": "12345"}  OR  {"order_id": "ord_..."}

    Efecto:
      1. Mueve el sample_task a apuntar a la orden real.
      2. Marca la orden real con sample=<estatus del task> si estaba vacío.
      3. Elimina la orden PROV- (o la manda a PAPELERA si preferimos auditar).
    """
    user = await require_sample_write(request)
    body = await request.json()
    real_num = (body.get("order_number") or "").strip()
    real_id = (body.get("order_id") or "").strip()

    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")
    if not task.get("is_provisional"):
        raise HTTPException(400, "El ejemplo no es provisional; nada que promover")

    real = None
    if real_id:
        real = await db.orders.find_one({"order_id": real_id}, {"_id": 0})
    if not real and real_num:
        real = await db.orders.find_one(
            {"order_number": {"$regex": f"^{re.escape(real_num)}$", "$options": "i"},
             "board": {"$ne": "PAPELERA DE RECICLAJE"}}, {"_id": 0})
    if not real:
        raise HTTPException(404, "Orden real no encontrada")
    if real.get("is_provisional"):
        raise HTTPException(400, "La orden destino también es PROV-, no se puede fusionar")

    prov_id = task["order_id"]
    # Se lee ANTES de archivarla: de la PROV- sale lo que se capturó durante la
    # vida del ejemplo y que la orden real puede no traer.
    prov_order = await db.orders.find_one({"order_id": prov_id}, _ORDER_FIELDS) or {}

    # 1) Repunta el task
    await db.sample_tasks.update_one(
        {"sample_task_id": task_id},
        {"$set": {
            "order_id": real["order_id"],
            "order_number": real["order_number"],
            "is_provisional": False,
            "updated_at": now_iso(),
        }})

    # 2) Lo trabajado mientras el ejemplo vivió SIN número de orden se vuelca a
    #    la orden real — pero sólo donde el CRM viene en blanco. Nunca pisa un
    #    dato que alguien ya capturó del lado del CRM: el ejemplo llena huecos,
    #    no corrige la orden.
    #    Los palomeos (arte/neck/trim/screens) NO se copian a
    #    art_sep_status/art_neck_status/trim_status/screens: son avance del
    #    EJEMPLO y contaminarían los KPIs del módulo de Arte, que miden trabajo
    #    de producción.
    llenar = {}
    if not (real.get("sample") or "").strip():
        aprobado = (task.get("approval") or APPROVAL_RECEIVED) == APPROVAL_APPROVED
        llenar["sample"] = "EJEMPLO APROBADO" if aprobado else (task.get("status") or "EJEMPLO PRIMERO")
    for campo in ("color", "client", "style"):
        if not (real.get(campo) or "").strip():
            valor = (prov_order or {}).get(campo)
            if valor:
                llenar[campo] = valor
    if llenar:
        await db.orders.update_one({"order_id": real["order_id"]}, {"$set": llenar})

    # 3) Manda la PROV- a la papelera (no borrar duro; deja rastro por auditoría)
    await db.orders.update_one(
        {"order_id": prov_id},
        {"$set": {"board": "PAPELERA DE RECICLAJE",
                  "promoted_to": real["order_id"],
                  "promoted_at": now_iso()}})

    await log_activity(user, "sample_promoted", {
        "sample_task_id": task_id,
        "from_prov": prov_id,
        "to_order_id": real["order_id"],
        "to_order_number": real["order_number"],
        "campos_llenados": list(llenar.keys()),
        "flags": task.get("flags") or {},
        "approval": task.get("approval") or APPROVAL_RECEIVED,
    })
    await _broadcast(real["order_id"])
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    await _enrich([task])
    return task


# ═══════════════ AVANCE (palomeos) Y APROBACIÓN ═════════════════════════════
@router.put("/tasks/{task_id}/flags")
async def set_flags(task_id: str, request: Request):
    """Palomea/despalomea arte, neck, trim o screens del ejemplo.

    Body: {"flags": {"arte": true, "neck": false}}  — parcial, se fusiona.

    Guarda además quién y cuándo palomeó cada uno: en el piso la pregunta que
    sigue a "¿ya está el arte?" siempre es "¿quién lo marcó?".
    """
    user = await require_sample_write(request)
    body = await request.json()
    entrantes = body.get("flags")
    if not isinstance(entrantes, dict) or not entrantes:
        raise HTTPException(400, "Body: {\"flags\": {\"arte\": true, ...}}")
    desconocidos = [k for k in entrantes if k not in SAMPLE_FLAGS]
    if desconocidos:
        raise HTTPException(400, f"Palomeo desconocido: {', '.join(desconocidos)}")

    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")

    flags = {f: bool((task.get("flags") or {}).get(f)) for f in SAMPLE_FLAGS}
    meta = dict(task.get("flags_meta") or {})
    ahora = now_iso()
    for k, v in entrantes.items():
        nuevo = bool(v)
        if flags[k] == nuevo:
            continue
        flags[k] = nuevo
        # Al despalomear se BORRA la firma: dejar el nombre de quien lo marcó en
        # un palomeo apagado haría creer que esa persona lo deshizo.
        meta[k] = {"by": user.get("user_id"), "by_name": user.get("name", ""), "at": ahora} if nuevo else None

    await db.sample_tasks.update_one(
        {"sample_task_id": task_id},
        {"$set": {"flags": flags, "flags_meta": meta, "updated_at": ahora}})
    await log_activity(user, "sample_flags", {
        "order_number": task.get("order_number"), "flags": entrantes})
    await _broadcast(task["order_id"])
    task.update({"flags": flags, "flags_meta": meta})
    await _enrich([task])
    return task


@router.put("/tasks/{task_id}/approval")
async def set_approval(task_id: str, request: Request):
    """Cambia la aprobación del ejemplo entre RECIBIDO y APROBADO.

    APROBADO lo manda a la pestaña Aprobados, tenga o no número de orden.
    """
    user = await require_sample_write(request)
    body = await request.json()
    valor = (body.get("approval") or "").strip().upper()
    if valor not in APPROVAL_VALUES:
        raise HTTPException(400, f"Aprobación inválida. Válidas: {', '.join(APPROVAL_VALUES)}")

    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")

    ahora = now_iso()
    upd = {"approval": valor, "updated_at": ahora}
    if valor == APPROVAL_APPROVED:
        upd.update({"approved_by": user.get("user_id"),
                    "approved_by_name": user.get("name", ""),
                    "approved_at": ahora})
        unset = {}
    else:
        # Regresar a RECIBIDO limpia la firma de aprobación, por lo mismo que
        # los palomeos: una firma vieja sobre un ejemplo no aprobado miente.
        unset = {"approved_by": "", "approved_by_name": "", "approved_at": ""}

    ops = {"$set": upd}
    if unset:
        ops["$unset"] = unset
    await db.sample_tasks.update_one({"sample_task_id": task_id}, ops)

    # Espeja al vocabulario legacy que lee el CRM en su columna `sample`.
    if valor == APPROVAL_APPROVED:
        await db.sample_tasks.update_one(
            {"sample_task_id": task_id}, {"$set": {"status": "EJEMPLO APROBADO"}})
        await db.orders.update_one(
            {"order_id": task["order_id"]}, {"$set": {"sample": "EJEMPLO APROBADO"}})

    await log_activity(user, "sample_approval", {
        "order_number": task.get("order_number"), "approval": valor})
    await _broadcast(task["order_id"])
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    await _enrich([task])
    return task


# ═══════════════ PESTAÑAS ═══════════════════════════════════════════════════
@router.get("/by-stage")
async def by_stage(request: Request, stage: str = STAGE_EJEMPLOS, limit: int = 500):
    """Ejemplos de una pestaña + el conteo de las tres.

    Los conteos van SIEMPRE los tres, no sólo el de la pestaña pedida: las
    cifras de las otras dos se ven en sus botones y quedarían rancias si sólo
    se refrescara la activa.
    """
    await require_auth(request)
    if stage not in STAGES:
        raise HTTPException(400, f"Pestaña inválida. Válidas: {', '.join(STAGES)}")
    limit = max(1, min(limit, 2000))

    # La etapa se deriva en Python (ver _stage_of): expresarla como query de
    # Mongo obligaría a repetir la misma regla en dos lenguajes y a que las dos
    # copias se mantuvieran iguales para siempre.
    todos = await db.sample_tasks.find({}, {"_id": 0}).sort("updated_at", -1).to_list(5000)
    for t in todos:
        t["approval"] = t.get("approval") or APPROVAL_RECEIVED
        t["stage"] = _stage_of(t)
    counts = {st: sum(1 for t in todos if t["stage"] == st) for st in STAGES}
    elegidos = [t for t in todos if t["stage"] == stage][:limit]
    await _enrich(elegidos)
    return {"stage": stage, "counts": counts, "total": counts[stage], "items": elegidos}


# ═══════════════ CAZA DE LA ORDEN REAL POR STYLE ════════════════════════════
@router.get("/tasks/{task_id}/order-matches")
async def order_matches(request: Request, task_id: str, limit: int = 20):
    """Órdenes reales que podrían ser este ejemplo, cazadas por STYLE.

    Es el camino inverso de /prov-suggestions (que parte de la orden y busca
    ejemplos): aquí se parte del ejemplo sin PO y se buscan sus órdenes, que es
    como se trabaja en el piso — el ejemplo existe primero y el número llega
    después.

    LA LLAVE ES EL NÚMERO DE DISEÑO, NO EL CAMPO `style`
    ────────────────────────────────────────────────────
    Medido sobre la base real (2026-08-18): los ejemplos guardan el número de
    diseño DENTRO de su campo `style` (WCC0023M5000), mientras que en las
    órdenes reales `style` viene basura (' ', '-') y el diseño vive en
    `design_#`, además con espacios al final. Cazar style contra style
    encontraba CERO de 36; cazar contra `design_#` encuentra 19.

    Por eso la comparación es contra `design_#` (y su typo heredado `desing_#`),
    insensible a mayúsculas y tolerante a los espacios sobrantes.

    El cliente y la cercanía de fechas sólo SUMAN puntos, no filtran: exigir que
    el cliente coincida deja fuera la orden buena cuando el ejemplo se capturó
    con el nombre corto del cliente.
    """
    await require_auth(request)
    task = await db.sample_tasks.find_one({"sample_task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Ejemplo no encontrado")

    prov = await db.orders.find_one(
        {"order_id": task["order_id"]},
        {**_ORDER_FIELDS, "design_#": 1, "desing_#": 1},
    ) or {}
    clave = (prov.get("design_#") or prov.get("desing_#") or prov.get("style") or "").strip().upper()
    if not clave:
        return {"design": "", "matches": [],
                "hint": "El ejemplo no tiene número de diseño; sin él no hay con qué cazar la orden."}

    # Los valores guardados traen espacios al final ('TS01340U1039 '), así que
    # la igualdad exacta falla; se compara con los espacios tolerados.
    rx = {"$regex": rf"^\s*{re.escape(clave)}\s*$", "$options": "i"}
    cliente = (prov.get("client") or "").strip().upper()
    reales = await db.orders.find(
        {"$or": [{"design_#": rx}, {"desing_#": rx}, {"style": rx}],
         "is_provisional": {"$ne": True},
         "order_id": {"$ne": task["order_id"]},
         "board": {"$nin": ["PAPELERA DE RECICLAJE", "EJEMPLOS"]}},
        {**_ORDER_FIELDS, "design_#": 1, "desing_#": 1},
    ).sort("created_at", -1).to_list(200)

    # Una orden que ya es el ejemplo de alguien más no se vuelve a ofrecer.
    tomadas = set()
    async for t in db.sample_tasks.find(
        {"order_id": {"$in": [o["order_id"] for o in reales]}}, {"_id": 0, "order_id": 1}
    ):
        tomadas.add(t["order_id"])

    prov_due = (prov.get("due_date") or prov.get("cancel_date") or "")[:10]
    out = []
    for o in reales:
        if o["order_id"] in tomadas:
            continue
        score = 50
        razones = ["mismo número de diseño"]
        if cliente and (o.get("client") or "").strip().upper() == cliente:
            score += 30
            razones.append("mismo cliente")
        if (prov.get("color") or "").strip().upper() and \
           (o.get("color") or "").strip().upper() == (prov.get("color") or "").strip().upper():
            score += 10
            razones.append("mismo color")
        o_due = (o.get("due_date") or o.get("cancel_date") or "")[:10]
        if prov_due and o_due:
            try:
                d1 = datetime.strptime(prov_due, "%Y-%m-%d").date()
                d2 = datetime.strptime(o_due, "%Y-%m-%d").date()
                dias = abs((d1 - d2).days)
                if dias <= 30:
                    score += max(0, 20 - dias // 2)
                    razones.append(f"fechas a {dias} día(s)")
            except ValueError:
                pass
        out.append({**o, "match_score": score, "match_reasons": razones})

    out.sort(key=lambda x: -x["match_score"])
    return {"design": clave, "matches": out[:limit]}


# ═══════════════ EXPORT EXCEL ═══════════════════════════════════════════════
@router.get("/export.xlsx")
async def export_xlsx(request: Request,
                      scope: str = "all",
                      filter: str = "all",
                      start: str = "",
                      end: str = ""):
    """Genera un XLSX con los ejemplos según:
      - scope: all | scheduled | backlog
      - filter: all | with_style | no_style
      - start, end: (opcional) rango de scheduled_date para acotar
    """
    user = await require_auth(request)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = {}
    if scope == "scheduled":
        q["scheduled_date"] = {"$ne": None, "$exists": True}
    elif scope == "backlog":
        q["$or"] = [{"scheduled_date": None}, {"scheduled_date": {"$exists": False}}]

    if start and end:
        q.setdefault("scheduled_date", {})
        if isinstance(q["scheduled_date"], dict):
            q["scheduled_date"].update({"$gte": start, "$lte": end})

    tasks = await db.sample_tasks.find(q, {"_id": 0}).to_list(10000)
    await _enrich(tasks)

    # Filtro post-enrich (necesita el order.style para segregar)
    if filter == "with_style":
        tasks = [t for t in tasks if (t.get("order") or {}).get("style")]
    elif filter == "no_style":
        tasks = [t for t in tasks if not (t.get("order") or {}).get("style")]

    tasks.sort(key=lambda t: (t.get("scheduled_date") or "9999-12-31", t.get("order_number") or ""))

    # ── Construcción del workbook ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Ejemplos"

    header = [
        "Orden", "PROV?", "Cliente", "Brand", "Style", "Color", "Cantidad",
        "Prioridad", "Cancel date", "Programada", "Ideal (-14d)", "Holgura (d)",
        "Estatus", "Operador", "HOT", "Vencido",
        "Playera OK", "Arte OK", "Neck OK", "Screens OK", "Tinta OK",
        "Notas",
    ]
    ws.append(header)

    # Estilos header
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c in ws[1]:
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
    ws.row_dimensions[1].height = 32

    # Colores de zebra + estatus
    hot_fill     = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # rojo suave
    overdue_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")   # naranja
    zebra_fill   = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")   # gris muy suave

    today = datetime.now(timezone.utc).date()
    for idx, t in enumerate(tasks):
        o = t.get("order") or {}
        r = t.get("resources") or {}
        cancel = (o.get("cancel_date") or "")[:10]
        sched  = (t.get("scheduled_date") or "")
        try:
            cd = datetime.strptime(cancel, "%Y-%m-%d").date() if cancel else None
        except ValueError:
            cd = None
        ideal = (cd - timedelta(days=14)).isoformat() if cd else ""
        try:
            sd = datetime.strptime(sched[:10], "%Y-%m-%d").date() if sched else None
        except ValueError:
            sd = None
        holgura = (cd - sd).days if (cd and sd) else ""

        def _yn(v):
            if v is True: return "SI"
            if v is False: return "NO"
            return "N/A"

        row = [
            t.get("order_number") or "",
            "PROV" if t.get("is_provisional") else "",
            o.get("client") or "",
            o.get("branding") or "",
            o.get("style") or "",
            o.get("color") or "",
            o.get("quantity") or "",
            o.get("priority") or "",
            cancel,
            sched or "",
            ideal,
            holgura,
            t.get("status") or "",
            t.get("operator") or "",
            "SI" if t.get("is_hot") else "",
            "SI" if t.get("overdue") else "",
            _yn(r.get("blank_ok")),
            _yn(r.get("art_ok")),
            _yn(r.get("neck_ok")),
            _yn(r.get("screens_ok")),
            _yn(r.get("paint_ok")),
            t.get("notes") or "",
        ]
        ws.append(row)

        excel_row = ws[ws.max_row]
        fill = None
        if t.get("is_hot"):    fill = hot_fill
        elif t.get("overdue"): fill = overdue_fill
        elif idx % 2 == 1:     fill = zebra_fill
        for c in excel_row:
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=False)
            if fill: c.fill = fill

    # Ancho de columnas
    widths = [10, 6, 22, 18, 14, 12, 9, 12, 12, 12, 12, 9, 20, 16, 6, 8, 10, 8, 8, 10, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze headers + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    # ── Hoja resumen por día ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Fecha", "Ejemplos", "HOT", "Vencidos"])
    for c in ws2[1]:
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = border
    day_map = {}
    for t in tasks:
        d = t.get("scheduled_date") or "SIN PROGRAMAR"
        day_map.setdefault(d, {"total": 0, "hot": 0, "overdue": 0})
        day_map[d]["total"] += 1
        if t.get("is_hot"): day_map[d]["hot"] += 1
        if t.get("overdue"): day_map[d]["overdue"] += 1
    for d in sorted(day_map.keys()):
        m = day_map[d]
        ws2.append([d, m["total"], m["hot"], m["overdue"]])
        for c in ws2[ws2.max_row]:
            c.border = border
    for i, w in enumerate([16, 12, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Serialización ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await log_activity(user, "samples_exported", {
        "count": len(tasks), "scope": scope, "filter": filter,
    })

    filename = f"ejemplos_{datetime.now(timezone.utc).strftime('%Y-%m-%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


